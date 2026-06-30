"""
Corporate Tools Suite - Flask Backend
Main application entry point with all routes.
"""

import os
import io
import json
import uuid
import time
import base64
import zipfile
import shutil
import threading
import traceback
import sys
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for, flash, session, send_from_directory, abort
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from flask_cors import CORS
import sqlite3
import functools

def get_db():
    db_path = os.path.join(BASE_DIR, 'users.db')
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            email TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            plan TEXT DEFAULT 'Free',
            is_active INTEGER DEFAULT 1,
            role TEXT DEFAULT 'user',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    conn.commit()
    conn.close()

# Import tool modules
from tools.file_manager import FileManager
from tools.pdf_tools import PDFTools
from tools.excel_tools import ExcelTools
from tools.gst_tools import GSTTools

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'corporate-tools-suite-dev-key-change-in-prod')
CORS(app)

# Configuration
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
OUTPUT_FOLDER = os.path.join(BASE_DIR, 'outputs')
MAX_CONTENT_LENGTH = 100 * 1024 * 1024  # 100MB

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['OUTPUT_FOLDER'] = OUTPUT_FOLDER
app.config['MAX_CONTENT_LENGTH'] = MAX_CONTENT_LENGTH

ALLOWED_EXTENSIONS = {
    'pdf': ['pdf'],
    'excel': ['xlsx', 'xls', 'csv'],
    'any': ['pdf', 'xlsx', 'xls', 'csv', 'txt', 'doc', 'docx', 'jpg', 'jpeg', 'png', 'zip']
}

# Ensure folders exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# Initialize DB
init_db()

# --- Decorators ---
def login_required(f):
    @functools.wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @functools.wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login_page'))
        if session.get('role') != 'admin':
            abort(403)
        return f(*args, **kwargs)
    return decorated_function

# Global dictionary to track background tasks
TASKS = {}

import time

def process_reco(task_id, pr_path, b2_path, out_path, client_name, cfg, email_cfg):
    try:
        from tools.reco_engine import GSTRecoEngine
        
        start_time = time.time()
        
        def on_progress(msg, pct):
            elapsed = time.time() - start_time
            eta = -1
            if pct > 0:
                total_est = elapsed / (pct / 100.0)
                eta = max(0, total_est - elapsed)
            
            TASKS[task_id]["status"] = "processing"
            TASKS[task_id]["message"] = msg
            TASKS[task_id]["percent"] = pct
            TASKS[task_id]["eta"] = int(eta)
            
        engine = GSTRecoEngine(cfg)
        engine.run(pr_path, b2_path, out_path, client_name=client_name, email_cfg=email_cfg, on_progress=on_progress)
        TASKS[task_id] = {"status": "done", "out_path": out_path}
    except Exception as e:
        import traceback
        traceback.print_exc()
        TASKS[task_id] = {"status": "error", "message": str(e)}

# Initialize tool handlers
file_mgr = FileManager(UPLOAD_FOLDER, OUTPUT_FOLDER)
pdf_tools = PDFTools(UPLOAD_FOLDER, OUTPUT_FOLDER)
excel_tools = ExcelTools(UPLOAD_FOLDER, OUTPUT_FOLDER)
gst_tools = GSTTools(UPLOAD_FOLDER, OUTPUT_FOLDER)

# ── Scan-to-PDF session store (in-memory) ────────────────
scan_sessions = {}  # {session_id: {images:[], created:float, pdf:None}}


def allowed_file(filename, file_type='any'):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS[file_type]


def save_uploaded_files(files, sub_folder=''):
    """Save uploaded files and return list of saved paths."""
    saved = []
    folder = os.path.join(UPLOAD_FOLDER, sub_folder) if sub_folder else UPLOAD_FOLDER
    os.makedirs(folder, exist_ok=True)
    for f in files:
        if f and f.filename:
            filename = secure_filename(f.filename)
            ts = datetime.now().strftime('%Y%m%d_%H%M%S_')
            filepath = os.path.join(folder, ts + filename)
            f.save(filepath)
            saved.append(filepath)
    return saved


# ─────────────────────────────────────────
# HOME / DASHBOARD
# ─────────────────────────────────────────
@app.route('/')
def index():
    return render_template('index.html')


@app.route('/laws')
def laws_page():
    return render_template('laws.html')


@app.route('/reco')
def reco_page():
    return render_template('reco_tool.html')


@app.route('/login')
def login_page():
    if 'user_id' in session:
        return redirect(url_for('dashboard_page'))
    return render_template('login.html')


@app.route('/advisory')
def advisory_page():
    return render_template('advisory.html')


@app.route('/dashboard')
@login_required
def dashboard_page():
    # Fetch latest user info
    conn = get_db()
    user = conn.execute('SELECT * FROM users WHERE id = ?', (session['user_id'],)).fetchone()
    conn.close()
    if not user:
        session.clear()
        return redirect(url_for('login_page'))
    return render_template('dashboard.html', user=dict(user))

@app.route('/admin')
@admin_required
def admin_page():
    return render_template('admin.html')

# ─────────────────────────────────────────
# AUTHENTICATION & ADMIN APIs
# ─────────────────────────────────────────
@app.route('/api/auth/register', methods=['POST'])
def api_register():
    data = request.json
    name = data.get('name', '').strip()
    email = data.get('email', '').strip()
    password = data.get('password', '')

    if not name or not email or not password:
        return jsonify({'success': False, 'error': 'All fields are required'})

    try:
        conn = get_db()
        c = conn.cursor()
        
        # Check if first user -> make them admin
        count = c.execute('SELECT COUNT(*) FROM users').fetchone()[0]
        role = 'admin' if count == 0 or email.lower() == 'youradvisor.ca@gmail.com' else 'user'
        
        pw_hash = generate_password_hash(password)
        c.execute('INSERT INTO users (name, email, password_hash, role) VALUES (?, ?, ?, ?)',
                  (name, email, pw_hash, role))
        conn.commit()
        
        user_id = c.lastrowid
        session['user_id'] = user_id
        session['name'] = name
        session['role'] = role
        session['plan'] = 'Free'
        
        return jsonify({'success': True})
    except sqlite3.IntegrityError:
        return jsonify({'success': False, 'error': 'Email already registered'})
    finally:
        conn.close()


@app.route('/api/auth/login', methods=['POST'])
def api_login():
    data = request.json
    email = data.get('email', '').strip()
    password = data.get('password', '')

    if not email or not password:
        return jsonify({'success': False, 'error': 'Email and password required'})

    conn = get_db()
    user = conn.execute('SELECT * FROM users WHERE email = ?', (email,)).fetchone()
    conn.close()

    if user and check_password_hash(user['password_hash'], password):
        if user['is_active'] == 0:
            return jsonify({'success': False, 'error': 'Account is blocked. Contact support.'})
            
        session['user_id'] = user['id']
        session['name'] = user['name']
        session['role'] = user['role']
        session['plan'] = user['plan']
        return jsonify({'success': True, 'role': user['role']})
    
    return jsonify({'success': False, 'error': 'Invalid credentials'})


@app.route('/api/auth/logout', methods=['POST'])
def api_logout():
    session.clear()
    return jsonify({'success': True})


@app.route('/api/admin/users', methods=['GET'])
@admin_required
def api_admin_users():
    conn = get_db()
    users = conn.execute('SELECT id, name, email, plan, is_active, role, created_at FROM users ORDER BY id DESC').fetchall()
    conn.close()
    return jsonify({'success': True, 'users': [dict(u) for u in users]})


@app.route('/api/admin/users/<int:user_id>/plan', methods=['POST'])
@admin_required
def api_admin_update_plan(user_id):
    plan = request.json.get('plan')
    conn = get_db()
    conn.execute('UPDATE users SET plan = ? WHERE id = ?', (plan, user_id))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/admin/users/<int:user_id>/status', methods=['POST'])
@admin_required
def api_admin_update_status(user_id):
    status = request.json.get('status')
    val = 1 if status == 'active' else 0
    conn = get_db()
    conn.execute('UPDATE users SET is_active = ? WHERE id = ?', (val, user_id))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/robots.txt')
def robots():
    return app.send_static_file('robots.txt')


@app.route('/sitemap.xml')
def sitemap():
    return app.send_static_file('sitemap.xml')


# ─────────────────────────────────────────
# FILE MANAGER
# ─────────────────────────────────────────
@app.route('/file-manager')
@login_required
def file_manager_page():
    return render_template('file_manager.html')


# ── New clean API routes (used by upgraded UI) ──────────────────────────

@app.route('/api/reco/run', methods=['POST'])
def api_reco_run():
    if 'pr_file' not in request.files and 'b2_file' not in request.files:
        return jsonify({'success': False, 'error': 'At least one file (PR or 2B) is required'}), 400
    
    pr_file = request.files.get('pr_file')
    b2_file = request.files.get('b2_file')
    
    if (not pr_file or not pr_file.filename) and (not b2_file or not b2_file.filename):
        return jsonify({'success': False, 'error': 'At least one valid file (PR or 2B) is required'}), 400
    
    import tempfile
    pr_path = os.path.join(tempfile.gettempdir(), 'pr_' + secure_filename(pr_file.filename)) if pr_file and pr_file.filename else None
    b2_path = os.path.join(tempfile.gettempdir(), '2b_' + secure_filename(b2_file.filename)) if b2_file and b2_file.filename else None
    out_path = os.path.join(tempfile.gettempdir(), f"GST_Reco_Output_{int(time.time())}.xlsx")
    
    if pr_path: pr_file.save(pr_path)
    if b2_path: b2_file.save(b2_path)
    
    try:
        cfg = {
            "amount_tolerance": float(request.form.get("amount_tolerance", 10)),
            "pct_tolerance": float(request.form.get("pct_tolerance", 5)),
            "bank_max_match_exhaust_side": request.form.get("bank_max_match_exhaust_side", "2B"),
            "bank_max_match_strategy": request.form.get("bank_max_match_strategy", "ALL_VS_ALL"),
            "bank_match_relax_head": request.form.get("bank_match_relax_head", "true") == "true",
            "bank_max_match_variance_pct": float(request.form.get("bank_max_match_variance_pct", 10000)),
            "features": {
                "user_knockoff": True,
                "pre_group": request.form.get("pre_group", "false") == "true",
                "pr_knockout": request.form.get("pr_knockout", "false") == "true",
                "grouped_match": request.form.get("grouped_match", "false") == "true",
                "rule_match": request.form.get("rule_match", "false") == "true",
                "fuzzy_invoice": request.form.get("fuzzy_invoice", "false") == "true",
                "vendor_fuzzy": request.form.get("vendor_fuzzy", "false") == "true",
                "month_year_match": request.form.get("month_year_match", "false") == "true",
                "bank_match": request.form.get("bank_match", "false") == "true",
                "bank_max_match": request.form.get("bank_max_match", "false") == "true",
                "2b_knockout": request.form.get("2b_knockout", "false") == "true"
            }
        }
        
        if request.form.get("match_rules"):
            try: cfg["match_rules"] = json.loads(request.form.get("match_rules"))
            except: pass
        if request.form.get("pr_ko_rules"):
            try: cfg["pr_ko_rules"] = json.loads(request.form.get("pr_ko_rules"))
            except: pass
        if request.form.get("b2_ko_rules"):
            try: cfg["b2_ko_rules"] = json.loads(request.form.get("b2_ko_rules"))
            except: pass

        
        email_cfg = None
        if request.form.get("email_enabled", "false") == "true":
            email_cfg = {
                "to": request.form.get("email_to", ""),
                "host": request.form.get("smtp_host", "smtp.gmail.com"),
                "port": int(request.form.get("smtp_port", 587)),
                "user": request.form.get("smtp_user", ""),
                "pass": request.form.get("smtp_pass", "")
            }

        client_name = request.form.get("client_name", "")
        task_id = str(uuid.uuid4())
        
        TASKS[task_id] = {"status": "processing"}
        
        thread = threading.Thread(target=process_reco, args=(task_id, pr_path, b2_path, out_path, client_name, cfg, email_cfg))
        thread.start()
        
        return jsonify({'success': True, 'task_id': task_id})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/reco/status/<task_id>', methods=['GET'])
def api_reco_status(task_id):
    if task_id not in TASKS:
        return jsonify({"status": "error", "message": "Server restarted. Please refresh the page and run again."}), 404
    return jsonify(TASKS[task_id])

@app.route('/api/reco/download/<task_id>', methods=['GET'])
def api_reco_download(task_id):
    if task_id not in TASKS or TASKS[task_id]["status"] != "done":
        abort(404)
    out_path = TASKS[task_id]["out_path"]
    return send_file(out_path, as_attachment=True, download_name='GST_Reco_Output.xlsx')

@app.route('/api/reco/template', methods=['GET'])
def api_reco_template():
    import pandas as pd
    import io
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    
    file_type = request.args.get('type', 'pr').lower()
    if file_type not in ['pr', '2b']:
        file_type = 'pr'
        
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df = pd.DataFrame(columns=[
            "Serial No", "GSTIN", "Pan", "Vendor Name", "Invoice Number", 
            "Invoice Date", "Taxable Value", "IGST", "CGST", "SGST", 
            "Compcess Value", "Total Tax", "Invoice Value", "Knock Off", 
            "State", "Remarks", "Is Bank"
        ])
        
        sheet_name = 'Purchase_Register' if file_type == 'pr' else 'GSTR_2B'
        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)
        
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        
        # Row 1: MYBBT Branding
        worksheet.merge_cells('A1:Q1')
        cell = worksheet.cell(row=1, column=1)
        cell.value = "MYBBT Business Bluetooth — Professional CA Firm Intelligence Engine"
        cell.font = Font(name='Segoe UI', size=16, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        worksheet.row_dimensions[1].height = 30
        
        # Row 2: Template Type
        worksheet.merge_cells('A2:Q2')
        cell2 = worksheet.cell(row=2, column=1)
        cell2.value = f"DATA TEMPLATE : {'PURCHASE REGISTER (PR)' if file_type == 'pr' else 'GSTR-2B'}"
        cell2.font = Font(name='Segoe UI', size=12, bold=True, color="333333")
        cell2.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        cell2.alignment = Alignment(horizontal="center", vertical="center")
        worksheet.row_dimensions[2].height = 20
        
        # Row 3: Headers
        header_font = Font(name='Segoe UI', bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for col_num, value in enumerate(df.columns.values, 1):
            c = worksheet.cell(row=3, column=col_num)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin_border
            
            # Auto-adjust column width
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = max(len(str(value)) + 2, 12)
            
    output.seek(0)
    filename = 'MYBBT_PR_Template.xlsx' if file_type == 'pr' else 'MYBBT_2B_Template.xlsx'
    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')




@app.route('/api/files/upload', methods=['POST'])
def files_upload():
    if 'files' not in request.files:
        return jsonify({'success': False, 'error': 'No files provided'}), 400
    files = request.files.getlist('files')
    saved = save_uploaded_files(files, 'file_manager')
    names = [os.path.basename(fp) for fp in saved]
    return jsonify({'success': True, 'uploaded': len(saved), 'files': names})


@app.route('/api/files/list')
def files_list():
    files = file_mgr.list_files()
    return jsonify({'success': True, 'files': files})


@app.route('/api/files/download/<filename>')
def files_download_single(filename):
    try:
        filepath = file_mgr.get_file_path(filename)
        return send_file(filepath, as_attachment=True)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 404


@app.route('/api/files/download-zip', methods=['POST'])
def files_download_zip():
    try:
        zip_path = file_mgr.create_zip()
        return send_file(zip_path, as_attachment=True, download_name='collected_files.zip')
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/files/delete/<filename>', methods=['DELETE'])
def files_delete(filename):
    result = file_mgr.delete_file(filename)
    return jsonify(result)


@app.route('/api/files/clear', methods=['DELETE'])
def files_clear_all():
    try:
        count = 0
        for f in file_mgr.list_files():
            file_mgr.delete_file(f['name'])
            count += 1
        return jsonify({'success': True, 'deleted': count})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/files/import-path', methods=['POST'])
def files_import_path():
    """
    Import files directly from a local folder path on the server machine.
    Works on localhost — the server and client are the same machine.
    """
    import json as _json
    data = request.get_json(silent=True) or {}
    folder_path = data.get('path', '').strip()
    recursive   = data.get('recursive', False)
    extensions  = [e.lower().lstrip('.') for e in (data.get('extensions') or [])]

    if not folder_path:
        return jsonify({'success': False, 'error': 'Folder path required'}), 400
    if not os.path.exists(folder_path):
        return jsonify({'success': False, 'error': f'Path not found: {folder_path}'}), 400
    if not os.path.isdir(folder_path):
        return jsonify({'success': False, 'error': f'Not a folder: {folder_path}'}), 400

    dest_folder = os.path.join(UPLOAD_FOLDER, 'file_manager')
    os.makedirs(dest_folder, exist_ok=True)

    imported = []
    skipped  = 0

    def process_dir(dirpath):
        nonlocal skipped
        try:
            entries = os.listdir(dirpath)
        except PermissionError:
            return
        for name in entries:
            full = os.path.join(dirpath, name)
            if os.path.isfile(full):
                ext = name.rsplit('.', 1)[-1].lower() if '.' in name else ''
                if extensions and ext not in extensions:
                    skipped += 1
                    continue
                # Safe copy with unique name if conflict
                dest_name = secure_filename(name)
                dest_path = os.path.join(dest_folder, dest_name)
                if os.path.abspath(full) == os.path.abspath(dest_path):
                    continue  # same file, skip
                # Avoid overwriting — add suffix
                base, dot_ext = (dest_name.rsplit('.', 1) if '.' in dest_name else (dest_name, ''))
                counter = 1
                while os.path.exists(dest_path):
                    dest_name = f"{base}_{counter}.{dot_ext}" if dot_ext else f"{base}_{counter}"
                    dest_path = os.path.join(dest_folder, dest_name)
                    counter += 1
                shutil.copy2(full, dest_path)
                imported.append(dest_name)
            elif os.path.isdir(full) and recursive:
                process_dir(full)

    process_dir(folder_path)

    return jsonify({
        'success': True,
        'imported': len(imported),
        'skipped': skipped,
        'files': imported
    })




# ── Old routes (kept for backward compat) ─────────────────────────────────
@app.route('/api/file-manager/upload', methods=['POST'])
def upload_files():
    return files_upload()

@app.route('/api/file-manager/list')
def list_files():
    return files_list()

@app.route('/api/file-manager/download-all')
def download_all_files():
    try:
        zip_path = file_mgr.create_zip()
        return send_file(zip_path, as_attachment=True, download_name='collected_files.zip')
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/file-manager/delete/<filename>', methods=['DELETE'])
def delete_file(filename):
    return files_delete(filename)

@app.route('/api/file-manager/download/<filename>')
def download_single_file(filename):
    return files_download_single(filename)




# ─────────────────────────────────────────
# PDF TOOLS
# ─────────────────────────────────────────
@app.route('/pdf-tools')
@login_required
def pdf_tools_page():
    return render_template('pdf_tools.html')


def _get_pdf(field='file'):
    f = request.files.get(field)
    if not f or not f.filename:
        return None, jsonify({'success': False, 'error': 'No PDF file provided'}), 400
    if not allowed_file(f.filename, 'pdf'):
        return None, jsonify({'success': False, 'error': 'Only PDF files allowed'}), 400
    saved = save_uploaded_files([f], 'pdf_ops')
    if not saved:
        return None, jsonify({'success': False, 'error': 'Failed to save file'}), 500
    return saved[0], None, None


@app.route('/api/pdf/info', methods=['POST'])
def pdf_info():
    path, err, code = _get_pdf()
    if err: return err, code
    return jsonify(pdf_tools.get_info(path))


@app.route('/api/pdf/compress', methods=['POST'])
def compress_pdf():
    path, err, code = _get_pdf()
    if err: return err, code
    quality = request.form.get('quality', 'medium')
    return jsonify(pdf_tools.compress(path, quality=quality))


@app.route('/api/pdf/merge', methods=['POST'])
def merge_pdfs():
    files = request.files.getlist('files')
    pdfs = [f for f in files if f and allowed_file(f.filename, 'pdf')]
    if len(pdfs) < 2:
        return jsonify({'success': False, 'error': 'Please upload at least 2 PDF files'}), 400
    saved = save_uploaded_files(pdfs, 'pdf_merge')
    return jsonify(pdf_tools.merge(saved))


@app.route('/api/pdf/split', methods=['POST'])
def split_pdf():
    path, err, code = _get_pdf()
    if err: return err, code
    mode      = request.form.get('mode', 'every')
    every_n   = int(request.form.get('every_n', 1))
    ranges_str = request.form.get('ranges', '')
    return jsonify(pdf_tools.split(path, mode=mode, every_n=every_n, ranges_str=ranges_str))


@app.route('/api/pdf/rotate', methods=['POST'])
def rotate_pdf():
    path, err, code = _get_pdf()
    if err: return err, code
    angle = int(request.form.get('angle', 90))
    pages = request.form.get('pages', 'all')
    return jsonify(pdf_tools.rotate(path, angle=angle, pages=pages))


@app.route('/api/pdf/to-images', methods=['POST'])
def pdf_to_images():
    path, err, code = _get_pdf()
    if err: return err, code
    fmt   = request.form.get('format', 'jpg')
    dpi   = int(request.form.get('dpi', 150))
    pages = request.form.get('pages', 'all')
    return jsonify(pdf_tools.pdf_to_images(path, fmt=fmt, dpi=dpi, pages=pages))


@app.route('/api/pdf/images-to-pdf', methods=['POST'])
def images_to_pdf():
    files = request.files.getlist('files')
    imgs  = [f for f in files if f and f.filename]
    if not imgs:
        return jsonify({'success': False, 'error': 'No images uploaded'}), 400
    saved = save_uploaded_files(imgs, 'pdf_imgs')
    return jsonify(pdf_tools.images_to_pdf(saved))


@app.route('/api/pdf/page-numbers', methods=['POST'])
def pdf_page_numbers():
    path, err, code = _get_pdf()
    if err: return err, code
    position   = request.form.get('position', 'bottom-center')
    start_from = int(request.form.get('start_from', 1))
    font_size  = int(request.form.get('font_size', 10))
    prefix     = request.form.get('prefix', '')
    suffix     = request.form.get('suffix', '')
    return jsonify(pdf_tools.add_page_numbers(path, position=position,
                   start_from=start_from, font_size=font_size,
                   prefix=prefix, suffix=suffix))


@app.route('/api/pdf/protect', methods=['POST'])
def protect_pdf():
    path, err, code = _get_pdf()
    if err: return err, code
    password = request.form.get('password', '')
    if not password:
        return jsonify({'success': False, 'error': 'Password required'}), 400
    allow_print = request.form.get('allow_print', 'true') == 'true'
    allow_copy  = request.form.get('allow_copy', 'false') == 'true'
    return jsonify(pdf_tools.protect(path, password=password,
                   allow_print=allow_print, allow_copy=allow_copy))


@app.route('/api/pdf/unlock', methods=['POST'])
def unlock_pdf():
    path, err, code = _get_pdf()
    if err: return err, code
    password = request.form.get('password', '')
    return jsonify(pdf_tools.unlock(path, password=password))


@app.route('/api/pdf/extract', methods=['POST'])
def extract_pages():
    path, err, code = _get_pdf()
    if err: return err, code
    pages = request.form.get('pages', '')
    if not pages:
        return jsonify({'success': False, 'error': 'Specify pages e.g. 1,3,5-8'}), 400
    return jsonify(pdf_tools.extract_pages(path, pages))


@app.route('/api/pdf/organize', methods=['POST'])
def organize_pdf():
    path, err, code = _get_pdf()
    if err: return err, code
    order = request.form.get('order', '')
    if not order:
        return jsonify({'success': False, 'error': 'Specify page order e.g. 3,1,2,4'}), 400
    return jsonify(pdf_tools.organize(path, order))


@app.route('/api/pdf/watermark', methods=['POST'])
def watermark_pdf():
    path, err, code = _get_pdf()
    if err: return err, code
    text     = request.form.get('text', 'CONFIDENTIAL')
    opacity  = float(request.form.get('opacity', 0.3))
    angle    = int(request.form.get('angle', 45))
    size     = int(request.form.get('font_size', 48))
    color    = request.form.get('color', 'gray')
    pages    = request.form.get('pages', 'all')
    return jsonify(pdf_tools.watermark(path, text=text, opacity=opacity,
                   angle=angle, font_size=size, color=color, pages=pages))


@app.route('/api/pdf/extract-text', methods=['POST'])
def pdf_extract_text():
    path, err, code = _get_pdf()
    if err: return err, code
    fmt   = request.form.get('format', 'txt')
    pages = request.form.get('pages', 'all')
    return jsonify(pdf_tools.extract_text(path, output_format=fmt, pages=pages))


@app.route('/api/pdf/repair', methods=['POST'])
def repair_pdf():
    path, err, code = _get_pdf()
    if err: return err, code
    return jsonify(pdf_tools.repair(path))


@app.route('/api/pdf/flatten', methods=['POST'])
def flatten_pdf():
    path, err, code = _get_pdf()
    if err: return err, code
    return jsonify(pdf_tools.flatten(path))


@app.route('/api/pdf/to-word', methods=['POST'])
def pdf_to_word():
    path, err, code = _get_pdf()
    if err: return err, code
    return jsonify(pdf_tools.pdf_to_word(path))


@app.route('/api/pdf/to-excel', methods=['POST'])
def pdf_to_excel():
    path, err, code = _get_pdf()
    if err: return err, code
    return jsonify(pdf_tools.pdf_to_excel(path))


@app.route('/api/pdf/office-to-pdf', methods=['POST'])
def office_to_pdf():
    f = request.files.get('file')
    if not f or not f.filename:
        return jsonify({'success': False, 'error': 'No file provided'}), 400
    allowed_exts = {'doc','docx','xls','xlsx','ppt','pptx','rtf','odt','odp','ods'}
    ext = f.filename.rsplit('.', 1)[-1].lower()
    if ext not in allowed_exts:
        return jsonify({'success': False, 'error': f'Unsupported: .{ext}. Use Word/Excel/PPT files'}), 400
    saved = save_uploaded_files([f], 'pdf_office')
    if not saved:
        return jsonify({'success': False, 'error': 'Upload failed'}), 500
    return jsonify(pdf_tools.office_to_pdf(saved[0]))


@app.route('/api/pdf/html-to-pdf', methods=['POST'])
def html_to_pdf():
    mode = request.form.get('mode', 'text')  # 'text', 'file', 'url'
    if mode == 'url':
        url = request.form.get('url', '')
        if not url:
            return jsonify({'success': False, 'error': 'URL required'}), 400
        return jsonify(pdf_tools.html_to_pdf(url=url))
    elif mode == 'file':
        f = request.files.get('file')
        if not f:
            return jsonify({'success': False, 'error': 'HTML file required'}), 400
        saved = save_uploaded_files([f], 'pdf_html')
        return jsonify(pdf_tools.html_to_pdf(html_path=saved[0]))
    else:
        html = request.form.get('html', '')
        if not html:
            return jsonify({'success': False, 'error': 'HTML content required'}), 400
        return jsonify(pdf_tools.html_to_pdf(html_content=html))


@app.route('/api/pdf/download/<filename>')
def download_pdf(filename):
    filepath = os.path.join(OUTPUT_FOLDER, filename)
    if os.path.exists(filepath):
        return send_file(filepath, as_attachment=True)
    return jsonify({'error': 'File not found'}), 404


# ─── Universal download route (used by all tools) ────────
@app.route('/download/<filename>')
def download_file(filename):
    """Universal download — checks outputs folder (and uploads as fallback)."""
    # Security: no path traversal
    filename = os.path.basename(filename)
    # Check output folder first
    filepath = os.path.join(OUTPUT_FOLDER, filename)
    if os.path.exists(filepath):
        return send_file(filepath, as_attachment=True, download_name=filename)
    # Fallback: uploads folder
    filepath2 = os.path.join(UPLOAD_FOLDER, filename)
    if os.path.exists(filepath2):
        return send_file(filepath2, as_attachment=True, download_name=filename)
    return jsonify({'error': 'File not found'}), 404




# ─────────────────────────────────────────
# EXCEL TOOLS
# ─────────────────────────────────────────
@app.route('/excel-tools')
@login_required
def excel_tools_page():
    return render_template('excel_tools.html')


@app.route('/api/excel/consolidate', methods=['POST'])
def consolidate_excel():
    if 'files' not in request.files:
        return jsonify({'success': False, 'error': 'No files provided'}), 400
    files = request.files.getlist('files')
    excels = [f for f in files if allowed_file(f.filename, 'excel')]
    if not excels:
        return jsonify({'success': False, 'error': 'Please upload Excel/CSV files'}), 400
    mode = request.form.get('mode', 'append')  # append or separate_sheets
    saved = save_uploaded_files(excels, 'excel_consolidate')
    try:
        result = excel_tools.consolidate(saved, mode=mode)
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/excel/download/<filename>')
def download_excel(filename):
    filepath = os.path.join(OUTPUT_FOLDER, filename)
    if os.path.exists(filepath):
        return send_file(filepath, as_attachment=True)
    return jsonify({'error': 'File not found'}), 404


# ─────────────────────────────────────────
# GST TOOLS
# ─────────────────────────────────────────
@app.route('/gst-tools')
@login_required
def gst_tools_page():
    return render_template('gst_tools.html')


def run_gst_task(task_id, func, *args):
    try:
        result = func(*args)
        if isinstance(result, dict) and result.get('success'):
            TASKS[task_id] = {
                "status": "done",
                "out_path": os.path.join(OUTPUT_FOLDER, result['filename']),
                "message": result.get('message', 'Completed')
            }
        else:
            TASKS[task_id] = {
                "status": "error",
                "error": result.get('error', 'Unknown error') if isinstance(result, dict) else str(result)
            }
    except Exception as e:
        import traceback
        traceback.print_exc()
        TASKS[task_id] = {"status": "error", "error": str(e)}


@app.route('/api/gst/gstr1', methods=['POST'])
def process_gstr1():
    files = request.files.getlist('files') or ([request.files['file']] if 'file' in request.files else [])
    if not files or not files[0].filename:
        return jsonify({'success': False, 'error': 'Please upload GSTR-1 PDF file(s)'}), 400
    saved = save_uploaded_files(files, 'gst')
    task_id = str(uuid.uuid4())
    TASKS[task_id] = {"status": "processing"}
    threading.Thread(target=run_gst_task, args=(task_id, gst_tools.process_gstr1, saved)).start()
    return jsonify({'success': True, 'task_id': task_id})


@app.route('/api/gst/gstr2b', methods=['POST'])
def process_gstr2b():
    files = request.files.getlist('files') or ([request.files['file']] if 'file' in request.files else [])
    if not files or not files[0].filename:
        return jsonify({'success': False, 'error': 'Please upload GSTR-2B Excel file(s)'}), 400
    saved = save_uploaded_files(files, 'gst')
    task_id = str(uuid.uuid4())
    TASKS[task_id] = {"status": "processing"}
    threading.Thread(target=run_gst_task, args=(task_id, gst_tools.process_gstr2b, saved)).start()
    return jsonify({'success': True, 'task_id': task_id})


@app.route('/api/gst/gstr2b-reco', methods=['POST'])
def process_gstr2b_reco():
    gstr2b_files = request.files.getlist('gstr2b_files')
    pr_files = request.files.getlist('pr_files')
    if not gstr2b_files or not gstr2b_files[0].filename:
        return jsonify({'success': False, 'error': 'GSTR-2B file(s) required'}), 400
    saved_2b = save_uploaded_files(gstr2b_files, 'gst')
    saved_pr = save_uploaded_files(pr_files, 'gst') if pr_files and pr_files[0].filename else []
    task_id = str(uuid.uuid4())
    TASKS[task_id] = {"status": "processing"}
    threading.Thread(target=run_gst_task, args=(task_id, gst_tools.reconcile_gstr2b, saved_2b, saved_pr)).start()
    return jsonify({'success': True, 'task_id': task_id})


@app.route('/api/gst/gstr3b', methods=['POST'])
def process_gstr3b():
    files = request.files.getlist('files') or ([request.files['file']] if 'file' in request.files else [])
    if not files or not files[0].filename:
        return jsonify({'success': False, 'error': 'Please upload GSTR-3B PDF file(s)'}), 400
    saved = save_uploaded_files(files, 'gst')
    task_id = str(uuid.uuid4())
    TASKS[task_id] = {"status": "processing"}
    threading.Thread(target=run_gst_task, args=(task_id, gst_tools.process_gstr3b, saved)).start()
    return jsonify({'success': True, 'task_id': task_id})


@app.route('/api/gst/gstr9', methods=['POST'])
def process_gstr9():
    files = request.files.getlist('files') or ([request.files['file']] if 'file' in request.files else [])
    if not files or not files[0].filename:
        return jsonify({'success': False, 'error': 'Please upload GSTR-9/9C PDF file(s)'}), 400
    saved = save_uploaded_files(files, 'gst')
    task_id = str(uuid.uuid4())
    TASKS[task_id] = {"status": "processing"}
    threading.Thread(target=run_gst_task, args=(task_id, gst_tools.process_gstr9, saved)).start()
    return jsonify({'success': True, 'task_id': task_id})


@app.route('/api/gst/tax-comparison', methods=['POST'])
def process_tax_comparison():
    files = request.files.getlist('files') or ([request.files['file']] if 'file' in request.files else [])
    if not files or not files[0].filename:
        return jsonify({'success': False, 'error': 'Please upload Tax Comparison Excel file(s)'}), 400
    saved = save_uploaded_files(files, 'gst')
    mode = request.form.get('mode', 'all')
    task_id = str(uuid.uuid4())
    TASKS[task_id] = {"status": "processing"}
    threading.Thread(target=run_gst_task, args=(task_id, gst_tools.process_tax_comparison, saved, mode)).start()
    return jsonify({'success': True, 'task_id': task_id})


@app.route('/api/gst/ecrrs', methods=['POST'])
def process_ecrrs():
    files = request.files.getlist('files') or ([request.files['file']] if 'file' in request.files else [])
    if not files or not files[0].filename:
        return jsonify({'success': False, 'error': 'Please upload ECRRS CSV file(s)'}), 400
    saved = save_uploaded_files(files, 'gst')
    task_id = str(uuid.uuid4())
    TASKS[task_id] = {"status": "processing"}
    threading.Thread(target=run_gst_task, args=(task_id, gst_tools.process_ecrrs, saved)).start()
    return jsonify({'success': True, 'task_id': task_id})


@app.route('/api/gst/credit-ledger', methods=['POST'])
def process_credit_ledger():
    files = request.files.getlist('files') or ([request.files['file']] if 'file' in request.files else [])
    if not files or not files[0].filename:
        return jsonify({'success': False, 'error': 'Please upload Credit Ledger CSV file(s)'}), 400
    saved = save_uploaded_files(files, 'gst')
    task_id = str(uuid.uuid4())
    TASKS[task_id] = {"status": "processing"}
    threading.Thread(target=run_gst_task, args=(task_id, gst_tools.process_credit_ledger, saved)).start()
    return jsonify({'success': True, 'task_id': task_id})


@app.route('/api/gst/cash-ledger', methods=['POST'])
def process_cash_ledger():
    files = request.files.getlist('files') or ([request.files['file']] if 'file' in request.files else [])
    if not files or not files[0].filename:
        return jsonify({'success': False, 'error': 'Please upload Cash Ledger CSV file(s)'}), 400
    saved = save_uploaded_files(files, 'gst')
    task_id = str(uuid.uuid4())
    TASKS[task_id] = {"status": "processing"}
    threading.Thread(target=run_gst_task, args=(task_id, gst_tools.process_cash_ledger, saved)).start()
    return jsonify({'success': True, 'task_id': task_id})


@app.route('/api/gst/pr-2b-reco', methods=['POST'])
def process_pr_2b_reco():
    gstr2b_files = request.files.getlist('gstr2b_files')
    pr_files = request.files.getlist('pr_files')
    if not gstr2b_files or not gstr2b_files[0].filename:
        return jsonify({'success': False, 'error': 'GSTR-2B file(s) required'}), 400
    saved_2b = save_uploaded_files(gstr2b_files, 'gst')
    saved_pr = save_uploaded_files(pr_files, 'gst') if pr_files and pr_files[0].filename else []
    try:
        result = gst_tools.reconcile_gstr2b(saved_2b, saved_pr)
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/gst/download/<filename>')
def download_gst_file(filename):
    filepath = os.path.join(OUTPUT_FOLDER, filename)
    if os.path.exists(filepath):
        return send_file(filepath, as_attachment=True)
    return jsonify({'error': 'File not found'}), 404


# ─────────────────────────────────────────
# UTILITY ROUTES
# ─────────────────────────────────────────
@app.route('/api/cleanup', methods=['POST'])
def cleanup_uploads():
    """Clean up old upload files."""
    try:
        count = 0
        for root, dirs, files in os.walk(UPLOAD_FOLDER):
            for f in files:
                fp = os.path.join(root, f)
                if os.path.getmtime(fp) < (datetime.now().timestamp() - 86400):  # 24 hours
                    os.remove(fp)
                    count += 1
        return jsonify({'success': True, 'cleaned': count})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


# ══════════════════════════════════════════════════════════
#  COMPLETE TOOL ROUTES — OCR, Crop, Sign, Redact, Compare,
#  PDF→PPTX, PDF→PDF/A, AI Summarize, Translate
# ══════════════════════════════════════════════════════════

@app.route('/api/pdf/ocr', methods=['POST'])
def pdf_ocr():
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'})
    f = request.files['file']
    if not f or not allowed_file(f.filename, 'pdf'):
        return jsonify({'success': False, 'error': 'Please upload a PDF file'})
    filename = secure_filename(f.filename)
    fpath = os.path.join(UPLOAD_FOLDER, filename)
    f.save(fpath)
    return jsonify(pdf_tools.ocr_pdf(fpath))


@app.route('/api/pdf/crop', methods=['POST'])
def pdf_crop():
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'})
    f = request.files['file']
    if not f or not allowed_file(f.filename, 'pdf'):
        return jsonify({'success': False, 'error': 'Please upload a PDF file'})
    filename = secure_filename(f.filename)
    fpath = os.path.join(UPLOAD_FOLDER, filename)
    f.save(fpath)
    top    = float(request.form.get('top', 0))
    bottom = float(request.form.get('bottom', 0))
    left   = float(request.form.get('left', 0))
    right  = float(request.form.get('right', 0))
    return jsonify(pdf_tools.crop_pdf(fpath, top, bottom, left, right))


@app.route('/api/pdf/sign', methods=['POST'])
def pdf_sign():
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'})
    f = request.files['file']
    sig_b64 = request.form.get('signature', '')
    if not f or not allowed_file(f.filename, 'pdf'):
        return jsonify({'success': False, 'error': 'Please upload a PDF file'})
    if not sig_b64:
        return jsonify({'success': False, 'error': 'No signature data received'})
    filename = secure_filename(f.filename)
    fpath = os.path.join(UPLOAD_FOLDER, filename)
    f.save(fpath)
    page_num  = request.form.get('page_num', None)
    position  = request.form.get('position', 'bottom-right')
    sig_width = request.form.get('sig_width', 200)
    return jsonify(pdf_tools.sign_pdf(fpath, sig_b64, page_num, position, sig_width))


@app.route('/api/pdf/redact', methods=['POST'])
def pdf_redact():
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'})
    f = request.files['file']
    search_text = request.form.get('text', '').strip()
    if not f or not allowed_file(f.filename, 'pdf'):
        return jsonify({'success': False, 'error': 'Please upload a PDF file'})
    if not search_text:
        return jsonify({'success': False, 'error': 'Please enter text to redact'})
    filename = secure_filename(f.filename)
    fpath = os.path.join(UPLOAD_FOLDER, filename)
    f.save(fpath)
    return jsonify(pdf_tools.redact_pdf(fpath, search_text))


@app.route('/api/pdf/compare', methods=['POST'])
def pdf_compare():
    files = request.files.getlist('files[]')
    if len(files) < 2:
        return jsonify({'success': False, 'error': 'Please upload exactly 2 PDF files'})
    saved = []
    for f in files[:2]:
        if f and allowed_file(f.filename, 'pdf'):
            fn = secure_filename(f.filename)
            fp = os.path.join(UPLOAD_FOLDER, fn)
            f.save(fp)
            saved.append(fp)
    if len(saved) < 2:
        return jsonify({'success': False, 'error': 'Need 2 valid PDF files'})
    return jsonify(pdf_tools.compare_pdf(saved[0], saved[1]))


@app.route('/api/pdf/to-pptx', methods=['POST'])
def pdf_to_pptx():
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'})
    f = request.files['file']
    if not f or not allowed_file(f.filename, 'pdf'):
        return jsonify({'success': False, 'error': 'Please upload a PDF file'})
    filename = secure_filename(f.filename)
    fpath = os.path.join(UPLOAD_FOLDER, filename)
    f.save(fpath)
    return jsonify(pdf_tools.to_pptx(fpath))


@app.route('/api/pdf/to-pdfa', methods=['POST'])
def pdf_to_pdfa():
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'})
    f = request.files['file']
    if not f or not allowed_file(f.filename, 'pdf'):
        return jsonify({'success': False, 'error': 'Please upload a PDF file'})
    filename = secure_filename(f.filename)
    fpath = os.path.join(UPLOAD_FOLDER, filename)
    f.save(fpath)
    return jsonify(pdf_tools.to_pdfa(fpath))


@app.route('/api/pdf/ai-summarize', methods=['POST'])
def pdf_ai_summarize():
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'})
    f = request.files['file']
    if not f or not allowed_file(f.filename, 'pdf'):
        return jsonify({'success': False, 'error': 'Please upload a PDF file'})
    api_key = request.form.get('api_key', '') or os.environ.get('GEMINI_API_KEY', '')
    if not api_key:
        return jsonify({'success': False,
                        'error': 'Gemini API key required. Add it in the field below.'})
    language = request.form.get('language', 'English')
    filename = secure_filename(f.filename)
    fpath = os.path.join(UPLOAD_FOLDER, filename)
    f.save(fpath)
    return jsonify(pdf_tools.ai_summarize(fpath, api_key, language))


@app.route('/api/pdf/translate', methods=['POST'])
def pdf_translate():
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'})
    f = request.files['file']
    if not f or not allowed_file(f.filename, 'pdf'):
        return jsonify({'success': False, 'error': 'Please upload a PDF file'})
    api_key = request.form.get('api_key', '') or os.environ.get('GEMINI_API_KEY', '')
    if not api_key:
        return jsonify({'success': False,
                        'error': 'Gemini API key required. Get it free at aistudio.google.com'})
    target_lang = request.form.get('target_lang', 'Hindi')
    filename = secure_filename(f.filename)
    fpath = os.path.join(UPLOAD_FOLDER, filename)
    f.save(fpath)
    return jsonify(pdf_tools.translate_pdf(fpath, api_key, target_lang))


# ══════════════════════════════════════════════════════════
#  NEW ROUTES — Page Organizer, Thumbnails, Remove Pages
# ══════════════════════════════════════════════════════════

@app.route('/api/pdf/thumbnails', methods=['POST'])
def pdf_thumbnails():
    """Generate page thumbnails for uploaded PDF."""
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'})
    f = request.files['file']
    if not f or not allowed_file(f.filename, 'pdf'):
        return jsonify({'success': False, 'error': 'Please upload a PDF file'})
    filename = secure_filename(f.filename)
    fpath = os.path.join(UPLOAD_FOLDER, filename)
    f.save(fpath)
    result = pdf_tools.get_thumbnails(fpath, dpi=72)
    result['filepath'] = filename  # send back so organizer can reference it
    return jsonify(result)


@app.route('/api/pdf/remove-pages', methods=['POST'])
def pdf_remove_pages():
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'})
    f = request.files['file']
    pages_str = request.form.get('pages', '')
    if not f or not allowed_file(f.filename, 'pdf'):
        return jsonify({'success': False, 'error': 'Please upload a PDF file'})
    if not pages_str:
        return jsonify({'success': False, 'error': 'Please specify pages to remove (e.g. 1,3,5-8)'})
    filename = secure_filename(f.filename)
    fpath = os.path.join(UPLOAD_FOLDER, filename)
    f.save(fpath)
    result = pdf_tools.remove_pages(fpath, pages_str)
    return jsonify(result)


@app.route('/api/pdf/organize-apply', methods=['POST'])
def pdf_organize_apply():
    """Apply page reorder + rotations from the organizer UI."""
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'})
    f = request.files['file']
    if not f or not allowed_file(f.filename, 'pdf'):
        return jsonify({'success': False, 'error': 'Please upload a PDF file'})
    try:
        new_order = json.loads(request.form.get('order', '[]'))
        rotations = json.loads(request.form.get('rotations', '{}'))
    except Exception:
        return jsonify({'success': False, 'error': 'Invalid order/rotation data'})
    filename = secure_filename(f.filename)
    fpath = os.path.join(UPLOAD_FOLDER, filename)
    f.save(fpath)
    result = pdf_tools.organize_pages(fpath, new_order, rotations)
    return jsonify(result)


@app.route('/api/pdf/merge-ordered', methods=['POST'])
def pdf_merge_ordered():
    """Merge multiple PDFs in custom order sent from frontend."""
    files = request.files.getlist('files[]')
    order_str = request.form.get('order', '')
    if len(files) < 2:
        return jsonify({'success': False, 'error': 'Please upload at least 2 PDF files'})
    saved = []
    for f in files:
        if f and allowed_file(f.filename, 'pdf'):
            fn = secure_filename(f.filename)
            fp = os.path.join(UPLOAD_FOLDER, fn)
            f.save(fp)
            saved.append((fn, fp))
    if len(saved) < 2:
        return jsonify({'success': False, 'error': 'Need at least 2 valid PDFs'})
    # Apply custom order if provided
    if order_str:
        try:
            order = json.loads(order_str)  # list of filenames in desired order
            ordered_paths = []
            name_map = {fn: fp for fn, fp in saved}
            for name in order:
                if name in name_map:
                    ordered_paths.append(name_map[name])
            if len(ordered_paths) == len(saved):
                saved = [(os.path.basename(p), p) for p in ordered_paths]
        except Exception:
            pass
    result = pdf_tools.merge([fp for _, fp in saved])
    return jsonify(result)


# ══════════════════════════════════════════════════════════
#  SCAN TO PDF ROUTES
# ══════════════════════════════════════════════════════════

@app.route('/scan/<session_id>')
def scan_mobile_page(session_id):
    """Mobile camera page — opened via QR code."""
    if session_id not in scan_sessions:
        return '<h2 style="font-family:sans-serif;padding:40px;color:#e2192c">⚠️ Session expired or invalid.<br>Please generate a new QR code.</h2>', 404
    return render_template('scan_mobile.html', session_id=session_id)


@app.route('/api/scan/create', methods=['POST'])
def scan_create():
    """Create a new scan session and return QR code."""
    # Clean old sessions (older than 1 hour)
    now = time.time()
    for sid in list(scan_sessions.keys()):
        if now - scan_sessions[sid]['created'] > 3600:
            folder = os.path.join(UPLOAD_FOLDER, f'scan_{sid}')
            if os.path.exists(folder):
                shutil.rmtree(folder, ignore_errors=True)
            del scan_sessions[sid]

    session_id = str(uuid.uuid4())[:10]
    scan_sessions[session_id] = {'images': [], 'created': now, 'pdf': None}
    os.makedirs(os.path.join(UPLOAD_FOLDER, f'scan_{session_id}'), exist_ok=True)

    # Build mobile URL (use request host)
    proto = 'https' if request.is_secure or 'onrender.com' in request.host else 'http'
    base_url = f"{proto}://{request.host}"
    mobile_url = f"{base_url}/scan/{session_id}"

    # Generate QR code
    try:
        import qrcode
        qr = qrcode.QRCode(version=1, box_size=8, border=2,
                           error_correction=qrcode.constants.ERROR_CORRECT_L)
        qr.add_data(mobile_url)
        qr.make(fit=True)
        img = qr.make_image(fill_color='#1f2937', back_color='white')
        buf = io.BytesIO()
        img.save(buf, format='PNG')
        qr_b64 = base64.b64encode(buf.getvalue()).decode()
        return jsonify({
            'success': True,
            'session_id': session_id,
            'mobile_url': mobile_url,
            'qr_code': f'data:image/png;base64,{qr_b64}'
        })
    except Exception as e:
        return jsonify({'success': False, 'error': f'QR generation failed: {str(e)}'})


@app.route('/api/scan/upload/<session_id>', methods=['POST'])
def scan_upload(session_id):
    """Mobile uploads scanned images here."""
    if session_id not in scan_sessions:
        return jsonify({'success': False, 'error': 'Session expired'})
    folder = os.path.join(UPLOAD_FOLDER, f'scan_{session_id}')
    os.makedirs(folder, exist_ok=True)
    files = request.files.getlist('images[]')
    if not files:
        single = request.files.get('image')
        if single:
            files = [single]
    saved = 0
    for f in files:
        if f and f.filename:
            idx = len(scan_sessions[session_id]['images']) + saved + 1
            ext = f.filename.rsplit('.', 1)[-1].lower() if '.' in f.filename else 'jpg'
            fn = f'scan_{idx:03d}.{ext}'
            fp = os.path.join(folder, fn)
            f.save(fp)
            scan_sessions[session_id]['images'].append(fp)
            saved += 1
    total = len(scan_sessions[session_id]['images'])
    return jsonify({'success': True, 'saved': saved, 'total': total})


@app.route('/api/scan/status/<session_id>')
def scan_status(session_id):
    """Desktop polls this to check how many images arrived."""
    if session_id not in scan_sessions:
        return jsonify({'success': False, 'error': 'Session expired'})
    s = scan_sessions[session_id]
    return jsonify({
        'success': True,
        'image_count': len(s['images']),
        'pdf': s.get('pdf')
    })


@app.route('/api/scan/generate/<session_id>', methods=['POST'])
def scan_generate(session_id):
    """Convert all scanned images to a single PDF."""
    if session_id not in scan_sessions:
        return jsonify({'success': False, 'error': 'Session expired'})
    s = scan_sessions[session_id]
    if not s['images']:
        return jsonify({'success': False, 'error': 'No images scanned yet! Scan at least one page.'})
    try:
        import fitz
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        out_name = f'scanned_{ts}.pdf'
        out_path = os.path.join(OUTPUT_FOLDER, out_name)
        doc = fitz.open()
        for img_path in sorted(s['images']):
            if not os.path.exists(img_path):
                continue
            try:
                img_doc = fitz.open(img_path)
                rect = img_doc[0].rect
                img_doc.close()
            except Exception:
                rect = fitz.Rect(0, 0, 595, 842)  # A4
            page = doc.new_page(width=rect.width, height=rect.height)
            page.insert_image(page.rect, filename=img_path)
        doc.save(out_path, garbage=4, deflate=True)
        doc.close()
        s['pdf'] = out_name
        size = os.path.getsize(out_path)
        size_str = f'{size/1024:.1f} KB' if size < 1024*1024 else f'{size/1024/1024:.2f} MB'
        return jsonify({
            'success': True, 'filename': out_name,
            'pages': len(s['images']), 'size_str': size_str,
            'message': f'Scanned PDF created! {len(s["images"])} page(s), {size_str}'
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


# ── Keep-alive ping endpoint ──────────────────────────────
@app.route('/ping')
def ping():
    return 'pong', 200


@app.errorhandler(404)
def not_found(e):
    return render_template('index.html'), 404

@app.errorhandler(413)
def file_too_large(e):
    return jsonify({'success': False, 'error': 'File too large. Maximum size is 100MB'}), 413


@app.route('/api/contact', methods=['POST'])
def contact_submit():
    """Handle contact form submissions from the frontend."""
    try:
        data = request.json
        fname = data.get('fname', '').strip()
        lname = data.get('lname', '').strip()
        email = data.get('email', '').strip()
        whatsapp = data.get('whatsapp', '').strip()
        service = data.get('service', '').strip()
        desc = data.get('desc', '').strip()
        source = data.get('source', 'unknown')

        if not fname or not email or not whatsapp:
            return jsonify({'success': False, 'error': 'Name, Email, and WhatsApp number are required.'})

        # Save to CSV
        csv_file = os.path.join(BASE_DIR, 'leads.csv')
        file_exists = os.path.isfile(csv_file)
        
        import csv
        with open(csv_file, 'a', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            if not file_exists:
                writer.writerow(['Timestamp', 'Source', 'First Name', 'Last Name', 'Email', 'WhatsApp', 'Service', 'Description'])
            
            writer.writerow([
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                source,
                fname,
                lname,
                email,
                whatsapp,
                service,
                desc
            ])
            
        return jsonify({'success': True, 'message': 'Request saved successfully.'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_ENV', 'development') == 'development'
    app.run(debug=debug, host='0.0.0.0', port=port)

