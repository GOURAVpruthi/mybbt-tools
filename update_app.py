import re

with open(r'C:\Users\hp\.gemini\antigravity\scratch\mybbt-tools\app.py', 'r', encoding='utf-8') as f:
    code = f.read()

new_route = '''@app.route('/api/reco/template', methods=['GET'])
def api_reco_template():
    import pandas as pd
    import io
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    
    file_type = request.args.get('type', 'pr').lower()
    if file_type not in ['pr', '2b']:
        file_type = 'pr'
        
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df = pd.DataFrame(columns=[
            "Serial No", "GSTIN", "Pan", "Vendor Name", "Invoice Number", 
            "Invoice Date", "Taxable Value", "IGST", "CGST", "SGST", 
            "Compcess Value", "Total Tax", "Invoice Value", "Knock Off", 
            "State", "Remarks", "Is Bank"
        ])
        
        sheet_name = 'Purchase_Register' if file_type == 'pr' else 'GSTR_2B'
        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)
        
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        
        # Row 1: MYBBT Branding
        worksheet.merge_cells('A1:Q1')
        cell = worksheet.cell(row=1, column=1)
        cell.value = "MYBBT Business Bluetooth — Professional CA Firm Intelligence Engine"
        cell.font = Font(name='Segoe UI', size=16, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        worksheet.row_dimensions[1].height = 30
        
        # Row 2: Template Type
        worksheet.merge_cells('A2:Q2')
        cell2 = worksheet.cell(row=2, column=1)
        cell2.value = f"DATA TEMPLATE : {'PURCHASE REGISTER (PR)' if file_type == 'pr' else 'GSTR-2B'}"
        cell2.font = Font(name='Segoe UI', size=12, bold=True, color="333333")
        cell2.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        cell2.alignment = Alignment(horizontal="center", vertical="center")
        worksheet.row_dimensions[2].height = 20
        
        # Row 3: Headers
        header_font = Font(name='Segoe UI', bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for col_num, value in enumerate(df.columns.values, 1):
            c = worksheet.cell(row=3, column=col_num)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin_border
            
            # Auto-adjust column width
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = max(len(str(value)) + 2, 12)
            
    output.seek(0)
    filename = 'MYBBT_PR_Template.xlsx' if file_type == 'pr' else 'MYBBT_2B_Template.xlsx'
    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
'''

pattern = r"@app\.route\('/api/reco/template', methods=\['GET'\]\).*?mimetype='application/vnd\.openxmlformats-officedocument\.spreadsheetml\.sheet'\)"
code = re.sub(pattern, new_route, code, flags=re.DOTALL)

with open(r'C:\Users\hp\.gemini\antigravity\scratch\mybbt-tools\app.py', 'w', encoding='utf-8') as f:
    f.write(code)

print('Updated app.py')
