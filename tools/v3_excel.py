import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def build_v3_excel(output_file, engine_self, pr_raw, b2_raw, all_m, pr_nil, b2_nil, unm_pr, unm_b2):
    """
    Constructs the exact 18-sheet v3.0 architecture.
    """
    engine_self.log("  [Excel V3] Compiling data...")
    
    with pd.ExcelWriter(output_file, engine="openpyxl") as wr:
        # 1. FINAL SUMMARY
        summary_data = {
            "Metric": ["Total Entries", "Total Tax (₹)", "Overall Variance (₹)"],
            "PR": [len(pr_raw), pr_raw["Total Tax"].sum() if "Total Tax" in pr_raw.columns else 0, ""],
            "2B": [len(b2_raw), b2_raw["Total Tax"].sum() if "Total Tax" in b2_raw.columns else 0, ""]
        }
        pd.DataFrame(summary_data).to_excel(wr, sheet_name="FINAL SUMMARY", index=False)
        
        # 2. PR SOURCE
        if not pr_raw.empty:
            pr_raw.to_excel(wr, sheet_name="PR SOURCE", index=False)
            
        # 3. 2B SOURCE
        if not b2_raw.empty:
            b2_raw.to_excel(wr, sheet_name="2B SOURCE", index=False)
            
        # 4. COMPARISON
        if not all_m.empty:
            # Reorder columns for comparison
            front = ["Match_Status", "Reco_Remark_Code", "Match_Remark", "Confidence"]
            for c in front:
                if c not in all_m.columns: all_m[c] = ""
            
            pr_cols = [c for c in all_m.columns if str(c).endswith("_PR")]
            b2_cols = [c for c in all_m.columns if str(c).endswith("_2B")]
            common = [c for c in all_m.columns if c not in front and c not in pr_cols and c not in b2_cols]
            
            ordered = front + common + pr_cols + b2_cols
            all_m[ordered].to_excel(wr, sheet_name="COMPARISON", index=False)
            
            # 6. Per Match-Type Sheets
            for code, grp in all_m.groupby("Reco_Remark_Code"):
                sheet_name = f"Matched_{code}"
                if len(sheet_name) > 31: sheet_name = sheet_name[:31]
                grp[ordered].to_excel(wr, sheet_name=sheet_name, index=False)
                
        # 5. ANALYSIS (Pivot summary)
        analysis_rows = []
        if not all_m.empty:
            for _, r in all_m.iterrows():
                analysis_rows.append({
                    "Match_Remark": r.get("Match_Remark", ""),
                    "PR_Amount": pd.to_numeric(r.get("Total Tax_PR", 0), errors='coerce') or 0,
                    "2B_Amount": pd.to_numeric(r.get("Total Tax_2B", 0), errors='coerce') or 0
                })
        if not unm_pr.empty:
            for _, r in unm_pr.iterrows():
                analysis_rows.append({
                    "Match_Remark": "Unmatched PR",
                    "PR_Amount": pd.to_numeric(r.get("Total Tax", 0), errors='coerce') or 0,
                    "2B_Amount": 0
                })
        if not unm_b2.empty:
            for _, r in unm_b2.iterrows():
                analysis_rows.append({
                    "Match_Remark": "Unmatched 2B",
                    "PR_Amount": 0,
                    "2B_Amount": pd.to_numeric(r.get("Total Tax", 0), errors='coerce') or 0
                })
        if analysis_rows:
            adf = pd.DataFrame(analysis_rows)
            pivot = adf.groupby("Match_Remark").sum().reset_index()
            pivot["Variance"] = pivot["PR_Amount"] - pivot["2B_Amount"]
            pivot.to_excel(wr, sheet_name="ANALYSIS", index=False)
            
        # 7. UNMATCHED PR
        if not unm_pr.empty:
            unm_pr.to_excel(wr, sheet_name="UNMATCHED PR", index=False)
            
        # 8. UNMATCHED 2B
        if not unm_b2.empty:
            unm_b2.to_excel(wr, sheet_name="UNMATCHED 2B", index=False)
            
        # 9. UNMATCHED ANALYSIS
        if not unm_pr.empty or not unm_b2.empty:
            um_list = []
            if not unm_pr.empty:
                for _, r in unm_pr.iterrows():
                    um_list.append({"Source": "PR", "Vendor": r.get("Vendor Name", ""), "Amount": pd.to_numeric(r.get("Total Tax", 0), errors='coerce') or 0})
            if not unm_b2.empty:
                for _, r in unm_b2.iterrows():
                    um_list.append({"Source": "2B", "Vendor": r.get("Vendor Name", ""), "Amount": pd.to_numeric(r.get("Total Tax", 0), errors='coerce') or 0})
            if um_list:
                um_df = pd.DataFrame(um_list)
                u_pivot = um_df.groupby(["Source", "Vendor"]).sum().reset_index()
                u_pivot.to_excel(wr, sheet_name="UNMATCHED ANALYSIS", index=False)
                
        # 10. NIL TAX ENTRIES
        nil_frames = []
        if not pr_nil.empty:
            pr_n = pr_nil.copy()
            pr_n["Source"] = "PR"
            nil_frames.append(pr_n)
        if not b2_nil.empty:
            b2_n = b2_nil.copy()
            b2_n["Source"] = "2B"
            nil_frames.append(b2_n)
        if nil_frames:
            pd.concat(nil_frames, ignore_index=True).to_excel(wr, sheet_name="NIL TAX ENTRIES", index=False)
            

                
    engine_self.log("  [Excel V3] Formatting...")
    try:
        format_v3_excel(output_file)
    except Exception as e:
        engine_self.log(f"  [Warning] Format issue: {e}")
    engine_self.log(f"  ✅ Saved V3 Output → {output_file}")


def format_v3_excel(output_file):
    wb = load_workbook(output_file)
    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = font
    wb.save(output_file)
