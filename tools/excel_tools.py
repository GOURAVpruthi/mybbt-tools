"""Excel Consolidator — merge multiple Excel/CSV files."""

import os
from datetime import datetime


class ExcelTools:
    def __init__(self, upload_folder, output_folder):
        self.upload_folder = upload_folder
        self.output_folder = output_folder

    def consolidate(self, input_paths, mode='append'):
        """
        Consolidate multiple Excel/CSV files.
        mode: 'append' — stack all rows into one sheet
              'separate_sheets' — each file becomes a separate sheet
        
        NOTE: Replace this implementation with your own Python code.
        """
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows

            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            out_name = f'consolidated_{ts}.xlsx'
            out_path = os.path.join(self.output_folder, out_name)

            dfs = []
            file_names = []
            total_rows = 0

            for path in input_paths:
                fname = os.path.basename(path)
                ext = fname.rsplit('.', 1)[-1].lower()
                try:
                    if ext == 'csv':
                        df = pd.read_csv(path, encoding='utf-8-sig')
                    else:
                        df = pd.read_excel(path, engine='openpyxl' if ext == 'xlsx' else 'xlrd')
                    df['_source_file'] = fname
                    dfs.append(df)
                    file_names.append(fname)
                    total_rows += len(df)
                except Exception as e:
                    pass  # Skip unreadable files

            if not dfs:
                return {'success': False, 'error': 'Could not read any files'}

            wb = Workbook()

            # Header style
            header_fill = PatternFill(start_color='1a56db', end_color='1a56db', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

            def style_sheet(ws, df):
                for row in dataframe_to_rows(df, index=False, header=True):
                    ws.append(row)
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = thin_border
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.border = thin_border
                for col in ws.columns:
                    max_len = max((len(str(cell.value or '')) for cell in col), default=10)
                    ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

            if mode == 'append':
                ws = wb.active
                ws.title = 'Consolidated'
                combined = pd.concat(dfs, ignore_index=True)
                style_sheet(ws, combined)
            else:
                wb.remove(wb.active)
                for df, name in zip(dfs, file_names):
                    sheet_name = name[:28].replace('/', '_').replace('\\', '_')
                    ws = wb.create_sheet(title=sheet_name)
                    style_sheet(ws, df)

            wb.save(out_path)

            return {
                'success': True,
                'filename': out_name,
                'files_merged': len(dfs),
                'total_rows': total_rows,
                'mode': mode,
                'size_str': self._fmt(os.path.getsize(out_path)),
                'message': f'Successfully consolidated {len(dfs)} files with {total_rows} total rows'
            }

        except ImportError as e:
            return {'success': False, 'error': f'Missing library: {e}. Run: pip install pandas openpyxl'}
        except Exception as e:
            return {'success': False, 'error': str(e)}

    def _fmt(self, size_bytes):
        if size_bytes < 1024:
            return f'{size_bytes} B'
        elif size_bytes < 1024 * 1024:
            return f'{size_bytes / 1024:.1f} KB'
        return f'{size_bytes / (1024 * 1024):.2f} MB'
