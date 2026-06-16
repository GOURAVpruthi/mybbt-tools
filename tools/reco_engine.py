#!/usr/bin/env python3
"""
GST Reconciliation Pro — v5.5  (Modern UI + All-vs-All MAX MATCH + Relax-Head)
============================================================================================
NEW in v5.5  (this build)
  • Max-Match strategy switched to ALL-VS-ALL by default (★) — within
    every common reconciliation group (both sides have residual rows),
    EVERY remaining PR row + EVERY remaining 2B row are matched together
    as one bulk match with the variance reported.  No "best-fit subset"
    leftover.  Legacy behaviour reachable via
    `bank_max_match_strategy: "BEST_FIT"`.
  • PASS 4 added (★ user-driven) — `PAN + FY` only (no Tax Head, no State).
    Runs after PASS 3 and absorbs cross-head residuals (e.g. 2B has IGST
    rows for a vendor + FY but PR booked the same as CGST+SGST).  Default
    on (`bank_match_relax_head: true`).

CARRIED FORWARD from v5.4
  • Tax-head classification fix — mixed-tax rows classified by dominant
    component (`_th`, `_head`, `_head_from_row`).

CARRIED FORWARD from v5.3
  • Bank MAX MATCH now ALWAYS EXHAUSTS THE 2B SIDE (★ user-requested rule):
       – Every common group (PAN/GSTIN + FY + Tax Head + State) that has
         entries on BOTH sides has all 2B entries matched, never leaving
         2B residual within such groups.
       – From the PR side, the subset whose sum is closest to 2B's total
         is picked.  When PR's total is below 2B's, all PR is taken
         (variance reported, residual stays in Only_in_PR).
       – Bank PR/2B totals balance overall (with PR usually carrying a
         small excess), so this matches reality: PR is more granular
         line-item data; 2B is the consolidated authority.
  • New config `bank_max_match_exhaust_side` (default "2B"; can be set to
    "PR" or "smaller" to revert to v112 behaviour).
  • Variance threshold default raised to 10000% so realistic groups
    flow through; absurd ratios (e.g. ₹5K target vs ₹4M only-row) are
    still rejected as "not a real match".
  • Empty-PR-subset cases (no PR row helps reduce variance) tagged as
    "Max Match (2B-Only)" and reported separately on Dashboard.

CARRIED FORWARD from v5.2
  • Bank MAX MATCH (★) — aggregate exhaustion pass after exact subset-sum
    strategies finish.  Greedy descending + 2-opt for big groups,
    exhaustive search for small groups.

CARRIED FORWARD from v5
  • Advanced Fuzzy Invoice — 10-rule deterministic normalization.
  • Bank-Entry Match — ITERATIVE subset-sum matcher on (GSTIN + FY +
    Tax Head + State) with PAN-based fallback pass.
  • Client Name banner, Send-by-Email, Refreshed UI.

CARRIED FORWARD from v4
  • Pre-Group Line Items, Custom remarks, full enrichment columns,
    tabbed Knockout config, free Spinbox max-combo.

EXE:  pyinstaller --onefile --windowed gst_reco_pro.py
Deps: pip install rapidfuzz openpyxl pandas pywin32
"""

import os, re, json, threading, traceback, datetime, itertools, smtplib, ssl
from email.message import EmailMessage
import pandas as pd
import numpy as np

try:
    from rapidfuzz import fuzz, process as rfp
    FUZZY_OK = True
except ImportError:
    FUZZY_OK = False

try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

try:
    import win32com.client
    OUTLOOK_OK = True
except ImportError:
    OUTLOOK_OK = False


# PALETTE
C = {
    "bg":"#F8F9FA","bg_alt":"#F1F3F4","sidebar":"#1F1F1F","sb_hover":"#2D2D2D",
    "sb_active":"#2563EB","sb_text":"#E8EAED","sb_text2":"#9AA0A6",
    "card":"#FFFFFF","card_alt":"#F8F9FA","hover":"#F1F3F4","active":"#E8F0FE",
    "border":"#DADCE0","border_strong":"#C4C7CC",
    "text":"#202124","text2":"#5F6368","text3":"#80868B",
    "accent":"#1A73E8","accent2":"#7C3AED","cyan":"#00BCD4",
    "green":"#0F9D58","red":"#EA4335","yellow":"#F9AB00","pink":"#EC4899","purple":"#673AB7",
    "tag_blue":"#E8F0FE","tag_green":"#E6F4EA","tag_grey":"#F1F3F4","tag_red":"#FCE8E6",
    "tag_pink":"#FCE7F3","tag_yellow":"#FEF7E0","tag_purple":"#F3E8FF",
}
FH1=("Segoe UI",22,"bold"); FH2=("Segoe UI",14,"bold"); FH3=("Segoe UI",11,"bold")
FTB=("Segoe UI",10,"bold"); FT=("Segoe UI",10); FSM=("Segoe UI",9)
FXS=("Segoe UI",8); FMO=("Consolas",9)


# DEFAULTS
KNOWN_COLS = [
    "GSTIN","Pan","Invoice Number","Invoice Date",
    "Total Tax","Total Tax Round","Invoice Prefix","Invoice Suffix",
    "Month Year","State","Vendor Name","IGST","CGST","SGST",
]
PREGROUP_DEFAULT_COLS = ["GSTIN","Invoice Number","Invoice Date"]

DEFAULT_MATCH_RULES = [
    {"cols":["GSTIN","Invoice Number","Invoice Date","Total Tax"],"label":"GSTIN+Invoice+Date+Tax",   "enabled":True, "remark":""},
    {"cols":["GSTIN","Invoice Number","Total Tax"],               "label":"GSTIN+Invoice+Tax",         "enabled":True, "remark":""},
    {"cols":["GSTIN","Invoice Date","Total Tax"],                 "label":"GSTIN+Date+Tax",            "enabled":True, "remark":""},
    {"cols":["GSTIN","Invoice Number"],                           "label":"GSTIN+Invoice",             "enabled":True, "remark":""},
    {"cols":["GSTIN","Total Tax"],                                "label":"GSTIN+Tax",                 "enabled":True, "remark":""},
    {"cols":["Pan","Invoice Number","Invoice Date","Total Tax"],  "label":"PAN+Invoice+Date+Tax",      "enabled":True, "remark":""},
    {"cols":["Pan","Invoice Number","Total Tax"],                 "label":"PAN+Invoice+Tax",           "enabled":True, "remark":""},
    {"cols":["Pan","Invoice Date","Total Tax"],                   "label":"PAN+Date+Tax",              "enabled":True, "remark":""},
    {"cols":["Pan","Invoice Number"],                             "label":"PAN+Invoice",               "enabled":True, "remark":""},
    {"cols":["Pan","Total Tax"],                                  "label":"PAN+Tax",                   "enabled":True, "remark":""},
    {"cols":["Invoice Number","Total Tax"],                       "label":"Invoice+Tax",               "enabled":True, "remark":""},
    {"cols":["GSTIN","Total Tax Round"],                          "label":"GSTIN+RoundedTax",          "enabled":True, "remark":""},
    {"cols":["Pan","Total Tax Round"],                            "label":"PAN+RoundedTax",            "enabled":True, "remark":""},
    {"cols":["Invoice Prefix","Invoice Suffix","Total Tax"],      "label":"Prefix+Suffix+Tax",         "enabled":True, "remark":""},
    {"cols":["GSTIN","Invoice Prefix","Invoice Suffix","Total Tax"],"label":"GSTIN+Prefix+Suffix+Tax","enabled":True, "remark":""},
    {"cols":["Pan","Invoice Prefix","Invoice Suffix","Total Tax"],"label":"PAN+Prefix+Suffix+Tax",    "enabled":True, "remark":""},
    {"cols":["GSTIN","Month Year","Total Tax"],                   "label":"GSTIN+MonthYear+Tax",       "enabled":True, "remark":""},
    {"cols":["Pan","Month Year","Total Tax"],                     "label":"PAN+MonthYear+Tax",         "enabled":True, "remark":""},
    {"cols":["GSTIN","State","Total Tax"],                        "label":"GSTIN+State+Tax",           "enabled":False,"remark":"Verify State"},
]

DEFAULT_KO_RULES = {
    "PR":[
        {"cols":["Pan","Invoice Number","State"],"label":"PAN+Invoice+State","enabled":True, "tax_head":True,"max_combo":3,"remark":""},
        {"cols":["Pan","Invoice Number"],         "label":"PAN+Invoice",      "enabled":True, "tax_head":True,"max_combo":3,"remark":""},
        {"cols":["GSTIN","Invoice Number"],       "label":"GSTIN+Invoice",    "enabled":True, "tax_head":True,"max_combo":3,"remark":""},
        {"cols":["Pan","State"],                  "label":"PAN+State",        "enabled":True, "tax_head":True,"max_combo":2,"remark":""},
    ],
    "2B":[
        {"cols":["GSTIN","State"],  "label":"GSTIN+State",  "enabled":True, "tax_head":True,"max_combo":3,"remark":""},
        {"cols":["GSTIN"],          "label":"GSTIN only",   "enabled":True, "tax_head":True,"max_combo":3,"remark":""},
        {"cols":["Pan","State"],    "label":"PAN+State",    "enabled":True, "tax_head":True,"max_combo":2,"remark":""},
        {"cols":["Pan"],            "label":"PAN only",     "enabled":False,"tax_head":True,"max_combo":2,"remark":""},
    ],
}

DEFAULT_FEATURES = {
    "user_knockoff":True, "pr_knockout":True, "2b_knockout":True,
    "grouped_match":True, "rule_match":True,  "fuzzy_invoice":True,
    "vendor_fuzzy":False, "month_year_match":True, "pct_tolerance":False,
    "pre_group":False,    "bank_match":False,   "bank_max_match":True,
}
RULES_FILE = os.path.join(os.path.expanduser("~"), ".gst_reco_v5.json")


# HELPERS
def get_fy(date):
    try:
        if pd.isna(date): return ""
        d = pd.Timestamp(date)
        return f"FY {d.year}-{str(d.year+1)[2:]}" if d.month>=4 else f"FY {d.year-1}-{str(d.year)[2:]}"
    except Exception: return ""


def enrich_merged(df, tol):
    for p,q in [("_PR","_2B"),("_x","_y")]:
        if f"Total Tax{p}" in df.columns and f"Total Tax{q}" in df.columns:
            df["Amount_Variance_(PR-2B)"] = (df[f"Total Tax{p}"]-df[f"Total Tax{q}"]).round(2)
            df["Excess_In"] = df["Amount_Variance_(PR-2B)"].apply(
                lambda v: "Balanced" if abs(v)<=tol else ("Excess in PR" if v>0 else "Excess in 2B"))
            break
    for p,q in [("_PR","_2B"),("_x","_y")]:
        dp=f"Invoice Date{p}"; dq=f"Invoice Date{q}"
        if dp in df.columns and dq in df.columns:
            dpr=pd.to_datetime(df[dp],errors="coerce"); db2=pd.to_datetime(df[dq],errors="coerce")
            df["Date_Variance_Days"]=(dpr-db2).dt.days.abs()
            df["Month_Match"]=(dpr.dt.to_period("M")==db2.dt.to_period("M")).map({True:"Yes",False:"No"})
            df["FY_Match"]=[("Yes" if get_fy(a)==get_fy(b) and get_fy(a) else "No") for a,b in zip(dpr,db2)]
            break
    for p,q in [("_PR","_2B"),("_x","_y")]:
        sp=f"State{p}"; sq=f"State{q}"
        if sp in df.columns and sq in df.columns:
            df["State_Match"]=(df[sp].astype(str).str.upper().str.strip()==
                              df[sq].astype(str).str.upper().str.strip()).map({True:"Yes",False:"No"})
            break
    def _head(igst, cgst, sgst):
        try:
            i=abs(float(igst or 0)); c=abs(float(cgst or 0)); s=abs(float(sgst or 0))
        except Exception: return "UNKNOWN"
        cs = c + s
        if i == 0 and cs == 0: return "UNKNOWN"
        if i > 0 and cs == 0:  return "IGST"
        if cs > 0 and i == 0:  return "CGST_SGST"
        return "IGST" if i > cs else "CGST_SGST"
    for p,q in [("_PR","_2B"),("_x","_y")]:
        if (f"IGST{p}" in df.columns and f"IGST{q}" in df.columns):
            pr_head = df.apply(lambda r: _head(r.get(f"IGST{p}",0), r.get(f"CGST{p}",0), r.get(f"SGST{p}",0)), axis=1)
            b2_head = df.apply(lambda r: _head(r.get(f"IGST{q}",0), r.get(f"CGST{q}",0), r.get(f"SGST{q}",0)), axis=1)
            df["Tax_Head_PR"] = pr_head; df["Tax_Head_2B"] = b2_head
            df["Head_Match"] = (pr_head == b2_head).map({True:"Yes",False:"No"})
            break
    for p,q in [("_PR","_2B"),("_x","_y")]:
        gp=f"GSTIN{p}"; gq=f"GSTIN{q}"
        if gp in df.columns and gq in df.columns:
            df["Transaction_Type"]=df.apply(
                lambda r: ("Interstate" if str(r.get(gp,""))[:2]!=str(r.get(gq,""))[:2]
                           and len(str(r.get(gp,"")))>=2 else "Intrastate"), axis=1)
            break
    return df


def enrich_side_df(df, tol):
    if df.empty or "_Side" not in df.columns or "Common_UID" not in df.columns:
        return df

    def _head_from_row(r):
        try:
            i=abs(float(r.get("IGST",0) or 0))
            c=abs(float(r.get("CGST",0) or 0))
            s=abs(float(r.get("SGST",0) or 0))
        except Exception: return "UNKNOWN"
        cs = c + s
        if i == 0 and cs == 0: return "UNKNOWN"
        if i > 0 and cs == 0:  return "IGST"
        if cs > 0 and i == 0:  return "CGST_SGST"
        return "IGST" if i > cs else "CGST_SGST"

    parts=[]
    for uid, grp in df.groupby("Common_UID"):
        pr=grp[grp["_Side"]=="PR"]; b2=grp[grp["_Side"]=="2B"]
        pt=pr["Total Tax"].sum() if not pr.empty else 0
        bt=b2["Total Tax"].sum() if not b2.empty else 0
        var=round(pt-bt,2)
        g=grp.copy()
        g["Amount_Variance_(PR-2B)"]=var
        g["Excess_In"]=("Balanced" if abs(var)<=tol else ("Excess in PR" if var>0 else "Excess in 2B"))
        d_pr=(pr["Invoice Date"].dropna().iloc[0] if not pr.empty and "Invoice Date" in pr.columns and not pr["Invoice Date"].dropna().empty else None)
        d_b2=(b2["Invoice Date"].dropna().iloc[0] if not b2.empty and "Invoice Date" in b2.columns and not b2["Invoice Date"].dropna().empty else None)
        if d_pr is not None and d_b2 is not None:
            try:
                dv=abs((pd.Timestamp(d_pr)-pd.Timestamp(d_b2)).days)
                mm="Yes" if pd.Timestamp(d_pr).to_period("M")==pd.Timestamp(d_b2).to_period("M") else "No"
                fm="Yes" if get_fy(d_pr)==get_fy(d_b2) else "No"
            except Exception: dv=mm=fm="N/A"
        else: dv=mm=fm="N/A"
        g["Date_Variance_Days"]=dv; g["Month_Match"]=mm; g["FY_Match"]=fm
        if "State" in grp.columns and not pr.empty and not b2.empty:
            s_pr = str(pr["State"].iloc[0] or "").upper().strip()
            s_b2 = str(b2["State"].iloc[0] or "").upper().strip()
            g["State_Match"] = "Yes" if (s_pr and s_b2 and s_pr==s_b2) else ("N/A" if not s_pr or not s_b2 else "No")
        else:
            g["State_Match"] = "N/A"
        if not pr.empty and not b2.empty:
            h_pr = _head_from_row(pr.iloc[0]); h_b2 = _head_from_row(b2.iloc[0])
            g["Tax_Head_PR"] = h_pr; g["Tax_Head_2B"] = h_b2
            g["Head_Match"] = "Yes" if (h_pr==h_b2 and h_pr!="UNKNOWN") else "No"
        else:
            g["Tax_Head_PR"] = g["Tax_Head_2B"] = g["Head_Match"] = "N/A"
        parts.append(g)
    return pd.concat(parts,ignore_index=True) if parts else df


# FUZZY INVOICE — 10-RULE DETERMINISTIC NORMALIZATION
_SEP_RE = re.compile(r'[-/\.\s:_]+')
_LOOKALIKE = str.maketrans({'S':'5','L':'1','I':'1','O':'0','Z':'2','B':'8'})

def _strip_zeros_per_segment(s):
    parts = re.split(r'([-/\.\s:_]+)', s)
    out = []
    for p in parts:
        if p.isdigit() and len(p) > 1:
            out.append(p.lstrip('0') or '0')
        else:
            out.append(p)
    return ''.join(out)

def _shorten_year(s):
    def repl(m):
        y1, y2 = m.group(1), m.group(2)
        y1s = y1[-2:]; y2s = y2[-2:]
        if int(y2s) == int(y1s)+1:
            return f"{y1s}-{y2s}"
        return m.group(0)
    return re.sub(r'(?<!\d)(20\d{2})-(20\d{2}|\d{2})(?!\d)', repl, s)

def _expand_year(s):
    def repl(m):
        y1, y2 = m.group(1), m.group(2)
        if int(y2) == int(y1)+1:
            return f"20{y1}-20{y2}"
        return m.group(0)
    return re.sub(r'(?<!\d)(\d{2})-(\d{2})(?!\d)', repl, s)

def norm_variants(inv):
    if inv is None: return set()
    s = str(inv).strip()
    if not s or s.lower() == 'nan': return set()
    V = set()
    u = s.upper()
    V.add(u)
    V.add(_SEP_RE.sub('', u))
    zz = _strip_zeros_per_segment(u)
    V.add(zz); V.add(_SEP_RE.sub('', zz))
    la = u.translate(_LOOKALIKE)
    V.add(la); V.add(_SEP_RE.sub('', la))
    la_zz = _strip_zeros_per_segment(la)
    V.add(la_zz); V.add(_SEP_RE.sub('', la_zz))
    for variant in {u, la, zz, la_zz}:
        sy = _shorten_year(variant)
        V.add(sy); V.add(_SEP_RE.sub('', sy))
        V.add(_strip_zeros_per_segment(sy))
        V.add(_SEP_RE.sub('', _strip_zeros_per_segment(sy)))
        ey = _expand_year(variant)
        V.add(ey); V.add(_SEP_RE.sub('', ey))
    V.discard('')
    return V

def _strip_trailing_fy(s):
    m = re.match(r'^(.+?)[-/\.\s:_](\d{2})-(\d{2}|\d{4})$', s)
    if not m: return None
    base, y1, y2 = m.group(1), m.group(2), m.group(3)
    if len(base) < 2: return None
    y2s = y2[-2:]
    try:
        if int(y2s) != int(y1)+1: return None
    except ValueError:
        return None
    return base

def _strip_glued_alpha_prefix(s):
    m = re.match(r'^([A-Z]{1,4})[-/\.\s:_]?(\d.*)$', s)
    if m: return m.group(2)
    return None

def asym_strip_variants(inv):
    if inv is None: return set()
    s = str(inv).strip().upper()
    if not s or s.lower() == 'nan': return set()
    out = set()
    base = _strip_trailing_fy(s)
    if base is not None:
        out |= norm_variants(base)
    base = _strip_glued_alpha_prefix(s)
    if base is not None:
        out |= norm_variants(base)
    return out

def _segment_equal(a_segs, b_segs):
    if len(a_segs) != len(b_segs): return False
    for sx, sy in zip(a_segs, b_segs):
        if sx == sy: continue
        nx = sx.lstrip('0') if sx.isdigit() else sx
        ny = sy.lstrip('0') if sy.isdigit() else sy
        if not nx: nx = '0'
        if not ny: ny = '0'
        if nx == ny: continue
        if nx.isdigit() and ny.isdigit():
            lng, shr = (nx, ny) if len(nx) >= len(ny) else (ny, nx)
            if lng.endswith(shr) and len(shr) >= max(1, len(lng)/2):
                continue
        return False
    return True

def segment_match(a, b):
    if not a or not b: return False
    a = str(a).upper().strip(); b = str(b).upper().strip()
    if a == b: return True
    a_segs = [x for x in re.split(r'[-/\.\s:_]+', a) if x]
    b_segs = [x for x in re.split(r'[-/\.\s:_]+', b) if x]
    if not a_segs or not b_segs: return False
    if len(a_segs) == len(b_segs):
        return _segment_equal(a_segs, b_segs)
    longer, shorter = (a_segs, b_segs) if len(a_segs) > len(b_segs) else (b_segs, a_segs)
    extra = len(longer) - len(shorter)
    prefix = longer[:extra]
    if not all(seg.isalpha() for seg in prefix):
        return False
    return _segment_equal(longer[extra:], shorter)

def fuzzy_invoice_match(a_raw, b_raw):
    if a_raw is None or b_raw is None: return (False, "")
    sa = str(a_raw).strip(); sb = str(b_raw).strip()
    if not sa or not sb or sa.lower()=='nan' or sb.lower()=='nan': return (False, "")
    va = norm_variants(sa); vb = norm_variants(sb)
    if va & vb:
        return (True, "Normalized")
    asa = asym_strip_variants(sa); asb = asym_strip_variants(sb)
    if asa & vb: return (True, "Strip-Prefix/Suffix (A→B)")
    if asb & va: return (True, "Strip-Prefix/Suffix (B→A)")
    if segment_match(sa, sb):
        return (True, "Segment Match")
    return (False, "")


# EMAIL SENDER
def send_email_outlook(recipient, subject, body, attachment_path):
    if not OUTLOOK_OK:
        return (False, "pywin32 not installed")
    try:
        outlook = win32com.client.Dispatch("Outlook.Application")
        mail = outlook.CreateItem(0)
        mail.To = recipient; mail.Subject = subject; mail.Body = body
        if attachment_path and os.path.exists(attachment_path):
            mail.Attachments.Add(os.path.abspath(attachment_path))
        mail.Display(False)
        return (True, "Outlook opened with message ready to send")
    except Exception as e:
        return (False, f"Outlook error: {e}")

def send_email_smtp(host, port, user, password, recipient, subject, body, attachment_path):
    try:
        msg = EmailMessage()
        msg["Subject"] = subject; msg["From"] = user; msg["To"] = recipient
        msg.set_content(body)
        if attachment_path and os.path.exists(attachment_path):
            with open(attachment_path, "rb") as fh:
                data = fh.read()
            msg.add_attachment(data, maintype="application",
                               subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               filename=os.path.basename(attachment_path))
        ctx = ssl.create_default_context()
        if int(port) == 465:
            with smtplib.SMTP_SSL(host, int(port), context=ctx) as s:
                s.login(user, password); s.send_message(msg)
        else:
            with smtplib.SMTP(host, int(port)) as s:
                s.starttls(context=ctx); s.login(user, password); s.send_message(msg)
        return (True, "Sent via SMTP")
    except Exception as e:
        return (False, f"SMTP error: {e}")


# ANIMATION HELPER
def _hex_to_rgb(h):
    h = h.lstrip('#')
    return (int(h[0:2],16), int(h[2:4],16), int(h[4:6],16))

def _rgb_to_hex(r,g,b):
    r=max(0,min(255,int(r))); g=max(0,min(255,int(g))); b=max(0,min(255,int(b)))
    return f"#{r:02X}{g:02X}{b:02X}"

def smooth_bg(widget, target, steps=6, delay=14, prop="bg"):
    try: widget.config(**{prop: target})
    except Exception: pass


class GSTRecoEngine:
    DEFAULT_BANK_KEYWORDS = [
        "BANK","HDFC","ICICI","AXIS","SBI","STATE BANK","KOTAK","YES BANK",
        "IDFC","INDUSIND","PNB","PUNJAB NATIONAL","BANK OF BARODA","BOB",
        "CANARA","UNION BANK","FEDERAL","RBL","IDBI","BANDHAN","CITI",
        "HSBC","STANDARD CHARTERED","DEUTSCHE","DBS","BARCLAYS","AMERICAN EXPRESS",
        "AMEX","AU SMALL FINANCE","UCO","INDIAN BANK","CENTRAL BANK","DCB",
        "SOUTH INDIAN","KARUR VYSYA","TAMILNAD MERCANTILE","JAMMU AND KASHMIR",
        "J&K BANK","BANK OF INDIA","BOI","BANK OF MAHARASHTRA","BOM"
    ]
    DEFAULT_EXCLUSION_KEYWORDS = [
        "INSURANCE","LOMBARD","LIFE","MUTUAL FUND","MUTUAL","AMC",
        "ASSET MANAGEMENT","BROKING","SECURITIES","CAPITAL",
        "CARDS","CREDIT CARD","PRUDENTIAL","GENERAL INSURANCE",
        "HEALTH INSURANCE","ERGO","LOMBARD","ALLIANZ","TOKIO",
        "STANDARD LIFE","METLIFE","REINSURANCE","MF","AMC LTD"
    ]

    def __init__(self,cfg,log_fn=None):
        self.tol   =float(cfg.get("amount_tolerance",10))
        self.pct   =float(cfg.get("pct_tolerance",5))
        self.fthr  =int(cfg.get("fuzzy_threshold",80))
        self.vthr  =int(cfg.get("vendor_threshold",75))
        self.mrules=[r for r in cfg.get("match_rules",DEFAULT_MATCH_RULES)    if r.get("enabled",True)]
        self.pr_ko =[r for r in cfg.get("pr_ko_rules",DEFAULT_KO_RULES["PR"]) if r.get("enabled",True)]
        self.b2_ko =[r for r in cfg.get("b2b_ko_rules",DEFAULT_KO_RULES["2B"])if r.get("enabled",True)]
        self.feat  ={**DEFAULT_FEATURES,**cfg.get("features",{})}
        self.pg_cols=cfg.get("pregroup_cols",PREGROUP_DEFAULT_COLS)
        self.bank_keywords = cfg.get("bank_keywords", self.DEFAULT_BANK_KEYWORDS)
        self.bank_exclusions = cfg.get("bank_exclusions", self.DEFAULT_EXCLUSION_KEYWORDS)
        self.bank_filter_on = bool(cfg.get("bank_filter_on", True))
        # Max-Match safety threshold: skip if abs(variance) > this fraction of smaller-side total.
        # 100.0 = picked sum may be up to 101x the target.  Default is intentionally loose so
        # realistic bank groups flow through; only absurd ratios get rejected.
        self.bank_max_match_var_pct = float(cfg.get("bank_max_match_variance_pct", 10000)) / 100.0
        # Minimum exhaust-side absolute total to consider Max-Match (avoids zero-target weirdness).
        # Default ₹1 — sub-rupee residuals are usually rounding noise the user is happy to absorb.
        self.bank_max_match_min_total = float(cfg.get("bank_max_match_min_total", 1))
        # Which side to exhaust fully in Max-Match (used by BEST_FIT strategy):
        # "2B" (default — no 2B leftover) / "PR" / "smaller" (= v112 behaviour).
        side = str(cfg.get("bank_max_match_exhaust_side", "2B")).strip().upper()
        if side not in ("2B", "PR", "SMALLER"):
            side = "2B"
        self.bank_max_match_side = side
        # Strategy: ALL_VS_ALL (default — match every PR row + every 2B row in each
        # common group, variance reported) or BEST_FIT (legacy — pick subset of
        # the larger side closest to the smaller side's total).
        strat = str(cfg.get("bank_max_match_strategy", "ALL_VS_ALL")).strip().upper()
        if strat not in ("ALL_VS_ALL", "BEST_FIT"):
            strat = "ALL_VS_ALL"
        self.bank_max_match_strategy = strat
        # PASS 4 — relax tax-head matching as a final fallback.  When True,
        # after PASS 3 finishes, an additional pass groups residual rows
        # by PAN + FY only (no Tax Head, no State) and runs Max-Match
        # all-vs-all.  This catches cases where 2B has IGST entries but
        # PR has CGST/SGST entries (or vice-versa) for the same vendor
        # and same financial year — typical of mixed-classification bank
        # charges and tax-engine quirks.
        self.bank_match_relax_head = bool(cfg.get("bank_match_relax_head", True))
        self.log   =log_fn or print
        self._ctr  =10000

    def _uid(self,t="M"):
        u=f"{t}_{self._ctr}"; self._ctr+=1; return u

    def _cc(self,df):
        df.columns=[re.sub(r"\s+"," ",str(c).strip()).replace("\xa0"," ") for c in df.columns]
        return df

    def _clean(self,df):
        df["Invoice Number"]=(df["Invoice Number"].astype(str).str.upper().str.replace(r"[^A-Z0-9]","",regex=True))
        df["Invoice Base"]  =df["Invoice Number"].str.extract(r"(\d{2,})",expand=False)
        df["Invoice Prefix"]=df["Invoice Number"].str[:3]
        df["Invoice Suffix"]=df["Invoice Number"].str[-3:]
        df["Total Tax"]     =pd.to_numeric(df["Total Tax"],errors="coerce").fillna(0).round(2)
        df["Total Tax Round"]=df["Total Tax"].round(-1)
        df["Invoice Date"]  =pd.to_datetime(df["Invoice Date"],errors="coerce")
        df["Month Year"]    =df["Invoice Date"].dt.to_period("M").astype(str)
        df["_FY"]           =df["Invoice Date"].apply(get_fy)
        for col in ["IGST","CGST","SGST"]:
            if col in df.columns:
                df[col]=pd.to_numeric(df[col],errors="coerce").fillna(0).round(2)
        if "State" not in df.columns: df["State"]=""
        df["State"]=df["State"].astype(str).str.strip().str.upper()
        if "Vendor Name" not in df.columns: df["Vendor Name"]=""
        df["Vendor Name"]=df["Vendor Name"].astype(str).str.strip().str.upper()
        return df

    def _pre_group(self, df, tag):
        key_cols = [c for c in self.pg_cols if c in df.columns]
        if not key_cols:
            self.log(f"  [PreGroup-{tag}] None of {self.pg_cols} found → skipped")
            return df
        sum_cols = [c for c in ["Total Tax","IGST","CGST","SGST"] if c in df.columns]
        other_cols = [c for c in df.columns if c not in key_cols+sum_cols+["UID","Total Tax Round"]]
        agg = {c:"sum" for c in sum_cols}
        agg.update({c:"first" for c in other_cols})
        grouped = df.groupby(key_cols,dropna=False,sort=False).agg(agg).reset_index()
        if "Total Tax" in grouped.columns:
            grouped["Total Tax Round"] = grouped["Total Tax"].round(-1)
        grouped["UID"] = tag+"_PG_"+grouped.index.astype(str)
        n_before = len(df); n_after = len(grouped)
        self.log(f"  [PreGroup-{tag}] {n_before} rows → {n_after} groups (saved {n_before-n_after} rows)")
        return grouped

    @staticmethod
    def _th(row):
        # Tax-head classification.  When both IGST and CGST/SGST are non-zero
        # (mixed-classification entries — common for some bank charges where
        # the books carry a small token IGST alongside the real CGST+SGST),
        # pick whichever side is dominant in absolute value.  Without this
        # the row gets bucketed by the wrong head and never joins its real
        # group (e.g. ₹5,167 IGST + ₹20L+₹20L CGST+SGST should clearly be
        # CGST_SGST, not IGST).
        try:
            i = abs(float(row.get("IGST", 0) or 0))
            c = abs(float(row.get("CGST", 0) or 0))
            s = abs(float(row.get("SGST", 0) or 0))
        except Exception:
            return "UNKNOWN"
        cs = c + s
        if i == 0 and cs == 0: return "UNKNOWN"
        if i > 0 and cs == 0:  return "IGST"
        if cs > 0 and i == 0:  return "CGST_SGST"
        # Both non-zero → dominant side wins (ties go to CGST_SGST since
        # bank charges typically split equally between CGST/SGST).
        return "IGST" if i > cs else "CGST_SGST"

    def _amt_ok(self,a,b):
        if self.feat.get("pct_tolerance",False):
            base=max(abs(a),abs(b),1)
            return abs(a-b)/base*100<=self.pct
        return abs(a-b)<=self.tol

    def _knockout(self,df,ko_rules,tag,uid_base):
        if df[df["Total Tax"]<0].empty:
            self.log(f"  [{tag} KO] No negatives → skipped"); return pd.DataFrame(),df
        ko_frames,matched=[],[]
        uc=uid_base
        for rule in ko_rules:
            cols=[c for c in rule["cols"] if c in df.columns]
            enf=rule.get("tax_head",True); mc=rule.get("max_combo",3); rmk=rule.get("remark","")
            if not cols: continue
            work=df[~df["UID"].isin(matched)].copy()
            if enf: work["_TH"]=work.apply(self._th,axis=1); key=cols+["_TH"]
            else: key=cols
            for c in key: work[c]=work[c].fillna("__").astype(str)
            for name,grp in work.groupby(key):
                pos=grp[grp["Total Tax"]>0].copy(); neg=grp[grp["Total Tax"]<0].copy()
                if pos.empty or neg.empty: continue
                if enf and str(name[-1])=="UNKNOWN": continue
                up,un=set(),set()
                pos=pos.sort_values("Total Tax"); neg=neg.sort_values("Total Tax")
                for _,nr in neg.iterrows():
                    if nr.name in un: continue
                    target=-nr["Total Tax"]; avail=pos[~pos.index.isin(up)]
                    if avail.empty: continue
                    combo=[]
                    for _,pr in avail.iterrows():
                        if self._amt_ok(pr["Total Tax"],target): combo=[pr]; break
                    if not combo:
                        pl=list(avail.iterrows())
                        for n in range(2,min(mc,len(pl))+1):
                            found=False
                            for indices in itertools.combinations(range(len(pl)),n):
                                s=sum(pl[i][1]["Total Tax"] for i in indices)
                                if self._amt_ok(s,target):
                                    combo=[pl[i][1] for i in indices]; found=True; break
                            if found: break
                    if combo:
                        uid=f"{tag}_KO_{uc}"; uc+=1
                        pair=pd.concat([r.to_frame().T for r in combo+[nr]],ignore_index=True)
                        pair["Match Status"]=f"{tag} Knockout"; pair["Matched On"]=rule["label"]
                        pair["Common_UID"]=uid
                        if rmk: pair["Custom_Remark"]=rmk
                        ko_frames.append(pair)
                        for r in combo: up.add(r.name); matched.append(r["UID"])
                        un.add(nr.name); matched.append(nr["UID"])
        ko_df=pd.concat(ko_frames,ignore_index=True) if ko_frames else pd.DataFrame()
        remaining=df[~df["UID"].isin(matched)].copy()
        self.log(f"  [{tag} KO] {len(matched)} knocked off, {len(remaining)} remain")
        return ko_df,remaining

    def _grouped(self,pr,b2):
        self.log(f"  [Grouped] {len(self.mrules)} rules …")
        frames,ps,bs=[],[],[]
        for rule in self.mrules:
            cols=rule["cols"]; rmk=rule.get("remark","")
            if not(set(cols)<=set(pr.columns) and set(cols)<=set(b2.columns)): continue
            pw=pr[~pr["UID"].isin(ps)]; bw=b2[~b2["UID"].isin(bs)]
            pg=pw.groupby(cols,dropna=False).agg(PT=("Total Tax","sum"),PU=("UID",list)).reset_index()
            bg=bw.groupby(cols,dropna=False).agg(BT=("Total Tax","sum"),BU=("UID",list)).reset_index()
            mg=pd.merge(pg,bg,on=cols,how="inner")
            if mg.empty: continue
            for _,row in mg.iterrows():
                ok=self._amt_ok(row["PT"],row["BT"]); uid=self._uid("GRP")
                for sub,side in [(pw[pw["UID"].isin(row["PU"])],"PR"),(bw[bw["UID"].isin(row["BU"])],"2B")]:
                    s=sub.copy()
                    s["Match Status"]="Grouped Match" if ok else "Grouped Amount Mismatch"
                    s["Matched On"]=rule["label"]; s["Common_UID"]=uid; s["_Side"]=side
                    if rmk: s["Custom_Remark"]=rmk
                    frames.append(s)
                ps.extend(row["PU"]); bs.extend(row["BU"])
        combined=pd.concat(frames,ignore_index=True) if frames else pd.DataFrame()
        if not combined.empty: combined=enrich_side_df(combined,self.tol)
        pr=pr[~pr["UID"].isin(ps)]; b2=b2[~b2["UID"].isin(bs)]
        self.log(f"  [Grouped] {len(combined)} rows"); return combined,pr,b2

    def _rules(self,pr,b2):
        self.log(f"  [Rules] {len(self.mrules)} rules …")
        frames,ps,bs=[],[],[]
        for rule in self.mrules:
            cols=rule["cols"]; rmk=rule.get("remark","")
            if not(set(cols)<=set(pr.columns) and set(cols)<=set(b2.columns)): continue
            pw=pr[~pr["UID"].isin(ps)]; bw=b2[~b2["UID"].isin(bs)]
            mg=pd.merge(pw,bw,on=cols,how="inner",suffixes=("_PR","_2B"))
            if mg.empty: continue
            if "Total Tax_PR" in mg.columns and "Total Tax_2B" in mg.columns:
                mg["Match Status"]=mg.apply(
                    lambda r: "Matched" if self._amt_ok(r["Total Tax_PR"],r["Total Tax_2B"]) else "Amount Mismatch",axis=1)
            else: mg["Match Status"]="Matched"
            mg["Matched On"]=rule["label"]
            mg["Common_UID"]=[self._uid("RULE") for _ in range(len(mg))]
            if rmk: mg["Custom_Remark"]=rmk
            mg=enrich_merged(mg,self.tol)
            frames.append(mg)
            if "UID_PR" in mg.columns: ps.extend(mg["UID_PR"].dropna().unique().tolist())
            if "UID_2B" in mg.columns: bs.extend(mg["UID_2B"].dropna().unique().tolist())
        combined=pd.concat(frames,ignore_index=True) if frames else pd.DataFrame()
        pr=pr[~pr["UID"].isin(ps)]; b2=b2[~b2["UID"].isin(bs)]
        self.log(f"  [Rules] {len(combined)} rows"); return combined,pr,b2

    def _fuzzy(self,pr,b2):
        self.log(f"  [Fuzzy v5] {len(pr)} PR rows …")
        frames,mp,mb=[],set(),set()
        b2=b2.copy()
        if "Invoice Number" not in b2.columns or "Invoice Number" not in pr.columns:
            self.log("  [Fuzzy] Invoice Number column missing → skipped")
            return pd.DataFrame(),pr,b2
        b2["_VB"] = b2["Invoice Number"].apply(norm_variants)
        b2["_AS"] = b2["Invoice Number"].apply(asym_strip_variants)

        def _score_pair(pr_row, b2_row):
            a = pr_row.get("Invoice Number",""); b = b2_row.get("Invoice Number","")
            va = pr_row.get("_VA", norm_variants(a))
            vb = b2_row.get("_VB", norm_variants(b))
            if va & vb:                    return ("Normalized", 3)
            asa = pr_row.get("_AS", asym_strip_variants(a))
            asb = b2_row.get("_AS", asym_strip_variants(b))
            if asa & vb or asb & va:       return ("Strip-Affix", 2)
            if segment_match(a, b):        return ("Segment Match", 1)
            return ("", 0)

        for _,row in pr.iterrows():
            if row["UID"] in mp: continue
            gstin=str(row.get("GSTIN",""))
            sub=b2[(b2["GSTIN"]==gstin)&(~b2["UID"].isin(mb))]
            if sub.empty: continue
            inv_pr = row.get("Invoice Number","")
            va = norm_variants(inv_pr); asa = asym_strip_variants(inv_pr)
            pt = row.get("Total Tax",0); my = str(row.get("Month Year",""))
            best = None
            for _, br in sub.iterrows():
                rule_label, rule_score = _score_pair(
                    {"_VA":va, "_AS":asa, "Invoice Number":inv_pr, "Total Tax":pt},
                    {"_VB":br.get("_VB",set()), "_AS":br.get("_AS",set()),
                     "Invoice Number":br.get("Invoice Number","")})
                if rule_score == 0: continue
                bt = br.get("Total Tax", 0); bmy = str(br.get("Month Year",""))
                amt_ok = self._amt_ok(pt, bt); month_ok = (my == bmy and my != "")
                bucket = (2 if amt_ok else 0) + (1 if month_ok else 0)
                key = (bucket, rule_score)
                if best is None or key > best[0]:
                    best = (key, br, rule_label, amt_ok, month_ok)
            if best is None: continue
            (_bucket, mr, rule_label, amt_ok, month_ok) = best
            bt = mr.get("Total Tax", 0); var = round(pt-bt, 2)
            uid = self._uid("FUZ")
            for part,side in [(row.to_frame().T.copy(),"PR"),(mr.to_frame().T.copy(),"2B")]:
                for c in ("_VB","_VA","_AS"):
                    if c in part.columns: part=part.drop(columns=c)
                part["Match Status"]="Fuzzy Match"
                part["Matched On"]=f"Fuzzy Invoice ({rule_label})"
                part["Common_UID"]=uid; part["_Side"]=side
                part["Fuzzy_Rule"]=rule_label
                part["Fuzzy_Amt_Match"]="Yes" if amt_ok else "No"
                part["Fuzzy_Month_Match"]="Yes" if month_ok else "No"
                part["Amount_Variance_(PR-2B)"]=var
                part["Excess_In"]=("Balanced" if abs(var)<=self.tol
                                   else ("Excess in PR" if var>0 else "Excess in 2B"))
                frames.append(part)
            mp.add(row["UID"]); mb.add(mr["UID"])
        combined=pd.concat(frames,ignore_index=True) if frames else pd.DataFrame()
        if not combined.empty: combined = enrich_side_df(combined, self.tol)
        for c in ("_VB","_AS"):
            if c in b2.columns: b2=b2.drop(columns=c)
        pr=pr[~pr["UID"].isin(mp)]; b2=b2[~b2["UID"].isin(mb)]
        self.log(f"  [Fuzzy v5] {len(mp)} pairs matched"); return combined,pr,b2

    # ═══════════════════════════════════════════════════════════
    # BANK-ENTRY MATCH — ITERATIVE SUBSET-SUM + MAX MATCH
    # ═══════════════════════════════════════════════════════════
    def _best_subset_to_target(self, rows_df, target, tol):
        """
        Find the subset of rows in rows_df whose Total Tax sums CLOSEST to
        `target`.  Returns (list of UIDs, actual_sum).

        Used by Max-Match (Strategy 4) to fully exhaust the smaller side
        of a group and pick the best-fitting subset on the larger side.

        Strategy:
          • Single-entry exact match → return immediately
          • All-or-nothing exact      → return all rows
          • n ≤ 18  → exhaustive search (smallest difference wins)
          • n  > 18 → greedy descending (largest values first that fit)
                       followed by 2-opt local refinement (add/remove/swap)
        Handles negative values too (refunds / reversals).
        """
        if rows_df.empty:
            return [], 0.0
        uids = rows_df["UID"].tolist()
        vals = [float(v) for v in rows_df["Total Tax"].tolist()]
        n = len(vals)

        # Single-entry exact match
        for i, v in enumerate(vals):
            if abs(v - target) <= tol:
                return [uids[i]], v

        # All-or-nothing exact
        total = sum(vals)
        if abs(total - target) <= tol:
            return uids[:], round(total, 2)

        # ── Small group: exhaustive subset-sum closest to target ──
        if n <= 18:
            # Start with EMPTY subset as the baseline candidate (covers the
            # case where every PR row is far above the target — picking nothing
            # is genuinely best).  Then evaluate every non-empty subset too.
            best_diff = abs(0 - target)
            best_set  = []
            best_sum  = 0.0
            # Also seed with "all" (so r=n is considered up front)
            if abs(total - target) < best_diff:
                best_diff = abs(total - target)
                best_set  = list(range(n))
                best_sum  = total
            for r in range(1, n):        # 1 … n-1  (n already considered above)
                for combo in itertools.combinations(range(n), r):
                    s = sum(vals[i] for i in combo)
                    d = abs(s - target)
                    if d < best_diff:
                        best_diff = d; best_set = list(combo); best_sum = s
                        if best_diff <= tol:
                            return [uids[i] for i in best_set], round(best_sum, 2)
            return [uids[i] for i in best_set], round(best_sum, 2)

        # ── Large group: greedy descending + 2-opt refinement ──
        sorted_idx = sorted(range(n), key=lambda i: -vals[i])
        picked = set()
        s = 0.0
        slack = max(tol, abs(target) * 0.005)
        for i in sorted_idx:
            v = vals[i]
            if v <= 0:
                # Negative value pulls sum down — only useful if currently overshooting
                if s > target:
                    picked.add(i); s += v
            else:
                if s + v <= target + slack:
                    picked.add(i); s += v

        # 2-opt local refinement: add / remove / swap to reduce |s - target|
        cur_diff = abs(s - target)
        for _ in range(80):
            improved = False
            # Try adding any unpicked element that improves things
            for i in range(n):
                if i in picked: continue
                ns = s + vals[i]; nd = abs(ns - target)
                if nd < cur_diff - 1e-6:
                    picked.add(i); s = ns; cur_diff = nd; improved = True; break
            if improved: continue
            # Try removing any picked element that improves things
            for i in list(picked):
                ns = s - vals[i]; nd = abs(ns - target)
                if nd < cur_diff - 1e-6:
                    picked.discard(i); s = ns; cur_diff = nd; improved = True; break
            if improved: continue
            # Try swapping one picked for one unpicked
            done = False
            for p in list(picked):
                for u in range(n):
                    if u in picked: continue
                    ns = s - vals[p] + vals[u]; nd = abs(ns - target)
                    if nd < cur_diff - 1e-6:
                        picked.discard(p); picked.add(u); s = ns; cur_diff = nd
                        improved = True; done = True; break
                if done: break
            if not improved: break

        return [uids[i] for i in picked], round(s, 2)

    def _identify_bank_rows(self, df, tag):
        target_names = {"is bank","bank","is_bank","isbank"}
        for col in df.columns:
            if str(col).strip().lower() in target_names:
                mask = df[col].astype(str).str.strip().str.upper().isin(["YES","Y","TRUE","1","BANK"])
                if mask.any():
                    self.log(f"    [BankID-{tag}] '{col}' column → {int(mask.sum())} rows marked as bank")
                    return df[mask].copy(), f"column:{col}"
                else:
                    self.log(f"    [BankID-{tag}] '{col}' column exists but no YES values found")
        if "Vendor Name" in df.columns and self.bank_keywords:
            kws = [re.escape(k.upper().strip()) for k in self.bank_keywords if k.strip()]
            if kws:
                pat = r"\b(" + "|".join(kws) + r")\b"
                vn = df["Vendor Name"].astype(str).str.upper()
                mask = vn.str.contains(pat, regex=True, na=False)
                excl_count = 0
                if self.bank_exclusions:
                    excl_kws = [re.escape(k.upper().strip()) for k in self.bank_exclusions if k.strip()]
                    if excl_kws:
                        excl_pat = r"\b(" + "|".join(excl_kws) + r")\b"
                        excl_mask = vn.str.contains(excl_pat, regex=True, na=False)
                        both = mask & excl_mask
                        excl_count = int(both.sum())
                        if excl_count:
                            self.log(f"    [BankID-{tag}] Excluded {excl_count} rows (bank+exclusion keyword)")
                            mask = mask & (~excl_mask)
                if mask.any():
                    self.log(f"    [BankID-{tag}] Vendor keyword match → {int(mask.sum())} rows (excluded: {excl_count})")
                    return df[mask].copy(), "vendor_keyword"
        self.log(f"    [BankID-{tag}] No bank signal found")
        return df.iloc[0:0].copy(), "no_signal"

    def _bank_match(self, pr, b2):
        """
        Iterative subset-sum bank matcher.

        For each (GSTIN + FY + Tax Head + State) group the engine loops
        until nothing more can be matched, trying three strategies per
        iteration:
          1. Full remaining balance  — |Σ PR_avail − Σ 2B_avail| ≤ tol
          2. Each PR entry  → find subset of 2B that sums to it
          3. Each 2B entry  → find subset of PR that sums to it

        A PAN-based second pass handles entries where GSTIN differs
        between PR and 2B for the same vendor.
        """
        self.log(f"  [BankMatch] {len(pr)} PR | {len(b2)} 2B unmatched")
        frames      = []
        matched_pr: set = set()
        matched_b2: set = set()

        # ── local constants ──────────────────────────────────────
        MAX_COMBO = 5
        MAX_ROWS  = 18   # skip subset search if group > this size
        MAX_ITER  = 60   # safety cap on iterations per group
        max_match_on = bool(self.feat.get("bank_max_match", True))
        if max_match_on:
            self.log(f"  [BankMatch] Max-Match ENABLED  "
                     f"(exhaust={self.bank_max_match_side}, "
                     f"variance threshold={self.bank_max_match_var_pct*100:.0f}%, "
                     f"min total ₹{self.bank_max_match_min_total})")
        else:
            self.log("  [BankMatch] Max-Match DISABLED — only exact subset-sum will run")

        # ── helpers ──────────────────────────────────────────────
        # Vendor-name normaliser used as the PRIMARY identifier in vendor-name-
        # based passes.  It absorbs corporate-suffix noise (Limited / Ltd / Pvt /
        # etc.), punctuation differences, and known bank aliases (e.g. "SBI"
        # collapses to "STATE BANK OF INDIA").  This lets entries with the same
        # vendor — but different PAN/GSTIN, or missing PAN — be grouped together.
        _SUFFIXES = (' LIMITED', ' LTD', ' PRIVATE', ' PVT',
                     ' COMPANY', ' INC', ' CORPORATION', ' CORP')
        def _norm_vendor(name):
            if pd.isna(name) or not str(name).strip():
                return ""
            s = str(name).upper()
            s = re.sub(r'[.,/&\-_()]', ' ', s)
            s = re.sub(r'\s+', ' ', s).strip()
            for suf in _SUFFIXES:
                while s.endswith(suf):
                    s = s[:-len(suf)].strip()
            # Bank aliases — extend as needed
            if 'STATE BANK OF INDIA' in s or s in ('SBI BANK', 'SBI', 'STATE BANK INDIA'):
                return 'STATE BANK OF INDIA'
            return s

        def with_th(df):
            if df.empty: return df
            d = df.copy()
            d["_TH"] = d.apply(self._th, axis=1)
            if "_FY" not in d.columns:
                d["_FY"] = d["Invoice Date"].apply(get_fy)
            if "State" not in d.columns: d["State"] = ""
            # Normalised vendor key — primary cross-PAN matching identifier.
            if "Vendor Name" in d.columns:
                d["_VendorKey"] = d["Vendor Name"].apply(_norm_vendor)
            else:
                d["_VendorKey"] = ""
            # Standardise State for grouping
            d["State"] = d["State"].astype(str).str.upper().str.strip()
            return d

        def find_subset(rows_df, target, tol):
            """Return list of index labels whose Total Tax sums to target,
            or None if not found, or 'TOO_MANY' if group too large."""
            vals = rows_df["Total Tax"].tolist()
            idx  = rows_df.index.tolist()
            if not vals: return None
            # single entry
            for i, v in enumerate(vals):
                if abs(v - target) <= tol: return [idx[i]]
            if len(vals) > MAX_ROWS: return "TOO_MANY"
            # combinations
            for n in range(2, min(MAX_COMBO, len(vals)) + 1):
                for combo in itertools.combinations(range(len(vals)), n):
                    s = sum(vals[i] for i in combo)
                    if abs(s - target) <= tol:
                        return [idx[i] for i in combo]
            return None

        def make_rows(df_side, side, uid, matched_on, kind):
            s = df_side.copy()
            s["Match Status"]    = "Bank Match"
            s["Matched On"]      = matched_on
            s["Common_UID"]      = uid
            s["_Side"]           = side
            s["Bank_Match_Kind"] = kind
            return s

        def process_group(pr_all, b2_all, label, do_max_match=True):
            """
            Iteratively match entries within one GSTIN/PAN group.
            Updates matched_pr / matched_b2 / frames in place.
            Returns dict of per-group stats.
            
            do_max_match: when False, runs only the exact subset-sum
            strategies (1-3) and skips the aggregate Max-Match (Strategy 4).
            Used so PASS 1 / PASS 2 only do exact matching — Max-Match
            runs once at the broadest scope (PASS 3), which lets it
            combine rows split across states or GSTINs.
            """
            lst = {"full": 0, "pr_to_b2": 0, "b2_to_pr": 0, "max_match": 0,
                   "max_skip": 0, "no_match": 0}

            for iteration in range(MAX_ITER):
                pr_avail = pr_all[~pr_all["UID"].isin(matched_pr)]
                b2_avail = b2_all[~b2_all["UID"].isin(matched_b2)]

                if pr_avail.empty or b2_avail.empty:
                    break

                pr_sum = round(pr_avail["Total Tax"].sum(), 2)
                b2_sum = round(b2_avail["Total Tax"].sum(), 2)
                found  = False

                # ── Strategy 1: full remaining balance ────────────
                if self._amt_ok(pr_sum, b2_sum):
                    uid = self._uid("BNK")
                    note = (f"{label} — Full Group "
                            f"(Σ PR=₹{pr_sum}, Σ 2B=₹{b2_sum})")
                    frames.append(make_rows(pr_avail, "PR", uid, note, "Full Group"))
                    frames.append(make_rows(b2_avail, "2B", uid, note, "Full Group"))
                    matched_pr.update(pr_avail["UID"].tolist())
                    matched_b2.update(b2_avail["UID"].tolist())
                    lst["full"] += 1
                    self.log(f"    ✓ FULL   {label}  "
                             f"PR={len(pr_avail)}×₹{pr_sum} = 2B={len(b2_avail)}×₹{b2_sum}")
                    break   # group exhausted

                # ── Strategy 2: each PR entry → find 2B subset ────
                if not found:
                    for _, pr_row in pr_avail.iterrows():
                        b2_rem = b2_avail[~b2_avail["UID"].isin(matched_b2)]
                        if b2_rem.empty: break
                        target = pr_row["Total Tax"]
                        if target == 0: continue
                        picked = find_subset(b2_rem, target, self.tol)
                        if picked and picked != "TOO_MANY":
                            picked_df = b2_avail.loc[picked]
                            uid  = self._uid("BNK")
                            note = (f"{label} — 1 PR→{len(picked)} 2B "
                                    f"(₹{target})")
                            frames.append(make_rows(
                                pr_row.to_frame().T.copy(), "PR", uid, note, "PR→2B Subset"))
                            frames.append(make_rows(
                                picked_df, "2B", uid, note, "PR→2B Subset"))
                            matched_pr.add(pr_row["UID"])
                            matched_b2.update(picked_df["UID"].tolist())
                            lst["pr_to_b2"] += 1
                            found = True
                            break   # restart while-loop with fresh avail sets

                # ── Strategy 3: each 2B entry → find PR subset ────
                if not found:
                    for _, b2_row in b2_avail.iterrows():
                        pr_rem = pr_avail[~pr_avail["UID"].isin(matched_pr)]
                        if pr_rem.empty: break
                        target = b2_row["Total Tax"]
                        if target == 0: continue
                        picked = find_subset(pr_rem, target, self.tol)
                        if picked and picked != "TOO_MANY":
                            picked_df = pr_avail.loc[picked]
                            uid  = self._uid("BNK")
                            note = (f"{label} — {len(picked)} PR→1 2B "
                                    f"(₹{target})")
                            frames.append(make_rows(
                                picked_df, "PR", uid, note, "PR Subset→2B"))
                            frames.append(make_rows(
                                b2_row.to_frame().T.copy(), "2B", uid, note, "PR Subset→2B"))
                            matched_pr.update(picked_df["UID"].tolist())
                            matched_b2.add(b2_row["UID"])
                            lst["b2_to_pr"] += 1
                            found = True
                            break

                if not found:
                    lst["no_match"] += 1
                    self.log(f"    ✗ NO MORE  {label}  "
                             f"PR_rem={len(pr_avail)} 2B_rem={len(b2_avail)}")
                    break   # nothing more matchable via exact strategies

            # ════════════════════════════════════════════════════════
            # ── Strategy 4: MAX MATCH — ALL-VS-ALL  (★ v5.5 user rule)
            #    Within each common group (both sides have residual rows),
            #    match EVERY remaining PR row + EVERY remaining 2B row as
            #    one bulk match.  Variance = PR_total - 2B_total is reported.
            #    Neither side keeps leftover within a common group — the
            #    user wants to see the discrepancy in the variance column,
            #    not as un-reconciled rows.
            #
            #    Legacy "best-fit subset" behaviour is preserved when
            #    bank_max_match_strategy == "BEST_FIT" (config opt-in).
            # ════════════════════════════════════════════════════════
            if max_match_on and do_max_match:
                pr_avail = pr_all[~pr_all["UID"].isin(matched_pr)]
                b2_avail = b2_all[~b2_all["UID"].isin(matched_b2)]
                if (not pr_avail.empty) and (not b2_avail.empty):
                    pr_sum = round(pr_avail["Total Tax"].sum(), 2)
                    b2_sum = round(b2_avail["Total Tax"].sum(), 2)

                    # Skip pure-noise groups
                    if abs(pr_sum) < self.bank_max_match_min_total and \
                       abs(b2_sum) < self.bank_max_match_min_total:
                        self.log(f"    ↷ MAX SKIP {label}  both totals "
                                 f"< ₹{self.bank_max_match_min_total} threshold")
                        lst["max_skip"] += 1
                    elif self.bank_max_match_strategy == "ALL_VS_ALL":
                        variance = round(pr_sum - b2_sum, 2)
                        uid  = self._uid("BNK")
                        kind = "Max Match"
                        note = (f"{label} — Max Match (All-vs-All) — "
                                f"All PR={len(pr_avail)}×₹{pr_sum} | "
                                f"All 2B={len(b2_avail)}×₹{b2_sum} "
                                f"(variance ₹{variance})")
                        frames.append(make_rows(pr_avail, "PR", uid, note, kind))
                        frames.append(make_rows(b2_avail, "2B", uid, note, kind))
                        matched_pr.update(pr_avail["UID"].tolist())
                        matched_b2.update(b2_avail["UID"].tolist())
                        lst["max_match"] += 1
                        self.log(
                            f"    ✓ MAX    {label}  All PR={len(pr_avail)}×₹{pr_sum} ↔ "
                            f"All 2B={len(b2_avail)}×₹{b2_sum}  (var ₹{variance})")
                    else:
                        # Legacy BEST_FIT behaviour (subset closest to target)
                        side_cfg = self.bank_max_match_side
                        if side_cfg == "2B":
                            small_df, small_total, small_side = b2_avail, b2_sum, "2B"
                            large_df, large_total, large_side = pr_avail, pr_sum, "PR"
                        elif side_cfg == "PR":
                            small_df, small_total, small_side = pr_avail, pr_sum, "PR"
                            large_df, large_total, large_side = b2_avail, b2_sum, "2B"
                        else:   # "SMALLER"
                            if abs(pr_sum) <= abs(b2_sum):
                                small_df, small_total, small_side = pr_avail, pr_sum, "PR"
                                large_df, large_total, large_side = b2_avail, b2_sum, "2B"
                            else:
                                small_df, small_total, small_side = b2_avail, b2_sum, "2B"
                                large_df, large_total, large_side = pr_avail, pr_sum, "PR"
                        if abs(small_total) < self.bank_max_match_min_total:
                            lst["max_skip"] += 1
                        else:
                            picked_uids, picked_sum = self._best_subset_to_target(
                                large_df, small_total, self.tol)
                            picked_df = (large_df[large_df["UID"].isin(picked_uids)]
                                         if picked_uids else large_df.iloc[0:0])
                            variance = round(picked_sum - small_total, 2)
                            uid  = self._uid("BNK")
                            kind = "Max Match" if not picked_df.empty else "Max Match (Side-Only)"
                            note = (f"{label} — Max Match (Best-Fit) — "
                                    f"All {small_side}={len(small_df)}×₹{small_total} | "
                                    f"{len(picked_df)}/{len(large_df)} {large_side}=₹{picked_sum} "
                                    f"(variance ₹{variance})")
                            if small_side == "2B":
                                if not picked_df.empty:
                                    frames.append(make_rows(picked_df, "PR", uid, note, kind))
                                frames.append(make_rows(small_df, "2B", uid, note, kind))
                                matched_b2.update(small_df["UID"].tolist())
                                if not picked_df.empty:
                                    matched_pr.update(picked_df["UID"].tolist())
                            else:
                                frames.append(make_rows(small_df, "PR", uid, note, kind))
                                if not picked_df.empty:
                                    frames.append(make_rows(picked_df, "2B", uid, note, kind))
                                matched_pr.update(small_df["UID"].tolist())
                                if not picked_df.empty:
                                    matched_b2.update(picked_df["UID"].tolist())
                            lst["max_match"] += 1

            return lst

        # ── bank identification ──────────────────────────────────
        pr = with_th(pr); b2 = with_th(b2)
        if pr.empty or b2.empty:
            self.log("  [BankMatch] One side empty → skipped")
            return pd.DataFrame(), pr, b2

        if self.bank_filter_on:
            self.log("  [BankMatch] Identifying bank entries (filter ON)…")
            pr_bank, _ = self._identify_bank_rows(pr, "PR")
            b2_bank, _ = self._identify_bank_rows(b2, "2B")
            if pr_bank.empty or b2_bank.empty:
                self.log("  [BankMatch] ⚠ No bank entries on one/both sides → skipped")
                self.log("    Tip: add 'Is Bank'=YES column, or check vendor keyword settings")
                for c in ("_TH",):
                    if c in pr.columns: pr = pr.drop(columns=c)
                    if c in b2.columns: b2 = b2.drop(columns=c)
                return pd.DataFrame(), pr, b2
            pr_full, b2_full = pr, b2
            pr, b2 = pr_bank, b2_bank
            self.log(f"  [BankMatch] Filtered → PR={len(pr)}  2B={len(b2)}")
        else:
            self.log("  [BankMatch] Filter OFF → matching ALL unmatched entries")
            pr_full, b2_full = pr, b2

        # ── common column prep ───────────────────────────────────
        gstin_keys = ["GSTIN", "_FY", "_TH", "State"]
        for c in gstin_keys:
            if c not in pr.columns: pr[c] = ""
            if c not in b2.columns: b2[c] = ""
            pr[c] = pr[c].fillna("").astype(str)
            b2[c] = b2[c].fillna("").astype(str)

        total_stats = {"full_group":0, "pr_to_b2":0, "b2_to_pr":0,
                       "max_match":0, "max_skip":0,
                       "no_match":0, "skip_unknown_th":0, "skip_empty":0}

        # ════════════════════════════════════════════════════════
        # PASS 1 — GSTIN-based groups
        # ════════════════════════════════════════════════════════
        pr_grps = pr.groupby(gstin_keys, dropna=False)
        b2_grps = b2.groupby(gstin_keys, dropna=False)
        common  = set(pr_grps.groups.keys()) & set(b2_grps.groups.keys())
        self.log(f"  [BankMatch-GSTIN] PR groups={len(pr_grps.groups)}  "
                 f"2B groups={len(b2_grps.groups)}  common={len(common)}")

        for key in common:
            if str(key[2]) == "UNKNOWN":
                total_stats["skip_unknown_th"] += 1; continue
            pr_all = pr_grps.get_group(key)
            b2_all = b2_grps.get_group(key)
            if (pr_all[~pr_all["UID"].isin(matched_pr)].empty or
                    b2_all[~b2_all["UID"].isin(matched_b2)].empty):
                total_stats["skip_empty"] += 1; continue
            gstin_s = (key[0][:10]+"…") if len(str(key[0]))>11 else key[0]
            label   = f"{gstin_s}|{key[1]}|{key[2]}|{key[3] or '∅'}"
            lst = process_group(pr_all, b2_all, label, do_max_match=False)
            total_stats["full_group"] += lst["full"]
            total_stats["pr_to_b2"]  += lst["pr_to_b2"]
            total_stats["b2_to_pr"]  += lst["b2_to_pr"]
            total_stats["max_match"] += lst["max_match"]
            total_stats["max_skip"]  += lst["max_skip"]
            total_stats["no_match"]  += lst["no_match"]

        # ════════════════════════════════════════════════════════
        # PASS 2 — PAN-based fallback for still-unmatched rows
        # (handles GSTIN format differences for the same vendor)
        # ════════════════════════════════════════════════════════
        has_pan = "Pan" in pr.columns and "Pan" in b2.columns
        if has_pan:
            # PASS 2 — VENDOR + state + FY + TH  (★ v5.6 — vendor-name based)
            # Uses normalized Vendor Name as primary identifier, so rows with
            # missing PAN, different PAN for the same vendor, or alias variants
            # (SBI / State Bank of India / SBI Bank Limited) are matched
            # together.  Falls back to PAN if vendor name is empty.
            vk_keys = ["_VendorKey", "_FY", "_TH", "State"]
            for c in vk_keys:
                if c not in pr.columns: pr[c] = ""
                if c not in b2.columns: b2[c] = ""
                pr[c] = pr[c].fillna("").astype(str)
                b2[c] = b2[c].fillna("").astype(str)

            pr_rem = pr[~pr["UID"].isin(matched_pr)]
            b2_rem = b2[~b2["UID"].isin(matched_b2)]

            if not pr_rem.empty and not b2_rem.empty:
                pan_pr_grp = pr_rem.groupby(vk_keys, dropna=False)
                pan_b2_grp = b2_rem.groupby(vk_keys, dropna=False)
                pan_common = (set(pan_pr_grp.groups.keys()) &
                              set(pan_b2_grp.groups.keys()))
                self.log(f"  [BankMatch-Vendor+State] common groups={len(pan_common)}")

                for key in pan_common:
                    if not str(key[0]).strip():       # blank vendor name → skip
                        continue
                    if str(key[2]) == "UNKNOWN": continue
                    pr_all = pan_pr_grp.get_group(key)
                    b2_all = pan_b2_grp.get_group(key)
                    if (pr_all[~pr_all["UID"].isin(matched_pr)].empty or
                            b2_all[~b2_all["UID"].isin(matched_b2)].empty):
                        continue
                    vk_s = str(key[0])[:18]
                    label = f"V:{vk_s}|{key[1]}|{key[2]}|{key[3] or '∅'}"
                    lst = process_group(pr_all, b2_all, label, do_max_match=False)
                    total_stats["full_group"] += lst["full"]
                    total_stats["pr_to_b2"]  += lst["pr_to_b2"]
                    total_stats["b2_to_pr"]  += lst["b2_to_pr"]
                    total_stats["max_match"] += lst["max_match"]
                    total_stats["max_skip"]  += lst["max_skip"]

        # ════════════════════════════════════════════════════════
        # PASS 3 — PAN + FY + Tax Head (NO State)  ★ NEW in v5.3
        # Catches state-mismatch cases (e.g. same GSTIN where PR books
        # show DELHI and 2B shows KARNATAKA for the same supplier).
        # Tax head still respected (per user rule "head matched").
        # ════════════════════════════════════════════════════════
        # PASS 3 — VENDOR + FY + Tax Head (no State)  ★ v5.6 vendor-key
        if has_pan and max_match_on:
            vk_no_state = ["_VendorKey", "_FY", "_TH"]
            pr_rem = pr[~pr["UID"].isin(matched_pr)]
            b2_rem = b2[~b2["UID"].isin(matched_b2)]
            if not pr_rem.empty and not b2_rem.empty:
                ns_pr = pr_rem.groupby(vk_no_state, dropna=False)
                ns_b2 = b2_rem.groupby(vk_no_state, dropna=False)
                ns_common = set(ns_pr.groups.keys()) & set(ns_b2.groups.keys())
                self.log(f"  [BankMatch-Vendor-NoState] common groups={len(ns_common)}")

                for key in ns_common:
                    if not str(key[0]).strip(): continue
                    if str(key[2]) == "UNKNOWN": continue
                    pr_all = ns_pr.get_group(key)
                    b2_all = ns_b2.get_group(key)
                    if (pr_all[~pr_all["UID"].isin(matched_pr)].empty or
                            b2_all[~b2_all["UID"].isin(matched_b2)].empty):
                        continue
                    vk_s = str(key[0])[:18]
                    label = f"V-NoState:{vk_s}|{key[1]}|{key[2]}"
                    lst = process_group(pr_all, b2_all, label, do_max_match=False)
                    total_stats["full_group"] += lst["full"]
                    total_stats["pr_to_b2"]  += lst["pr_to_b2"]
                    total_stats["b2_to_pr"]  += lst["b2_to_pr"]
                    total_stats["max_match"] += lst["max_match"]
                    total_stats["max_skip"]  += lst["max_skip"]

        # ════════════════════════════════════════════════════════
        # PASS 4 — PAN + FY only  (no Tax Head, no State)  ★ NEW in v5.5
        # Final relax-head fallback: catches cases where 2B has, say,
        # IGST entries but PR booked the same vendor + FY as CGST+SGST
        # (mixed-classification bank charges, tax-engine quirks).
        # Off by default conceptually since it ignores head, but
        # defaulted ON because real-world bank data has these quirks
        # and the user explicitly wants minimal residual.  Disable via
        # config `bank_match_relax_head: false`.
        # ════════════════════════════════════════════════════════
        # PASS 4 — VENDOR + FY only  (no Tax Head, no State)  ★ v5.6
        # Final aggressive pass — broadest reasonable scope.  Catches
        # cross-head residuals (2B has IGST while PR booked CGST+SGST
        # for same vendor/FY) plus any remaining state-tag drift.
        if has_pan and max_match_on and self.bank_match_relax_head:
            vk_fy = ["_VendorKey", "_FY"]
            pr_rem = pr[~pr["UID"].isin(matched_pr)]
            b2_rem = b2[~b2["UID"].isin(matched_b2)]
            if not pr_rem.empty and not b2_rem.empty:
                pf_pr = pr_rem.groupby(vk_fy, dropna=False)
                pf_b2 = b2_rem.groupby(vk_fy, dropna=False)
                pf_common = set(pf_pr.groups.keys()) & set(pf_b2.groups.keys())
                self.log(f"  [BankMatch-Vendor-NoHead] common groups={len(pf_common)}")
                for key in pf_common:
                    if not str(key[0]).strip(): continue
                    pr_all = pf_pr.get_group(key)
                    b2_all = pf_b2.get_group(key)
                    if (pr_all[~pr_all["UID"].isin(matched_pr)].empty or
                            b2_all[~b2_all["UID"].isin(matched_b2)].empty):
                        continue
                    vk_s = str(key[0])[:18]
                    label = f"V-NoHead:{vk_s}|{key[1]}"
                    lst = process_group(pr_all, b2_all, label)
                    total_stats["full_group"] += lst["full"]
                    total_stats["pr_to_b2"]  += lst["pr_to_b2"]
                    total_stats["b2_to_pr"]  += lst["b2_to_pr"]
                    total_stats["max_match"] += lst["max_match"]
                    total_stats["max_skip"]  += lst["max_skip"]

        self.log(
            f"  [BankMatch] Results: Full={total_stats['full_group']}  "
            f"PR→2B={total_stats['pr_to_b2']}  2B→PR={total_stats['b2_to_pr']}  "
            f"MaxMatch={total_stats['max_match']}  "
            f"MaxSkip={total_stats['max_skip']}  "
            f"NoMatch={total_stats['no_match']}  "
            f"Skip(UnknownTH)={total_stats['skip_unknown_th']}"
        )

        combined = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
        if not combined.empty:
            combined = enrich_side_df(combined, self.tol)

        pr_out = pr_full[~pr_full["UID"].isin(matched_pr)]
        b2_out = b2_full[~b2_full["UID"].isin(matched_b2)]
        for c in ("_TH",):
            if c in pr_out.columns: pr_out = pr_out.drop(columns=c)
            if c in b2_out.columns: b2_out = b2_out.drop(columns=c)

        self.log(f"  [BankMatch] {len(matched_pr)} PR + {len(matched_b2)} 2B matched")
        return combined, pr_out, b2_out

    def _vendor_fuzzy(self,pr,b2):
        if not FUZZY_OK: self.log("  [VendorFuzzy] rapidfuzz missing → skipped"); return pd.DataFrame(),pr,b2
        if "Vendor Name" not in pr.columns or "Vendor Name" not in b2.columns:
            self.log("  [VendorFuzzy] Vendor Name col missing → skipped"); return pd.DataFrame(),pr,b2
        self.log(f"  [VendorFuzzy] {len(pr)} PR rows …")
        frames,mp,mb=[],set(),set()
        for _,row in pr.iterrows():
            if row["UID"] in mp: continue
            vn=str(row.get("Vendor Name","")).strip()
            if not vn or vn in("nan","","-"): continue
            pt=row.get("Total Tax",0)
            sub=b2[~b2["UID"].isin(mb)].copy()
            sub=sub[sub["Total Tax"].apply(lambda t:self._amt_ok(pt,t))]
            if sub.empty: continue
            res=rfp.extractOne(vn,sub["Vendor Name"].astype(str),scorer=fuzz.token_sort_ratio)
            if res is None or res[1]<self.vthr: continue
            mr=sub[sub["Vendor Name"].astype(str)==res[0]]
            if mr.empty: continue
            mr=mr.iloc[0]; uid=self._uid("VND")
            var=round(pt-mr.get("Total Tax",0),2)
            for part,side in [(row.to_frame().T.copy(),"PR"),(mr.to_frame().T.copy(),"2B")]:
                part["Match Status"]="Vendor Fuzzy Match"; part["Matched On"]=f"Vendor Fuzzy (score={res[1]})"
                part["Common_UID"]=uid; part["_Side"]=side; part["Fuzzy_Score"]=res[1]
                part["Amount_Variance_(PR-2B)"]=var
                part["Excess_In"]=("Balanced" if abs(var)<=self.tol else ("Excess in PR" if var>0 else "Excess in 2B"))
                frames.append(part)
            mp.add(row["UID"]); mb.add(mr["UID"])
        combined=pd.concat(frames,ignore_index=True) if frames else pd.DataFrame()
        if not combined.empty: combined = enrich_side_df(combined, self.tol)
        pr=pr[~pr["UID"].isin(mp)]; b2=b2[~b2["UID"].isin(mb)]
        self.log(f"  [VendorFuzzy] {len(mp)} pairs"); return combined,pr,b2

    def _unmatched(self,pr,b2):
        if "Pan" not in pr.columns or "Pan" not in b2.columns: return pd.DataFrame()
        comp=pd.merge(pr,b2,on=["Pan"],how="outer",suffixes=("_PR","_2B"),indicator=True)
        def rm(r):
            m=r.get("_merge","")
            if m=="left_only":  return "Only in PR"
            if m=="right_only": return "Only in GSTR2B"
            if str(r.get("Invoice Number_PR",""))!=str(r.get("Invoice Number_2B","")): return "Invoice mismatch under PAN"
            if str(r.get("GSTIN_PR",""))!=str(r.get("GSTIN_2B","")): return "Different GSTIN for same PAN"
            return "Mismatch"
        comp["Match Status"]=comp.apply(rm,axis=1)
        want=["Pan","GSTIN_PR","GSTIN_2B","Vendor Name_PR","Vendor Name_2B",
              "Invoice Number_PR","Invoice Number_2B","Total Tax_PR","Total Tax_2B","Match Status"]
        return comp[[c for c in want if c in comp.columns]]

    def run(self,pr_file,gstr2b_file,output_file,client_name="",email_cfg=None):
        f=self.feat
        self.client_name = (client_name or "").strip()
        self.log("═"*56+"\n  GST RECONCILIATION PRO  v5\n"+"═"*56)
        if self.client_name:
            self.log(f"  Client: {self.client_name}")
        self.log("\n📂 Loading …")
        pr=self._cc(pd.read_excel(pr_file,    header=2))
        b2=self._cc(pd.read_excel(gstr2b_file,header=2))
        pr["Source"]="PR"; b2["Source"]="GSTR2B"
        pr["UID"]="PR_"+pr.index.astype(str); b2["UID"]="2B_"+b2.index.astype(str)

        def get_ko(df):
            if "Knock Off" in df.columns:
                return df[df["Knock Off"].astype(str).str.upper()=="YES"].copy()
            return pd.DataFrame(columns=df.columns)
        ko_pr=get_ko(pr); ko_b2=get_ko(b2)
        if f.get("user_knockoff",True):
            pr=pr[~pr["UID"].isin(ko_pr["UID"])]; b2=b2[~b2["UID"].isin(ko_b2["UID"])]
            self.log(f"  User knockoffs → PR:{len(ko_pr)}  2B:{len(ko_b2)}")
        self.log(f"  PR={len(pr)} | 2B={len(b2)}")
        self.log("\n🧹 Cleaning + deriving columns …")
        pr=self._clean(pr); b2=self._clean(b2)

        if f.get("pre_group",False):
            self.log("\n🔄 Step 0 — Pre-Group Line Items")
            pr=self._pre_group(pr,"PR"); b2=self._pre_group(b2,"2B")

        t_pr=len(pr)+len(ko_pr); t_b2=len(b2)+len(ko_b2)

        pr_ko=pd.DataFrame()
        if f.get("pr_knockout",True) and self.pr_ko:
            self.log("\n🔄 Step 1 — PR Knockout")
            pr_ko,pr=self._knockout(pr,self.pr_ko,"PR",50000); t_pr+=len(pr_ko)
        else: self.log("\n⏭ Step 1 — PR Knockout DISABLED")

        grp=pd.DataFrame()
        if f.get("grouped_match",True):
            self.log("\n🔄 Step 2 — Grouped Match"); grp,pr,b2=self._grouped(pr,b2)
        else: self.log("\n⏭ Step 2 — Grouped DISABLED")

        rule=pd.DataFrame()
        if f.get("rule_match",True):
            self.log("\n🔄 Step 3 — Rule Match"); rule,pr,b2=self._rules(pr,b2)
        else: self.log("\n⏭ Step 3 — Rule Match DISABLED")

        fuz=pd.DataFrame()
        if f.get("fuzzy_invoice",True):
            self.log("\n🔄 Step 4 — Fuzzy Invoice"); fuz,pr,b2=self._fuzzy(pr,b2)
        else: self.log("\n⏭ Step 4 — Fuzzy DISABLED")

        vnd=pd.DataFrame()
        if f.get("vendor_fuzzy",False):
            self.log("\n🔄 Step 5 — Vendor Fuzzy"); vnd,pr,b2=self._vendor_fuzzy(pr,b2)
        else: self.log("\n⏭ Step 5 — Vendor Fuzzy DISABLED")

        bnk=pd.DataFrame()
        if f.get("bank_match",False):
            self.log("\n🔄 Step 6 — Bank-Entry Match (iterative subset-sum + Max Match)")
            bnk,pr,b2=self._bank_match(pr,b2)
        else:
            self.log("\n⏭ Step 6 — Bank Match DISABLED")

        b2_ko=pd.DataFrame()
        if f.get("2b_knockout",True) and self.b2_ko:
            self.log("\n🔄 Step 7 — 2B Knockout")
            b2_ko,b2=self._knockout(b2,self.b2_ko,"2B",60000); t_b2+=len(b2_ko)
        else: self.log("\n⏭ Step 7 — 2B Knockout DISABLED")

        grp_pr=len(grp[grp["_Side"]=="PR"]) if not grp.empty and "_Side" in grp.columns else 0
        grp_b2=len(grp[grp["_Side"]=="2B"]) if not grp.empty and "_Side" in grp.columns else 0
        fp=len(fuz[fuz["_Side"]=="PR"])  if not fuz.empty and "_Side" in fuz.columns else 0
        vp=len(vnd[vnd["_Side"]=="PR"])  if not vnd.empty and "_Side" in vnd.columns else 0
        bp=len(bnk[bnk["_Side"]=="PR"])  if not bnk.empty and "_Side" in bnk.columns else 0
        bb=len(bnk[bnk["_Side"]=="2B"])  if not bnk.empty and "_Side" in bnk.columns else 0
        # Max-Match breakdown for dashboard
        if not bnk.empty and "Bank_Match_Kind" in bnk.columns:
            mm_pr = len(bnk[(bnk["_Side"]=="PR") & (bnk["Bank_Match_Kind"].str.startswith("Max Match", na=False))])
            mm_b2 = len(bnk[(bnk["_Side"]=="2B") & (bnk["Bank_Match_Kind"].str.startswith("Max Match", na=False))])
            so_b2 = len(bnk[(bnk["_Side"]=="2B") & (bnk["Bank_Match_Kind"]=="Max Match (Side-Only)")])
            ex_pr = bp - mm_pr; ex_b2 = bb - mm_b2
        else:
            mm_pr = mm_b2 = so_b2 = ex_pr = ex_b2 = 0
        t_pr+=grp_pr+len(rule)+fp+vp+bp+len(pr); t_b2+=grp_b2+len(rule)+fp+vp+bb+len(b2)
        dash={"Client":self.client_name or "—",
              "Total PR Entries":t_pr,"Total GSTR-2B Entries":t_b2,"——":"",
              "Grouped Matches (rows)":f"{grp_pr} PR + {grp_b2} 2B","Rule-Based Matches":len(rule),
              "Fuzzy Matches":fp,"Vendor Fuzzy Matches":vp,
              "Bank-Entry Matches (total)":f"{bp} PR + {bb} 2B",
              "  • Bank Exact Subset-Sum":f"{ex_pr} PR + {ex_b2} 2B",
              "  • Bank MAX MATCH (paired) ★":f"{mm_pr} PR + {mm_b2-so_b2} 2B",
              "  • Bank MAX MATCH (2B side-only)":f"0 PR + {so_b2} 2B",
              "PR Internal Knockout (rows)":len(pr_ko),"2B Internal Knockout (rows)":len(b2_ko),
              "User Knockoff PR":len(ko_pr),"User Knockoff 2B":len(ko_b2),
              "———":"","Only in PR":len(pr),"Only in GSTR-2B":len(b2)}
        dash_df=pd.DataFrame(list(dash.items()),columns=["Metric","Value"])
        unm=self._unmatched(pr,b2)
        all_m=pd.concat([df for df in [grp,rule,fuz,vnd,bnk] if not df.empty],ignore_index=True)

        self.log("\n💾 Writing Excel …")
        with pd.ExcelWriter(output_file,engine="openpyxl") as wr:
            if not all_m.empty:  all_m.to_excel(wr,sheet_name="All_Matched",     index=False)
            if not grp.empty:    grp.to_excel(wr,  sheet_name="Grouped_Matched",  index=False)
            if not rule.empty:   rule.to_excel(wr, sheet_name="Rule_Matched",     index=False)
            if not fuz.empty:    fuz.to_excel(wr,  sheet_name="Fuzzy_Matched",    index=False)
            if not vnd.empty:    vnd.to_excel(wr,  sheet_name="Vendor_Fuzzy",     index=False)
            if not bnk.empty:    bnk.to_excel(wr,  sheet_name="Bank_Matched",     index=False)
            if not pr_ko.empty:  pr_ko.to_excel(wr,sheet_name="PR_Knockout",      index=False)
            if not b2_ko.empty:  b2_ko.to_excel(wr,sheet_name="2B_Knockout",      index=False)
            pr.to_excel(wr,sheet_name="Only_in_PR",      index=False)
            b2.to_excel(wr,sheet_name="Only_in_GSTR2B",  index=False)
            if not ko_pr.empty: ko_pr.to_excel(wr,sheet_name="KnockOff_PR",index=False)
            if not ko_b2.empty: ko_b2.to_excel(wr,sheet_name="KnockOff_2B",index=False)
            if not unm.empty:   unm.to_excel(wr,sheet_name="Unmatched_Summary",index=False)
            dash_df.to_excel(wr,sheet_name="Dashboard",index=False)
        if OPENPYXL_OK: self._fmt_wb(output_file)
        self.log(f"\n✅ Saved → {output_file}")

        if email_cfg and email_cfg.get("enabled") and email_cfg.get("to"):
            self.log(f"\n📧 Emailing to {email_cfg['to']} …")
            subj = email_cfg.get("subject") or f"GST Reconciliation — {self.client_name or 'Report'}"
            body = email_cfg.get("body") or (
                f"Hello,\n\nPlease find attached the GST reconciliation report"
                f"{(' for ' + self.client_name) if self.client_name else ''}.\n\n"
                f"Generated: {datetime.datetime.now().strftime('%d %b %Y %H:%M')}\n\n"
                f"Summary:\n" + "\n".join(f"  • {k}: {v}" for k,v in dash.items() if v != "")
            )
            mode = email_cfg.get("mode","outlook")
            if mode == "smtp":
                ok, msg = send_email_smtp(email_cfg.get("smtp_host"), email_cfg.get("smtp_port",587),
                                          email_cfg.get("smtp_user"), email_cfg.get("smtp_pass"),
                                          email_cfg["to"], subj, body, output_file)
            else:
                ok, msg = send_email_outlook(email_cfg["to"], subj, body, output_file)
            self.log(("  ✅ " if ok else "  ❌ ") + msg)

        self.log("\n"+"═"*56)
        for k,v in dash.items(): self.log(f"  {k:<44} {v}")
        return dash

    def _fmt_wb(self,path):
        cn = getattr(self, "client_name", "") or ""
        wb=load_workbook(path)
        for sn in wb.sheetnames:
            ws=wb[sn]
            if sn=="Dashboard": _fmt_dashboard(ws, cn)
            else: _fmt_sheet(ws, status_col="Match Status", client_name=cn)
        clr={"All_Matched":"059669","Grouped_Matched":"4F46E5","Rule_Matched":"059669",
             "Fuzzy_Matched":"06B6D4","Vendor_Fuzzy":"EC4899","Bank_Matched":"D97706",
             "PR_Knockout":"6B7280","2B_Knockout":"6B7280",
             "Only_in_PR":"DC2626","Only_in_GSTR2B":"4F46E5",
             "Unmatched_Summary":"D97706","Dashboard":"1E1B4B"}
        for sn,c in clr.items():
            if sn in wb.sheetnames: wb[sn].sheet_properties.tabColor=c
        if "Dashboard" in wb.sheetnames:
            wb.move_sheet("Dashboard",offset=-wb.sheetnames.index("Dashboard"))
        wb.save(path)


