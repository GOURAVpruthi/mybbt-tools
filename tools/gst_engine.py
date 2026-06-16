"""
GST Tools Suite
===============
File Collector  +  PDF Compressor

Run by double-clicking this file, or:
    python gst_tools_suite.py
"""

# ── auto-install missing packages BEFORE any other import ───────
import subprocess, sys, os

def _install(pkg):
    try:
        subprocess.run([sys.executable, "-m", "pip", "install", pkg, "--quiet"],
                       capture_output=True, timeout=240)
    except Exception:
        pass

# Try each package. pikepdf is OPTIONAL — app works without it.
for pip_name, import_name in [("pillow", "PIL"),
                              ("PyMuPDF", "fitz"),
                              ("pikepdf", "pikepdf"),
                              ("openpyxl", "openpyxl"),
                              ("pdfplumber", "pdfplumber")]:
    try:
        __import__(import_name)
    except Exception:
        _install(pip_name)

# ── now import what we can, with fallbacks ──────────────────────
import re
import io
import shutil
import threading
import traceback

try:
    import fitz
    HAS_PYMUPDF = True
except Exception:
    HAS_PYMUPDF = False

try:
    from PIL import Image
    HAS_PILLOW = True
except Exception:
    HAS_PILLOW = False

try:
    import pikepdf
    HAS_PIKEPDF = True
except Exception:
    HAS_PIKEPDF = False

try:
    import openpyxl as _openpyxl_check
    HAS_OPENPYXL = True
except Exception:
    HAS_OPENPYXL = False

try:
    import pdfplumber as _pdfplumber_check
    HAS_PDFPLUMBER = True
except Exception:
    HAS_PDFPLUMBER = False


# ════════════════════════════════════════════════════════════════
#  PALETTE
# ════════════════════════════════════════════════════════════════
BG       = "#1e1e2e"
CARD     = "#2a2a3e"
ACCENT   = "#6c63ff"
ACCENT2  = "#a78bfa"
TEXT     = "#e2e8f0"
SUBTEXT  = "#94a3b8"
SUCCESS  = "#22c55e"
WARNING  = "#f59e0b"
ERROR    = "#ef4444"
ENTRY_BG = "#16213e"
BORDER   = "#3d3d5c"


# ════════════════════════════════════════════════════════════════
#  HELPERS
# ════════════════════════════════════════════════════════════════
_SEPARATORS_RE = re.compile(r"[\s\-_\.]+")

def _normalize_for_match(s):
    return _SEPARATORS_RE.sub("", s.lower())

def _keyword_matches(filename, keyword):
    return _normalize_for_match(keyword) in _normalize_for_match(filename)

EXT_PRESETS = {
    "Any (all formats)":      None,
    "PDF (.pdf)":             (".pdf",),
    "Excel (.xlsx, .xls)":    (".xlsx", ".xls", ".xlsm", ".xlsb"),
    "CSV (.csv)":             (".csv",),
    "JSON (.json)":           (".json",),
    "Word (.docx, .doc)":     (".docx", ".doc"),
    "Text (.txt)":            (".txt",),
    "Custom...":              "CUSTOM",
}

def _ext_matches(filename, ext_tuple):
    if ext_tuple is None: return True
    fn = filename.lower()
    return any(fn.endswith(e) for e in ext_tuple)


# ════════════════════════════════════════════════════════════════
#  PDF COMPRESSION ENGINES
# ════════════════════════════════════════════════════════════════
def compress_lossless(in_path, out_path):
    doc = fitz.open(in_path)
    try:
        doc.save(out_path, garbage=4, deflate=True, deflate_images=True,
                 deflate_fonts=True, clean=True)
    finally:
        doc.close()


def compress_images_preserve_text(in_path, out_path, jpeg_quality, downscale=1.0):
    if not HAS_PIKEPDF or not HAS_PILLOW:
        raise RuntimeError("pikepdf or pillow not available")
    pdf = pikepdf.Pdf.open(in_path)
    try:
        for page in pdf.pages:
            try: images = page.images
            except Exception: continue
            for name in list(images.keys()):
                raw_image = images[name]
                try:
                    pdfimg = pikepdf.PdfImage(raw_image)
                    pil = pdfimg.as_pil_image()
                    if pil.mode in ("RGBA", "P", "LA"):
                        pil = pil.convert("RGB")
                    elif pil.mode == "CMYK":
                        pil = pil.convert("RGB")
                    if downscale < 1.0:
                        w, h = pil.size
                        pil = pil.resize((max(1, int(w * downscale)),
                                          max(1, int(h * downscale))),
                                         Image.LANCZOS)
                    buf = io.BytesIO()
                    pil.save(buf, format="JPEG", quality=jpeg_quality, optimize=True)
                    new_bytes = buf.getvalue()
                    try: raw_size = len(bytes(raw_image.read_raw_bytes()))
                    except Exception: raw_size = 10**9
                    if len(new_bytes) >= raw_size: continue
                    cs = pikepdf.Name("/DeviceGray") if pil.mode == "L" \
                         else pikepdf.Name("/DeviceRGB")
                    raw_image.write(new_bytes, filter=pikepdf.Name("/DCTDecode"))
                    raw_image.Width = pil.width
                    raw_image.Height = pil.height
                    raw_image.BitsPerComponent = 8
                    raw_image.ColorSpace = cs
                    for k in ("/DecodeParms", "/Decode"):
                        if k in raw_image: del raw_image[k]
                except Exception: continue
        pdf.save(out_path, compress_streams=True,
                 object_stream_mode=pikepdf.ObjectStreamMode.generate,
                 recompress_flate=True, linearize=False)
    finally:
        pdf.close()


def compress_rasterize(in_path, out_path, dpi, jpeg_quality):
    doc = fitz.open(in_path)
    new_doc = fitz.open()
    try:
        for page in doc:
            pix = page.get_pixmap(dpi=dpi, alpha=False)
            img_bytes = pix.tobytes("jpeg", jpg_quality=jpeg_quality)
            new_page = new_doc.new_page(width=page.rect.width, height=page.rect.height)
            new_page.insert_image(new_page.rect, stream=img_bytes)
        new_doc.save(out_path, garbage=4, deflate=True)
    finally:
        new_doc.close()
        doc.close()


def smart_compress(in_path, out_path, target_mb, allow_text_loss, log):
    # SAFETY: never let output be the same path as input
    if os.path.abspath(in_path) == os.path.abspath(out_path):
        raise ValueError(
            f"Input and output paths are the same:\n  {in_path}\n"
            "Pick a different output folder, or this tool will refuse "
            "to overwrite the source file.")

    target_bytes = int(target_mb * 1024 * 1024)
    orig_bytes = os.path.getsize(in_path)

    if orig_bytes <= target_bytes:
        shutil.copy2(in_path, out_path)
        log(f"   ℹ  Already ≤ target — copied as-is.", "dim")
        return orig_bytes, "already optimal", True

    best_size, best_strategy = orig_bytes, "original"
    try:
        compress_lossless(in_path, out_path)
        sz = os.path.getsize(out_path)
        pct = (1 - sz / orig_bytes) * 100
        if sz <= target_bytes:
            log(f"   ✓  Lossless: {sz/1024/1024:.2f} MB  (-{pct:.0f}%)  ✅ Target hit", "ok")
            return sz, "lossless", True
        log(f"   ↺  Lossless: {sz/1024/1024:.2f} MB  (-{pct:.0f}%), > target", "ren")
        best_size, best_strategy = sz, "lossless"
    except Exception as e:
        log(f"   ⚠  Lossless step failed: {e}", "err")
        # Fall back to copy of original as starting baseline
        try:
            shutil.copy2(in_path, out_path)
        except Exception as e2:
            log(f"   ⚠  Could not even copy original: {e2}", "err")
            raise

    if HAS_PIKEPDF and HAS_PILLOW:
        for q, scale in [(80, 1.0), (70, 1.0), (60, 0.9),
                         (50, 0.8), (40, 0.7), (30, 0.6)]:
            try:
                tmp = out_path + ".t2.tmp.pdf"
                compress_images_preserve_text(in_path, tmp, q, scale)
                sz = os.path.getsize(tmp)
                if sz < best_size:
                    shutil.move(tmp, out_path)
                    best_size = sz
                    best_strategy = f"images q={q} scale={scale:.2f}"
                else:
                    try: os.remove(tmp)
                    except: pass
                if sz <= target_bytes:
                    pct = (1 - sz / orig_bytes) * 100
                    log(f"   ✓  Images q={q} s={scale}: "
                        f"{sz/1024/1024:.2f} MB (-{pct:.0f}%) ✅", "ok")
                    return sz, best_strategy, True
            except Exception:
                continue
    else:
        log("   ⚠  pikepdf not available — image-recompress tier skipped.", "dim")

    if not allow_text_loss:
        log(f"   ⚠  Target not reached. Best: {best_size/1024/1024:.2f} MB "
            f"({best_strategy}). Switch to 'Allow text loss' to go smaller.", "dim")
        return best_size, best_strategy, False

    for dpi, q in [(200, 80), (150, 75), (120, 70), (100, 60),
                   (85, 55), (72, 50), (60, 40), (50, 30)]:
        try:
            tmp = out_path + ".t3.tmp.pdf"
            compress_rasterize(in_path, tmp, dpi, q)
            sz = os.path.getsize(tmp)
            if sz < best_size:
                shutil.move(tmp, out_path)
                best_size = sz
                best_strategy = f"rasterized @ {dpi}dpi q{q} (TEXT LOST)"
            else:
                try: os.remove(tmp)
                except: pass
            if sz <= target_bytes:
                pct = (1 - sz / orig_bytes) * 100
                log(f"   ⚠  Rasterize @ {dpi}dpi q{q}: "
                    f"{sz/1024/1024:.2f} MB ⚠ TEXT LOST", "ren")
                return sz, best_strategy, True
        except Exception:
            continue

    for sfx in (".t2.tmp.pdf", ".t3.tmp.pdf"):
        p = out_path + sfx
        if os.path.exists(p):
            try: os.remove(p)
            except: pass
    log(f"   ⚠  Target unreachable. Best: {best_size/1024/1024:.2f} MB", "ren")
    return best_size, best_strategy, False


# ════════════════════════════════════════════════════════════════
#  GSTR-1 / GSTR-3B  EXTRACTOR ENGINES  (embedded)
# ════════════════════════════════════════════════════════════════
# The full GSTR-1 and GSTR-3B PDF-to-Excel extractor engines are
# embedded here as base64-encoded source.  They run in their own
# isolated namespaces so their constants (STATE_CODES, NUM_FMT,
# etc.) don't collide with the rest of the suite.

import base64 as _base64

_GSTR1_ENGINE_B64 = (
    "U1RBVEVfQ09ERVMgPSB7CiAgICAiMDEiOiAiSmFtbXUgJiBLYXNobWlyIiwgIjAyIjogIkhpbWFj"
    "aGFsIFByYWRlc2giLCAiMDMiOiAiUHVuamFiIiwKICAgICIwNCI6ICJDaGFuZGlnYXJoIiwgIjA1"
    "IjogIlV0dGFyYWtoYW5kIiwgIjA2IjogIkhhcnlhbmEiLCAiMDciOiAiRGVsaGkiLAogICAgIjA4"
    "IjogIlJhamFzdGhhbiIsICIwOSI6ICJVdHRhciBQcmFkZXNoIiwgIjEwIjogIkJpaGFyIiwgIjEx"
    "IjogIlNpa2tpbSIsCiAgICAiMTIiOiAiQXJ1bmFjaGFsIFByYWRlc2giLCAiMTMiOiAiTmFnYWxh"
    "bmQiLCAiMTQiOiAiTWFuaXB1ciIsICIxNSI6ICJNaXpvcmFtIiwKICAgICIxNiI6ICJUcmlwdXJh"
    "IiwgIjE3IjogIk1lZ2hhbGF5YSIsICIxOCI6ICJBc3NhbSIsICIxOSI6ICJXZXN0IEJlbmdhbCIs"
    "CiAgICAiMjAiOiAiSmhhcmtoYW5kIiwgIjIxIjogIk9kaXNoYSIsICIyMiI6ICJDaGhhdHRpc2dh"
    "cmgiLCAiMjMiOiAiTWFkaHlhIFByYWRlc2giLAogICAgIjI0IjogIkd1amFyYXQiLCAiMjUiOiAi"
    "RGFtYW4gJiBEaXUiLCAiMjYiOiAiRGFkcmEgJiBOYWdhciBIYXZlbGkgYW5kIERhbWFuICYgRGl1"
    "IiwKICAgICIyNyI6ICJNYWhhcmFzaHRyYSIsICIyOCI6ICJBbmRocmEgUHJhZGVzaCAoT2xkKSIs"
    "ICIyOSI6ICJLYXJuYXRha2EiLAogICAgIjMwIjogIkdvYSIsICIzMSI6ICJMYWtzaGFkd2VlcCIs"
    "ICIzMiI6ICJLZXJhbGEiLCAiMzMiOiAiVGFtaWwgTmFkdSIsCiAgICAiMzQiOiAiUHVkdWNoZXJy"
    "eSIsICIzNSI6ICJBbmRhbWFuICYgTmljb2JhciBJc2xhbmRzIiwgIjM2IjogIlRlbGFuZ2FuYSIs"
    "CiAgICAiMzciOiAiQW5kaHJhIFByYWRlc2giLCAiMzgiOiAiTGFkYWtoIiwgIjk3IjogIk90aGVy"
    "IFRlcnJpdG9yeSIsICI5OSI6ICJDZW50cmUgSnVyaXNkaWN0aW9uIiwKfQoKTU9OVEhfQUJCUiA9"
    "IHsKICAgICJKYW51YXJ5IjogIkphbiIsICJGZWJydWFyeSI6ICJGZWIiLCAiTWFyY2giOiAiTWFy"
    "IiwgIkFwcmlsIjogIkFwciIsCiAgICAiTWF5IjogICAgICJNYXkiLCAiSnVuZSI6ICAgICAiSnVu"
    "IiwgIkp1bHkiOiAgIkp1bCIsICJBdWd1c3QiOiAgICJBdWciLAogICAgIlNlcHRlbWJlciI6ICJT"
    "ZXAiLCAiT2N0b2JlciI6ICJPY3QiLCAiTm92ZW1iZXIiOiAiTm92IiwgIkRlY2VtYmVyIjogIkRl"
    "YyIsCn0KCiMgQ29sdW1uIHNjaGVtYXMg4oCUIHdoaWNoIG51bWVyaWMgY29sdW1ucyBlYWNoIHZh"
    "bHVlLWxpbmUgY2FycmllcyAoaW4gb3JkZXIpClNDSEVNQV81ID0gWyJ0YXhhYmxlIiwgImlnc3Qi"
    "LCAiY2dzdCIsICJzZ3N0IiwgImNlc3MiXSAgICMgaW50cmEtc3RhdGUgZnVsbCBzZXQKU0NIRU1B"
    "XzMgPSBbInRheGFibGUiLCAiaWdzdCIsICJjZXNzIl0gICAgICAgICAgICAgICAgICAgICMgaW50"
    "ZXItc3RhdGUgb25seSAobm8gQ0dTVC9TR1NUKQpTQ0hFTUFfMSA9IFsidGF4YWJsZSJdICAgICAg"
    "ICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgIyB2YWx1ZS1vbmx5IHJvd3MKCgojIFRhYmxl"
    "IGV4dHJhY3Rpb24gcnVsZXMuCiMgRWFjaCBydWxlIHNheXM6IGxvY2F0ZSB0aGlzIGhlYWRpbmcs"
    "IHRoZW4gZmluZCB0aGUgZGF0YSBsaW5lIHRoYXQgYmVnaW5zCiMgd2l0aCB0aGlzIG1hcmtlciwg"
    "dGhlbiBtYXAgdGhlIG51bWJlcnMgZm91bmQgdGhlcmUgdG8gdGhlc2UgY29sdW1ucy4KVEFSR0VU"
    "UyA9IFsKICAgICMgLS0tLSBDdXJyZW50IHBlcmlvZCBvdXR3YXJkIHN1cHBsaWVzIC0tLS0KICAg"
    "ICgiNEEiLCAgIkIyQiBSZWd1bGFyIiwKICAgICAiNEEgLSBUYXhhYmxlIG91dHdhcmQgc3VwcGxp"
    "ZXMgbWFkZSB0byByZWdpc3RlcmVkIHBlcnNvbnMgKG90aGVyIHRoYW4iLAogICAgICJUb3RhbCAi"
    "LCBTQ0hFTUFfNSksCiAgICAoIjRCIiwgICJCMkIgUmV2ZXJzZSBDaGFyZ2UiLAogICAgICI0QiAt"
    "IFRheGFibGUgb3V0d2FyZCBzdXBwbGllcyBtYWRlIHRvIHJlZ2lzdGVyZWQgcGVyc29ucyBhdHRy"
    "YWN0aW5nIHRheCBvbiByZXZlcnNlIiwKICAgICAiVG90YWwgIiwgU0NIRU1BXzUpLAogICAgKCI1"
    "IiwgICAiQjJDTCAoTGFyZ2UpIiwKICAgICAiNSAtIFRheGFibGUgb3V0d2FyZCBpbnRlci1zdGF0"
    "ZSBzdXBwbGllcyBtYWRlIHRvIHVucmVnaXN0ZXJlZCBwZXJzb25zIiwKICAgICAiVG90YWwgIiwg"
    "U0NIRU1BXzMpLAoKICAgICMgNkEgRXhwb3J0cyDigJQgVG90YWwgcm93IChhZ2dyZWdhdGVzIEVY"
    "UFdQICsgRVhQV09QKSArIHN1Yi1yb3dzCiAgICAoIjZBIiwgICJFeHBvcnRzIC0gVG90YWwgKDZB"
    "KSIsCiAgICAgIjZBIOKAkyBFeHBvcnRzIiwgIlRvdGFsICIsIFNDSEVNQV8zKSwKICAgICgiNkEi"
    "LCAgIkV4cG9ydHMgLSBFWFBXUCAod2l0aCBwYXltZW50KSIsCiAgICAgIjZBIOKAkyBFeHBvcnRz"
    "IiwgIi0gRVhQV1AiLCBTQ0hFTUFfMyksCiAgICAoIjZBIiwgICJFeHBvcnRzIC0gRVhQV09QICh3"
    "aXRob3V0IHBheW1lbnQpIiwKICAgICAiNkEg4oCTIEV4cG9ydHMiLCAiLSBFWFBXT1AiLCBTQ0hF"
    "TUFfMSksCgogICAgIyA2QiBTRVog4oCUIFRvdGFsIHJvdyArIHN1Yi1yb3dzCiAgICAoIjZCIiwg"
    "ICJTRVogLSBUb3RhbCAoNkIpIiwKICAgICAiNkIgLSBTdXBwbGllcyBtYWRlIHRvIFNFWiIsICJU"
    "b3RhbCAiLCBTQ0hFTUFfMyksCiAgICAoIjZCIiwgICJTRVogLSBTRVpXUCAod2l0aCBwYXltZW50"
    "KSIsCiAgICAgIjZCIC0gU3VwcGxpZXMgbWFkZSB0byBTRVoiLCAiLSBTRVpXUCIsIFNDSEVNQV8z"
    "KSwKICAgICgiNkIiLCAgIlNFWiAtIFNFWldPUCAod2l0aG91dCBwYXltZW50KSIsCiAgICAgIjZC"
    "IC0gU3VwcGxpZXMgbWFkZSB0byBTRVoiLCAiLSBTRVpXT1AiLCBTQ0hFTUFfMSksCgogICAgKCI2"
    "QyIsICAiRGVlbWVkIEV4cG9ydHMgKERFKSIsCiAgICAgIjZDIC0gRGVlbWVkIEV4cG9ydHMiLCAi"
    "VG90YWwgIiwgU0NIRU1BXzUpLAoKICAgICgiNyIsICAgIkIyQ1MgKE90aGVycykiLAogICAgICI3"
    "LSBUYXhhYmxlIHN1cHBsaWVzIiwgIlRvdGFsICIsIFNDSEVNQV81KSwKCiAgICAjIDgg4oCUIHN1"
    "Yi1yb3dzCiAgICAoIjgiLCAgICJOaWwgcmF0ZWQiLAogICAgICI4IC0gTmlsIHJhdGVkLCBleGVt"
    "cHRlZCIsICItIE5pbCIsIFNDSEVNQV8xKSwKICAgICgiOCIsICAgIkV4ZW1wdGVkIiwKICAgICAi"
    "OCAtIE5pbCByYXRlZCwgZXhlbXB0ZWQiLCAiLSBFeGVtcHRlZCIsIFNDSEVNQV8xKSwKICAgICgi"
    "OCIsICAgIk5vbi1HU1QiLAogICAgICI4IC0gTmlsIHJhdGVkLCBleGVtcHRlZCIsICItIE5vbi1H"
    "U1QiLCBTQ0hFTUFfMSksCgogICAgIyAtLS0tIDlBIEFtZW5kbWVudHMgLS0tLQogICAgKCI5QSIs"
    "ICAiQW1lbmRtZW50IC0gQjJCIFJlZ3VsYXIiLAogICAgICI5QSAtIEFtZW5kbWVudCB0byB0YXhh"
    "YmxlIG91dHdhcmQgc3VwcGxpZXMgbWFkZSB0byByZWdpc3RlcmVkIHBlcnNvbiBpbiByZXR1cm5z"
    "IG9mIGVhcmxpZXIgdGF4IHBlcmlvZHMgaW4gdGFibGUgNCAtIEIyQiBSZWd1bGFyIiwKICAgICAi"
    "TmV0IGRpZmZlcmVudGlhbCBhbW91bnQiLCBTQ0hFTUFfNSksCiAgICAoIjlBIiwgICJBbWVuZG1l"
    "bnQgLSBCMkIgUmV2ZXJzZSBDaGFyZ2UiLAogICAgICI5QSAtIEFtZW5kbWVudCB0byB0YXhhYmxl"
    "IG91dHdhcmQgc3VwcGxpZXMgbWFkZSB0byByZWdpc3RlcmVkIHBlcnNvbiBpbiByZXR1cm5zIG9m"
    "IGVhcmxpZXIgdGF4IHBlcmlvZHMgaW4gdGFibGUgNCAtIEIyQiBSZXZlcnNlIGNoYXJnZSIsCiAg"
    "ICAgIk5ldCBkaWZmZXJlbnRpYWwgYW1vdW50IiwgU0NIRU1BXzUpLAogICAgKCI5QSIsICAiQW1l"
    "bmRtZW50IC0gQjJDTCIsCiAgICAgIjlBIC0gQW1lbmRtZW50IHRvIEludGVyLVN0YXRlIHN1cHBs"
    "aWVzIG1hZGUgdG8gdW5yZWdpc3RlcmVkIHBlcnNvbiIsCiAgICAgIk5ldCBkaWZmZXJlbnRpYWwg"
    "YW1vdW50IiwgU0NIRU1BXzMpLAogICAgKCI5QSIsICAiQW1lbmRtZW50IC0gRXhwb3J0cyBFWFBX"
    "UCIsCiAgICAgIjlBIC0gQW1lbmRtZW50IHRvIEV4cG9ydCBzdXBwbGllcyIsICItIEVYUFdQIiwg"
    "U0NIRU1BXzMpLAogICAgKCI5QSIsICAiQW1lbmRtZW50IC0gRXhwb3J0cyBFWFBXT1AiLAogICAg"
    "ICI5QSAtIEFtZW5kbWVudCB0byBFeHBvcnQgc3VwcGxpZXMiLCAiLSBFWFBXT1AiLCBTQ0hFTUFf"
    "MSksCiAgICAoIjlBIiwgICJBbWVuZG1lbnQgLSBTRVogU0VaV1AiLAogICAgICI5QSAtIEFtZW5k"
    "bWVudCB0byBzdXBwbGllcyBtYWRlIHRvIFNFWiIsICItIFNFWldQIiwgU0NIRU1BXzMpLAogICAg"
    "KCI5QSIsICAiQW1lbmRtZW50IC0gU0VaIFNFWldPUCIsCiAgICAgIjlBIC0gQW1lbmRtZW50IHRv"
    "IHN1cHBsaWVzIG1hZGUgdG8gU0VaIiwgIi0gU0VaV09QIiwgU0NIRU1BXzEpLAogICAgKCI5QSIs"
    "ICAiQW1lbmRtZW50IC0gRGVlbWVkIEV4cG9ydHMiLAogICAgICI5QSAtIEFtZW5kbWVudCB0byBE"
    "ZWVtZWQgRXhwb3J0cyIsCiAgICAgIk5ldCBkaWZmZXJlbnRpYWwgYW1vdW50IiwgU0NIRU1BXzUp"
    "LAoKICAgICMgLS0tLSA5QiBDcmVkaXQvRGViaXQgTm90ZXMgLS0tLQogICAgKCI5QiIsICAiQ0RO"
    "UiAtIENyZWRpdC9EZWJpdCBOb3RlcyAoUmVnaXN0ZXJlZCkiLAogICAgICI5QiAtIENyZWRpdC9E"
    "ZWJpdCBOb3RlcyAoUmVnaXN0ZXJlZCkiLAogICAgICJUb3RhbCAtIE5ldCBvZmYgZGViaXQvY3Jl"
    "ZGl0IG5vdGVzIiwgU0NIRU1BXzUpLAogICAgKCI5QiIsICAiQ0ROVVIgLSBCMkNMIiwKICAgICAi"
    "OUIgLSBDcmVkaXQvRGViaXQgTm90ZXMgKFVucmVnaXN0ZXJlZCkiLCAiLSBCMkNMIiwgU0NIRU1B"
    "XzMpLAogICAgKCI5QiIsICAiQ0ROVVIgLSBFWFBXUCIsCiAgICAgIjlCIC0gQ3JlZGl0L0RlYml0"
    "IE5vdGVzIChVbnJlZ2lzdGVyZWQpIiwgIi0gRVhQV1AiLCBTQ0hFTUFfMyksCiAgICAoIjlCIiwg"
    "ICJDRE5VUiAtIEVYUFdPUCIsCiAgICAgIjlCIC0gQ3JlZGl0L0RlYml0IE5vdGVzIChVbnJlZ2lz"
    "dGVyZWQpIiwgIi0gRVhQV09QIiwgU0NIRU1BXzEpLAoKICAgICMgLS0tLSA5QyBBbWVuZGVkIENE"
    "TiAtLS0tCiAgICAoIjlDIiwgICJDRE5SQSAtIEFtZW5kZWQgQ0ROIChSZWdpc3RlcmVkKSIsCiAg"
    "ICAgIjlDIC0gQW1lbmRlZCBDcmVkaXQvRGViaXQgTm90ZXMgKFJlZ2lzdGVyZWQpIiwKICAgICAi"
    "TmV0IERpZmZlcmVudGlhbCBhbW91bnQiLCBTQ0hFTUFfNSksCiAgICAoIjlDIiwgICJDRE5VUkEg"
    "LSBCMkNMIiwKICAgICAiOUMgLSBBbWVuZGVkIENyZWRpdC9EZWJpdCBOb3RlcyAoVW5yZWdpc3Rl"
    "cmVkKSIsICItIEIyQ0wiLCBTQ0hFTUFfMyksCiAgICAoIjlDIiwgICJDRE5VUkEgLSBFWFBXUCIs"
    "CiAgICAgIjlDIC0gQW1lbmRlZCBDcmVkaXQvRGViaXQgTm90ZXMgKFVucmVnaXN0ZXJlZCkiLCAi"
    "LSBFWFBXUCIsIFNDSEVNQV8zKSwKICAgICgiOUMiLCAgIkNETlVSQSAtIEVYUFdPUCIsCiAgICAg"
    "IjlDIC0gQW1lbmRlZCBDcmVkaXQvRGViaXQgTm90ZXMgKFVucmVnaXN0ZXJlZCkiLCAiLSBFWFBX"
    "T1AiLCBTQ0hFTUFfMSksCgogICAgKCIxMCIsICAiQW1lbmRtZW50IHRvIEIyQyAoT3RoZXJzKSIs"
    "CiAgICAgIjEwIC0gQW1lbmRtZW50IHRvIHRheGFibGUgb3V0d2FyZCBzdXBwbGllcyBtYWRlIHRv"
    "IHVucmVnaXN0ZXJlZCBwZXJzb24iLAogICAgICJOZXQgZGlmZmVyZW50aWFsIGFtb3VudCIsIFND"
    "SEVNQV81KSwKCiAgICAjIC0tLS0gMTEgQWR2YW5jZXMgLS0tLQogICAgKCIxMUEiLCAiQWR2YW5j"
    "ZXMgcmVjZWl2ZWQiLAogICAgICIxMUEoMSksIDExQSgyKSAtIEFkdmFuY2VzIHJlY2VpdmVkIiwg"
    "IlRvdGFsICIsIFNDSEVNQV81KSwKICAgICgiMTFCIiwgIkFkdmFuY2VzIGFkanVzdGVkIGFnYWlu"
    "c3QgY3VycmVudC1wZXJpb2Qgc3VwcGxpZXMiLAogICAgICIxMUIoMSksIDExQigyKSAtIEFkdmFu"
    "Y2UgYW1vdW50IHJlY2VpdmVkIGluIGVhcmxpZXIgdGF4IHBlcmlvZCIsCiAgICAgIlRvdGFsICIs"
    "IFNDSEVNQV81KSwKICAgICgiMTFBIiwgIkFtZW5kbWVudCB0byBBZHZhbmNlcyByZWNlaXZlZCIs"
    "CiAgICAgIjExQSAtIEFtZW5kbWVudCB0byBhZHZhbmNlcyByZWNlaXZlZCIsICJUb3RhbCAiLCBT"
    "Q0hFTUFfNSksCiAgICAoIjExQiIsICJBbWVuZG1lbnQgdG8gQWR2YW5jZXMgYWRqdXN0ZWQiLAog"
    "ICAgICIxMUIgLSBBbWVuZG1lbnQgdG8gYWR2YW5jZXMgYWRqdXN0ZWQiLCAiVG90YWwgIiwgU0NI"
    "RU1BXzUpLAoKICAgICMgLS0tLSAxMiBIU04gLS0tLQogICAgKCIxMiIsICAiSFNOLXdpc2Ugc3Vt"
    "bWFyeSIsCiAgICAgIjEyIC0gSFNOLXdpc2Ugc3VtbWFyeSIsICJUb3RhbCAiLCBTQ0hFTUFfNSks"
    "CgogICAgIyAtLS0tIDEzIERvY3VtZW50cyAoY291bnQgb25seSkgLS0tLQogICAgKCIxMyIsICAi"
    "RG9jdW1lbnRzIGlzc3VlZCAoY291bnQgb25seSkiLAogICAgICIxMyAtIERvY3VtZW50cyBpc3N1"
    "ZWQiLCAiTmV0IGlzc3VlZCBkb2N1bWVudHMiLCBbXSksCgogICAgIyAtLS0tIDE0IEVDTyAtLS0t"
    "CiAgICAoIjE0IiwgICJTdXBwbGllcyB0aHJvdWdoIEUtQ29tbWVyY2UgT3BlcmF0b3JzIiwKICAg"
    "ICAiMTQgLSBTdXBwbGllcyBtYWRlIHRocm91Z2ggRS1Db21tZXJjZSBPcGVyYXRvcnMiLCAiVG90"
    "YWwgIiwgU0NIRU1BXzUpLAogICAgKCIxNEEiLCAiQW1lbmRlZCBTdXBwbGllcyB0aHJvdWdoIEVD"
    "TyIsCiAgICAgIjE0QSAtIEFtZW5kZWQgU3VwcGxpZXMgbWFkZSB0aHJvdWdoIEUtQ29tbWVyY2Ug"
    "T3BlcmF0b3JzIiwKICAgICAiTmV0IGRpZmZlcmVudGlhbCBhbW91bnQiLCBTQ0hFTUFfNSksCgog"
    "ICAgIyAtLS0tIDE1IFN1cHBsaWVzIHUvcyA5KDUpIC0tLS0KICAgICgiMTUiLCAgIlN1cHBsaWVz"
    "IHUvcyA5KDUpIiwKICAgICAiMTUgLSBTdXBwbGllcyBVL3MgOSg1KSIsICJUb3RhbCAiLCBTQ0hF"
    "TUFfNSksCiAgICAoIjE1QShJKSIsICAiQW1lbmRlZCA5KDUpIC0gUmVnaXN0ZXJlZCBSZWNpcGll"
    "bnRzIiwKICAgICAiMTVBIChJKSAtIEFtZW5kZWQgU3VwcGxpZXMgVS9zIDkoNSkiLAogICAgICJO"
    "ZXQgZGlmZmVyZW50aWFsIGFtb3VudCIsIFNDSEVNQV81KSwKICAgICgiMTVBKElJKSIsICJBbWVu"
    "ZGVkIDkoNSkgLSBVbnJlZ2lzdGVyZWQgUmVjaXBpZW50cyIsCiAgICAgIjE1QSAoSUkpIC0gQW1l"
    "bmRlZCBTdXBwbGllcyIsCiAgICAgIk5ldCBkaWZmZXJlbnRpYWwgYW1vdW50IiwgU0NIRU1BXzUp"
    "LApdCgoKIyAtLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0t"
    "LS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0gIwojICBQYXJzaW5nIGhlbHBlcnMKIyAtLS0tLS0t"
    "LS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0t"
    "LS0tLS0tLS0tLS0gIwoKIyBOdW1iZXIgcmVnZXgg4oCUIGNhcHR1cmVzOgojICAgMSkgICgxMiwz"
    "NDUuNjcpICBvciAgKCAxMiwzNDUuNjcgKSAgICAgICDihpIgcGFyZW50aGVzaXNlZCBuZWdhdGl2"
    "ZSAgKEdTVFItMSBzdGFuZGFyZCkKIyAgIDIpICAtMTIsMzQ1LjY3ICAgICAgICAgICAgICAgICAg"
    "ICAgICAgICAgIOKGkiBsZWFkaW5nIG1pbnVzCiMgICAzKSAgMTIsMzQ1LjY3LSAgICAgICAgICAg"
    "ICAgICAgICAgICAgICAgICDihpIgdHJhaWxpbmcgbWludXMgKHJhcmUpCiMgICA0KSAgMTIsMzQ1"
    "LjY3ICAgICAgICAgICAgICAgICAgICAgICAgICAgICDihpIgcG9zaXRpdmUKIyBJbmRpYW4gY29t"
    "bWEgZ3JvdXBpbmcgKGxha2gvY3JvcmU6IDEsMjMsNDUsNjc4LjkwKSBpcyBzdXBwb3J0ZWQuCk5V"
    "TV9SRSA9IHJlLmNvbXBpbGUoCiAgICByIiIiCiAgICAoP1A8cGFyZW4+XChccypcZHsxLDN9KD86"
    "LFxkezIsM30pKlwuXGR7Mn1ccypcKSkgICAgICAjICgxMiwzNDUuNjcpCiAgICB8CiAgICAoP1A8"
    "bGVhZD4tXGR7MSwzfSg/OixcZHsyLDN9KSpcLlxkezJ9KSAgICAgICAgICAgICAgICAjIC0xMiwz"
    "NDUuNjcKICAgIHwKICAgICg/UDx0cmFpbD5cZHsxLDN9KD86LFxkezIsM30pKlwuXGR7Mn0tKSAg"
    "ICAgICAgICAgICAgICMgMTIsMzQ1LjY3LQogICAgfAogICAgKD9QPHBvcz5cZHsxLDN9KD86LFxk"
    "ezIsM30pKlwuXGR7Mn0pICAgICAgICAgICAgICAgICAgIyAxMiwzNDUuNjcgIG9yICAxMjMuNDUK"
    "ICAgICIiIiwKICAgIHJlLlZFUkJPU0UsCikKV1NfUkUgID0gcmUuY29tcGlsZShyIlxzKyIpCgoK"
    "ZGVmIG5vcm0ocyk6CiAgICAiIiJDb2xsYXBzZSB3aGl0ZXNwYWNlOyBsb3dlcmNhc2UtZnJlZS4i"
    "IiIKICAgIHJldHVybiBXU19SRS5zdWIoIiAiLCBzKS5zdHJpcCgpCgoKZGVmIF90b19mbG9hdCh0"
    "b2tlbik6CiAgICAiIiJDb252ZXJ0IGEgbWF0Y2hlZCBudW1lcmljIHRva2VuIChhbnkgb2YgdGhl"
    "IDQgZm9ybWF0cyBhYm92ZSkgdG8gYSBmbG9hdC4iIiIKICAgIHQgPSB0b2tlbi5zdHJpcCgpCiAg"
    "ICBzaWduID0gMS4wCiAgICBpZiB0LnN0YXJ0c3dpdGgoIigiKSBhbmQgdC5lbmRzd2l0aCgiKSIp"
    "OgogICAgICAgIHNpZ24gPSAtMS4wCiAgICAgICAgdCA9IHRbMTotMV0uc3RyaXAoKQogICAgZWxp"
    "ZiB0LnN0YXJ0c3dpdGgoIi0iKToKICAgICAgICBzaWduID0gLTEuMAogICAgICAgIHQgPSB0WzE6"
    "XQogICAgZWxpZiB0LmVuZHN3aXRoKCItIik6CiAgICAgICAgc2lnbiA9IC0xLjAKICAgICAgICB0"
    "ID0gdFs6LTFdCiAgICByZXR1cm4gc2lnbiAqIGZsb2F0KHQucmVwbGFjZSgiLCIsICIiKSkKCgoj"
    "IFdhdGVybWFyayBsZXR0ZXJzIHRoYXQgdGhlIEdTVCBwb3J0YWwgc29tZXRpbWVzIG92ZXJsYXlz"
    "IGRpYWdvbmFsbHkgb24gdGhlIFBERiwKIyB3aGljaCBwZGZwbHVtYmVyIG9jY2FzaW9uYWxseSBp"
    "bnNlcnRzIG1pZC10ZXh0LiBTdHJpcHBpbmcgaXNvbGF0ZWQgc2luZ2xlIGxldHRlcnMKIyAobm90"
    "IHByZWNlZGVkL2ZvbGxvd2VkIGJ5IGEgd29yZCBjaGFyYWN0ZXIpIG1ha2VzIG51bWVyaWMgcGFy"
    "c2luZyByb2J1c3QuCldBVEVSTUFSS19MRVRURVJTID0gcmUuY29tcGlsZShyIig/PCFbQS1aYS16"
    "MC05XSlbRklOQUxdKD8hW0EtWmEtejAtOV0pIikKCgpkZWYgY2xlYW5fbGluZV9mb3JfbnVtYmVy"
    "cyhsaW5lKToKICAgICIiIgogICAgUmVtb3ZlIGlzb2xhdGVkIHNpbmdsZS1sZXR0ZXIgd2F0ZXJt"
    "YXJrIGZyYWdtZW50cyAoJ0YnLCAnSScsICdOJywgJ0EnLCAnTCcpCiAgICB0aGF0IHRoZSAnRklO"
    "QUwnIGRpYWdvbmFsIHdhdGVybWFyayBzb21ldGltZXMgaW5qZWN0cyBpbnRvIGEgbGluZS4KICAg"
    "IE9ubHkgcmVtb3ZlcyB0aGVtIHdoZW4gdGhleSBhcmUgY2xlYXJseSBub3QgcGFydCBvZiBhIHdv"
    "cmQuCiAgICAiIiIKICAgICMgTXVsdGlwbGUgcGFzc2VzIGJlY2F1c2UgcmVtb3ZhbCBjYW4gZXhw"
    "b3NlIG5ldyBpc29sYXRlZCBsZXR0ZXJzCiAgICBwcmV2ID0gTm9uZQogICAgY3VyID0gbGluZQog"
    "ICAgd2hpbGUgcHJldiAhPSBjdXI6CiAgICAgICAgcHJldiA9IGN1cgogICAgICAgIGN1ciA9IFdB"
    "VEVSTUFSS19MRVRURVJTLnN1YigiIiwgY3VyKQogICAgIyBDb2xsYXBzZSB3aGl0ZXNwYWNlIGlu"
    "dHJvZHVjZWQgYnkgcmVtb3ZhbAogICAgcmV0dXJuIFdTX1JFLnN1YigiICIsIGN1cikuc3RyaXAo"
    "KQoKCmRlZiBleHRyYWN0X251bWJlcnMobGluZSk6CiAgICAiIiIKICAgIFJldHVybiBsaXN0IG9m"
    "IGZsb2F0cyBmcm9tIGEgbGluZS4KICAgIEhhbmRsZXMgSW5kaWFuIGNvbW1hIGZvcm1hdCBBTkQg"
    "bmVnYXRpdmUgdmFsdWVzIHdyaXR0ZW4gYXMKICAgIHBhcmVudGhlc2VzLCBsZWFkaW5nIG1pbnVz"
    "LCBvciB0cmFpbGluZyBtaW51cy4KICAgIFN0cmlwcyBkaWFnb25hbC13YXRlcm1hcmsgc2luZ2xl"
    "LWxldHRlciBmcmFnbWVudHMgYmVmb3JlIHBhcnNpbmcuCiAgICAiIiIKICAgIGNsZWFuZWQgPSBj"
    "bGVhbl9saW5lX2Zvcl9udW1iZXJzKGxpbmUpCiAgICBvdXQgPSBbXQogICAgZm9yIG0gaW4gTlVN"
    "X1JFLmZpbmRpdGVyKGNsZWFuZWQpOgogICAgICAgIG91dC5hcHBlbmQoX3RvX2Zsb2F0KG0uZ3Jv"
    "dXAoMCkpKQogICAgcmV0dXJuIG91dAoKCmRlZiBleHRyYWN0X3JlY29yZHMobGluZSwgdmFsdWVf"
    "bWFya2VyKToKICAgICIiIgogICAgVHJ5IHRvIHB1bGwgdGhlIGludGVnZXIgcmVjb3JkIGNvdW50"
    "IG91dCBvZiBhIHZhbHVlIGxpbmUuCiAgICBGb3JtYXQgZXhhbXBsZTogJ1RvdGFsIDEgSW52b2lj"
    "ZSAxNSw0MSw4MDEuMDAgLi4uJwogICAgICAgICAgICAgICAgICAgICdOZXQgVG90YWwgKERlYml0"
    "IG5vdGVzIOKAkyBDcmVkaXQgbm90ZXMpIDAgTm90ZSAwLjAwIC4uLicKICAgIFJldHVybnMgaW50"
    "IG9yIDAgaWYgbm90IGZvdW5kIC8gdW5wYXJzZWFibGUuCiAgICBUaGUgd2F0ZXJtYXJrIHNvbWV0"
    "aW1lcyBwcmVmaXhlcyBhIGxldHRlciB0byB0aGUgZGlnaXQgKGUuZy4gIkkwIiwgIkYwIiksCiAg"
    "ICBzbyB3ZSBzdHJpcCBub24tZGlnaXRzLgogICAgIiIiCiAgICAjIFRha2UgdGhlIHBvcnRpb24g"
    "YWZ0ZXIgdGhlIG1hcmtlcgogICAgaWR4ID0gbGluZS5sb3dlcigpLmZpbmQodmFsdWVfbWFya2Vy"
    "LnN0cmlwKCkubG93ZXIoKSkKICAgIGlmIGlkeCA9PSAtMToKICAgICAgICByZXR1cm4gMAogICAg"
    "dGFpbCA9IGxpbmVbaWR4ICsgbGVuKHZhbHVlX21hcmtlcik6XS5zdHJpcCgpCiAgICAjIEZpcnN0"
    "IHdoaXRlc3BhY2UtdG9rZW4gaXMgdGhlIHJlY29yZHMgY291bnQgKHBvc3NpYmx5IHByZWZpeGVk"
    "IGJ5IHdhdGVybWFyaykKICAgIGZpcnN0X3RvayA9IHRhaWwuc3BsaXQoKVswXSBpZiB0YWlsLnNw"
    "bGl0KCkgZWxzZSAiIgogICAgZGlnaXRzID0gcmUuc3ViKHIiXEQiLCAiIiwgZmlyc3RfdG9rKQog"
    "ICAgcmV0dXJuIGludChkaWdpdHMpIGlmIGRpZ2l0cyBlbHNlIDAKCgpkZWYgZXh0cmFjdF9wZGZf"
    "dGV4dChwZGZfcGF0aCk6CiAgICAiIiJDb25jYXRlbmF0ZSB0ZXh0IGZyb20gZXZlcnkgcGFnZS4i"
    "IiIKICAgIHBhcnRzID0gW10KICAgIHdpdGggcGRmcGx1bWJlci5vcGVuKHBkZl9wYXRoKSBhcyBw"
    "ZGY6CiAgICAgICAgZm9yIHBhZ2UgaW4gcGRmLnBhZ2VzOgogICAgICAgICAgICB0ID0gcGFnZS5l"
    "eHRyYWN0X3RleHQoKSBvciAiIgogICAgICAgICAgICBwYXJ0cy5hcHBlbmQodCkKICAgIHJldHVy"
    "biAiXG4iLmpvaW4ocGFydHMpCgoKZGVmIHBhcnNlX21ldGEodGV4dCk6CiAgICAiIiJQdWxsIEdT"
    "VElOLCBwZXJpb2QsIEZZLCBsZWdhbCBuYW1lLCBBUk4sIEFSTiBkYXRlLiIiIgogICAgZ3N0aW5f"
    "bSAgICA9IHJlLnNlYXJjaChyIkdTVElOXHMrKFswLTldezJ9W0EtWjAtOV17MTN9KSIsIHRleHQp"
    "CiAgICBwZXJpb2RfbSAgID0gcmUuc2VhcmNoKHIiVGF4XHMqcGVyaW9kXHMrKFtBLVphLXpdKyki"
    "LCB0ZXh0KQogICAgZnlfbSAgICAgICA9IHJlLnNlYXJjaChyIkZpbmFuY2lhbFxzKnllYXJccyso"
    "XGR7NH0tXGR7Mn0pIiwgdGV4dCkKICAgIGxlZ2FsX20gICAgPSByZS5zZWFyY2gociJMZWdhbCBu"
    "YW1lIG9mIHRoZSByZWdpc3RlcmVkIHBlcnNvblxzKyguKykiLCB0ZXh0KQogICAgYXJuX20gICAg"
    "ICA9IHJlLnNlYXJjaChyIlwoY1wpXHMqQVJOXHMrKFxTKykiLCB0ZXh0KQogICAgYXJuX2RhdGVf"
    "bSA9IHJlLnNlYXJjaChyIlwoZFwpXHMqQVJOXHMqZGF0ZVxzKyhcUyspIiwgdGV4dCkKCiAgICBn"
    "c3RpbiA9IGdzdGluX20uZ3JvdXAoMSkgaWYgZ3N0aW5fbSBlbHNlICIiCiAgICBzdGF0ZV9jb2Rl"
    "ID0gZ3N0aW5bOjJdIGlmIGdzdGluIGVsc2UgIiIKICAgIHN0YXRlX25hbWUgPSBTVEFURV9DT0RF"
    "Uy5nZXQoc3RhdGVfY29kZSwgIlVua25vd24iKQogICAgcGVyaW9kID0gcGVyaW9kX20uZ3JvdXAo"
    "MSkgaWYgcGVyaW9kX20gZWxzZSAiIgogICAgZnkgPSBmeV9tLmdyb3VwKDEpIGlmIGZ5X20gZWxz"
    "ZSAiIgogICAgIyBJbmRpYW4gRlk6IEFwcuKAk0RlYyA9IGZpcnN0IHllYXIgKHl5MSksIEphbuKA"
    "k01hciA9IHNlY29uZCB5ZWFyICh5eTIpCiAgICB5eSA9ICIiCiAgICBpZiBmeSBhbmQgIi0iIGlu"
    "IGZ5OgogICAgICAgIHl5MSA9IGZ5LnNwbGl0KCItIilbMF1bLTI6XQogICAgICAgIHl5MiA9IGZ5"
    "LnNwbGl0KCItIilbMV0KICAgICAgICBpZiBwZXJpb2QgaW4gKCJKYW51YXJ5IiwgIkZlYnJ1YXJ5"
    "IiwgIk1hcmNoIik6CiAgICAgICAgICAgIHl5ID0geXkyCiAgICAgICAgZWxzZToKICAgICAgICAg"
    "ICAgeXkgPSB5eTEKICAgIG1vbnRoID0gZiJ7TU9OVEhfQUJCUi5nZXQocGVyaW9kLCBwZXJpb2Rb"
    "OjNdKX0te3l5fSIgaWYgcGVyaW9kIGVsc2UgIiIKCiAgICByZXR1cm4gewogICAgICAgICJnc3Rp"
    "biI6ICAgICAgZ3N0aW4sCiAgICAgICAgInN0YXRlX2NvZGUiOiBzdGF0ZV9jb2RlLAogICAgICAg"
    "ICJzdGF0ZV9uYW1lIjogc3RhdGVfbmFtZSwKICAgICAgICAibW9udGgiOiAgICAgIG1vbnRoLAog"
    "ICAgICAgICJmeSI6ICAgICAgICAgZnksCiAgICAgICAgInRheF9wZXJpb2QiOiBwZXJpb2QsCiAg"
    "ICAgICAgImxlZ2FsX25hbWUiOiBsZWdhbF9tLmdyb3VwKDEpLnN0cmlwKCkgaWYgbGVnYWxfbSBl"
    "bHNlICIiLAogICAgICAgICJhcm4iOiAgICAgICAgYXJuX20uZ3JvdXAoMSkgaWYgYXJuX20gZWxz"
    "ZSAiIiwKICAgICAgICAiYXJuX2RhdGUiOiAgIGFybl9kYXRlX20uZ3JvdXAoMSkgaWYgYXJuX2Rh"
    "dGVfbSBlbHNlICIiLAogICAgfQoKCkhFQURfUkUgPSByZS5jb21waWxlKHIiXlxzKlxkK1tBLVpd"
    "P1xzKig/OlwoW0lWWF0rXCkpP1xzKlst4oCTXSIpCgoKREFTSF9SRSA9IHJlLmNvbXBpbGUociJb"
    "XHUyMDEwXHUyMDExXHUyMDEyXHUyMDEzXHUyMDE0XHUyMDE1XC1dIikgICMgYW55IGtpbmQgb2Yg"
    "ZGFzaAoKCmRlZiBub3JtX2Rhc2hlcyhzKToKICAgICIiIlJlcGxhY2UgYWxsIGRhc2ggdmFyaWFu"
    "dHMgKGVuLWRhc2gsIGVtLWRhc2gsIGh5cGhlbi1taW51cywgZXRjKSB3aXRoIHNpbXBsZSAnLScu"
    "IiIiCiAgICByZXR1cm4gREFTSF9SRS5zdWIoIi0iLCBzKQoKCmRlZiBmaW5kX2hlYWRpbmdfaWR4"
    "KGxpbmVzLCBoZWFkaW5nX3N1YnN0ciwgc3RhcnQ9MCk6CiAgICAiIiJSZXR1cm4gaW5kZXggb2Yg"
    "Zmlyc3QgbGluZSB0aGF0IGNvbnRhaW5zIGhlYWRpbmdfc3Vic3RyIChkYXNoLSBhbmQgd3Mtbm9y"
    "bWFsaXplZCkuIiIiCiAgICBuZWVkbGUgPSBub3JtX2Rhc2hlcyhub3JtKGhlYWRpbmdfc3Vic3Ry"
    "KSkKICAgIGZvciBpIGluIHJhbmdlKHN0YXJ0LCBsZW4obGluZXMpKToKICAgICAgICBpZiBuZWVk"
    "bGUgaW4gbm9ybV9kYXNoZXMobm9ybShsaW5lc1tpXSkpOgogICAgICAgICAgICByZXR1cm4gaQog"
    "ICAgcmV0dXJuIC0xCgoKZGVmIGZpbmRfdmFsdWVfaWR4KGxpbmVzLCBzdGFydF9pZHgsIG1hcmtl"
    "ciwgbWF4X2xpbmVzPTIwKToKICAgICIiIgogICAgRmluZCBhIGRhdGEgbGluZSBhZnRlciBzdGFy"
    "dF9pZHggd2hvc2Ugbm9ybWFsaXplZCB0ZXh0IGJlZ2lucwogICAgd2l0aCAob3IgY29udGFpbnMs"
    "IGZvciBzdWItcm93IG1hcmtlcnMgbGlrZSAnLSBFWFBXUCcpIHRoZSBtYXJrZXIuCiAgICBTdG9w"
    "cyBpZiBhIG5ldyB0YWJsZSBoZWFkaW5nIGFwcGVhcnMgaW4gYmV0d2Vlbi4KCiAgICBUb2xlcmF0"
    "ZXMgd2F0ZXJtYXJrIGZyYWdtZW50cyAoc2luZ2xlLWNoYXIgJ0YnLCAnSScsICdOJywgJ0EnLCAn"
    "TCcgbGluZXMpCiAgICBhbmQgcGFnaW5hdGlvbiBub2lzZSBsaWtlICdJUCBBZGRyZXNzOicsICdE"
    "ZXNjcmlwdGlvbiBOby4gb2YuLi4nIGNvbHVtbnMKICAgIHRoYXQgbWF5IGludGVybGVhdmUgYmV0"
    "d2VlbiBoZWFkaW5nIGFuZCBkYXRhIG9uIG11bHRpLXBhZ2UgcmV0dXJucy4KICAgICIiIgogICAg"
    "bmVlZGxlID0gbm9ybV9kYXNoZXMobm9ybShtYXJrZXIpKQogICAgZm9yIGkgaW4gcmFuZ2Uoc3Rh"
    "cnRfaWR4ICsgMSwgbWluKHN0YXJ0X2lkeCArIDEgKyBtYXhfbGluZXMsIGxlbihsaW5lcykpKToK"
    "ICAgICAgICByYXcgPSBsaW5lc1tpXQogICAgICAgIGxpbmVfbm9ybSA9IG5vcm1fZGFzaGVzKG5v"
    "cm0ocmF3KSkKCiAgICAgICAgIyBTa2lwIGVtcHR5IC8gd2F0ZXJtYXJrLWZyYWdtZW50IC8gcGFn"
    "aW5hdGlvbiBub2lzZSBsaW5lcyAoY29udGludWUgc2Nhbm5pbmcpCiAgICAgICAgaWYgbm90IGxp"
    "bmVfbm9ybToKICAgICAgICAgICAgY29udGludWUKICAgICAgICBpZiBsZW4obGluZV9ub3JtKSA8"
    "PSAyOiAgICAgICAgICAgICAgICAgICAgICAgIyAnRicsICdJTicsIGV0YyDigJQgd2F0ZXJtYXJr"
    "IGRlYnJpcwogICAgICAgICAgICBjb250aW51ZQogICAgICAgIGlmIGxpbmVfbm9ybS5zdGFydHN3"
    "aXRoKCJJUCBBZGRyZXNzIik6CiAgICAgICAgICAgIGNvbnRpbnVlCiAgICAgICAgaWYgbGluZV9u"
    "b3JtLnN0YXJ0c3dpdGgoIkRlc2NyaXB0aW9uIE5vLiBvZiIpOiAgIyBoZWFkZXIgcm93IHJlcGVh"
    "dHMgb24gcGFnZSBicmVhawogICAgICAgICAgICBjb250aW51ZQogICAgICAgIGlmIGxpbmVfbm9y"
    "bS5zdGFydHN3aXRoKCJyZWNvcmRzIERvY3VtZW50IFR5cGUiKToKICAgICAgICAgICAgY29udGlu"
    "dWUKCiAgICAgICAgIyBTdG9wIG9ubHkgaWYgd2UgaGl0IGEgTkVXIGhlYWRpbmcgcm93IHRoYXQg"
    "ZG9lc24ndCBpdHNlbGYgbWF0Y2ggb3VyIG1hcmtlcgogICAgICAgIGlmIEhFQURfUkUubWF0Y2go"
    "cmF3KSBhbmQgbm90IGxpbmVfbm9ybS5zdGFydHN3aXRoKG5lZWRsZSk6CiAgICAgICAgICAgIHJl"
    "dHVybiAtMQoKICAgICAgICBpZiBsaW5lX25vcm0uc3RhcnRzd2l0aChuZWVkbGUpIG9yICgiICIg"
    "KyBuZWVkbGUgKyAiICIpIGluICgiICIgKyBsaW5lX25vcm0gKyAiICIpOgogICAgICAgICAgICBy"
    "ZXR1cm4gaQogICAgcmV0dXJuIC0xCgoKZGVmIHBhcnNlX3BkZihwZGZfcGF0aCk6CiAgICAiIiJS"
    "ZXR1cm4gKG1ldGFfZGljdCwgbGlzdF9vZl90YWJsZV9yb3dfZGljdHMpIG9yIHJhaXNlIG9uIGhh"
    "cmQgZmFpbHVyZS4iIiIKICAgIHRleHQgPSBleHRyYWN0X3BkZl90ZXh0KHBkZl9wYXRoKQogICAg"
    "aWYgbm90IHRleHQuc3RyaXAoKToKICAgICAgICByYWlzZSBWYWx1ZUVycm9yKCJQREYgaGFzIG5v"
    "IGV4dHJhY3RhYmxlIHRleHQgKHNjYW5uZWQgLyBpbWFnZS1vbmx5KS4iKQoKICAgIG1ldGEgPSBw"
    "YXJzZV9tZXRhKHRleHQpCiAgICBpZiBub3QgbWV0YVsiZ3N0aW4iXToKICAgICAgICByYWlzZSBW"
    "YWx1ZUVycm9yKCJDb3VsZCBub3QgbG9jYXRlIEdTVElOIGluIFBERi4iKQoKICAgIGxpbmVzID0g"
    "dGV4dC5zcGxpdCgiXG4iKQogICAgb3V0X3Jvd3MgPSBbXQoKICAgIGZvciB0YWJsZV9ubywgZGVz"
    "YywgaGVhZGluZywgbWFya2VyLCBzY2hlbWEgaW4gVEFSR0VUUzoKICAgICAgICBoX2lkeCA9IGZp"
    "bmRfaGVhZGluZ19pZHgobGluZXMsIGhlYWRpbmcpCiAgICAgICAgcm93ID0gewogICAgICAgICAg"
    "ICAidGFibGVfbm8iOiAgIHRhYmxlX25vLAogICAgICAgICAgICAiZGVzY3JpcHRpb24iOiBkZXNj"
    "LAogICAgICAgICAgICAicmVjb3JkcyI6ICAgIDAsCiAgICAgICAgICAgICJ0YXhhYmxlIjogICAg"
    "MC4wLAogICAgICAgICAgICAiaWdzdCI6ICAgICAgIDAuMCwKICAgICAgICAgICAgImNnc3QiOiAg"
    "ICAgICAwLjAsCiAgICAgICAgICAgICJzZ3N0IjogICAgICAgMC4wLAogICAgICAgICAgICAiY2Vz"
    "cyI6ICAgICAgIDAuMCwKICAgICAgICAgICAgImZvdW5kIjogICAgICBGYWxzZSwKICAgICAgICB9"
    "CiAgICAgICAgaWYgaF9pZHggPT0gLTE6CiAgICAgICAgICAgIG91dF9yb3dzLmFwcGVuZChyb3cp"
    "CiAgICAgICAgICAgIGNvbnRpbnVlCgogICAgICAgIHZfaWR4ID0gZmluZF92YWx1ZV9pZHgobGlu"
    "ZXMsIGhfaWR4LCBtYXJrZXIpCiAgICAgICAgaWYgdl9pZHggPT0gLTE6CiAgICAgICAgICAgIG91"
    "dF9yb3dzLmFwcGVuZChyb3cpCiAgICAgICAgICAgIGNvbnRpbnVlCgogICAgICAgIGxpbmUgPSBs"
    "aW5lc1t2X2lkeF0KICAgICAgICByb3dbImZvdW5kIl0gPSBUcnVlCgogICAgICAgICMgUmVjb3Jk"
    "cyAoYmVzdC1lZmZvcnQpCiAgICAgICAgdHJ5OgogICAgICAgICAgICByb3dbInJlY29yZHMiXSA9"
    "IGV4dHJhY3RfcmVjb3JkcyhsaW5lLCBtYXJrZXIpCiAgICAgICAgZXhjZXB0IEV4Y2VwdGlvbjoK"
    "ICAgICAgICAgICAgcm93WyJyZWNvcmRzIl0gPSAwCgogICAgICAgICMgSWYgdGhpcyBpcyB0aGUg"
    "ZG9jdW1lbnRzLWNvdW50IGxpbmUgKFRhYmxlIDEzKSwgcmVjb3Jkcy1vbmx5CiAgICAgICAgaWYg"
    "bm90IHNjaGVtYToKICAgICAgICAgICAgb3V0X3Jvd3MuYXBwZW5kKHJvdykKICAgICAgICAgICAg"
    "Y29udGludWUKCiAgICAgICAgIyBQdWxsIG51bWVyaWMgdmFsdWVzIGZyb20gdGhlIGxpbmUKICAg"
    "ICAgICBudW1zID0gZXh0cmFjdF9udW1iZXJzKGxpbmUpCiAgICAgICAgZm9yIGNvbF9uYW1lLCB2"
    "YWwgaW4gemlwKHNjaGVtYSwgbnVtcyk6CiAgICAgICAgICAgIHJvd1tjb2xfbmFtZV0gPSB2YWwK"
    "CiAgICAgICAgb3V0X3Jvd3MuYXBwZW5kKHJvdykKCiAgICByZXR1cm4gbWV0YSwgb3V0X3Jvd3MK"
    "CgojIC0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0t"
    "LS0tLS0tLS0tLS0tLS0tLS0tLS0tLSAjCiMgIEV4Y2VsIHdyaXRlcgojIC0tLS0tLS0tLS0tLS0t"
    "LS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0t"
    "LS0tLSAjCgpIRFJfRklMTCAgID0gUGF0dGVybkZpbGwoInNvbGlkIiwgc3RhcnRfY29sb3I9IjFG"
    "NEU3OCIpClNVQl9GSUxMICAgPSBQYXR0ZXJuRmlsbCgic29saWQiLCBzdGFydF9jb2xvcj0iMkU3"
    "NUI2IikKVE9UQUxfRklMTCA9IFBhdHRlcm5GaWxsKCJzb2xpZCIsIHN0YXJ0X2NvbG9yPSJGRkU2"
    "OTkiKQpHUk9VUF9GSUxMID0gUGF0dGVybkZpbGwoInNvbGlkIiwgc3RhcnRfY29sb3I9IkRERUJG"
    "NyIpCkVSUl9GSUxMICAgPSBQYXR0ZXJuRmlsbCgic29saWQiLCBzdGFydF9jb2xvcj0iRkZDN0NF"
    "IikKTUVUQV9GSUxMICA9IFBhdHRlcm5GaWxsKCJzb2xpZCIsIHN0YXJ0X2NvbG9yPSJGMkYyRjIi"
    "KQoKV0hJVEUgPSBGb250KG5hbWU9IkFyaWFsIiwgYm9sZD1UcnVlLCBjb2xvcj0iRkZGRkZGIiwg"
    "c2l6ZT0xMSkKQk9MRCAgPSBGb250KG5hbWU9IkFyaWFsIiwgYm9sZD1UcnVlLCBzaXplPTEwKQpS"
    "RUcgICA9IEZvbnQobmFtZT0iQXJpYWwiLCBzaXplPTEwKQpUSVRMRSA9IEZvbnQobmFtZT0iQXJp"
    "YWwiLCBib2xkPVRydWUsIHNpemU9MTQsIGNvbG9yPSJGRkZGRkYiKQoKdGhpbiA9IFNpZGUoYm9y"
    "ZGVyX3N0eWxlPSJ0aGluIiwgY29sb3I9IkI0QjRCNCIpCkJPUkRFUiA9IEJvcmRlcihsZWZ0PXRo"
    "aW4sIHJpZ2h0PXRoaW4sIHRvcD10aGluLCBib3R0b209dGhpbikKQ0VOVEVSID0gQWxpZ25tZW50"
    "KGhvcml6b250YWw9ImNlbnRlciIsIHZlcnRpY2FsPSJjZW50ZXIiLCB3cmFwX3RleHQ9VHJ1ZSkK"
    "TEVGVCAgID0gQWxpZ25tZW50KGhvcml6b250YWw9ImxlZnQiLCAgIHZlcnRpY2FsPSJjZW50ZXIi"
    "LCB3cmFwX3RleHQ9VHJ1ZSkKUklHSFQgID0gQWxpZ25tZW50KGhvcml6b250YWw9InJpZ2h0Iiwg"
    "IHZlcnRpY2FsPSJjZW50ZXIiKQoKTlVNX0ZNVCA9ICcjLCMjMC4wMDsoIywjIzAuMDApOyItIicK"
    "CgpkZWYgd3JpdGVfZXhjZWwocmV0dXJucywgb3V0cHV0X3BhdGgpOgogICAgIiIiCiAgICByZXR1"
    "cm5zOiBsaXN0IG9mIGRpY3RzOgogICAgICAgIHsgIm1ldGEiOiB7Li4ufSwgInJvd3MiOiBbIHsu"
    "Li59LCAuLi4gXSwgInNvdXJjZV9maWxlIjogIm5hbWUucGRmIiB9CiAgICAiIiIKICAgIHdiID0g"
    "V29ya2Jvb2soKQoKICAgICMgLS0tLSBTaGVldCAxOiBDb25zb2xpZGF0ZWQgLS0tLS0tLS0tLS0t"
    "LS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0KICAgIHdzID0gd2IuYWN0aXZlCiAgICB3cy50"
    "aXRsZSA9ICJDb25zb2xpZGF0ZWQiCgogICAgd3MubWVyZ2VfY2VsbHMoIkExOk8xIikKICAgIHdz"
    "WyJBMSJdID0gIkdTVFItMSBDT05TT0xJREFURUQg4oCUIFNUQVRFLVdJU0UgLyBNT05USC1XSVNF"
    "IFRBQkxFIFNVTU1BUlkiCiAgICB3c1siQTEiXS5mb250ID0gVElUTEUKICAgIHdzWyJBMSJdLmZp"
    "bGwgPSBIRFJfRklMTAogICAgd3NbIkExIl0uYWxpZ25tZW50ID0gQ0VOVEVSCiAgICB3cy5yb3df"
    "ZGltZW5zaW9uc1sxXS5oZWlnaHQgPSAyOAoKICAgIGhlYWRlcnMgPSBbCiAgICAgICAgIlNyIE5v"
    "IiwgIk1vbnRoIiwgIkZZIiwgIlN0YXRlIENvZGUiLCAiU3RhdGUgTmFtZSIsICJHU1RJTiIsCiAg"
    "ICAgICAgIkxlZ2FsIE5hbWUiLCAiQVJOIiwKICAgICAgICAiVGFibGUgTm8uIiwgIlRhYmxlIERl"
    "c2NyaXB0aW9uIiwKICAgICAgICAiTm8uIG9mIFJlY29yZHMiLAogICAgICAgICJUYXhhYmxlIFZh"
    "bHVlICjigrkpIiwgIklHU1QgKOKCuSkiLCAiQ0dTVCAo4oK5KSIsICJTR1NUL1VUR1NUICjigrkp"
    "IiwgIkNFU1MgKOKCuSkiCiAgICBdCiAgICBIRFJfUk9XID0gMwogICAgZm9yIGksIGggaW4gZW51"
    "bWVyYXRlKGhlYWRlcnMsIHN0YXJ0PTEpOgogICAgICAgIGMgPSB3cy5jZWxsKHJvdz1IRFJfUk9X"
    "LCBjb2x1bW49aSwgdmFsdWU9aCkKICAgICAgICBjLmZvbnQgPSBXSElURTsgYy5maWxsID0gU1VC"
    "X0ZJTEw7IGMuYWxpZ25tZW50ID0gQ0VOVEVSOyBjLmJvcmRlciA9IEJPUkRFUgogICAgd3Mucm93"
    "X2RpbWVuc2lvbnNbSERSX1JPV10uaGVpZ2h0ID0gMzYKCiAgICBzciA9IDAKICAgIHIgPSBIRFJf"
    "Uk9XICsgMQogICAgZGF0YV9zdGFydCA9IHIKCiAgICBmb3IgcmV0IGluIHJldHVybnM6CiAgICAg"
    "ICAgbSA9IHJldFsibWV0YSJdCiAgICAgICAgZm9yIHJvdyBpbiByZXRbInJvd3MiXToKICAgICAg"
    "ICAgICAgc3IgKz0gMQogICAgICAgICAgICB2YWxzID0gWwogICAgICAgICAgICAgICAgc3IsIG1b"
    "Im1vbnRoIl0sIG1bImZ5Il0sIG1bInN0YXRlX2NvZGUiXSwgbVsic3RhdGVfbmFtZSJdLCBtWyJn"
    "c3RpbiJdLAogICAgICAgICAgICAgICAgbVsibGVnYWxfbmFtZSJdLCBtWyJhcm4iXSwKICAgICAg"
    "ICAgICAgICAgIHJvd1sidGFibGVfbm8iXSwgcm93WyJkZXNjcmlwdGlvbiJdLAogICAgICAgICAg"
    "ICAgICAgcm93WyJyZWNvcmRzIl0sCiAgICAgICAgICAgICAgICByb3dbInRheGFibGUiXSwgcm93"
    "WyJpZ3N0Il0sIHJvd1siY2dzdCJdLCByb3dbInNnc3QiXSwgcm93WyJjZXNzIl0sCiAgICAgICAg"
    "ICAgIF0KICAgICAgICAgICAgZm9yIGksIHYgaW4gZW51bWVyYXRlKHZhbHMsIHN0YXJ0PTEpOgog"
    "ICAgICAgICAgICAgICAgY2VsbCA9IHdzLmNlbGwocm93PXIsIGNvbHVtbj1pLCB2YWx1ZT12KQog"
    "ICAgICAgICAgICAgICAgY2VsbC5mb250ID0gUkVHOyBjZWxsLmJvcmRlciA9IEJPUkRFUgogICAg"
    "ICAgICAgICAgICAgaWYgaSBpbiAoMSwgMiwgMywgNCwgOSwgMTEpOgogICAgICAgICAgICAgICAg"
    "ICAgIGNlbGwuYWxpZ25tZW50ID0gQ0VOVEVSCiAgICAgICAgICAgICAgICBlbGlmIGkgaW4gKDUs"
    "IDYsIDcsIDgsIDEwKToKICAgICAgICAgICAgICAgICAgICBjZWxsLmFsaWdubWVudCA9IExFRlQK"
    "ICAgICAgICAgICAgICAgIGVsc2U6CiAgICAgICAgICAgICAgICAgICAgY2VsbC5hbGlnbm1lbnQg"
    "PSBSSUdIVAogICAgICAgICAgICAgICAgICAgIGNlbGwubnVtYmVyX2Zvcm1hdCA9IE5VTV9GTVQK"
    "ICAgICAgICAgICAgciArPSAxCgogICAgZGF0YV9lbmQgPSByIC0gMQoKICAgIHdpZHRocyA9IHsi"
    "QSI6NywiQiI6OSwiQyI6OSwiRCI6NiwiRSI6MTgsIkYiOjIwLCJHIjoyNCwiSCI6MTgsCiAgICAg"
    "ICAgICAgICAgIkkiOjksIkoiOjQ2LCJLIjoxMSwiTCI6MTgsIk0iOjE2LCJOIjoxNiwiTyI6MTgs"
    "IlAiOjEyfQogICAgZm9yIGNvbCwgdyBpbiB3aWR0aHMuaXRlbXMoKToKICAgICAgICB3cy5jb2x1"
    "bW5fZGltZW5zaW9uc1tjb2xdLndpZHRoID0gdwoKICAgIHdzLmZyZWV6ZV9wYW5lcyA9ICJLNCIK"
    "ICAgIGlmIGRhdGFfZW5kID49IGRhdGFfc3RhcnQ6CiAgICAgICAgd3MuYXV0b19maWx0ZXIucmVm"
    "ID0gZiJBe0hEUl9ST1d9OlB7ZGF0YV9lbmR9IgoKICAgICMgLS0tLSBTaGVldCAyOiBTdGF0ZS1N"
    "b250aCBQaXZvdCBTdW1tYXJ5IC0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tCiAgICB3czIg"
    "PSB3Yi5jcmVhdGVfc2hlZXQoIlN0YXRlLU1vbnRoIFRvdGFscyIpCiAgICB3czIubWVyZ2VfY2Vs"
    "bHMoIkExOkgxIikKICAgIHdzMlsiQTEiXSA9ICJUT1RBTCBMSUFCSUxJVFkgUEVSIFJFVFVSTiAo"
    "VGFibGVzIDTigJMxMSwgZXhjbHVkZXMgMTIvMTMpIgogICAgd3MyWyJBMSJdLmZvbnQgPSBUSVRM"
    "RTsgd3MyWyJBMSJdLmZpbGwgPSBIRFJfRklMTDsgd3MyWyJBMSJdLmFsaWdubWVudCA9IENFTlRF"
    "UgogICAgd3MyLnJvd19kaW1lbnNpb25zWzFdLmhlaWdodCA9IDI4CgogICAgaDIgPSBbIlNyIE5v"
    "IiwgIk1vbnRoIiwgIlN0YXRlIENvZGUiLCAiU3RhdGUgTmFtZSIsICJHU1RJTiIsCiAgICAgICAg"
    "ICAiVGF4YWJsZSBWYWx1ZSAo4oK5KSIsICJJR1NUICjigrkpIiwgIkNHU1QgKOKCuSkiLCAiU0dT"
    "VC9VVEdTVCAo4oK5KSIsICJDRVNTICjigrkpIl0KICAgIGZvciBpLCBoIGluIGVudW1lcmF0ZSho"
    "Miwgc3RhcnQ9MSk6CiAgICAgICAgYyA9IHdzMi5jZWxsKHJvdz0zLCBjb2x1bW49aSwgdmFsdWU9"
    "aCkKICAgICAgICBjLmZvbnQgPSBXSElURTsgYy5maWxsID0gU1VCX0ZJTEw7IGMuYWxpZ25tZW50"
    "ID0gQ0VOVEVSOyBjLmJvcmRlciA9IEJPUkRFUgogICAgd3MyLnJvd19kaW1lbnNpb25zWzNdLmhl"
    "aWdodCA9IDM2CgogICAgc3IgPSAwOyByciA9IDQKICAgIEVYQ0xVREVfVEFCTEVTID0geyIxMiIs"
    "ICIxMyJ9ICAjIEhTTiBzdW1tYXJ5IGFuZCBEb2N1bWVudHMgPSBkdXBsaWNhdGVzIG9mIHN1cHBs"
    "aWVzIGFib3ZlCiAgICAjIEFsc28gc2tpcCB0aGUgNkEvNkIgZ3JvdXAtVG90YWwgcm93cyBiZWNh"
    "dXNlIHdlIGFscmVhZHkgY291bnQgdGhlaXIgRVhQV1AvRVhQV09QLwogICAgIyBTRVpXUC9TRVpX"
    "T1Agc3ViLXJvd3MgaW5kaXZpZHVhbGx5IOKAlCBpbmNsdWRpbmcgdGhlIFRvdGFsIHdvdWxkIGRv"
    "dWJsZS1jb3VudC4KICAgIEVYQ0xVREVfREVTQ1JJUFRJT05TID0gewogICAgICAgICJFeHBvcnRz"
    "IC0gVG90YWwgKDZBKSIsCiAgICAgICAgIlNFWiAtIFRvdGFsICg2QikiLAogICAgfQogICAgZm9y"
    "IHJldCBpbiByZXR1cm5zOgogICAgICAgIG0gPSByZXRbIm1ldGEiXQogICAgICAgIHRvdGFscyA9"
    "IHsidGF4YWJsZSI6MC4wLCJpZ3N0IjowLjAsImNnc3QiOjAuMCwic2dzdCI6MC4wLCJjZXNzIjow"
    "LjB9CiAgICAgICAgZm9yIHJvdyBpbiByZXRbInJvd3MiXToKICAgICAgICAgICAgaWYgcm93WyJ0"
    "YWJsZV9ubyJdIGluIEVYQ0xVREVfVEFCTEVTOgogICAgICAgICAgICAgICAgY29udGludWUKICAg"
    "ICAgICAgICAgaWYgcm93WyJkZXNjcmlwdGlvbiJdIGluIEVYQ0xVREVfREVTQ1JJUFRJT05TOgog"
    "ICAgICAgICAgICAgICAgY29udGludWUKICAgICAgICAgICAgZm9yIGsgaW4gdG90YWxzOgogICAg"
    "ICAgICAgICAgICAgdG90YWxzW2tdICs9IHJvd1trXQogICAgICAgIHNyICs9IDEKICAgICAgICBv"
    "dXQgPSBbc3IsIG1bIm1vbnRoIl0sIG1bInN0YXRlX2NvZGUiXSwgbVsic3RhdGVfbmFtZSJdLCBt"
    "WyJnc3RpbiJdLAogICAgICAgICAgICAgICB0b3RhbHNbInRheGFibGUiXSwgdG90YWxzWyJpZ3N0"
    "Il0sIHRvdGFsc1siY2dzdCJdLCB0b3RhbHNbInNnc3QiXSwgdG90YWxzWyJjZXNzIl1dCiAgICAg"
    "ICAgZm9yIGksIHYgaW4gZW51bWVyYXRlKG91dCwgc3RhcnQ9MSk6CiAgICAgICAgICAgIGNlbGwg"
    "PSB3czIuY2VsbChyb3c9cnIsIGNvbHVtbj1pLCB2YWx1ZT12KQogICAgICAgICAgICBjZWxsLmZv"
    "bnQgPSBSRUc7IGNlbGwuYm9yZGVyID0gQk9SREVSCiAgICAgICAgICAgIGlmIGkgPD0gNToKICAg"
    "ICAgICAgICAgICAgIGNlbGwuYWxpZ25tZW50ID0gQ0VOVEVSIGlmIGkgIT0gNCBlbHNlIExFRlQK"
    "ICAgICAgICAgICAgZWxzZToKICAgICAgICAgICAgICAgIGNlbGwuYWxpZ25tZW50ID0gUklHSFQ7"
    "IGNlbGwubnVtYmVyX2Zvcm1hdCA9IE5VTV9GTVQKICAgICAgICByciArPSAxCgogICAgd3MyLmNv"
    "bHVtbl9kaW1lbnNpb25zWyJBIl0ud2lkdGggPSA3CiAgICB3czIuY29sdW1uX2RpbWVuc2lvbnNb"
    "IkIiXS53aWR0aCA9IDEwCiAgICB3czIuY29sdW1uX2RpbWVuc2lvbnNbIkMiXS53aWR0aCA9IDEx"
    "CiAgICB3czIuY29sdW1uX2RpbWVuc2lvbnNbIkQiXS53aWR0aCA9IDIyCiAgICB3czIuY29sdW1u"
    "X2RpbWVuc2lvbnNbIkUiXS53aWR0aCA9IDIyCiAgICBmb3IgY29sIGluICgiRiIsIkciLCJIIiwi"
    "SSIsIkoiKToKICAgICAgICB3czIuY29sdW1uX2RpbWVuc2lvbnNbY29sXS53aWR0aCA9IDE4CiAg"
    "ICB3czIuZnJlZXplX3BhbmVzID0gIkE0IgogICAgaWYgcnIgPiA0OgogICAgICAgIHdzMi5hdXRv"
    "X2ZpbHRlci5yZWYgPSBmIkEzOkp7cnItMX0iCgogICAgIyAtLS0tIFNoZWV0IDM6IFByb2Nlc3Np"
    "bmcgTG9nIC0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0KICAgIHdzMyA9"
    "IHdiLmNyZWF0ZV9zaGVldCgiUHJvY2Vzc2luZyBMb2ciKQogICAgd3MzLmFwcGVuZChbIiMiLCAi"
    "U291cmNlIEZpbGUiLCAiU3RhdHVzIiwgIkdTVElOIiwgIlN0YXRlIiwgIk1vbnRoIiwgIlRvdGFs"
    "IFJvd3MiLCAiTm90ZXMiXSkKICAgIGZvciBjIGluIHJhbmdlKDEsIDkpOgogICAgICAgIGNlbGwg"
    "PSB3czMuY2VsbChyb3c9MSwgY29sdW1uPWMpCiAgICAgICAgY2VsbC5mb250ID0gV0hJVEU7IGNl"
    "bGwuZmlsbCA9IFNVQl9GSUxMOyBjZWxsLmFsaWdubWVudCA9IENFTlRFUjsgY2VsbC5ib3JkZXIg"
    "PSBCT1JERVIKICAgIGZvciBpLCByZXQgaW4gZW51bWVyYXRlKHJldHVybnMsIHN0YXJ0PTEpOgog"
    "ICAgICAgIG0gPSByZXRbIm1ldGEiXQogICAgICAgIG5vdGVzID0gcmV0LmdldCgibm90ZXMiLCAi"
    "IikKICAgICAgICBzdGF0dXMgPSByZXQuZ2V0KCJzdGF0dXMiLCAiT0siKQogICAgICAgIG91dCA9"
    "IFtpLCByZXQuZ2V0KCJzb3VyY2VfZmlsZSIsIiIpLCBzdGF0dXMsCiAgICAgICAgICAgICAgIG0u"
    "Z2V0KCJnc3RpbiIsIiIpLCBtLmdldCgic3RhdGVfbmFtZSIsIiIpLCBtLmdldCgibW9udGgiLCIi"
    "KSwKICAgICAgICAgICAgICAgbGVuKHJldC5nZXQoInJvd3MiLFtdKSksIG5vdGVzXQogICAgICAg"
    "IGZvciBjLCB2IGluIGVudW1lcmF0ZShvdXQsIHN0YXJ0PTEpOgogICAgICAgICAgICBjZWxsID0g"
    "d3MzLmNlbGwocm93PWkrMSwgY29sdW1uPWMsIHZhbHVlPXYpCiAgICAgICAgICAgIGNlbGwuZm9u"
    "dCA9IFJFRzsgY2VsbC5ib3JkZXIgPSBCT1JERVI7IGNlbGwuYWxpZ25tZW50ID0gTEVGVAogICAg"
    "ICAgICAgICBpZiBzdGF0dXMgIT0gIk9LIjoKICAgICAgICAgICAgICAgIGNlbGwuZmlsbCA9IEVS"
    "Ul9GSUxMCiAgICBmb3IgY29sLCB3IGluIFsoIkEiLDYpLCgiQiIsNDIpLCgiQyIsMTApLCgiRCIs"
    "MjIpLCgiRSIsMjIpLCgiRiIsMTApLCgiRyIsMTEpLCgiSCIsNTApXToKICAgICAgICB3czMuY29s"
    "dW1uX2RpbWVuc2lvbnNbY29sXS53aWR0aCA9IHcKICAgIHdzMy5mcmVlemVfcGFuZXMgPSAiQTIi"
    "CgogICAgd2Iuc2F2ZShvdXRwdXRfcGF0aCkKCgojIC0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0t"
    "LS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLSAjCiMgIE1h"
    "aW4KIyAtLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0t"
    "LS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0gIwoKZGVmIHJlc29sdmVfb3V0cHV0X3BhdGgocmF3KToK"
    "ICAgICIiIgogICAgTWFrZSB0aGUgb3V0cHV0IHBhdGggZm9yZ2l2aW5nOgogICAgICAtIElmIHVz"
    "ZXIgcGFzc2VkIGFuIGV4aXN0aW5nIGRpcmVjdG9yeSAgICAgICDihpIgYXBwZW5kIGRlZmF1bHQg"
    "ZmlsZW5hbWUKICAgICAgLSBJZiBwYXRoIGhhcyBubyBleHRlbnNpb24gICAgICAgICAgICAgICAg"
    "ICAg4oaSIGFwcGVuZCAueGxzeAogICAgICAtIElmIGV4dGVuc2lvbiBpcyBub3QgLnhsc3ggICAg"
    "ICAgICAgICAgICAgICDihpIgcmVwbGFjZSB3aXRoIC54bHN4CiAgICAgIC0gSWYgcGFyZW50IGRp"
    "cmVjdG9yeSBkb2Vzbid0IGV4aXN0ICAgICAgICAgIOKGkiBjcmVhdGUgaXQKICAgIFJldHVybnMg"
    "YSBQYXRoIHBvaW50aW5nIHRvIGEgd3JpdGFibGUgLnhsc3ggZmlsZSBsb2NhdGlvbi4KICAgICIi"
    "IgogICAgcCA9IFBhdGgocmF3KQogICAgREVGQVVMVF9OQU1FID0gIkdTVFIxX0NvbnNvbGlkYXRl"
    "ZC54bHN4IgoKICAgIGlmIHAuZXhpc3RzKCkgYW5kIHAuaXNfZGlyKCk6CiAgICAgICAgcCA9IHAg"
    "LyBERUZBVUxUX05BTUUKICAgIGVsaWYgcC5zdWZmaXgubG93ZXIoKSAhPSAiLnhsc3giOgogICAg"
    "ICAgIHAgPSBwLndpdGhfc3VmZml4KCIueGxzeCIpCgogICAgIyBNYWtlIHN1cmUgcGFyZW50IGZv"
    "bGRlciBleGlzdHMKICAgIHAucGFyZW50Lm1rZGlyKHBhcmVudHM9VHJ1ZSwgZXhpc3Rfb2s9VHJ1"
    "ZSkKCiAgICAjIElmIHRoZSByZXNvbHZlZCBmaWxlIGFscmVhZHkgZXhpc3RzIGFuZCBpcyBsb2Nr"
    "ZWQgKG9wZW4gaW4gRXhjZWwpLCB3YXJuIGVhcmx5CiAgICBpZiBwLmV4aXN0cygpOgogICAgICAg"
    "IHRyeToKICAgICAgICAgICAgd2l0aCBvcGVuKHAsICJhYiIpOgogICAgICAgICAgICAgICAgcGFz"
    "cwogICAgICAgIGV4Y2VwdCBQZXJtaXNzaW9uRXJyb3I6CiAgICAgICAgICAgIHJhaXNlIFBlcm1p"
    "c3Npb25FcnJvcigKICAgICAgICAgICAgICAgIGYiT3V0cHV0IGZpbGUgaXMgbG9ja2VkIChwcm9i"
    "YWJseSBvcGVuIGluIEV4Y2VsKTpcbiAge3B9XG4iCiAgICAgICAgICAgICAgICAiQ2xvc2UgaXQg"
    "aW4gRXhjZWwgYW5kIHJlLXJ1bi4iCiAgICAgICAgICAgICkKICAgIHJldHVybiBwCgoKZGVmIHBy"
    "b2Nlc3NfcGRmcyhpbl9mb2xkZXIsIG91dF9wYXRoLCBvbl9wcm9ncmVzcz1Ob25lLCBvbl9sb2c9"
    "Tm9uZSk6CiAgICAiIiIKICAgIENvcmUgcHJvY2Vzc2luZyBwaXBlbGluZSByZXVzYWJsZSBieSBD"
    "TEkgYW5kIEdVSS4KCiAgICBDYWxsYmFja3M6CiAgICAgICAgb25fcHJvZ3Jlc3MoaWR4LCB0b3Rh"
    "bCkgICAgICAgICAg4oaSIGZpcmVzIGFmdGVyIGV2ZXJ5IFBERiAoZm9yIHByb2dyZXNzIGJhcikK"
    "ICAgICAgICBvbl9sb2cobGluZSwgbGV2ZWwpICAgICAgICAgICAgICDihpIgZmlyZXMgZm9yIGV2"
    "ZXJ5IHRleHR1YWwgdXBkYXRlOwogICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAg"
    "ICAgICAgIGxldmVsIGlzICJvayIgLyAiZmFpbCIgLyAiaW5mbyIKCiAgICBSZXR1cm5zOiAob2tf"
    "Y291bnQsIGZhaWxfY291bnQsIHRvdGFsX2NvdW50LCBvdXRfcGF0aCkKICAgICIiIgogICAgcGRm"
    "cyA9IHNvcnRlZChbcCBmb3IgcCBpbiBpbl9mb2xkZXIucmdsb2IoIioucGRmIildKQogICAgaWYg"
    "bm90IHBkZnM6CiAgICAgICAgcmFpc2UgRmlsZU5vdEZvdW5kRXJyb3IoZiJObyBQREZzIGZvdW5k"
    "IHVuZGVyOiB7aW5fZm9sZGVyfSIpCgogICAgaWYgb25fbG9nOgogICAgICAgIG9uX2xvZyhmIkZv"
    "dW5kIHtsZW4ocGRmcyl9IFBERiBmaWxlKHMpLiAgUHJvY2Vzc2luZy4uLiIsICJpbmZvIikKCiAg"
    "ICByZXR1cm5zID0gW10KICAgIG9rID0gMAogICAgZmFpbCA9IDAKCiAgICBmb3IgaWR4LCBwZGZf"
    "cGF0aCBpbiBlbnVtZXJhdGUocGRmcywgc3RhcnQ9MSk6CiAgICAgICAgdHJ5OgogICAgICAgICAg"
    "ICByZWwgPSBwZGZfcGF0aC5yZWxhdGl2ZV90byhpbl9mb2xkZXIpCiAgICAgICAgZXhjZXB0IFZh"
    "bHVlRXJyb3I6CiAgICAgICAgICAgIHJlbCA9IHBkZl9wYXRoLm5hbWUKCiAgICAgICAgdHJ5Ogog"
    "ICAgICAgICAgICBtZXRhLCByb3dzID0gcGFyc2VfcGRmKHBkZl9wYXRoKQogICAgICAgICAgICBy"
    "ZXR1cm5zLmFwcGVuZCh7CiAgICAgICAgICAgICAgICAibWV0YSI6IG1ldGEsICJyb3dzIjogcm93"
    "cywKICAgICAgICAgICAgICAgICJzb3VyY2VfZmlsZSI6IHN0cihyZWwpLAogICAgICAgICAgICAg"
    "ICAgInN0YXR1cyI6ICJPSyIsICJub3RlcyI6ICIiLAogICAgICAgICAgICB9KQogICAgICAgICAg"
    "ICBvayArPSAxCiAgICAgICAgICAgIGlmIG9uX2xvZzoKICAgICAgICAgICAgICAgIG9uX2xvZygK"
    "ICAgICAgICAgICAgICAgICAgICBmIlt7aWR4Oj40fS97bGVuKHBkZnMpfV0gIE9LICAgICIKICAg"
    "ICAgICAgICAgICAgICAgICBmInttZXRhWydzdGF0ZV9uYW1lJ106PDIyfSB7bWV0YVsnbW9udGgn"
    "XTo8OH0gICIKICAgICAgICAgICAgICAgICAgICBmInttZXRhWydnc3RpbiddfSAgKHtwZGZfcGF0"
    "aC5uYW1lfSkiLAogICAgICAgICAgICAgICAgICAgICJvayIsCiAgICAgICAgICAgICAgICApCiAg"
    "ICAgICAgZXhjZXB0IEV4Y2VwdGlvbiBhcyBlOgogICAgICAgICAgICBmYWlsICs9IDEKICAgICAg"
    "ICAgICAgcmV0dXJucy5hcHBlbmQoewogICAgICAgICAgICAgICAgIm1ldGEiOiB7ImdzdGluIjoi"
    "Iiwic3RhdGVfY29kZSI6IiIsInN0YXRlX25hbWUiOiIiLCJtb250aCI6IiIsCiAgICAgICAgICAg"
    "ICAgICAgICAgICAgICAiZnkiOiIiLCJsZWdhbF9uYW1lIjoiIiwiYXJuIjoiIn0sCiAgICAgICAg"
    "ICAgICAgICAicm93cyI6IFtdLAogICAgICAgICAgICAgICAgInNvdXJjZV9maWxlIjogc3RyKHJl"
    "bCksCiAgICAgICAgICAgICAgICAic3RhdHVzIjogIkZBSUwiLAogICAgICAgICAgICAgICAgIm5v"
    "dGVzIjogZiJ7dHlwZShlKS5fX25hbWVfX306IHtlfSIsCiAgICAgICAgICAgIH0pCiAgICAgICAg"
    "ICAgIGlmIG9uX2xvZzoKICAgICAgICAgICAgICAgIG9uX2xvZygKICAgICAgICAgICAgICAgICAg"
    "ICBmIlt7aWR4Oj40fS97bGVuKHBkZnMpfV0gIEZBSUwgIHtwZGZfcGF0aC5uYW1lfSAg4oaSICB7"
    "ZX0iLAogICAgICAgICAgICAgICAgICAgICJmYWlsIiwKICAgICAgICAgICAgICAgICkKCiAgICAg"
    "ICAgaWYgb25fcHJvZ3Jlc3M6CiAgICAgICAgICAgIG9uX3Byb2dyZXNzKGlkeCwgbGVuKHBkZnMp"
    "KQoKICAgIGlmIG9uX2xvZzoKICAgICAgICBvbl9sb2coZiJXcml0aW5nIGNvbnNvbGlkYXRlZCBF"
    "eGNlbDogIHtvdXRfcGF0aH0iLCAiaW5mbyIpCiAgICB3cml0ZV9leGNlbChyZXR1cm5zLCBvdXRf"
    "cGF0aCkKICAgIHJldHVybiBvaywgZmFpbCwgbGVuKHBkZnMpLCBvdXRfcGF0aAoKCiMgLS0tLS0t"
    "LS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0t"
    "LS0tLS0tLS0tLS0tICMKIyAgR1VJICh0a2ludGVyIOKAlCBidW5kbGVkIHdpdGggUHl0aG9uIG9u"
    "IFdpbmRvd3MgLyBNYWMpCiMgLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0t"
    "LS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tICMKCg=="
)

_GSTR3B_ENGINE_B64 = (
    "U1RBVEVfQ09ERVMgPSB7CiAgICAiMDEiOiAiSmFtbXUgJiBLYXNobWlyIiwgIjAyIjogIkhpbWFj"
    "aGFsIFByYWRlc2giLCAiMDMiOiAiUHVuamFiIiwKICAgICIwNCI6ICJDaGFuZGlnYXJoIiwgIjA1"
    "IjogIlV0dGFyYWtoYW5kIiwgIjA2IjogIkhhcnlhbmEiLCAiMDciOiAiRGVsaGkiLAogICAgIjA4"
    "IjogIlJhamFzdGhhbiIsICIwOSI6ICJVdHRhciBQcmFkZXNoIiwgIjEwIjogIkJpaGFyIiwgIjEx"
    "IjogIlNpa2tpbSIsCiAgICAiMTIiOiAiQXJ1bmFjaGFsIFByYWRlc2giLCAiMTMiOiAiTmFnYWxh"
    "bmQiLCAiMTQiOiAiTWFuaXB1ciIsICIxNSI6ICJNaXpvcmFtIiwKICAgICIxNiI6ICJUcmlwdXJh"
    "IiwgIjE3IjogIk1lZ2hhbGF5YSIsICIxOCI6ICJBc3NhbSIsICIxOSI6ICJXZXN0IEJlbmdhbCIs"
    "CiAgICAiMjAiOiAiSmhhcmtoYW5kIiwgIjIxIjogIk9kaXNoYSIsICIyMiI6ICJDaGhhdHRpc2dh"
    "cmgiLCAiMjMiOiAiTWFkaHlhIFByYWRlc2giLAogICAgIjI0IjogIkd1amFyYXQiLCAiMjUiOiAi"
    "RGFtYW4gJiBEaXUiLCAiMjYiOiAiRGFkcmEgJiBOYWdhciBIYXZlbGkgYW5kIERhbWFuICYgRGl1"
    "IiwKICAgICIyNyI6ICJNYWhhcmFzaHRyYSIsICIyOCI6ICJBbmRocmEgUHJhZGVzaCAoT2xkKSIs"
    "ICIyOSI6ICJLYXJuYXRha2EiLAogICAgIjMwIjogIkdvYSIsICIzMSI6ICJMYWtzaGFkd2VlcCIs"
    "ICIzMiI6ICJLZXJhbGEiLCAiMzMiOiAiVGFtaWwgTmFkdSIsCiAgICAiMzQiOiAiUHVkdWNoZXJy"
    "eSIsICIzNSI6ICJBbmRhbWFuICYgTmljb2JhciBJc2xhbmRzIiwgIjM2IjogIlRlbGFuZ2FuYSIs"
    "CiAgICAiMzciOiAiQW5kaHJhIFByYWRlc2giLCAiMzgiOiAiTGFkYWtoIiwgIjk3IjogIk90aGVy"
    "IFRlcnJpdG9yeSIsICI5OSI6ICJDZW50cmUgSnVyaXNkaWN0aW9uIiwKfQoKTU9OVEhfQUJCUiA9"
    "IHsKICAgICJKYW51YXJ5IjogIkphbiIsICJGZWJydWFyeSI6ICJGZWIiLCAiTWFyY2giOiAiTWFy"
    "IiwgIkFwcmlsIjogIkFwciIsCiAgICAiTWF5IjogICAgICJNYXkiLCAiSnVuZSI6ICAgICAiSnVu"
    "IiwgIkp1bHkiOiAgIkp1bCIsICJBdWd1c3QiOiAgICJBdWciLAogICAgIlNlcHRlbWJlciI6ICJT"
    "ZXAiLCAiT2N0b2JlciI6ICJPY3QiLCAiTm92ZW1iZXIiOiAiTm92IiwgIkRlY2VtYmVyIjogIkRl"
    "YyIsCn0KCgojIC0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0t"
    "LS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLSAjCiMgIENlbGwgY2xlYW5lcnMgYW5kIG51bWJl"
    "ciBwYXJzaW5nCiMgLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0t"
    "LS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tICMKCldTX1JFID0gcmUuY29tcGlsZShyIlxz"
    "KyIpCiMgV2F0ZXJtYXJrICJGSUxFRCIgZGlhZ29uYWxseSBvdmVybGF5cyBlYWNoIHBhZ2UuIHBk"
    "ZnBsdW1iZXIgc29tZXRpbWVzIGluamVjdHMKIyBpc29sYXRlZCBzaW5nbGUgdXBwZXJjYXNlIGxl"
    "dHRlcnMgKEYsIEksIEwsIEUsIEQpIGFzIGNlbGwgcHJlZml4ZXMgb3IgaW4gdGhlCiMgbWlkZGxl"
    "IG9mIG51bWVyaWMgdG9rZW5zLiBXZSBzdHJpcCB0aG9zZSBkZWZlbnNpdmVseS4KTEVBRElOR19M"
    "RVRURVJfUkUgPSByZS5jb21waWxlKHIiXltBLVpdXHMrIikKTEVBRElOR19MRVRURVJTX05VTV9S"
    "RSA9IHJlLmNvbXBpbGUociJeW0EtWmEtel0rKD89W1xkXChcLV0pIikKCgpkZWYgY2xlYW5fY2Vs"
    "bChzKToKICAgICIiIk5vcm1hbGl6ZSB3aGl0ZXNwYWNlLCBzdHJpcCB3YXRlcm1hcmsgc2luZ2xl"
    "LWxldHRlciBwcmVmaXhlcy4iIiIKICAgIGlmIHMgaXMgTm9uZToKICAgICAgICByZXR1cm4gIiIK"
    "ICAgIHMgPSBzdHIocykKICAgICMgQ29sbGFwc2UgYWxsIHdoaXRlc3BhY2UgKGluY2wuIG5ld2xp"
    "bmVzIGluc2lkZSBjZWxscyBmcm9tIG11bHRpLWxpbmUgd3JhcHMpCiAgICBzID0gV1NfUkUuc3Vi"
    "KCIgIiwgcykuc3RyaXAoKQogICAgIyBTdHJpcCBsZWFkaW5nICdYICcgKHNpbmdsZSB1cHBlcmNh"
    "c2UgbGV0dGVyICsgc3BhY2UpIOKAlCB3YXRlcm1hcmsgYXJ0aWZhY3QKICAgIHMgPSBMRUFESU5H"
    "X0xFVFRFUl9SRS5zdWIoIiIsIHMpCiAgICByZXR1cm4gcy5zdHJpcCgpCgoKZGVmIHBhcnNlX251"
    "bV9jZWxsKHMpOgogICAgIiIiCiAgICBDb252ZXJ0IGEgbnVtZXJpYyBjZWxsIHRvIGZsb2F0Lgog"
    "ICAgSGFuZGxlczoKICAgICAgLSBlbXB0eSAvICctJyAvIE5vbmUgICAgICAgIOKGkiAwLjAKICAg"
    "ICAgLSAnTDM2NjA5MDIuMDAnICAgICAgICAgICAgICDihpIgMzY2MDkwMi4wMCAgKHN0cmlwIHdh"
    "dGVybWFyayBsZXR0ZXIgcHJlZml4KQogICAgICAtICcyNzE5OTc2OSAuMDAnICAgICAgICAgICAg"
    "IOKGkiAyNzE5OTc2OS4wMCAoc3RyaXAgaW50ZXJuYWwgd2hpdGVzcGFjZSBmcm9tIHdyYXBwZWQg"
    "Y29sKQogICAgICAtICcxMTc3MDIuMCAwJyAgICAgICAgICAgICAgIOKGkiAxMTc3MDIuMDAgICAo"
    "c2FtZSkKICAgICAgLSAnKDUwLDAwMC4wMCknICAgICAgICAgICAgICDihpIgLTUwMDAwLjAwICAg"
    "KHBhcmVuIG5lZ2F0aXZlKQogICAgICAtICctMTIsMzQ1LjAwJyAvICcxMiwzNDUuMDAtJ+KGkiAt"
    "MTIzNDUuMDAgICAobGVhZC90cmFpbCBtaW51cykKICAgICAgLSAnMSwyMyw0NTYuNzgnIG9yICcx"
    "MjM0NTYuNzgnIChJbmRpYW4gb3IgcGxhaW4gY29tbWEgZm9ybWF0KSDihpIgMTIzNDU2Ljc4CiAg"
    "ICBBbnl0aGluZyB0aGF0IHdvbid0IHBhcnNlIGNsZWFubHkgcmV0dXJucyAwLjAuCiAgICAiIiIK"
    "ICAgIGlmIHMgaXMgTm9uZToKICAgICAgICByZXR1cm4gMC4wCiAgICBzID0gY2xlYW5fY2VsbChz"
    "KQogICAgaWYgcyA9PSAiIiBvciBzID09ICItIjoKICAgICAgICByZXR1cm4gMC4wCiAgICAjIFN0"
    "cmlwIEFMTCBpbnRlcm5hbCB3aGl0ZXNwYWNlIOKAlCBoYW5kbGVzIGZyYWdtZW50ZWQgbnVtYmVy"
    "cyBsaWtlICIxMTc3MDIuMCAwIgogICAgcyA9IHJlLnN1YihyIlxzKyIsICIiLCBzKQogICAgIyBT"
    "dHJpcCBsZWFkaW5nIG5vbi1udW1lcmljIHdhdGVybWFyayBsZXR0ZXJzIChlLmcuLCAnTDM2NjA5"
    "MDIuMDAnKQogICAgcyA9IExFQURJTkdfTEVUVEVSU19OVU1fUkUuc3ViKCIiLCBzKQogICAgaWYg"
    "cyA9PSAiIiBvciBzID09ICItIjoKICAgICAgICByZXR1cm4gMC4wCiAgICAjIERldGVjdCBzaWdu"
    "CiAgICBzaWduID0gMS4wCiAgICBpZiBzLnN0YXJ0c3dpdGgoIigiKSBhbmQgcy5lbmRzd2l0aCgi"
    "KSIpOgogICAgICAgIHNpZ24gPSAtMS4wCiAgICAgICAgcyA9IHNbMTotMV0KICAgIGVsaWYgcy5z"
    "dGFydHN3aXRoKCItIik6CiAgICAgICAgc2lnbiA9IC0xLjAKICAgICAgICBzID0gc1sxOl0KICAg"
    "IGVsaWYgcy5lbmRzd2l0aCgiLSIpOgogICAgICAgIHNpZ24gPSAtMS4wCiAgICAgICAgcyA9IHNb"
    "Oi0xXQogICAgIyBSZW1vdmUgdGhvdXNhbmRzIHNlcGFyYXRvcnMgKEluZGlhbiBvciB3ZXN0ZXJu"
    "KQogICAgcyA9IHMucmVwbGFjZSgiLCIsICIiKQogICAgdHJ5OgogICAgICAgIHJldHVybiBzaWdu"
    "ICogZmxvYXQocykKICAgIGV4Y2VwdCBWYWx1ZUVycm9yOgogICAgICAgIHJldHVybiAwLjAKCgpk"
    "ZWYgY2VsbF9zdGFydHNfd2l0aChjZWxsLCBwcmVmaXgpOgogICAgIiIiQ2FzZS1zZW5zaXRpdmUg"
    "cHJlZml4IGNoZWNrIG9uIGNsZWFuZWQgY2VsbC4iIiIKICAgIHJldHVybiBjbGVhbl9jZWxsKGNl"
    "bGwpLnN0YXJ0c3dpdGgocHJlZml4KQoKCmRlZiByb3dfdGV4dChyb3cpOgogICAgIiIiSm9pbiBh"
    "IHRhYmxlIHJvdydzIGNlbGxzIGludG8gb25lIHNwYWNlLXNlcGFyYXRlZCBzdHJpbmcgZm9yIGNv"
    "bnRlbnQgbWF0Y2hpbmcuIiIiCiAgICByZXR1cm4gIiAiLmpvaW4oY2xlYW5fY2VsbChjKSBmb3Ig"
    "YyBpbiByb3cgaWYgYyBpcyBub3QgTm9uZSkKCgpkZWYgdGFibGVfdGV4dCh0YWJsZSk6CiAgICAi"
    "IiJBbGwgdGV4dCBpbiBhIHRhYmxlLCBsb3dlcmNhc2VkIOKAlCBmb3Igc2VjdGlvbiBpZGVudGlm"
    "aWNhdGlvbi4iIiIKICAgIHJldHVybiAiICIuam9pbihyb3dfdGV4dChyKSBmb3IgciBpbiB0YWJs"
    "ZSkubG93ZXIoKQoKCiMgLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0t"
    "LS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tICMKIyAgU2VjdGlvbiBpZGVudGlmaWNh"
    "dGlvbiAmIHBhcnNpbmcKIyAtLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0t"
    "LS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0gIwoKZGVmIGlkZW50aWZ5X3NlY3Rp"
    "b24odGFibGUpOgogICAgIiIiSWRlbnRpZnkgd2hpY2ggR1NUUi0zQiBzZWN0aW9uIHRoaXMgZXh0"
    "cmFjdGVkIHRhYmxlIGJlbG9uZ3MgdG8uIiIiCiAgICB0ZXh0ID0gdGFibGVfdGV4dCh0YWJsZSkK"
    "CiAgICBpZiAidGF4IHBhaWQgdGhyb3VnaCBpdGMiIGluIHRleHQgb3IgKCJkZXNjcmlwdGkiIGlu"
    "IHRleHQgYW5kICJwYXlhYmxlIiBpbiB0ZXh0KToKICAgICAgICByZXR1cm4gIjYuMSIKICAgIGlm"
    "ICJzeXN0ZW0gY29tcHV0ZWQiIGluIHRleHQgYW5kICgiaW50ZXJlc3QiIGluIHRleHQgb3IgImxh"
    "dGUgZmVlIiBpbiB0ZXh0KToKICAgICAgICByZXR1cm4gIjUuMSIKICAgIGlmICJpbnRlci0gc3Rh"
    "dGUiIGluIHRleHQgb3IgImludGVyLXN0YXRlIHN1cHBsaWVzIiBpbiB0ZXh0OgogICAgICAgIGlm"
    "ICJjb21wb3NpdGlvbiBzY2hlbWUiIGluIHRleHQgb3IgIm5vbiBnc3Qgc3VwcGx5IiBpbiB0ZXh0"
    "IG9yICJuaWwgcmF0ZWQgc3VwcGx5IiBpbiB0ZXh0OgogICAgICAgICAgICByZXR1cm4gIjUiCiAg"
    "ICBpZiAiKGEpIG91dHdhcmQgdGF4YWJsZSIgaW4gdGV4dDoKICAgICAgICByZXR1cm4gIjMuMSIK"
    "ICAgIGlmICJlbGVjdHJvbmljIGNvbW1lcmNlIG9wZXJhdG9yIiBpbiB0ZXh0IGFuZCAiOSg1KSIg"
    "aW4gdGV4dDoKICAgICAgICByZXR1cm4gIjMuMS4xIgogICAgaWYgInVucmVnaXN0ZXJlZCBwZXJz"
    "b25zIiBpbiB0ZXh0IGFuZCAoImNvbXBvc2l0aW9uIHRheGFibGUiIGluIHRleHQgb3IgInVpbiBo"
    "b2xkZXJzIiBpbiB0ZXh0KToKICAgICAgICByZXR1cm4gIjMuMiIKICAgIGlmIGFueShrIGluIHRl"
    "eHQgZm9yIGsgaW4gKAogICAgICAgICJpdGMgYXZhaWxhYmxlIiwgIml0YyByZXZlcnNlZCIsICJu"
    "ZXQgaXRjIGF2YWlsYWJsZSIsCiAgICAgICAgImltcG9ydCBvZiBnb29kcyIsICJpbXBvcnQgb2Yg"
    "c2VydmljZXMiLCAiYWxsIG90aGVyIGl0YyIsCiAgICAgICAgImluZWxpZ2libGUgaXRjIiwgIml0"
    "YyByZWNsYWltZWQiLAogICAgKSk6CiAgICAgICAgcmV0dXJuICI0IgogICAgaWYgImJyZWFrdXAg"
    "b2YgdGF4IGxpYWJpbGl0eSIgaW4gdGV4dDoKICAgICAgICByZXR1cm4gImJyZWFrdXAiCiAgICAj"
    "IEJyZWFrdXAgc2VjdGlvbiBkYXRhIHJvdyBjYW4gYmUgYSBzdGFuZGFsb25lIHNpbmdsZS1yb3cg"
    "dGFibGUgb24gYSBsYXRlciBwYWdlCiAgICAjIHdpdGggZmlyc3QgY2VsbCBzaGFwZWQgbGlrZSAi"
    "SmFudWFyeSAyMDI2Ii4gRGV0ZWN0IHRoYXQgcGF0dGVybiBleHBsaWNpdGx5LgogICAgaWYgdGFi"
    "bGUgYW5kIGxlbih0YWJsZSkgPD0gMjoKICAgICAgICBmaXJzdCA9IGNsZWFuX2NlbGwodGFibGVb"
    "MF1bMF0pIGlmIHRhYmxlWzBdIGVsc2UgIiIKICAgICAgICBpZiByZS5tYXRjaChyIl5bQS1aXVth"
    "LXpdK1xzK1xkezR9JCIsIGZpcnN0KToKICAgICAgICAgICAgcmV0dXJuICJicmVha3VwIgogICAg"
    "IyBCcmVha3VwIGhlYWRlci1vbmx5IHRhYmxlIChQZXJpb2QgfCBJbnRlZ3JhdGVkIHRheCB8IENl"
    "bnRyYWwgdGF4IHwgLi4uKQogICAgaWYgInBlcmlvZCIgaW4gdGV4dCBhbmQgImludGVncmF0ZWQg"
    "dGF4IiBpbiB0ZXh0IGFuZCBsZW4odGFibGUpIDw9IDI6CiAgICAgICAgcmV0dXJuICJicmVha3Vw"
    "IgogICAgcmV0dXJuIE5vbmUKCgpkZWYgcGFyc2VfM18xKHRhYmxlKToKICAgICIiIlBhcnNlIFRh"
    "YmxlIDMuMSDigJQgNSByb3dzIChhKS0oZSkuIiIiCiAgICBtYXJrZXJzID0gWwogICAgICAgICgi"
    "KGEpIiwgICIzLjEoYSkiLCAiT3V0d2FyZCB0YXhhYmxlIHN1cHBsaWVzIChvdGhlciB0aGFuIHpl"
    "cm8vbmlsL2V4ZW1wdGVkKSIpLAogICAgICAgICgiKGIpIiwgICIzLjEoYikiLCAiT3V0d2FyZCB0"
    "YXhhYmxlIHN1cHBsaWVzICh6ZXJvIHJhdGVkKSIpLAogICAgICAgICgiKGMiLCAgICIzLjEoYyki"
    "LCAiT3RoZXIgb3V0d2FyZCBzdXBwbGllcyAobmlsIHJhdGVkLCBleGVtcHRlZCkiKSwgICMgUERG"
    "IHVzZXMgIihjICkiIHdpdGggc3BhY2UKICAgICAgICAoIihkKSIsICAiMy4xKGQpIiwgIklud2Fy"
    "ZCBzdXBwbGllcyAobGlhYmxlIHRvIHJldmVyc2UgY2hhcmdlKSIpLAogICAgICAgICgiKGUpIiwg"
    "ICIzLjEoZSkiLCAiTm9uLUdTVCBvdXR3YXJkIHN1cHBsaWVzIiksCiAgICBdCiAgICBvdXQgPSBb"
    "XQogICAgZm9yIHJvdyBpbiB0YWJsZToKICAgICAgICBpZiBub3Qgcm93OgogICAgICAgICAgICBj"
    "b250aW51ZQogICAgICAgIGZpcnN0ID0gY2xlYW5fY2VsbChyb3dbMF0pCiAgICAgICAgZm9yIHBy"
    "ZWZpeCwgc2VjdGlvbiwgZGVzYyBpbiBtYXJrZXJzOgogICAgICAgICAgICBpZiBmaXJzdC5zdGFy"
    "dHN3aXRoKHByZWZpeCk6CiAgICAgICAgICAgICAgICAjIEV4cGVjdCA1IG51bWVyaWMgY2VsbHMg"
    "YWZ0ZXIgZGVzY3JpcHRpb24KICAgICAgICAgICAgICAgIHZhbHMgPSBbcGFyc2VfbnVtX2NlbGwo"
    "cm93W2ldKSBpZiBpIDwgbGVuKHJvdykgZWxzZSAwLjAgZm9yIGkgaW4gcmFuZ2UoMSwgNildCiAg"
    "ICAgICAgICAgICAgICBvdXQuYXBwZW5kKHsKICAgICAgICAgICAgICAgICAgICAic2VjdGlvbiI6"
    "IHNlY3Rpb24sICJkZXNjcmlwdGlvbiI6IGRlc2MsCiAgICAgICAgICAgICAgICAgICAgInRheGFi"
    "bGUiOiB2YWxzWzBdLCAiaWdzdCI6IHZhbHNbMV0sCiAgICAgICAgICAgICAgICAgICAgImNnc3Qi"
    "OiB2YWxzWzJdLCAic2dzdCI6IHZhbHNbM10sICJjZXNzIjogdmFsc1s0XSwKICAgICAgICAgICAg"
    "ICAgIH0pCiAgICAgICAgICAgICAgICBicmVhawogICAgcmV0dXJuIG91dAoKCmRlZiBwYXJzZV8z"
    "XzFfMSh0YWJsZSk6CiAgICAiIiJQYXJzZSBUYWJsZSAzLjEuMSDigJQgMiByb3dzOiAoaSkgYW5k"
    "IChpaSkuIiIiCiAgICBtYXJrZXJzID0gWwogICAgICAgICgiKGkpIiwgICIzLjEuMShpKSIsICAi"
    "VGF4YWJsZSBzdXBwbGllcyBvbiB3aGljaCBFQ08gcGF5cyB0YXggdS9zIDkoNSkiKSwKICAgICAg"
    "ICAoIihpaSkiLCAiMy4xLjEoaWkpIiwgIlRheGFibGUgc3VwcGxpZXMgbWFkZSBieSBSUCB0aHJv"
    "dWdoIEVDTyB1L3MgOSg1KSIpLAogICAgXQogICAgb3V0ID0gW10KICAgIGZvciByb3cgaW4gdGFi"
    "bGU6CiAgICAgICAgaWYgbm90IHJvdzoKICAgICAgICAgICAgY29udGludWUKICAgICAgICBmaXJz"
    "dCA9IGNsZWFuX2NlbGwocm93WzBdKQogICAgICAgIGZvciBwcmVmaXgsIHNlY3Rpb24sIGRlc2Mg"
    "aW4gbWFya2VyczoKICAgICAgICAgICAgIyBOZWVkIHRvIGJlIGNhcmVmdWw6ICIoaWkpIiBtdXN0"
    "IG1hdGNoIGJlZm9yZSAiKGkpIiBiZWNhdXNlICIoaSkiIGlzIGEgcHJlZml4IG9mICIoaWkpIgog"
    "ICAgICAgICAgICAjIE9yZGVyIHRoZW0gY2FyZWZ1bGx5IGJlbG93CiAgICAgICAgICAgIHBhc3MK"
    "ICAgICAgICAjIFJlLXRlc3QgaW4gY29ycmVjdCBwcmVjZWRlbmNlIG9yZGVyCiAgICAgICAgaWYg"
    "Zmlyc3Quc3RhcnRzd2l0aCgiKGlpKSIpIG9yICIgKGlpKSAiIGluICgiICIgKyBmaXJzdCArICIg"
    "Iik6CiAgICAgICAgICAgIHZhbHMgPSBbcGFyc2VfbnVtX2NlbGwocm93W2ldKSBpZiBpIDwgbGVu"
    "KHJvdykgZWxzZSAwLjAgZm9yIGkgaW4gcmFuZ2UoMSwgNildCiAgICAgICAgICAgIG91dC5hcHBl"
    "bmQoewogICAgICAgICAgICAgICAgInNlY3Rpb24iOiAiMy4xLjEoaWkpIiwgImRlc2NyaXB0aW9u"
    "IjogIlRheGFibGUgc3VwcGxpZXMgbWFkZSBieSBSUCB0aHJvdWdoIEVDTyB1L3MgOSg1KSIsCiAg"
    "ICAgICAgICAgICAgICAidGF4YWJsZSI6IHZhbHNbMF0sICJpZ3N0IjogdmFsc1sxXSwKICAgICAg"
    "ICAgICAgICAgICJjZ3N0IjogdmFsc1syXSwgInNnc3QiOiB2YWxzWzNdLCAiY2VzcyI6IHZhbHNb"
    "NF0sCiAgICAgICAgICAgIH0pCiAgICAgICAgZWxpZiBmaXJzdC5zdGFydHN3aXRoKCIoaSkiKSBv"
    "ciAiIChpKSAiIGluICgiICIgKyBmaXJzdCArICIgIik6CiAgICAgICAgICAgIHZhbHMgPSBbcGFy"
    "c2VfbnVtX2NlbGwocm93W2ldKSBpZiBpIDwgbGVuKHJvdykgZWxzZSAwLjAgZm9yIGkgaW4gcmFu"
    "Z2UoMSwgNildCiAgICAgICAgICAgIG91dC5hcHBlbmQoewogICAgICAgICAgICAgICAgInNlY3Rp"
    "b24iOiAiMy4xLjEoaSkiLCAiZGVzY3JpcHRpb24iOiAiVGF4YWJsZSBzdXBwbGllcyBvbiB3aGlj"
    "aCBFQ08gcGF5cyB0YXggdS9zIDkoNSkiLAogICAgICAgICAgICAgICAgInRheGFibGUiOiB2YWxz"
    "WzBdLCAiaWdzdCI6IHZhbHNbMV0sCiAgICAgICAgICAgICAgICAiY2dzdCI6IHZhbHNbMl0sICJz"
    "Z3N0IjogdmFsc1szXSwgImNlc3MiOiB2YWxzWzRdLAogICAgICAgICAgICB9KQogICAgcmV0dXJu"
    "IG91dAoKCmRlZiBwYXJzZV8zXzIodGFibGUpOgogICAgIiIiUGFyc2UgVGFibGUgMy4yIOKAlCBp"
    "bnRlci1zdGF0ZSBicmVha2Rvd24uIDMgcm93cywgMiBudW1lcmljIGNvbHMgKFRheGFibGUgKyBJ"
    "R1NUKS4iIiIKICAgIG1hcmtlcnMgPSBbCiAgICAgICAgKCJ1bnJlZ2lzdGVyZWQgcGVyc29ucyIs"
    "ICAgICAgICAiMy4yIChVbnJlZ2lzdGVyZWQpIiwgICJJbnRlci1zdGF0ZSBzdXBwbGllcyAtIFVu"
    "cmVnaXN0ZXJlZCBQZXJzb25zIiksCiAgICAgICAgKCJjb21wb3NpdGlvbiB0YXhhYmxlIiwgICAg"
    "ICAgICAiMy4yIChDb21wb3NpdGlvbikiLCAgICJJbnRlci1zdGF0ZSBzdXBwbGllcyAtIENvbXBv"
    "c2l0aW9uIFRheGFibGUgUGVyc29ucyIpLAogICAgICAgICgidWluIGhvbGRlcnMiLCAgICAgICAg"
    "ICAgICAgICAgIjMuMiAoVUlOKSIsICAgICAgICAgICAiSW50ZXItc3RhdGUgc3VwcGxpZXMgLSBV"
    "SU4gaG9sZGVycyIpLAogICAgXQogICAgb3V0ID0gW10KICAgIGZvciByb3cgaW4gdGFibGU6CiAg"
    "ICAgICAgaWYgbm90IHJvdzoKICAgICAgICAgICAgY29udGludWUKICAgICAgICBmaXJzdCA9IGNs"
    "ZWFuX2NlbGwocm93WzBdKS5sb3dlcigpCiAgICAgICAgZm9yIGtleSwgc2VjdGlvbiwgZGVzYyBp"
    "biBtYXJrZXJzOgogICAgICAgICAgICBpZiBrZXkgaW4gZmlyc3Q6CiAgICAgICAgICAgICAgICB0"
    "YXhhYmxlID0gcGFyc2VfbnVtX2NlbGwocm93WzFdKSBpZiBsZW4ocm93KSA+IDEgZWxzZSAwLjAK"
    "ICAgICAgICAgICAgICAgIGlnc3QgPSBwYXJzZV9udW1fY2VsbChyb3dbMl0pIGlmIGxlbihyb3cp"
    "ID4gMiBlbHNlIDAuMAogICAgICAgICAgICAgICAgb3V0LmFwcGVuZCh7CiAgICAgICAgICAgICAg"
    "ICAgICAgInNlY3Rpb24iOiBzZWN0aW9uLCAiZGVzY3JpcHRpb24iOiBkZXNjLAogICAgICAgICAg"
    "ICAgICAgICAgICJ0YXhhYmxlIjogdGF4YWJsZSwgImlnc3QiOiBpZ3N0LAogICAgICAgICAgICAg"
    "ICAgICAgICJjZ3N0IjogMC4wLCAic2dzdCI6IDAuMCwgImNlc3MiOiAwLjAsCiAgICAgICAgICAg"
    "ICAgICB9KQogICAgICAgICAgICAgICAgYnJlYWsKICAgIHJldHVybiBvdXQKCgpkZWYgcGFyc2Vf"
    "NCh0YWJsZSk6CiAgICAiIiJQYXJzZSBUYWJsZSA0IElUQy4gNCBudW1lcmljIGNvbHMgKElHU1Qs"
    "IENHU1QsIFNHU1QsIENlc3MpIHBlciByb3cuIE5vIHRheGFibGUgdmFsdWUuIiIiCiAgICAjIDQo"
    "QSkgc3ViLXJvd3MgYXJlICgxKS0oNSkgQUZURVIgIkEuIElUQyBBdmFpbGFibGUiIGhlYWRlcgog"
    "ICAgIyA0KEIpIHN1Yi1yb3dzIGFyZSAoMSktKDIpIEFGVEVSICJCLiBJVEMgUmV2ZXJzZWQiIGhl"
    "YWRlcgogICAgIyA0KEMpIGlzICJDLiBOZXQgSVRDIGF2YWlsYWJsZSIKICAgICMgNChEKSBpcyAi"
    "KEQpIE90aGVyIERldGFpbHMiIHRoZW4gc3ViLXJvd3MgKDEpLSgyKQogICAgb3V0ID0gW10KICAg"
    "ICMgRGVmYXVsdCB0byAnQScgYmVjYXVzZSBUYWJsZSA0IGFsd2F5cyBzdGFydHMgd2l0aCBTZWN0"
    "aW9uIEEg4oCUIGltcG9ydGFudCBmb3IKICAgICMgbXVsdGktcGFnZSB0YWJsZXMgd2hlcmUgcGFn"
    "ZSAyIGJlZ2lucyBkaXJlY3RseSB3aXRoIHJvdyAiKDQpIiB3aXRob3V0IGFueQogICAgIyBwcmVj"
    "ZWRpbmcgJ0EuIElUQyBBdmFpbGFibGUnIGhlYWRlciBvbiB0aGF0IHBhZ2UuCiAgICBjdXJyZW50"
    "X3NlY3Rpb24gPSAiQSIKCiAgICBhX2Rlc2NzID0gewogICAgICAgICIoMSkiOiAoIjQoQSkoMSki"
    "LCAiSW1wb3J0IG9mIGdvb2RzIiksCiAgICAgICAgIigyKSI6ICgiNChBKSgyKSIsICJJbXBvcnQg"
    "b2Ygc2VydmljZXMiKSwKICAgICAgICAiKDMpIjogKCI0KEEpKDMpIiwgIklud2FyZCBzdXBwbGll"
    "cyBsaWFibGUgdG8gUkNNIChvdGhlciB0aGFuIDEgJiAyKSIpLAogICAgICAgICIoNCkiOiAoIjQo"
    "QSkoNCkiLCAiSW53YXJkIHN1cHBsaWVzIGZyb20gSVNEIiksCiAgICAgICAgIig1KSI6ICgiNChB"
    "KSg1KSIsICJBbGwgb3RoZXIgSVRDIiksCiAgICB9CiAgICBiX2Rlc2NzID0gewogICAgICAgICIo"
    "MSkiOiAoIjQoQikoMSkiLCAiQXMgcGVyIHJ1bGVzIDM4LCA0MiAmIDQzIGFuZCBzZWN0aW9uIDE3"
    "KDUpIiksCiAgICAgICAgIigyKSI6ICgiNChCKSgyKSIsICJPdGhlcnMiKSwKICAgIH0KICAgIGRf"
    "ZGVzY3MgPSB7CiAgICAgICAgIigxKSI6ICgiNChEKSgxKSIsICJJVEMgcmVjbGFpbWVkIHdoaWNo"
    "IHdhcyByZXZlcnNlZCB1bmRlciBUYWJsZSA0KEIpKDIpIGVhcmxpZXIiKSwKICAgICAgICAiKDIp"
    "IjogKCI0KEQpKDIpIiwgIkluZWxpZ2libGUgSVRDIHVuZGVyIHNlY3Rpb24gMTYoNCkgJiBQb1Mg"
    "cmVzdHJpY3Rpb24iKSwKICAgIH0KCiAgICBmb3Igcm93IGluIHRhYmxlOgogICAgICAgIGlmIG5v"
    "dCByb3c6CiAgICAgICAgICAgIGNvbnRpbnVlCiAgICAgICAgZmlyc3QgPSBjbGVhbl9jZWxsKHJv"
    "d1swXSkKICAgICAgICBmaXJzdF9sb3dlciA9IGZpcnN0Lmxvd2VyKCkKCiAgICAgICAgIyBTZWN0"
    "aW9uIGhlYWRlcnMg4oCUIHN3aXRjaCBjb250ZXh0CiAgICAgICAgaWYgZmlyc3RfbG93ZXIuc3Rh"
    "cnRzd2l0aCgiYS4gaXRjIGF2YWlsYWJsZSIpOgogICAgICAgICAgICBjdXJyZW50X3NlY3Rpb24g"
    "PSAiQSIKICAgICAgICAgICAgY29udGludWUKICAgICAgICBpZiBmaXJzdF9sb3dlci5zdGFydHN3"
    "aXRoKCJiLiBpdGMgcmV2ZXJzZWQiKToKICAgICAgICAgICAgY3VycmVudF9zZWN0aW9uID0gIkIi"
    "CiAgICAgICAgICAgIGNvbnRpbnVlCiAgICAgICAgaWYgZmlyc3RfbG93ZXIuc3RhcnRzd2l0aCgi"
    "Yy4gbmV0IGl0YyBhdmFpbGFibGUiKToKICAgICAgICAgICAgIyA0KEMpIGlzIGEgc2luZ2xlIGRh"
    "dGEgcm93IHdpdGggNCBudW1lcmljIHZhbHVlcwogICAgICAgICAgICB2YWxzID0gW3BhcnNlX251"
    "bV9jZWxsKHJvd1tpXSkgaWYgaSA8IGxlbihyb3cpIGVsc2UgMC4wIGZvciBpIGluIHJhbmdlKDEs"
    "IDUpXQogICAgICAgICAgICBvdXQuYXBwZW5kKHsKICAgICAgICAgICAgICAgICJzZWN0aW9uIjog"
    "IjQoQykiLCAiZGVzY3JpcHRpb24iOiAiTmV0IElUQyBhdmFpbGFibGUgKEEgLSBCKSIsCiAgICAg"
    "ICAgICAgICAgICAidGF4YWJsZSI6IE5vbmUsICJpZ3N0IjogdmFsc1swXSwKICAgICAgICAgICAg"
    "ICAgICJjZ3N0IjogdmFsc1sxXSwgInNnc3QiOiB2YWxzWzJdLCAiY2VzcyI6IHZhbHNbM10sCiAg"
    "ICAgICAgICAgIH0pCiAgICAgICAgICAgIGN1cnJlbnRfc2VjdGlvbiA9IE5vbmUKICAgICAgICAg"
    "ICAgY29udGludWUKICAgICAgICBpZiBmaXJzdF9sb3dlci5zdGFydHN3aXRoKCIoZCkgb3RoZXIg"
    "ZGV0YWlscyIpIG9yIGZpcnN0X2xvd2VyLnN0YXJ0c3dpdGgoImQuIG90aGVyIGRldGFpbHMiKToK"
    "ICAgICAgICAgICAgIyA0KEQpIGlzIGEgaGVhZGVyIHRoYXQgQUxTTyBoYXMgdG90YWwgdmFsdWVz"
    "IG9uIHRoZSBzYW1lIHJvdwogICAgICAgICAgICB2YWxzID0gW3BhcnNlX251bV9jZWxsKHJvd1tp"
    "XSkgaWYgaSA8IGxlbihyb3cpIGVsc2UgMC4wIGZvciBpIGluIHJhbmdlKDEsIDUpXQogICAgICAg"
    "ICAgICBvdXQuYXBwZW5kKHsKICAgICAgICAgICAgICAgICJzZWN0aW9uIjogIjQoRCkiLCAiZGVz"
    "Y3JpcHRpb24iOiAiT3RoZXIgRGV0YWlscyAodG90YWwpIiwKICAgICAgICAgICAgICAgICJ0YXhh"
    "YmxlIjogTm9uZSwgImlnc3QiOiB2YWxzWzBdLAogICAgICAgICAgICAgICAgImNnc3QiOiB2YWxz"
    "WzFdLCAic2dzdCI6IHZhbHNbMl0sICJjZXNzIjogdmFsc1szXSwKICAgICAgICAgICAgfSkKICAg"
    "ICAgICAgICAgY3VycmVudF9zZWN0aW9uID0gIkQiCiAgICAgICAgICAgIGNvbnRpbnVlCgogICAg"
    "ICAgICMgU3ViLXJvd3Mg4oCUIGRlcGVuZCBvbiBjdXJyZW50X3NlY3Rpb24KICAgICAgICBpZiBj"
    "dXJyZW50X3NlY3Rpb24gaW4gKCJBIiwgIkIiLCAiRCIpOgogICAgICAgICAgICAjIERldGVjdCAo"
    "MSktKDUpIG1hcmtlcnMKICAgICAgICAgICAgbWFya2VyX21hdGNoID0gTm9uZQogICAgICAgICAg"
    "ICBmb3Iga2V5IGluICgiKDEpIiwgIigyKSIsICIoMykiLCAiKDQpIiwgIig1KSIpOgogICAgICAg"
    "ICAgICAgICAgaWYgZmlyc3Quc3RhcnRzd2l0aChrZXkpOgogICAgICAgICAgICAgICAgICAgIG1h"
    "cmtlcl9tYXRjaCA9IGtleQogICAgICAgICAgICAgICAgICAgIGJyZWFrCiAgICAgICAgICAgIGlm"
    "IG1hcmtlcl9tYXRjaDoKICAgICAgICAgICAgICAgIGRlc2NfbWFwID0geyJBIjogYV9kZXNjcywg"
    "IkIiOiBiX2Rlc2NzLCAiRCI6IGRfZGVzY3N9W2N1cnJlbnRfc2VjdGlvbl0KICAgICAgICAgICAg"
    "ICAgIGlmIG1hcmtlcl9tYXRjaCBpbiBkZXNjX21hcDoKICAgICAgICAgICAgICAgICAgICBzZWN0"
    "aW9uLCBkZXNjID0gZGVzY19tYXBbbWFya2VyX21hdGNoXQogICAgICAgICAgICAgICAgICAgIHZh"
    "bHMgPSBbcGFyc2VfbnVtX2NlbGwocm93W2ldKSBpZiBpIDwgbGVuKHJvdykgZWxzZSAwLjAgZm9y"
    "IGkgaW4gcmFuZ2UoMSwgNSldCiAgICAgICAgICAgICAgICAgICAgb3V0LmFwcGVuZCh7CiAgICAg"
    "ICAgICAgICAgICAgICAgICAgICJzZWN0aW9uIjogc2VjdGlvbiwgImRlc2NyaXB0aW9uIjogZGVz"
    "YywKICAgICAgICAgICAgICAgICAgICAgICAgInRheGFibGUiOiBOb25lLCAiaWdzdCI6IHZhbHNb"
    "MF0sCiAgICAgICAgICAgICAgICAgICAgICAgICJjZ3N0IjogdmFsc1sxXSwgInNnc3QiOiB2YWxz"
    "WzJdLCAiY2VzcyI6IHZhbHNbM10sCiAgICAgICAgICAgICAgICAgICAgfSkKICAgIHJldHVybiBv"
    "dXQKCgpkZWYgcGFyc2VfNSh0YWJsZSk6CiAgICAiIiIKICAgIFBhcnNlIFRhYmxlIDUg4oCUIGV4"
    "ZW1wdC9uaWwtcmF0ZWQgYW5kIG5vbi1HU1QgaW53YXJkIHN1cHBsaWVzLgogICAgMiBjb2x1bW5z"
    "OiBJbnRlci1TdGF0ZSwgSW50cmEtU3RhdGUuCiAgICBFYWNoIFBERiByb3cgYmVjb21lcyBUV08g"
    "b3V0cHV0IHJvd3MgKEludGVyICsgSW50cmEpLgogICAgIiIiCiAgICBvdXQgPSBbXQogICAgdGFy"
    "Z2V0cyA9IFsKICAgICAgICAoImNvbXBvc2l0aW9uIHNjaGVtZSIsICAgIjUgKENvbXBvc2l0aW9u"
    "L0V4ZW1wdC9OaWwpIiwgIkNvbXBvc2l0aW9uIC8gRXhlbXB0IC8gTmlsIHJhdGVkIHN1cHBseSAo"
    "aW53YXJkKSIpLAogICAgICAgICgibm9uIGdzdCBzdXBwbHkiLCAgICAgICAiNSAoTm9uLUdTVCki"
    "LCAgICAgICAgICAgICAgICAgIk5vbi1HU1Qgc3VwcGx5IChpbndhcmQpIiksCiAgICBdCiAgICBm"
    "b3Igcm93IGluIHRhYmxlOgogICAgICAgIGlmIG5vdCByb3c6CiAgICAgICAgICAgIGNvbnRpbnVl"
    "CiAgICAgICAgZmlyc3QgPSBjbGVhbl9jZWxsKHJvd1swXSkubG93ZXIoKQogICAgICAgIGZvciBr"
    "ZXksIHNlY3Rpb24sIGRlc2NfYmFzZSBpbiB0YXJnZXRzOgogICAgICAgICAgICBpZiBrZXkgaW4g"
    "Zmlyc3Q6CiAgICAgICAgICAgICAgICBpbnRlciA9IHBhcnNlX251bV9jZWxsKHJvd1sxXSkgaWYg"
    "bGVuKHJvdykgPiAxIGVsc2UgMC4wCiAgICAgICAgICAgICAgICBpbnRyYSA9IHBhcnNlX251bV9j"
    "ZWxsKHJvd1syXSkgaWYgbGVuKHJvdykgPiAyIGVsc2UgMC4wCiAgICAgICAgICAgICAgICBvdXQu"
    "YXBwZW5kKHsKICAgICAgICAgICAgICAgICAgICAic2VjdGlvbiI6IHNlY3Rpb24sICJkZXNjcmlw"
    "dGlvbiI6IGRlc2NfYmFzZSArICIg4oCUIEludGVyLVN0YXRlIiwKICAgICAgICAgICAgICAgICAg"
    "ICAidGF4YWJsZSI6IGludGVyLCAiaWdzdCI6IDAuMCwgImNnc3QiOiAwLjAsICJzZ3N0IjogMC4w"
    "LCAiY2VzcyI6IDAuMCwKICAgICAgICAgICAgICAgIH0pCiAgICAgICAgICAgICAgICBvdXQuYXBw"
    "ZW5kKHsKICAgICAgICAgICAgICAgICAgICAic2VjdGlvbiI6IHNlY3Rpb24sICJkZXNjcmlwdGlv"
    "biI6IGRlc2NfYmFzZSArICIg4oCUIEludHJhLVN0YXRlIiwKICAgICAgICAgICAgICAgICAgICAi"
    "dGF4YWJsZSI6IGludHJhLCAiaWdzdCI6IDAuMCwgImNnc3QiOiAwLjAsICJzZ3N0IjogMC4wLCAi"
    "Y2VzcyI6IDAuMCwKICAgICAgICAgICAgICAgIH0pCiAgICAgICAgICAgICAgICBicmVhawogICAg"
    "cmV0dXJuIG91dAoKCmRlZiBwYXJzZV81XzEodGFibGUpOgogICAgIiIiUGFyc2UgVGFibGUgNS4x"
    "IOKAlCBJbnRlcmVzdCBhbmQgTGF0ZSBmZWUuIDQgdGF4IGNvbHVtbnMuIiIiCiAgICBvdXQgPSBb"
    "XQogICAgbWFya2VycyA9IFsKICAgICAgICAoInN5c3RlbSBjb21wdXRlZCIsICI1LjEgKFN5c3Rl"
    "bSBJbnRlcmVzdCkiLCAiU3lzdGVtIGNvbXB1dGVkIEludGVyZXN0IiksCiAgICAgICAgKCJpbnRl"
    "cmVzdCBwYWlkIiwgICAgIjUuMSAoSW50ZXJlc3QgUGFpZCkiLCAgICJJbnRlcmVzdCBQYWlkIiks"
    "CiAgICAgICAgKCJsYXRlIGZlZSIsICAgICAgICAgIjUuMSAoTGF0ZSBGZWUpIiwgICAgICAgICJM"
    "YXRlIGZlZSIpLAogICAgXQogICAgZm9yIHJvdyBpbiB0YWJsZToKICAgICAgICBpZiBub3Qgcm93"
    "OgogICAgICAgICAgICBjb250aW51ZQogICAgICAgIGZpcnN0ID0gY2xlYW5fY2VsbChyb3dbMF0p"
    "Lmxvd2VyKCkKICAgICAgICBmb3Iga2V5LCBzZWN0aW9uLCBkZXNjIGluIG1hcmtlcnM6CiAgICAg"
    "ICAgICAgIGlmIGtleSBpbiBmaXJzdDoKICAgICAgICAgICAgICAgIHZhbHMgPSBbcGFyc2VfbnVt"
    "X2NlbGwocm93W2ldKSBpZiBpIDwgbGVuKHJvdykgZWxzZSAwLjAgZm9yIGkgaW4gcmFuZ2UoMSwg"
    "NSldCiAgICAgICAgICAgICAgICBvdXQuYXBwZW5kKHsKICAgICAgICAgICAgICAgICAgICAic2Vj"
    "dGlvbiI6IHNlY3Rpb24sICJkZXNjcmlwdGlvbiI6IGRlc2MsCiAgICAgICAgICAgICAgICAgICAg"
    "InRheGFibGUiOiBOb25lLCAiaWdzdCI6IHZhbHNbMF0sCiAgICAgICAgICAgICAgICAgICAgImNn"
    "c3QiOiB2YWxzWzFdLCAic2dzdCI6IHZhbHNbMl0sICJjZXNzIjogdmFsc1szXSwKICAgICAgICAg"
    "ICAgICAgIH0pCiAgICAgICAgICAgICAgICBicmVhawogICAgcmV0dXJuIG91dAoKCmRlZiBwYXJz"
    "ZV82XzEodGFibGUpOgogICAgIiIiCiAgICBQYXJzZSBUYWJsZSA2LjEg4oCUIFBheW1lbnQgb2Yg"
    "dGF4LiAxMCBudW1lcmljIGNvbHVtbnMgcGVyIHRheC10eXBlIHJvdy4KICAgIFJldHVybnMgcm93"
    "cyB3aXRoIGEgZGlmZmVyZW50IHNjaGVtYSAodXNlZCBpbiBhIHNlcGFyYXRlIHNoZWV0KS4KICAg"
    "IENvbHVtbiBsYXlvdXQgKGFmdGVyIERlc2NyaXB0aW9uIGNvbCk6CiAgICAgICAwOiBUYXggcGF5"
    "YWJsZQogICAgICAgMTogQWRqdXN0bWVudCBvZiBuZWdhdGl2ZSBsaWFiaWxpdHkgKHByZXZpb3Vz"
    "IHBlcmlvZCkKICAgICAgIDI6IE5ldCBUYXggUGF5YWJsZQogICAgICAgMy02OiBUYXggcGFpZCB0"
    "aHJvdWdoIElUQyAoSUdTVCwgQ0dTVCwgU0dTVCwgQ2VzcykKICAgICAgIDc6IFRheCBwYWlkIGlu"
    "IGNhc2gKICAgICAgIDg6IEludGVyZXN0IHBhaWQgaW4gY2FzaAogICAgICAgOTogTGF0ZSBmZWUg"
    "cGFpZCBpbiBjYXNoCiAgICAiIiIKICAgIG91dCA9IFtdCiAgICBjdXJyZW50X3N1YnNlY3Rpb24g"
    "PSBOb25lICAjICdBJyBvciAnQicKCiAgICB0YXhfdHlwZXMgPSBbImludGVncmF0ZWQgdGF4Iiwg"
    "ImNlbnRyYWwgdGF4IiwgInN0YXRlL3V0IHRheCIsICJjZXNzIl0KCiAgICBmb3Igcm93IGluIHRh"
    "YmxlOgogICAgICAgIGlmIG5vdCByb3c6CiAgICAgICAgICAgIGNvbnRpbnVlCiAgICAgICAgZmly"
    "c3RfbG93ZXIgPSBjbGVhbl9jZWxsKHJvd1swXSkubG93ZXIoKQogICAgICAgIGpvaW5lZF9sb3dl"
    "ciA9IHJvd190ZXh0KHJvdykubG93ZXIoKQoKICAgICAgICAjIERldGVjdCAoQSkgYW5kIChCKSBz"
    "dWItc2VjdGlvbiBoZWFkZXJzCiAgICAgICAgaWYgIihhKSIgaW4gZmlyc3RfbG93ZXIgYW5kICJv"
    "dGhlciB0aGFuIHJldmVyc2UiIGluIGpvaW5lZF9sb3dlcjoKICAgICAgICAgICAgY3VycmVudF9z"
    "dWJzZWN0aW9uID0gIkEiCiAgICAgICAgICAgIGNvbnRpbnVlCiAgICAgICAgaWYgIihiKSIgaW4g"
    "Zmlyc3RfbG93ZXIgYW5kICgicmV2ZXJzZSBjaGFyZ2UiIGluIGpvaW5lZF9sb3dlciBvciAiOSg1"
    "KSIgaW4gam9pbmVkX2xvd2VyKToKICAgICAgICAgICAgY3VycmVudF9zdWJzZWN0aW9uID0gIkIi"
    "CiAgICAgICAgICAgIGNvbnRpbnVlCgogICAgICAgIGlmIGN1cnJlbnRfc3Vic2VjdGlvbiBpcyBO"
    "b25lOgogICAgICAgICAgICBjb250aW51ZQoKICAgICAgICAjIE1hdGNoIHRheC10eXBlIHJvd3MK"
    "ICAgICAgICB0YXhfdHlwZSA9IE5vbmUKICAgICAgICBmb3IgdHQgaW4gdGF4X3R5cGVzOgogICAg"
    "ICAgICAgICBpZiB0dCBpbiBmaXJzdF9sb3dlcjoKICAgICAgICAgICAgICAgIHRheF90eXBlID0g"
    "dHQudGl0bGUoKS5yZXBsYWNlKCIvVXQiLCAiL1VUIikKICAgICAgICAgICAgICAgIGJyZWFrCiAg"
    "ICAgICAgaWYgdGF4X3R5cGUgaXMgTm9uZToKICAgICAgICAgICAgY29udGludWUKCiAgICAgICAg"
    "IyBFeHRyYWN0IDEwIG51bWVyaWMgdmFsdWVzIGZyb20gY29sdW1ucyAxLi4xMAogICAgICAgIG51"
    "bXMgPSBbcGFyc2VfbnVtX2NlbGwocm93W2ldKSBpZiBpIDwgbGVuKHJvdykgZWxzZSAwLjAgZm9y"
    "IGkgaW4gcmFuZ2UoMSwgMTEpXQogICAgICAgIG91dC5hcHBlbmQoewogICAgICAgICAgICAic3Vi"
    "c2VjdGlvbiI6IGYiNi4xKHtjdXJyZW50X3N1YnNlY3Rpb259KSIsCiAgICAgICAgICAgICJzdWJz"
    "ZWN0aW9uX2Rlc2MiOiAiT3RoZXIgdGhhbiByZXZlcnNlIGNoYXJnZSIgaWYgY3VycmVudF9zdWJz"
    "ZWN0aW9uID09ICJBIgogICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgZWxzZSAiUmV2ZXJz"
    "ZSBjaGFyZ2UgYW5kIHN1cHBsaWVzIHUvcyA5KDUpIiwKICAgICAgICAgICAgInRheF90eXBlIjog"
    "dGF4X3R5cGUsCiAgICAgICAgICAgICJ0YXhfcGF5YWJsZSI6ICAgbnVtc1swXSwKICAgICAgICAg"
    "ICAgImFkanVzdG1lbnQiOiAgICBudW1zWzFdLAogICAgICAgICAgICAibmV0X3RheF9wYXlhYmxl"
    "IjogbnVtc1syXSwKICAgICAgICAgICAgIml0Y19pZ3N0IjogICAgICBudW1zWzNdLAogICAgICAg"
    "ICAgICAiaXRjX2Nnc3QiOiAgICAgIG51bXNbNF0sCiAgICAgICAgICAgICJpdGNfc2dzdCI6ICAg"
    "ICAgbnVtc1s1XSwKICAgICAgICAgICAgIml0Y19jZXNzIjogICAgICBudW1zWzZdLAogICAgICAg"
    "ICAgICAidGF4X2luX2Nhc2giOiAgIG51bXNbN10sCiAgICAgICAgICAgICJpbnRlcmVzdF9jYXNo"
    "IjogbnVtc1s4XSwKICAgICAgICAgICAgImxhdGVfZmVlX2Nhc2giOiBudW1zWzldLAogICAgICAg"
    "IH0pCiAgICByZXR1cm4gb3V0CgoKZGVmIHBhcnNlX2JyZWFrdXAodGFibGUsIGhlYWRlcl90YWJs"
    "ZT1Ob25lKToKICAgICIiIgogICAgUGFyc2UgJ0JyZWFrdXAgb2YgdGF4IGxpYWJpbGl0eSBkZWNs"
    "YXJlZCcgKHBlcmlvZC13aXNlKS4KICAgIFNpbmdsZSBkYXRhIHJvdzogPHBlcmlvZD4gPGlnc3Q+"
    "IDxjZ3N0PiA8c2dzdD4gPGNlc3M+CiAgICAiIiIKICAgIG91dCA9IFtdCiAgICBmb3Igcm93IGlu"
    "IHRhYmxlOgogICAgICAgIGlmIG5vdCByb3c6CiAgICAgICAgICAgIGNvbnRpbnVlCiAgICAgICAg"
    "Zmlyc3QgPSBjbGVhbl9jZWxsKHJvd1swXSkKICAgICAgICAjIFNraXAgaGVhZGVyIHJvdwogICAg"
    "ICAgIGlmIGZpcnN0Lmxvd2VyKCkgPT0gInBlcmlvZCI6CiAgICAgICAgICAgIGNvbnRpbnVlCiAg"
    "ICAgICAgIyBEYXRhIHJvdyDigJQgaGFzIGEgbW9udGgveWVhciB0ZXh0IGluIGZpcnN0IGNvbAog"
    "ICAgICAgICMgZS5nLiwgJ0phbnVhcnkgMjAyNicKICAgICAgICBpZiByZS5zZWFyY2gociJbQS1a"
    "YS16XStccytcZHs0fSIsIGZpcnN0KToKICAgICAgICAgICAgdmFscyA9IFtwYXJzZV9udW1fY2Vs"
    "bChyb3dbaV0pIGlmIGkgPCBsZW4ocm93KSBlbHNlIDAuMCBmb3IgaSBpbiByYW5nZSgxLCA1KV0K"
    "ICAgICAgICAgICAgb3V0LmFwcGVuZCh7CiAgICAgICAgICAgICAgICAic2VjdGlvbiI6ICJCcmVh"
    "a3VwIiwgImRlc2NyaXB0aW9uIjogZiJUYXggbGlhYmlsaXR5IGJyZWFrdXAg4oCUIHtmaXJzdH0i"
    "LAogICAgICAgICAgICAgICAgInRheGFibGUiOiBOb25lLCAiaWdzdCI6IHZhbHNbMF0sCiAgICAg"
    "ICAgICAgICAgICAiY2dzdCI6IHZhbHNbMV0sICJzZ3N0IjogdmFsc1syXSwgImNlc3MiOiB2YWxz"
    "WzNdLAogICAgICAgICAgICB9KQogICAgcmV0dXJuIG91dAoKCiMgLS0tLS0tLS0tLS0tLS0tLS0t"
    "LS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0t"
    "ICMKIyAgTWV0YSBleHRyYWN0aW9uCiMgLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0t"
    "LS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tICMKCmRlZiBwYXJzZV9t"
    "ZXRhKHRleHQpOgogICAgZ3N0aW5fbSAgPSByZS5zZWFyY2gociJHU1RJTlxzK29mXHMrdGhlXHMr"
    "c3VwcGxpZXJccysoWzAtOV17Mn1bQS1aMC05XXsxM30pIiwgdGV4dCkKICAgIGlmIG5vdCBnc3Rp"
    "bl9tOgogICAgICAgIGdzdGluX20gPSByZS5zZWFyY2gociJcYihbMC05XXsyfVtBLVpdezV9XGR7"
    "NH1bQS1aXVtBLVowLTldWltBLVowLTldKVxiIiwgdGV4dCkKICAgIHBlcmlvZF9tID0gcmUuc2Vh"
    "cmNoKHIiUGVyaW9kXHMrKFtBLVphLXpdKykiLCB0ZXh0KQogICAgZnlfbSAgICAgPSByZS5zZWFy"
    "Y2gociJZZWFyXHMrKFxkezR9LVxkezJ9KSIsIHRleHQpCiAgICBsZWdhbF9tICA9IHJlLnNlYXJj"
    "aChyIkxlZ2FsXHMrbmFtZVxzK29mXHMrdGhlXHMrcmVnaXN0ZXJlZFxzK3BlcnNvblxzKyguKyki"
    "LCB0ZXh0KQogICAgYXJuX20gICAgPSByZS5zZWFyY2gociIyXChjXClcLlxzKkFSTlxzKyhcUysp"
    "IiwgdGV4dCkKICAgIGFybl9kYXRlX20gPSByZS5zZWFyY2gociIyXChkXClcLlxzKkRhdGVccytv"
    "ZlxzK0FSTlxzKyhcUyspIiwgdGV4dCkKCiAgICBnc3RpbiA9IGdzdGluX20uZ3JvdXAoMSkgaWYg"
    "Z3N0aW5fbSBlbHNlICIiCiAgICBzdGF0ZV9jb2RlID0gZ3N0aW5bOjJdIGlmIGdzdGluIGVsc2Ug"
    "IiIKICAgIHN0YXRlX25hbWUgPSBTVEFURV9DT0RFUy5nZXQoc3RhdGVfY29kZSwgIlVua25vd24i"
    "KQogICAgcGVyaW9kID0gcGVyaW9kX20uZ3JvdXAoMSkgaWYgcGVyaW9kX20gZWxzZSAiIgogICAg"
    "ZnkgPSBmeV9tLmdyb3VwKDEpIGlmIGZ5X20gZWxzZSAiIgogICAgIyBJbmRpYW4gRlk6IEFwcuKA"
    "k0RlYyA9IGZpcnN0IHllYXIsIEphbuKAk01hciA9IHNlY29uZCB5ZWFyCiAgICB5eSA9ICIiCiAg"
    "ICBpZiBmeSBhbmQgIi0iIGluIGZ5OgogICAgICAgIHl5MSA9IGZ5LnNwbGl0KCItIilbMF1bLTI6"
    "XQogICAgICAgIHl5MiA9IGZ5LnNwbGl0KCItIilbMV0KICAgICAgICB5eSA9IHl5MiBpZiBwZXJp"
    "b2QgaW4gKCJKYW51YXJ5IiwgIkZlYnJ1YXJ5IiwgIk1hcmNoIikgZWxzZSB5eTEKICAgIG1vbnRo"
    "ID0gZiJ7TU9OVEhfQUJCUi5nZXQocGVyaW9kLCBwZXJpb2RbOjNdKX0te3l5fSIgaWYgcGVyaW9k"
    "IGVsc2UgIiIKCiAgICByZXR1cm4gewogICAgICAgICJnc3RpbiI6ICAgICAgZ3N0aW4sCiAgICAg"
    "ICAgInN0YXRlX2NvZGUiOiBzdGF0ZV9jb2RlLAogICAgICAgICJzdGF0ZV9uYW1lIjogc3RhdGVf"
    "bmFtZSwKICAgICAgICAibW9udGgiOiAgICAgIG1vbnRoLAogICAgICAgICJmeSI6ICAgICAgICAg"
    "ZnksCiAgICAgICAgInRheF9wZXJpb2QiOiBwZXJpb2QsCiAgICAgICAgImxlZ2FsX25hbWUiOiBj"
    "bGVhbl9jZWxsKGxlZ2FsX20uZ3JvdXAoMSkpIGlmIGxlZ2FsX20gZWxzZSAiIiwKICAgICAgICAi"
    "YXJuIjogICAgICAgIGFybl9tLmdyb3VwKDEpIGlmIGFybl9tIGVsc2UgIiIsCiAgICAgICAgImFy"
    "bl9kYXRlIjogICBhcm5fZGF0ZV9tLmdyb3VwKDEpIGlmIGFybl9kYXRlX20gZWxzZSAiIiwKICAg"
    "IH0KCgojIC0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0t"
    "LS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLSAjCiMgIE1haW4gcGVyLVBERiBwYXJzZXIKIyAtLS0t"
    "LS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0t"
    "LS0tLS0tLS0tLS0tLS0gIwoKZGVmIHBhcnNlX3BkZihwZGZfcGF0aCk6CiAgICAiIiJSZXR1cm4g"
    "KG1ldGEsIGNvbnNvbGlkYXRlZF9yb3dzLCBwYXltZW50X3Jvd3MpLiBSYWlzZXMgb24gaGFyZCBm"
    "YWlsdXJlLiIiIgogICAgZnVsbF90ZXh0ID0gIiIKICAgIGFsbF90YWJsZXMgPSBbXQogICAgd2l0"
    "aCBwZGZwbHVtYmVyLm9wZW4ocGRmX3BhdGgpIGFzIHBkZjoKICAgICAgICBmb3IgcGFnZSBpbiBw"
    "ZGYucGFnZXM6CiAgICAgICAgICAgIHQgPSBwYWdlLmV4dHJhY3RfdGV4dCgpIG9yICIiCiAgICAg"
    "ICAgICAgIGZ1bGxfdGV4dCArPSB0ICsgIlxuIgogICAgICAgICAgICBmb3IgdGFibGUgaW4gcGFn"
    "ZS5leHRyYWN0X3RhYmxlcygpIG9yIFtdOgogICAgICAgICAgICAgICAgY2xlYW5lZCA9IFtbY2xl"
    "YW5fY2VsbChjKSBpZiBjIGVsc2UgIiIgZm9yIGMgaW4gcm93XSBmb3Igcm93IGluIHRhYmxlXQog"
    "ICAgICAgICAgICAgICAgYWxsX3RhYmxlcy5hcHBlbmQoY2xlYW5lZCkKCiAgICBpZiBub3QgZnVs"
    "bF90ZXh0LnN0cmlwKCk6CiAgICAgICAgcmFpc2UgVmFsdWVFcnJvcigiUERGIGhhcyBubyBleHRy"
    "YWN0YWJsZSB0ZXh0IChzY2FubmVkIC8gaW1hZ2Utb25seSkuIikKCiAgICBtZXRhID0gcGFyc2Vf"
    "bWV0YShmdWxsX3RleHQpCiAgICBpZiBub3QgbWV0YVsiZ3N0aW4iXToKICAgICAgICByYWlzZSBW"
    "YWx1ZUVycm9yKCJDb3VsZCBub3QgbG9jYXRlIEdTVElOIGluIFBERi4iKQoKICAgIGNvbnNvbGlk"
    "YXRlZCA9IFtdICAgIyByb3dzIGZvciB0aGUgbWFpbiBDb25zb2xpZGF0ZWQgc2hlZXQKICAgIHBh"
    "eW1lbnQgICAgICA9IFtdICAgIyByb3dzIGZvciB0aGUgNi4xIFRheCBQYXltZW50IHNoZWV0CiAg"
    "ICBzZWVuX3NlY3Rpb25zID0gc2V0KCkgICMgYXZvaWQgZHVwbGljYXRlIHBhcnNlcyBpZiB0YWJs"
    "ZSBhcHBlYXJzIHR3aWNlCgogICAgZm9yIHRhYmxlIGluIGFsbF90YWJsZXM6CiAgICAgICAgc2Vj"
    "dGlvbl90eXBlID0gaWRlbnRpZnlfc2VjdGlvbih0YWJsZSkKICAgICAgICBpZiBzZWN0aW9uX3R5"
    "cGUgaXMgTm9uZToKICAgICAgICAgICAgY29udGludWUKCiAgICAgICAgaWYgc2VjdGlvbl90eXBl"
    "ID09ICIzLjEiOgogICAgICAgICAgICBmb3IgciBpbiBwYXJzZV8zXzEodGFibGUpOgogICAgICAg"
    "ICAgICAgICAgY29uc29saWRhdGVkLmFwcGVuZChyKQogICAgICAgIGVsaWYgc2VjdGlvbl90eXBl"
    "ID09ICIzLjEuMSI6CiAgICAgICAgICAgIGZvciByIGluIHBhcnNlXzNfMV8xKHRhYmxlKToKICAg"
    "ICAgICAgICAgICAgIGNvbnNvbGlkYXRlZC5hcHBlbmQocikKICAgICAgICBlbGlmIHNlY3Rpb25f"
    "dHlwZSA9PSAiMy4yIjoKICAgICAgICAgICAgZm9yIHIgaW4gcGFyc2VfM18yKHRhYmxlKToKICAg"
    "ICAgICAgICAgICAgIGNvbnNvbGlkYXRlZC5hcHBlbmQocikKICAgICAgICBlbGlmIHNlY3Rpb25f"
    "dHlwZSA9PSAiNCI6CiAgICAgICAgICAgIGZvciByIGluIHBhcnNlXzQodGFibGUpOgogICAgICAg"
    "ICAgICAgICAga2V5ID0gKHJbInNlY3Rpb24iXSwgclsiZGVzY3JpcHRpb24iXSkKICAgICAgICAg"
    "ICAgICAgIGlmIGtleSBpbiBzZWVuX3NlY3Rpb25zOgogICAgICAgICAgICAgICAgICAgIGNvbnRp"
    "bnVlCiAgICAgICAgICAgICAgICBzZWVuX3NlY3Rpb25zLmFkZChrZXkpCiAgICAgICAgICAgICAg"
    "ICBjb25zb2xpZGF0ZWQuYXBwZW5kKHIpCiAgICAgICAgZWxpZiBzZWN0aW9uX3R5cGUgPT0gIjUi"
    "OgogICAgICAgICAgICBmb3IgciBpbiBwYXJzZV81KHRhYmxlKToKICAgICAgICAgICAgICAgIGNv"
    "bnNvbGlkYXRlZC5hcHBlbmQocikKICAgICAgICBlbGlmIHNlY3Rpb25fdHlwZSA9PSAiNS4xIjoK"
    "ICAgICAgICAgICAgZm9yIHIgaW4gcGFyc2VfNV8xKHRhYmxlKToKICAgICAgICAgICAgICAgIGNv"
    "bnNvbGlkYXRlZC5hcHBlbmQocikKICAgICAgICBlbGlmIHNlY3Rpb25fdHlwZSA9PSAiNi4xIjoK"
    "ICAgICAgICAgICAgZm9yIHIgaW4gcGFyc2VfNl8xKHRhYmxlKToKICAgICAgICAgICAgICAgIHBh"
    "eW1lbnQuYXBwZW5kKHIpCiAgICAgICAgZWxpZiBzZWN0aW9uX3R5cGUgPT0gImJyZWFrdXAiOgog"
    "ICAgICAgICAgICBmb3IgciBpbiBwYXJzZV9icmVha3VwKHRhYmxlKToKICAgICAgICAgICAgICAg"
    "IGtleSA9IChyWyJzZWN0aW9uIl0sIHJbImRlc2NyaXB0aW9uIl0pCiAgICAgICAgICAgICAgICBp"
    "ZiBrZXkgaW4gc2Vlbl9zZWN0aW9uczoKICAgICAgICAgICAgICAgICAgICBjb250aW51ZQogICAg"
    "ICAgICAgICAgICAgc2Vlbl9zZWN0aW9ucy5hZGQoa2V5KQogICAgICAgICAgICAgICAgY29uc29s"
    "aWRhdGVkLmFwcGVuZChyKQoKICAgIHJldHVybiBtZXRhLCBjb25zb2xpZGF0ZWQsIHBheW1lbnQK"
    "CgojIC0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0t"
    "LS0tLS0tLS0tLS0tLS0tLS0tLS0tLSAjCiMgIEV4Y2VsIHdyaXRlcgojIC0tLS0tLS0tLS0tLS0t"
    "LS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0t"
    "LS0tLSAjCgpIRFJfRklMTCAgID0gUGF0dGVybkZpbGwoInNvbGlkIiwgc3RhcnRfY29sb3I9IjFG"
    "NEU3OCIpClNVQl9GSUxMICAgPSBQYXR0ZXJuRmlsbCgic29saWQiLCBzdGFydF9jb2xvcj0iMkU3"
    "NUI2IikKVE9UQUxfRklMTCA9IFBhdHRlcm5GaWxsKCJzb2xpZCIsIHN0YXJ0X2NvbG9yPSJGRkU2"
    "OTkiKQpHUk9VUF9GSUxMID0gUGF0dGVybkZpbGwoInNvbGlkIiwgc3RhcnRfY29sb3I9IkRERUJG"
    "NyIpCkVSUl9GSUxMICAgPSBQYXR0ZXJuRmlsbCgic29saWQiLCBzdGFydF9jb2xvcj0iRkZDN0NF"
    "IikKCldISVRFID0gRm9udChuYW1lPSJBcmlhbCIsIGJvbGQ9VHJ1ZSwgY29sb3I9IkZGRkZGRiIs"
    "IHNpemU9MTEpCkJPTEQgID0gRm9udChuYW1lPSJBcmlhbCIsIGJvbGQ9VHJ1ZSwgc2l6ZT0xMCkK"
    "UkVHICAgPSBGb250KG5hbWU9IkFyaWFsIiwgc2l6ZT0xMCkKVElUTEUgPSBGb250KG5hbWU9IkFy"
    "aWFsIiwgYm9sZD1UcnVlLCBzaXplPTE0LCBjb2xvcj0iRkZGRkZGIikKCnRoaW4gPSBTaWRlKGJv"
    "cmRlcl9zdHlsZT0idGhpbiIsIGNvbG9yPSJCNEI0QjQiKQpCT1JERVIgPSBCb3JkZXIobGVmdD10"
    "aGluLCByaWdodD10aGluLCB0b3A9dGhpbiwgYm90dG9tPXRoaW4pCkNFTlRFUiA9IEFsaWdubWVu"
    "dChob3Jpem9udGFsPSJjZW50ZXIiLCB2ZXJ0aWNhbD0iY2VudGVyIiwgd3JhcF90ZXh0PVRydWUp"
    "CkxFRlQgICA9IEFsaWdubWVudChob3Jpem9udGFsPSJsZWZ0IiwgICB2ZXJ0aWNhbD0iY2VudGVy"
    "Iiwgd3JhcF90ZXh0PVRydWUpClJJR0hUICA9IEFsaWdubWVudChob3Jpem9udGFsPSJyaWdodCIs"
    "ICB2ZXJ0aWNhbD0iY2VudGVyIikKCk5VTV9GTVQgPSAnIywjIzAuMDA7KCMsIyMwLjAwKTsiLSIn"
    "CgoKZGVmIF9zdHlsZV9oZWFkZXIod3MsIHJvdywgbl9jb2xzKToKICAgIGZvciBjIGluIHJhbmdl"
    "KDEsIG5fY29scyArIDEpOgogICAgICAgIGNlbGwgPSB3cy5jZWxsKHJvdz1yb3csIGNvbHVtbj1j"
    "KQogICAgICAgIGNlbGwuZm9udCA9IFdISVRFOyBjZWxsLmZpbGwgPSBTVUJfRklMTAogICAgICAg"
    "IGNlbGwuYWxpZ25tZW50ID0gQ0VOVEVSOyBjZWxsLmJvcmRlciA9IEJPUkRFUgoKCmRlZiB3cml0"
    "ZV9leGNlbChyZXR1cm5zLCBvdXRwdXRfcGF0aCk6CiAgICAiIiIKICAgIHJldHVybnM6IGxpc3Qg"
    "b2YgZGljdHM6CiAgICAgICAgeyAibWV0YSI6IHsuLi59LCAiY29uc29saWRhdGVkIjogWy4uLl0s"
    "ICJwYXltZW50IjogWy4uLl0sCiAgICAgICAgICAic291cmNlX2ZpbGUiOiAiLi4uIiwgInN0YXR1"
    "cyI6ICJPSyJ8IkZBSUwiLCAibm90ZXMiOiAiLi4uIiB9CiAgICAiIiIKICAgIHdiID0gV29ya2Jv"
    "b2soKQoKICAgICMgPT09PT09PT09PSBTaGVldCAxOiBDb25zb2xpZGF0ZWQgPT09PT09PT09PT09"
    "PT09PT09PT09PT09PT09CiAgICB3cyA9IHdiLmFjdGl2ZQogICAgd3MudGl0bGUgPSAiQ29uc29s"
    "aWRhdGVkIgoKICAgIHdzLm1lcmdlX2NlbGxzKCJBMTpOMSIpCiAgICB3c1siQTEiXSA9ICJHU1RS"
    "LTNCIENPTlNPTElEQVRFRCDigJQgU1RBVEUtV0lTRSAvIE1PTlRILVdJU0UgLyBTRUNUSU9OLVdJ"
    "U0UiCiAgICB3c1siQTEiXS5mb250ID0gVElUTEU7IHdzWyJBMSJdLmZpbGwgPSBIRFJfRklMTDsg"
    "d3NbIkExIl0uYWxpZ25tZW50ID0gQ0VOVEVSCiAgICB3cy5yb3dfZGltZW5zaW9uc1sxXS5oZWln"
    "aHQgPSAyOAoKICAgIGhlYWRlcnMgPSBbCiAgICAgICAgIlNyIE5vIiwgIk1vbnRoIiwgIkZZIiwg"
    "IlN0YXRlIENvZGUiLCAiU3RhdGUgTmFtZSIsICJHU1RJTiIsCiAgICAgICAgIkxlZ2FsIE5hbWUi"
    "LCAiQVJOIiwKICAgICAgICAiU2VjdGlvbiIsICJEZXNjcmlwdGlvbiIsCiAgICAgICAgIlRheGFi"
    "bGUgVmFsdWUgKOKCuSkiLCAiSUdTVCAo4oK5KSIsICJDR1NUICjigrkpIiwgIlNHU1QvVVRHU1Qg"
    "KOKCuSkiLCAiQ0VTUyAo4oK5KSIKICAgIF0KICAgIEhEUiA9IDMKICAgIGZvciBpLCBoIGluIGVu"
    "dW1lcmF0ZShoZWFkZXJzLCBzdGFydD0xKToKICAgICAgICBjID0gd3MuY2VsbChyb3c9SERSLCBj"
    "b2x1bW49aSwgdmFsdWU9aCkKICAgICAgICBjLmZvbnQgPSBXSElURTsgYy5maWxsID0gU1VCX0ZJ"
    "TEw7IGMuYWxpZ25tZW50ID0gQ0VOVEVSOyBjLmJvcmRlciA9IEJPUkRFUgogICAgd3Mucm93X2Rp"
    "bWVuc2lvbnNbSERSXS5oZWlnaHQgPSAzNgoKICAgIHNyID0gMAogICAgciA9IEhEUiArIDEKICAg"
    "IGRhdGFfc3RhcnQgPSByCiAgICBmb3IgcmV0IGluIHJldHVybnM6CiAgICAgICAgbSA9IHJldFsi"
    "bWV0YSJdCiAgICAgICAgZm9yIHJvdyBpbiByZXQuZ2V0KCJjb25zb2xpZGF0ZWQiLCBbXSk6CiAg"
    "ICAgICAgICAgIHNyICs9IDEKICAgICAgICAgICAgdmFscyA9IFsKICAgICAgICAgICAgICAgIHNy"
    "LCBtWyJtb250aCJdLCBtWyJmeSJdLCBtWyJzdGF0ZV9jb2RlIl0sIG1bInN0YXRlX25hbWUiXSwg"
    "bVsiZ3N0aW4iXSwKICAgICAgICAgICAgICAgIG1bImxlZ2FsX25hbWUiXSwgbVsiYXJuIl0sCiAg"
    "ICAgICAgICAgICAgICByb3dbInNlY3Rpb24iXSwgcm93WyJkZXNjcmlwdGlvbiJdLAogICAgICAg"
    "ICAgICAgICAgcm93LmdldCgidGF4YWJsZSIpLCByb3cuZ2V0KCJpZ3N0IiwgMC4wKSwgcm93Lmdl"
    "dCgiY2dzdCIsIDAuMCksCiAgICAgICAgICAgICAgICByb3cuZ2V0KCJzZ3N0IiwgMC4wKSwgcm93"
    "LmdldCgiY2VzcyIsIDAuMCksCiAgICAgICAgICAgIF0KICAgICAgICAgICAgZm9yIGksIHYgaW4g"
    "ZW51bWVyYXRlKHZhbHMsIHN0YXJ0PTEpOgogICAgICAgICAgICAgICAgY2VsbCA9IHdzLmNlbGwo"
    "cm93PXIsIGNvbHVtbj1pLCB2YWx1ZT12KQogICAgICAgICAgICAgICAgY2VsbC5mb250ID0gUkVH"
    "OyBjZWxsLmJvcmRlciA9IEJPUkRFUgogICAgICAgICAgICAgICAgaWYgaSBpbiAoMSwgMiwgMywg"
    "NCwgOSk6CiAgICAgICAgICAgICAgICAgICAgY2VsbC5hbGlnbm1lbnQgPSBDRU5URVIKICAgICAg"
    "ICAgICAgICAgIGVsaWYgaSBpbiAoNSwgNiwgNywgOCwgMTApOgogICAgICAgICAgICAgICAgICAg"
    "IGNlbGwuYWxpZ25tZW50ID0gTEVGVAogICAgICAgICAgICAgICAgZWxzZToKICAgICAgICAgICAg"
    "ICAgICAgICBjZWxsLmFsaWdubWVudCA9IFJJR0hUCiAgICAgICAgICAgICAgICAgICAgY2VsbC5u"
    "dW1iZXJfZm9ybWF0ID0gTlVNX0ZNVAogICAgICAgICAgICByICs9IDEKICAgIGRhdGFfZW5kID0g"
    "ciAtIDEKCiAgICB3aWR0aHMgPSB7IkEiOjcsIkIiOjksIkMiOjksIkQiOjYsIkUiOjE4LCJGIjoy"
    "MCwiRyI6MjIsIkgiOjE4LAogICAgICAgICAgICAgICJJIjoxNCwiSiI6NTIsIksiOjE4LCJMIjox"
    "OCwiTSI6MTYsIk4iOjE2LCJPIjoxOH0KICAgIGZvciBjb2wsIHcgaW4gd2lkdGhzLml0ZW1zKCk6"
    "CiAgICAgICAgd3MuY29sdW1uX2RpbWVuc2lvbnNbY29sXS53aWR0aCA9IHcKICAgIHdzLmZyZWV6"
    "ZV9wYW5lcyA9ICJLNCIKICAgIGlmIGRhdGFfZW5kID49IGRhdGFfc3RhcnQ6CiAgICAgICAgd3Mu"
    "YXV0b19maWx0ZXIucmVmID0gZiJBe0hEUn06T3tkYXRhX2VuZH0iCgogICAgIyA9PT09PT09PT09"
    "IFNoZWV0IDI6IFRheCBQYXltZW50ICg2LjEpID09PT09PT09PT09PT09PT09PT09PT0KICAgIHdz"
    "MiA9IHdiLmNyZWF0ZV9zaGVldCgiVGF4IFBheW1lbnQgKDYuMSkiKQogICAgd3MyLm1lcmdlX2Nl"
    "bGxzKCJBMTpRMSIpCiAgICB3czJbIkExIl0gPSAiR1NUUi0zQiAgVEFCTEUgNi4xICDigJQgUEFZ"
    "TUVOVCBPRiBUQVgiCiAgICB3czJbIkExIl0uZm9udCA9IFRJVExFOyB3czJbIkExIl0uZmlsbCA9"
    "IEhEUl9GSUxMOyB3czJbIkExIl0uYWxpZ25tZW50ID0gQ0VOVEVSCiAgICB3czIucm93X2RpbWVu"
    "c2lvbnNbMV0uaGVpZ2h0ID0gMjgKCiAgICBoMiA9IFsKICAgICAgICAiU3IgTm8iLCAiTW9udGgi"
    "LCAiRlkiLCAiU3RhdGUgQ29kZSIsICJTdGF0ZSBOYW1lIiwgIkdTVElOIiwKICAgICAgICAiU3Vi"
    "LXNlY3Rpb24iLCAiU3ViLXNlY3Rpb24gRGVzYy4iLCAiVGF4IFR5cGUiLAogICAgICAgICJUYXgg"
    "UGF5YWJsZSAo4oK5KSIsICJBZGp1c3RtZW50ICjigrkpIiwgIk5ldCBUYXggUGF5YWJsZSAo4oK5"
    "KSIsCiAgICAgICAgIlBhaWQgdmlhIElUQyAtIElHU1QgKOKCuSkiLCAiUGFpZCB2aWEgSVRDIC0g"
    "Q0dTVCAo4oK5KSIsCiAgICAgICAgIlBhaWQgdmlhIElUQyAtIFNHU1QvVVQgKOKCuSkiLCAiUGFp"
    "ZCB2aWEgSVRDIC0gQ2VzcyAo4oK5KSIsCiAgICAgICAgIlRheCBQYWlkIGluIENhc2ggKOKCuSki"
    "LCAiSW50ZXJlc3QgaW4gQ2FzaCAo4oK5KSIsICJMYXRlIEZlZSBpbiBDYXNoICjigrkpIiwKICAg"
    "IF0KICAgIGZvciBpLCBoIGluIGVudW1lcmF0ZShoMiwgc3RhcnQ9MSk6CiAgICAgICAgYyA9IHdz"
    "Mi5jZWxsKHJvdz0zLCBjb2x1bW49aSwgdmFsdWU9aCkKICAgICAgICBjLmZvbnQgPSBXSElURTsg"
    "Yy5maWxsID0gU1VCX0ZJTEw7IGMuYWxpZ25tZW50ID0gQ0VOVEVSOyBjLmJvcmRlciA9IEJPUkRF"
    "UgogICAgd3MyLnJvd19kaW1lbnNpb25zWzNdLmhlaWdodCA9IDQyCgogICAgc3IgPSAwOyByciA9"
    "IDQKICAgIGZvciByZXQgaW4gcmV0dXJuczoKICAgICAgICBtID0gcmV0WyJtZXRhIl0KICAgICAg"
    "ICBmb3Igcm93IGluIHJldC5nZXQoInBheW1lbnQiLCBbXSk6CiAgICAgICAgICAgIHNyICs9IDEK"
    "ICAgICAgICAgICAgdmFscyA9IFsKICAgICAgICAgICAgICAgIHNyLCBtWyJtb250aCJdLCBtWyJm"
    "eSJdLCBtWyJzdGF0ZV9jb2RlIl0sIG1bInN0YXRlX25hbWUiXSwgbVsiZ3N0aW4iXSwKICAgICAg"
    "ICAgICAgICAgIHJvd1sic3Vic2VjdGlvbiJdLCByb3dbInN1YnNlY3Rpb25fZGVzYyJdLCByb3db"
    "InRheF90eXBlIl0sCiAgICAgICAgICAgICAgICByb3dbInRheF9wYXlhYmxlIl0sIHJvd1siYWRq"
    "dXN0bWVudCJdLCByb3dbIm5ldF90YXhfcGF5YWJsZSJdLAogICAgICAgICAgICAgICAgcm93WyJp"
    "dGNfaWdzdCJdLCByb3dbIml0Y19jZ3N0Il0sIHJvd1siaXRjX3Nnc3QiXSwgcm93WyJpdGNfY2Vz"
    "cyJdLAogICAgICAgICAgICAgICAgcm93WyJ0YXhfaW5fY2FzaCJdLCByb3dbImludGVyZXN0X2Nh"
    "c2giXSwgcm93WyJsYXRlX2ZlZV9jYXNoIl0sCiAgICAgICAgICAgIF0KICAgICAgICAgICAgZm9y"
    "IGksIHYgaW4gZW51bWVyYXRlKHZhbHMsIHN0YXJ0PTEpOgogICAgICAgICAgICAgICAgY2VsbCA9"
    "IHdzMi5jZWxsKHJvdz1yciwgY29sdW1uPWksIHZhbHVlPXYpCiAgICAgICAgICAgICAgICBjZWxs"
    "LmZvbnQgPSBSRUc7IGNlbGwuYm9yZGVyID0gQk9SREVSCiAgICAgICAgICAgICAgICBpZiBpIGlu"
    "ICgxLCAyLCAzLCA0LCA3KToKICAgICAgICAgICAgICAgICAgICBjZWxsLmFsaWdubWVudCA9IENF"
    "TlRFUgogICAgICAgICAgICAgICAgZWxpZiBpIGluICg1LCA2LCA4LCA5KToKICAgICAgICAgICAg"
    "ICAgICAgICBjZWxsLmFsaWdubWVudCA9IExFRlQKICAgICAgICAgICAgICAgIGVsc2U6CiAgICAg"
    "ICAgICAgICAgICAgICAgY2VsbC5hbGlnbm1lbnQgPSBSSUdIVAogICAgICAgICAgICAgICAgICAg"
    "IGNlbGwubnVtYmVyX2Zvcm1hdCA9IE5VTV9GTVQKICAgICAgICAgICAgcnIgKz0gMQoKICAgIGZv"
    "ciBjb2wsIHcgaW4gWygiQSIsNiksKCJCIiw5KSwoIkMiLDkpLCgiRCIsNiksKCJFIiwxOCksKCJG"
    "IiwyMCksCiAgICAgICAgICAgICAgICAgICAoIkciLDEwKSwoIkgiLDMwKSwoIkkiLDE2KSwKICAg"
    "ICAgICAgICAgICAgICAgICgiSiIsMTYpLCgiSyIsMTQpLCgiTCIsMTgpLAogICAgICAgICAgICAg"
    "ICAgICAgKCJNIiwxOCksKCJOIiwxOCksKCJPIiwxOCksKCJQIiwxNiksCiAgICAgICAgICAgICAg"
    "ICAgICAoIlEiLDE4KSwoIlIiLDE4KSwoIlMiLDE4KV06CiAgICAgICAgd3MyLmNvbHVtbl9kaW1l"
    "bnNpb25zW2NvbF0ud2lkdGggPSB3CiAgICB3czIuZnJlZXplX3BhbmVzID0gIko0IgogICAgaWYg"
    "cnIgPiA0OgogICAgICAgIHdzMi5hdXRvX2ZpbHRlci5yZWYgPSBmIkEzOlN7cnItMX0iCgogICAg"
    "IyA9PT09PT09PT09IFNoZWV0IDM6IFN0YXRlLU1vbnRoIFN1bW1hcnkgPT09PT09PT09PT09PT09"
    "PT09PT0KICAgIHdzMyA9IHdiLmNyZWF0ZV9zaGVldCgiU3RhdGUtTW9udGggU3VtbWFyeSIpCiAg"
    "ICB3czMubWVyZ2VfY2VsbHMoIkExOk4xIikKICAgIHdzM1siQTEiXSA9ICJLRVkgVE9UQUxTIFBF"
    "UiBSRVRVUk4iCiAgICB3czNbIkExIl0uZm9udCA9IFRJVExFOyB3czNbIkExIl0uZmlsbCA9IEhE"
    "Ul9GSUxMOyB3czNbIkExIl0uYWxpZ25tZW50ID0gQ0VOVEVSCiAgICB3czMucm93X2RpbWVuc2lv"
    "bnNbMV0uaGVpZ2h0ID0gMjgKCiAgICBoMyA9IFsKICAgICAgICAiU3IgTm8iLCAiTW9udGgiLCAi"
    "U3RhdGUgQ29kZSIsICJTdGF0ZSBOYW1lIiwgIkdTVElOIiwKICAgICAgICAiMy4xKGEpIFRheGFi"
    "bGUgKOKCuSkiLCAiMy4xKGEpIElHU1QgKOKCuSkiLCAiMy4xKGEpIENHU1QgKOKCuSkiLCAiMy4x"
    "KGEpIFNHU1QgKOKCuSkiLAogICAgICAgICJUb3RhbCBJVEMgQXZhaWwuIOKAlCBBICjigrkpIiwK"
    "ICAgICAgICAiVG90YWwgSVRDIFJldmVyc2VkIOKAlCBCICjigrkpIiwKICAgICAgICAiTmV0IElU"
    "QyDigJQgQyAo4oK5KSIsCiAgICAgICAgIkNhc2ggUGFpZCAoNi4xIOKAlCBJR1NUK0NHU1QrU0dT"
    "VCtDZXNzKSAo4oK5KSIsCiAgICAgICAgIlRvdGFsIExpYWJpbGl0eSBCcmVha3VwICjigrkpIiwK"
    "ICAgIF0KICAgIGZvciBpLCBoIGluIGVudW1lcmF0ZShoMywgc3RhcnQ9MSk6CiAgICAgICAgYyA9"
    "IHdzMy5jZWxsKHJvdz0zLCBjb2x1bW49aSwgdmFsdWU9aCkKICAgICAgICBjLmZvbnQgPSBXSElU"
    "RTsgYy5maWxsID0gU1VCX0ZJTEw7IGMuYWxpZ25tZW50ID0gQ0VOVEVSOyBjLmJvcmRlciA9IEJP"
    "UkRFUgogICAgd3MzLnJvd19kaW1lbnNpb25zWzNdLmhlaWdodCA9IDUwCgogICAgc3IgPSAwOyBy"
    "ciA9IDQKICAgIGZvciByZXQgaW4gcmV0dXJuczoKICAgICAgICBtID0gcmV0WyJtZXRhIl0KICAg"
    "ICAgICBjb25zID0gcmV0LmdldCgiY29uc29saWRhdGVkIiwgW10pCiAgICAgICAgcGF5ID0gcmV0"
    "LmdldCgicGF5bWVudCIsIFtdKQoKICAgICAgICBkZWYgZmluZChzZWN0aW9uX3ByZWZpeCk6CiAg"
    "ICAgICAgICAgIHJldHVybiBuZXh0KChyb3cgZm9yIHJvdyBpbiBjb25zIGlmIHJvd1sic2VjdGlv"
    "biJdID09IHNlY3Rpb25fcHJlZml4KSwgTm9uZSkKCiAgICAgICAgcl8zXzFfYSA9IGZpbmQoIjMu"
    "MShhKSIpCiAgICAgICAgIyBTdW0gNChBKSgxLi41KQogICAgICAgIGFfaWdzdCA9IHN1bShyb3db"
    "Imlnc3QiXSBmb3Igcm93IGluIGNvbnMgaWYgcm93WyJzZWN0aW9uIl0uc3RhcnRzd2l0aCgiNChB"
    "KSgiKSkKICAgICAgICBhX2Nnc3QgPSBzdW0ocm93WyJjZ3N0Il0gZm9yIHJvdyBpbiBjb25zIGlm"
    "IHJvd1sic2VjdGlvbiJdLnN0YXJ0c3dpdGgoIjQoQSkoIikpCiAgICAgICAgYV9zZ3N0ID0gc3Vt"
    "KHJvd1sic2dzdCJdIGZvciByb3cgaW4gY29ucyBpZiByb3dbInNlY3Rpb24iXS5zdGFydHN3aXRo"
    "KCI0KEEpKCIpKQogICAgICAgIGFfY2VzcyA9IHN1bShyb3dbImNlc3MiXSBmb3Igcm93IGluIGNv"
    "bnMgaWYgcm93WyJzZWN0aW9uIl0uc3RhcnRzd2l0aCgiNChBKSgiKSkKICAgICAgICB0b3RhbF9h"
    "ID0gYV9pZ3N0ICsgYV9jZ3N0ICsgYV9zZ3N0ICsgYV9jZXNzCgogICAgICAgICMgU3VtIDQoQiko"
    "MS4uMikKICAgICAgICBiX2lnc3QgPSBzdW0ocm93WyJpZ3N0Il0gZm9yIHJvdyBpbiBjb25zIGlm"
    "IHJvd1sic2VjdGlvbiJdLnN0YXJ0c3dpdGgoIjQoQikoIikpCiAgICAgICAgYl9jZ3N0ID0gc3Vt"
    "KHJvd1siY2dzdCJdIGZvciByb3cgaW4gY29ucyBpZiByb3dbInNlY3Rpb24iXS5zdGFydHN3aXRo"
    "KCI0KEIpKCIpKQogICAgICAgIGJfc2dzdCA9IHN1bShyb3dbInNnc3QiXSBmb3Igcm93IGluIGNv"
    "bnMgaWYgcm93WyJzZWN0aW9uIl0uc3RhcnRzd2l0aCgiNChCKSgiKSkKICAgICAgICBiX2Nlc3Mg"
    "PSBzdW0ocm93WyJjZXNzIl0gZm9yIHJvdyBpbiBjb25zIGlmIHJvd1sic2VjdGlvbiJdLnN0YXJ0"
    "c3dpdGgoIjQoQikoIikpCiAgICAgICAgdG90YWxfYiA9IGJfaWdzdCArIGJfY2dzdCArIGJfc2dz"
    "dCArIGJfY2VzcwoKICAgICAgICByXzRfYyA9IGZpbmQoIjQoQykiKQogICAgICAgIGNfdG90YWwg"
    "PSAwLjAKICAgICAgICBpZiByXzRfYzoKICAgICAgICAgICAgY190b3RhbCA9IChyXzRfYy5nZXQo"
    "Imlnc3QiKSBvciAwKSArIChyXzRfYy5nZXQoImNnc3QiKSBvciAwKSArIChyXzRfYy5nZXQoInNn"
    "c3QiKSBvciAwKSArIChyXzRfYy5nZXQoImNlc3MiKSBvciAwKQoKICAgICAgICAjIENhc2ggcGFp"
    "ZCB0b3RhbAogICAgICAgIGNhc2hfcGFpZCA9IHN1bShyb3dbInRheF9pbl9jYXNoIl0gZm9yIHJv"
    "dyBpbiBwYXkpCgogICAgICAgICMgQnJlYWt1cCB0b3RhbAogICAgICAgIGJyZWFrdXBfdG90YWwg"
    "PSAwLjAKICAgICAgICBmb3Igcm93IGluIGNvbnM6CiAgICAgICAgICAgIGlmIHJvd1sic2VjdGlv"
    "biJdID09ICJCcmVha3VwIjoKICAgICAgICAgICAgICAgIGJyZWFrdXBfdG90YWwgPSAocm93Lmdl"
    "dCgiaWdzdCIpIG9yIDApICsgKHJvdy5nZXQoImNnc3QiKSBvciAwKSArIChyb3cuZ2V0KCJzZ3N0"
    "Iikgb3IgMCkgKyAocm93LmdldCgiY2VzcyIpIG9yIDApCiAgICAgICAgICAgICAgICBicmVhawoK"
    "ICAgICAgICBzciArPSAxCiAgICAgICAgdmFscyA9IFsKICAgICAgICAgICAgc3IsIG1bIm1vbnRo"
    "Il0sIG1bInN0YXRlX2NvZGUiXSwgbVsic3RhdGVfbmFtZSJdLCBtWyJnc3RpbiJdLAogICAgICAg"
    "ICAgICAocl8zXzFfYSBvciB7fSkuZ2V0KCJ0YXhhYmxlIiwgMC4wKSwKICAgICAgICAgICAgKHJf"
    "M18xX2Egb3Ige30pLmdldCgiaWdzdCIsIDAuMCksCiAgICAgICAgICAgIChyXzNfMV9hIG9yIHt9"
    "KS5nZXQoImNnc3QiLCAwLjApLAogICAgICAgICAgICAocl8zXzFfYSBvciB7fSkuZ2V0KCJzZ3N0"
    "IiwgMC4wKSwKICAgICAgICAgICAgdG90YWxfYSwgdG90YWxfYiwgY190b3RhbCwgY2FzaF9wYWlk"
    "LCBicmVha3VwX3RvdGFsLAogICAgICAgIF0KICAgICAgICBmb3IgaSwgdiBpbiBlbnVtZXJhdGUo"
    "dmFscywgc3RhcnQ9MSk6CiAgICAgICAgICAgIGNlbGwgPSB3czMuY2VsbChyb3c9cnIsIGNvbHVt"
    "bj1pLCB2YWx1ZT12KQogICAgICAgICAgICBjZWxsLmZvbnQgPSBSRUc7IGNlbGwuYm9yZGVyID0g"
    "Qk9SREVSCiAgICAgICAgICAgIGlmIGkgPD0gNDoKICAgICAgICAgICAgICAgIGNlbGwuYWxpZ25t"
    "ZW50ID0gQ0VOVEVSIGlmIGkgIT0gNCBlbHNlIExFRlQKICAgICAgICAgICAgZWxpZiBpID09IDU6"
    "CiAgICAgICAgICAgICAgICBjZWxsLmFsaWdubWVudCA9IExFRlQKICAgICAgICAgICAgZWxzZToK"
    "ICAgICAgICAgICAgICAgIGNlbGwuYWxpZ25tZW50ID0gUklHSFQKICAgICAgICAgICAgICAgIGNl"
    "bGwubnVtYmVyX2Zvcm1hdCA9IE5VTV9GTVQKICAgICAgICByciArPSAxCgogICAgZm9yIGNvbCwg"
    "dyBpbiBbKCJBIiw2KSwoIkIiLDkpLCgiQyIsNiksKCJEIiwxOCksKCJFIiwyMiksCiAgICAgICAg"
    "ICAgICAgICAgICAoIkYiLDE4KSwoIkciLDE2KSwoIkgiLDE2KSwoIkkiLDE2KSwKICAgICAgICAg"
    "ICAgICAgICAgICgiSiIsMjApLCgiSyIsMjApLCgiTCIsMTgpLCgiTSIsMjYpLCgiTiIsMjIpXToK"
    "ICAgICAgICB3czMuY29sdW1uX2RpbWVuc2lvbnNbY29sXS53aWR0aCA9IHcKICAgIHdzMy5mcmVl"
    "emVfcGFuZXMgPSAiQTQiCiAgICBpZiByciA+IDQ6CiAgICAgICAgd3MzLmF1dG9fZmlsdGVyLnJl"
    "ZiA9IGYiQTM6Tntyci0xfSIKCiAgICAjID09PT09PT09PT0gU2hlZXQgNDogUHJvY2Vzc2luZyBM"
    "b2cgPT09PT09PT09PT09PT09PT09PT09PT09PQogICAgd3M0ID0gd2IuY3JlYXRlX3NoZWV0KCJQ"
    "cm9jZXNzaW5nIExvZyIpCiAgICBsb2dfaGVhZGVycyA9IFsiIyIsICJTb3VyY2UgRmlsZSIsICJT"
    "dGF0dXMiLCAiR1NUSU4iLCAiU3RhdGUiLCAiTW9udGgiLAogICAgICAgICAgICAgICAgICAgIkNv"
    "bnNvbGlkYXRlZCBSb3dzIiwgIlBheW1lbnQgUm93cyIsICJOb3RlcyJdCiAgICB3czQuYXBwZW5k"
    "KGxvZ19oZWFkZXJzKQogICAgZm9yIGMgaW4gcmFuZ2UoMSwgbGVuKGxvZ19oZWFkZXJzKSArIDEp"
    "OgogICAgICAgIGNlbGwgPSB3czQuY2VsbChyb3c9MSwgY29sdW1uPWMpCiAgICAgICAgY2VsbC5m"
    "b250ID0gV0hJVEU7IGNlbGwuZmlsbCA9IFNVQl9GSUxMOyBjZWxsLmFsaWdubWVudCA9IENFTlRF"
    "UjsgY2VsbC5ib3JkZXIgPSBCT1JERVIKICAgIGZvciBpLCByZXQgaW4gZW51bWVyYXRlKHJldHVy"
    "bnMsIHN0YXJ0PTEpOgogICAgICAgIG0gPSByZXRbIm1ldGEiXQogICAgICAgIHN0YXR1cyA9IHJl"
    "dC5nZXQoInN0YXR1cyIsICJPSyIpCiAgICAgICAgb3V0ID0gW2ksIHJldC5nZXQoInNvdXJjZV9m"
    "aWxlIiwiIiksIHN0YXR1cywKICAgICAgICAgICAgICAgbS5nZXQoImdzdGluIiwiIiksIG0uZ2V0"
    "KCJzdGF0ZV9uYW1lIiwiIiksIG0uZ2V0KCJtb250aCIsIiIpLAogICAgICAgICAgICAgICBsZW4o"
    "cmV0LmdldCgiY29uc29saWRhdGVkIiwgW10pKSwgbGVuKHJldC5nZXQoInBheW1lbnQiLCBbXSkp"
    "LAogICAgICAgICAgICAgICByZXQuZ2V0KCJub3RlcyIsIiIpXQogICAgICAgIGZvciBjLCB2IGlu"
    "IGVudW1lcmF0ZShvdXQsIHN0YXJ0PTEpOgogICAgICAgICAgICBjZWxsID0gd3M0LmNlbGwocm93"
    "PWkrMSwgY29sdW1uPWMsIHZhbHVlPXYpCiAgICAgICAgICAgIGNlbGwuZm9udCA9IFJFRzsgY2Vs"
    "bC5ib3JkZXIgPSBCT1JERVI7IGNlbGwuYWxpZ25tZW50ID0gTEVGVAogICAgICAgICAgICBpZiBz"
    "dGF0dXMgIT0gIk9LIjoKICAgICAgICAgICAgICAgIGNlbGwuZmlsbCA9IEVSUl9GSUxMCiAgICBm"
    "b3IgY29sLCB3IGluIFsoIkEiLDYpLCgiQiIsNDIpLCgiQyIsMTApLCgiRCIsMjIpLCgiRSIsMjIp"
    "LCgiRiIsMTApLAogICAgICAgICAgICAgICAgICAgKCJHIiwxOCksKCJIIiwxMyksKCJJIiw1MCld"
    "OgogICAgICAgIHdzNC5jb2x1bW5fZGltZW5zaW9uc1tjb2xdLndpZHRoID0gdwogICAgd3M0LmZy"
    "ZWV6ZV9wYW5lcyA9ICJBMiIKCiAgICB3Yi5zYXZlKG91dHB1dF9wYXRoKQoKCiMgLS0tLS0tLS0t"
    "LS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0t"
    "LS0tLS0tLS0tICMKIyAgT3V0cHV0IHBhdGggaGVscGVyCiMgLS0tLS0tLS0tLS0tLS0tLS0tLS0t"
    "LS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tICMK"
    "CmRlZiByZXNvbHZlX291dHB1dF9wYXRoKHJhdyk6CiAgICBwID0gUGF0aChyYXcpCiAgICBERUZB"
    "VUxUX05BTUUgPSAiR1NUUjNCX0NvbnNvbGlkYXRlZC54bHN4IgogICAgaWYgcC5leGlzdHMoKSBh"
    "bmQgcC5pc19kaXIoKToKICAgICAgICBwID0gcCAvIERFRkFVTFRfTkFNRQogICAgZWxpZiBwLnN1"
    "ZmZpeC5sb3dlcigpICE9ICIueGxzeCI6CiAgICAgICAgcCA9IHAud2l0aF9zdWZmaXgoIi54bHN4"
    "IikKICAgIHAucGFyZW50Lm1rZGlyKHBhcmVudHM9VHJ1ZSwgZXhpc3Rfb2s9VHJ1ZSkKICAgIGlm"
    "IHAuZXhpc3RzKCk6CiAgICAgICAgdHJ5OgogICAgICAgICAgICB3aXRoIG9wZW4ocCwgImFiIik6"
    "CiAgICAgICAgICAgICAgICBwYXNzCiAgICAgICAgZXhjZXB0IFBlcm1pc3Npb25FcnJvcjoKICAg"
    "ICAgICAgICAgcmFpc2UgUGVybWlzc2lvbkVycm9yKAogICAgICAgICAgICAgICAgZiJPdXRwdXQg"
    "ZmlsZSBpcyBsb2NrZWQgKHByb2JhYmx5IG9wZW4gaW4gRXhjZWwpOlxuICB7cH1cbiIKICAgICAg"
    "ICAgICAgICAgICJDbG9zZSBpdCBpbiBFeGNlbCBhbmQgcmUtcnVuLiIKICAgICAgICAgICAgKQog"
    "ICAgcmV0dXJuIHAKCgojIC0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0t"
    "LS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLSAjCiMgIFByb2Nlc3NpbmcgcGlwZWxp"
    "bmUKIyAtLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0t"
    "LS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0gIwoKZGVmIHByb2Nlc3NfcGRmcyhpbl9mb2xkZXIsIG91"
    "dF9wYXRoLCBvbl9wcm9ncmVzcz1Ob25lLCBvbl9sb2c9Tm9uZSk6CiAgICBwZGZzID0gc29ydGVk"
    "KFtwIGZvciBwIGluIGluX2ZvbGRlci5yZ2xvYigiKi5wZGYiKV0pCiAgICBpZiBub3QgcGRmczoK"
    "ICAgICAgICByYWlzZSBGaWxlTm90Rm91bmRFcnJvcihmIk5vIFBERnMgZm91bmQgdW5kZXI6IHtp"
    "bl9mb2xkZXJ9IikKCiAgICBpZiBvbl9sb2c6CiAgICAgICAgb25fbG9nKGYiRm91bmQge2xlbihw"
    "ZGZzKX0gUERGIGZpbGUocykuICBQcm9jZXNzaW5nLi4uIiwgImluZm8iKQoKICAgIHJldHVybnMg"
    "PSBbXQogICAgb2sgPSAwOyBmYWlsID0gMAoKICAgIGZvciBpZHgsIHBkZl9wYXRoIGluIGVudW1l"
    "cmF0ZShwZGZzLCBzdGFydD0xKToKICAgICAgICB0cnk6CiAgICAgICAgICAgIHJlbCA9IHBkZl9w"
    "YXRoLnJlbGF0aXZlX3RvKGluX2ZvbGRlcikKICAgICAgICBleGNlcHQgVmFsdWVFcnJvcjoKICAg"
    "ICAgICAgICAgcmVsID0gcGRmX3BhdGgubmFtZQoKICAgICAgICB0cnk6CiAgICAgICAgICAgIG1l"
    "dGEsIGNvbnNvbGlkYXRlZCwgcGF5bWVudCA9IHBhcnNlX3BkZihwZGZfcGF0aCkKICAgICAgICAg"
    "ICAgcmV0dXJucy5hcHBlbmQoewogICAgICAgICAgICAgICAgIm1ldGEiOiBtZXRhLCAiY29uc29s"
    "aWRhdGVkIjogY29uc29saWRhdGVkLCAicGF5bWVudCI6IHBheW1lbnQsCiAgICAgICAgICAgICAg"
    "ICAic291cmNlX2ZpbGUiOiBzdHIocmVsKSwgInN0YXR1cyI6ICJPSyIsICJub3RlcyI6ICIiLAog"
    "ICAgICAgICAgICB9KQogICAgICAgICAgICBvayArPSAxCiAgICAgICAgICAgIGlmIG9uX2xvZzoK"
    "ICAgICAgICAgICAgICAgIG9uX2xvZygKICAgICAgICAgICAgICAgICAgICBmIlt7aWR4Oj40fS97"
    "bGVuKHBkZnMpfV0gIE9LICAgICIKICAgICAgICAgICAgICAgICAgICBmInttZXRhWydzdGF0ZV9u"
    "YW1lJ106PDIyfSB7bWV0YVsnbW9udGgnXTo8OH0gICIKICAgICAgICAgICAgICAgICAgICBmIntt"
    "ZXRhWydnc3RpbiddfSAgKHtwZGZfcGF0aC5uYW1lfSkiLAogICAgICAgICAgICAgICAgICAgICJv"
    "ayIsCiAgICAgICAgICAgICAgICApCiAgICAgICAgZXhjZXB0IEV4Y2VwdGlvbiBhcyBlOgogICAg"
    "ICAgICAgICBmYWlsICs9IDEKICAgICAgICAgICAgcmV0dXJucy5hcHBlbmQoewogICAgICAgICAg"
    "ICAgICAgIm1ldGEiOiB7ImdzdGluIjoiIiwic3RhdGVfY29kZSI6IiIsInN0YXRlX25hbWUiOiIi"
    "LCJtb250aCI6IiIsCiAgICAgICAgICAgICAgICAgICAgICAgICAiZnkiOiIiLCJsZWdhbF9uYW1l"
    "IjoiIiwiYXJuIjoiIn0sCiAgICAgICAgICAgICAgICAiY29uc29saWRhdGVkIjogW10sICJwYXlt"
    "ZW50IjogW10sCiAgICAgICAgICAgICAgICAic291cmNlX2ZpbGUiOiBzdHIocmVsKSwKICAgICAg"
    "ICAgICAgICAgICJzdGF0dXMiOiAiRkFJTCIsCiAgICAgICAgICAgICAgICAibm90ZXMiOiBmInt0"
    "eXBlKGUpLl9fbmFtZV9ffToge2V9IiwKICAgICAgICAgICAgfSkKICAgICAgICAgICAgaWYgb25f"
    "bG9nOgogICAgICAgICAgICAgICAgb25fbG9nKGYiW3tpZHg6PjR9L3tsZW4ocGRmcyl9XSAgRkFJ"
    "TCAge3BkZl9wYXRoLm5hbWV9ICDihpIgIHtlfSIsICJmYWlsIikKCiAgICAgICAgaWYgb25fcHJv"
    "Z3Jlc3M6CiAgICAgICAgICAgIG9uX3Byb2dyZXNzKGlkeCwgbGVuKHBkZnMpKQoKICAgIGlmIG9u"
    "X2xvZzoKICAgICAgICBvbl9sb2coZiJXcml0aW5nIGNvbnNvbGlkYXRlZCBFeGNlbDogIHtvdXRf"
    "cGF0aH0iLCAiaW5mbyIpCiAgICB3cml0ZV9leGNlbChyZXR1cm5zLCBvdXRfcGF0aCkKICAgIHJl"
    "dHVybiBvaywgZmFpbCwgbGVuKHBkZnMpLCBvdXRfcGF0aAoKCiMgLS0tLS0tLS0tLS0tLS0tLS0t"
    "LS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0t"
    "ICMKIyAgR1VJCiMgLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0t"
    "LS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tICMKCg=="
)

_GSTR1_NS = {"__name__": "_gstr1_engine"}
_GSTR3B_NS = {"__name__": "_gstr3b_engine"}

def _load_extractor_engines():
    """Decode and exec the embedded engines into their isolated namespaces.
    Called lazily — only when the user clicks Run on those tabs."""
    if not HAS_PDFPLUMBER:
        raise RuntimeError(
            "pdfplumber not installed. In CMD:  pip install pdfplumber openpyxl")
    if "parse_pdf" not in _GSTR1_NS:
        src1 = _base64.b64decode("".join(_GSTR1_ENGINE_B64)).decode("utf-8")
        # Provide pre-imported modules so engines don't re-import
        _GSTR1_NS["os"] = os
        _GSTR1_NS["re"] = re
        _GSTR1_NS["sys"] = sys
        _GSTR1_NS["traceback"] = traceback
        from pathlib import Path as _P
        _GSTR1_NS["Path"] = _P
        import pdfplumber as _pp
        _GSTR1_NS["pdfplumber"] = _pp
        from openpyxl import Workbook as _WB
        from openpyxl.styles import Font as _F, PatternFill as _PF, Alignment as _A, Border as _B, Side as _S
        from openpyxl.utils import get_column_letter as _gcl
        _GSTR1_NS["Workbook"] = _WB
        _GSTR1_NS["Font"] = _F; _GSTR1_NS["PatternFill"] = _PF
        _GSTR1_NS["Alignment"] = _A; _GSTR1_NS["Border"] = _B; _GSTR1_NS["Side"] = _S
        _GSTR1_NS["get_column_letter"] = _gcl
        exec(src1, _GSTR1_NS)

    if "parse_pdf" not in _GSTR3B_NS:
        src3 = _base64.b64decode("".join(_GSTR3B_ENGINE_B64)).decode("utf-8")
        _GSTR3B_NS["os"] = os
        _GSTR3B_NS["re"] = re
        _GSTR3B_NS["sys"] = sys
        from pathlib import Path as _P
        _GSTR3B_NS["Path"] = _P
        import pdfplumber as _pp
        _GSTR3B_NS["pdfplumber"] = _pp
        from openpyxl import Workbook as _WB
        from openpyxl.styles import Font as _F, PatternFill as _PF, Alignment as _A, Border as _B, Side as _S
        _GSTR3B_NS["Workbook"] = _WB
        _GSTR3B_NS["Font"] = _F; _GSTR3B_NS["PatternFill"] = _PF
        _GSTR3B_NS["Alignment"] = _A; _GSTR3B_NS["Border"] = _B; _GSTR3B_NS["Side"] = _S
        exec(src3, _GSTR3B_NS)


# ════════════════════════════════════════════════════════════════
#  TAX COMPARISON CONSOLIDATOR ENGINE
# ════════════════════════════════════════════════════════════════
# Reference for state codes
_STATE_CODES = {
    "01": "Jammu & Kashmir", "02": "Himachal Pradesh", "03": "Punjab",
    "04": "Chandigarh", "05": "Uttarakhand", "06": "Haryana", "07": "Delhi",
    "08": "Rajasthan", "09": "Uttar Pradesh", "10": "Bihar", "11": "Sikkim",
    "12": "Arunachal Pradesh", "13": "Nagaland", "14": "Manipur", "15": "Mizoram",
    "16": "Tripura", "17": "Meghalaya", "18": "Assam", "19": "West Bengal",
    "20": "Jharkhand", "21": "Odisha", "22": "Chhattisgarh", "23": "Madhya Pradesh",
    "24": "Gujarat", "25": "Daman & Diu", "26": "Dadra & NH and Daman & Diu",
    "27": "Maharashtra", "28": "Andhra Pradesh (Old)", "29": "Karnataka",
    "30": "Goa", "31": "Lakshadweep", "32": "Kerala", "33": "Tamil Nadu",
    "34": "Puducherry", "35": "Andaman & Nicobar Islands", "36": "Telangana",
    "37": "Andhra Pradesh", "38": "Ladakh", "97": "Other Territory",
    "99": "Centre Jurisdiction",
}

# Each comparison file from GSTN portal contains these 8 sheets with known structure.
# (header_row_for_columns, data_start_row, data_end_marker_col_A='Total', max_cols)
_TC_SHEET_LAYOUT = {
    "Tax Liability Summary": {
        "main_header_row":  8,    # column-group labels (rowspan with row 7)
        "sub_header_row":   9,    # IGST/CGST/SGST/CESS sub-labels
        "data_start_row":   10,
        "expected_cols":    27,
    },
    "Comparison Summary": {
        "main_header_row":  8,
        "sub_header_row":   9,
        "data_start_row":   10,
        "expected_cols":    15,
    },
    "Tax liability": {
        "main_header_row":  5,
        "sub_header_row":   6,
        "data_start_row":   7,
        "expected_cols":    37,
    },
    "Reverse charge": {
        "main_header_row":  5,
        "sub_header_row":   6,
        "data_start_row":   7,
        "expected_cols":    37,
    },
    "Export and SEZ": {
        "main_header_row":  5,
        "sub_header_row":   6,
        "data_start_row":   7,
        "expected_cols":    11,
    },
    "ITC (Other than IMPG)": {
        "main_header_row":  5,
        "sub_header_row":   6,
        "data_start_row":   7,
        "expected_cols":    37,
    },
    "ITC (IMPG)": {
        "main_header_row":  5,
        "sub_header_row":   6,
        "data_start_row":   7,
        "expected_cols":    11,
    },
    "RCM_LIABILITY_ITC": {
        "main_header_row":  5,
        "sub_header_row":   6,
        "data_start_row":   7,
        "expected_cols":    37,
    },
}


def parse_tax_comparison_file(filepath):
    """
    Parse a single GSTN portal 'Tax liability & ITC comparison' xlsx.
    Returns: {
        "meta": {gstin, fy, state_code, state_name, legal_name, trade_name, report_dt, source_file},
        "sheets": {sheet_name: {"main_headers": [...], "sub_headers": [...], "rows": [{month, values:[...]}, ...]}}
    }
    """
    import openpyxl, re
    import warnings
    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=False)

    # --- Extract meta from filename + first sheet ---
    fname = os.path.basename(filepath)
    m = re.search(r"(\d{4}-\d{2})_(\d{2}[A-Z0-9]{13})_", fname)
    fy_from_name = m.group(1) if m else ""
    gstin_from_name = m.group(2) if m else ""

    meta = {
        "source_file": fname,
        "gstin": gstin_from_name, "fy": fy_from_name,
        "state_code": gstin_from_name[:2] if gstin_from_name else "",
        "state_name": _STATE_CODES.get(gstin_from_name[:2], "Unknown") if gstin_from_name else "Unknown",
        "legal_name": "", "trade_name": "", "report_dt": "",
    }

    # Pull legal/trade names from any sheet (typically rows 4-5)
    for sn in wb.sheetnames:
        ws = wb[sn]
        for r in range(1, min(ws.max_row + 1, 8)):
            for c in range(1, min(ws.max_column + 1, 15)):
                v = ws.cell(r, c).value
                if not isinstance(v, str): continue
                t = v.strip()
                if t.startswith("GSTIN:"):
                    val = t.split(":", 1)[1].strip()
                    if val and not meta["gstin"]: meta["gstin"] = val
                elif t.startswith("Legal name:"):
                    if not meta["legal_name"]:
                        meta["legal_name"] = t.split(":", 1)[1].strip()
                elif t.startswith("Trade name"):
                    if not meta["trade_name"]:
                        meta["trade_name"] = t.split(":", 1)[1].strip()
                elif t.startswith("Financial Year:"):
                    if not meta["fy"]:
                        meta["fy"] = t.split(":", 1)[1].strip()
                elif t.startswith("Report generated"):
                    if not meta["report_dt"]:
                        meta["report_dt"] = t.split(":", 1)[1].strip() \
                            if ":" in t else t
        if meta["legal_name"]: break

    # Refill state info if GSTIN was only inside file
    if meta["gstin"] and not meta["state_code"]:
        meta["state_code"] = meta["gstin"][:2]
        meta["state_name"] = _STATE_CODES.get(meta["state_code"], "Unknown")

    # --- Extract each sheet ---
    sheets = {}
    for sn, layout in _TC_SHEET_LAYOUT.items():
        if sn not in wb.sheetnames:
            sheets[sn] = {"missing": True}
            continue
        ws = wb[sn]
        # Collect headers — first read main_header row (with merged-cell expansion),
        # then sub_header row. We capture column-by-column to preserve order.
        mh = []
        sh = []
        max_c = min(ws.max_column, layout["expected_cols"])
        for c in range(1, max_c + 1):
            # main header — resolve merged cells: if the cell is in a merged range,
            # take the top-left value of that range
            mh.append(_resolve_merged_value(ws, layout["main_header_row"], c))
            sh.append(_resolve_merged_value(ws, layout["sub_header_row"], c))
        # Data rows — read until "Total" found in column A
        rows = []
        r = layout["data_start_row"]
        while r <= ws.max_row:
            month = ws.cell(r, 1).value
            if month is None or (isinstance(month, str) and not month.strip()):
                r += 1
                continue
            values = []
            for c in range(2, max_c + 1):
                values.append(ws.cell(r, c).value)
            rows.append({
                "month":    str(month).strip(),
                "is_total": (str(month).strip().lower() == "total"),
                "values":   values,
            })
            if str(month).strip().lower() == "total":
                break
            r += 1
        sheets[sn] = {
            "main_headers": mh,
            "sub_headers":  sh,
            "rows":         rows,
        }
    wb.close()
    return {"meta": meta, "sheets": sheets}


def _resolve_merged_value(ws, row, col):
    """Return the cell value; if the cell is inside a merged range, return the
    top-left cell's value of that range."""
    cell = ws.cell(row, col)
    val = cell.value
    if val is not None and (not isinstance(val, str) or val.strip()):
        return val
    # Check merged ranges
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            return ws.cell(mr.min_row, mr.min_col).value
    return None


# Pretty short labels for output sheets (Excel limit 31 chars)
_SHORT_NAMES = {
    "Tax Liability Summary":  "Tax Liability Summary",
    "Comparison Summary":     "Comparison Summary",
    "Tax liability":          "1. Tax Liability",
    "Reverse charge":         "2. Reverse Charge",
    "Export and SEZ":         "3. Export & SEZ",
    "ITC (Other than IMPG)":  "4. ITC (Other than IMPG)",
    "ITC (IMPG)":             "5. ITC (IMPG)",
    "RCM_LIABILITY_ITC":      "6. RCM Liability & ITC",
}


def write_consolidated_comparison(all_data, out_path, mode):
    """
    all_data: list of dicts returned by parse_tax_comparison_file
    mode: 'single' | 'multi' | 'both'
    """
    import openpyxl
    from openpyxl.styles import Font as XLFont, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    HDR_FILL = PatternFill("solid", start_color="1F4E78")
    SUB_FILL = PatternFill("solid", start_color="2E75B6")
    META_FILL = PatternFill("solid", start_color="FFE699")
    TOTAL_FILL = PatternFill("solid", start_color="FFF2CC")
    GROUP_FILL = PatternFill("solid", start_color="DDEBF7")

    WHITE_B = XLFont(name="Calibri", bold=True, color="FFFFFF", size=10)
    BOLD = XLFont(name="Calibri", bold=True, size=10)
    REG = XLFont(name="Calibri", size=10)
    TITLE = XLFont(name="Calibri", bold=True, color="FFFFFF", size=14)

    thin = Side(border_style="thin", color="B4B4B4")
    BORDER_ALL = Border(left=thin, right=thin, top=thin, bottom=thin)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
    RIGHT = Alignment(horizontal="right", vertical="center")

    NUM_FMT = '#,##0.00;(#,##0.00);"-"'

    # ── 1. Cover / Index sheet ──────────────────────────────
    ws_idx = wb.create_sheet("Cover")
    ws_idx.merge_cells("A1:F1")
    ws_idx["A1"] = "GST Tax Liability & ITC Comparison — Consolidated"
    ws_idx["A1"].font = TITLE
    ws_idx["A1"].fill = HDR_FILL
    ws_idx["A1"].alignment = CENTER
    ws_idx.row_dimensions[1].height = 30

    ws_idx["A3"] = "Total states/GSTINs consolidated:"
    ws_idx["B3"] = len(all_data)
    ws_idx["A3"].font = BOLD
    ws_idx["B3"].font = BOLD

    headers = ["S.No.", "State Code", "State Name", "GSTIN", "Legal Name", "Source File"]
    for c, h in enumerate(headers, 1):
        cell = ws_idx.cell(row=5, column=c, value=h)
        cell.font = WHITE_B
        cell.fill = SUB_FILL
        cell.alignment = CENTER
        cell.border = BORDER_ALL
    ws_idx.row_dimensions[5].height = 22

    for i, data in enumerate(all_data, 1):
        m = data["meta"]
        for c, val in enumerate([i, m["state_code"], m["state_name"], m["gstin"],
                                 m["legal_name"], m["source_file"]], 1):
            cell = ws_idx.cell(row=5 + i, column=c, value=val)
            cell.font = REG
            cell.alignment = LEFT if c >= 3 else CENTER
            cell.border = BORDER_ALL

    for col, w in zip("ABCDEF", [7, 10, 22, 22, 28, 42]):
        ws_idx.column_dimensions[col].width = w
    ws_idx.freeze_panes = "A6"

    # ── 2. Multi-sheet output (one sheet per original section) ──
    if mode in ("multi", "both"):
        for sheet_key in _TC_SHEET_LAYOUT.keys():
            ws_name = _SHORT_NAMES[sheet_key][:31]
            ws = wb.create_sheet(ws_name)

            # Title
            ws["A1"] = f"{sheet_key}  —  All States Consolidated"
            ws["A1"].font = TITLE
            ws["A1"].fill = HDR_FILL
            ws["A1"].alignment = CENTER

            # Determine column count from first available data set
            ncols = 0
            for data in all_data:
                ss = data["sheets"].get(sheet_key)
                if ss and not ss.get("missing"):
                    ncols = max(ncols, len(ss["main_headers"]))
            if ncols == 0:
                ws["A3"] = f"(Sheet '{sheet_key}' was missing from all input files.)"
                ws["A3"].font = REG
                continue

            ws.merge_cells(start_row=1, start_column=1,
                           end_row=1, end_column=ncols + 4)
            ws.row_dimensions[1].height = 28

            # Multi-row header:  S.No | State Code | State Name | GSTIN | <orig cols>
            meta_cols = ["S.No.", "State Code", "State Name", "GSTIN"]

            # Row 3: main headers
            for c, lbl in enumerate(meta_cols, 1):
                cell = ws.cell(row=3, column=c, value=lbl)
                cell.font = WHITE_B; cell.fill = SUB_FILL
                cell.alignment = CENTER; cell.border = BORDER_ALL
            # use the FIRST data's main_headers as canonical
            canon_main = []
            canon_sub = []
            for data in all_data:
                ss = data["sheets"].get(sheet_key)
                if ss and not ss.get("missing"):
                    canon_main = ss["main_headers"]
                    canon_sub = ss["sub_headers"]
                    break
            for c in range(1, ncols + 1):
                mv = canon_main[c-1] if c-1 < len(canon_main) else None
                sv = canon_sub[c-1] if c-1 < len(canon_sub) else None
                cell = ws.cell(row=3, column=4 + c, value=mv)
                cell.font = WHITE_B; cell.fill = SUB_FILL
                cell.alignment = CENTER; cell.border = BORDER_ALL
                cell = ws.cell(row=4, column=4 + c, value=sv)
                cell.font = BOLD; cell.fill = GROUP_FILL
                cell.alignment = CENTER; cell.border = BORDER_ALL

            for c in range(1, 5):
                ws.cell(row=4, column=c).fill = SUB_FILL
                ws.cell(row=4, column=c).border = BORDER_ALL
            ws.row_dimensions[3].height = 38
            ws.row_dimensions[4].height = 22

            # Add a "Month" column on the left side of each row
            # Restructure: S.No | State Code | State Name | GSTIN | Month | <orig cols from col 2>
            # Actually the original col 1 (Tax Period) IS the month. Re-layout below.
            # Easier: put Month as 5th meta col, then values start from col 6.
            ws.cell(row=3, column=5, value="Tax Period").font = WHITE_B
            ws.cell(row=3, column=5).fill = SUB_FILL
            ws.cell(row=3, column=5).alignment = CENTER
            ws.cell(row=3, column=5).border = BORDER_ALL
            ws.cell(row=4, column=5).fill = SUB_FILL
            ws.cell(row=4, column=5).border = BORDER_ALL
            # Shift main headers right by 1: re-write cols 6..ncols+4
            for c in range(1, ncols):
                mv = canon_main[c] if c < len(canon_main) else None
                sv = canon_sub[c] if c < len(canon_sub) else None
                cell = ws.cell(row=3, column=5 + c, value=mv)
                cell.font = WHITE_B; cell.fill = SUB_FILL
                cell.alignment = CENTER; cell.border = BORDER_ALL
                cell = ws.cell(row=4, column=5 + c, value=sv)
                cell.font = BOLD; cell.fill = GROUP_FILL
                cell.alignment = CENTER; cell.border = BORDER_ALL
            # Clear out the last col we double-wrote
            cur_last = 4 + ncols
            if cur_last >= 5 + (ncols - 1):
                # The col index 4+ncols overwrites the bottom of original layout — clear:
                ws.cell(row=3, column=cur_last).value = None
                ws.cell(row=4, column=cur_last).value = None
                ws.cell(row=3, column=cur_last).fill = PatternFill()
                ws.cell(row=4, column=cur_last).fill = PatternFill()
                ws.cell(row=3, column=cur_last).border = Border()
                ws.cell(row=4, column=cur_last).border = Border()

            # Write data rows
            r = 5
            sn = 0
            for data in all_data:
                ss = data["sheets"].get(sheet_key)
                if not ss or ss.get("missing"):
                    continue
                m = data["meta"]
                for row_data in ss["rows"]:
                    sn += 1
                    is_total = row_data["is_total"]
                    fill = TOTAL_FILL if is_total else None
                    font = BOLD if is_total else REG

                    meta_vals = [sn, m["state_code"], m["state_name"],
                                 m["gstin"], row_data["month"]]
                    for c, v in enumerate(meta_vals, 1):
                        cell = ws.cell(row=r, column=c, value=v)
                        cell.font = font
                        cell.border = BORDER_ALL
                        cell.alignment = LEFT if c in (3, 4) else CENTER
                        if c == 1:
                            cell.number_format = "0"  # integer S.No.
                        if fill: cell.fill = fill

                    for c, v in enumerate(row_data["values"], 1):
                        # Skip the first value which corresponds to column 2 in orig
                        # (already we shifted: original col 2 → output col 6 = 5 + 1)
                        if c >= ncols: break
                        cell = ws.cell(row=r, column=5 + c, value=v)
                        cell.font = font
                        cell.border = BORDER_ALL
                        if isinstance(v, (int, float)):
                            cell.alignment = RIGHT
                            cell.number_format = NUM_FMT
                        else:
                            cell.alignment = LEFT
                        if fill: cell.fill = fill
                    r += 1

            # Column widths
            ws.column_dimensions["A"].width = 7
            ws.column_dimensions["B"].width = 7
            ws.column_dimensions["C"].width = 20
            ws.column_dimensions["D"].width = 18
            ws.column_dimensions["E"].width = 10
            for c in range(6, 5 + ncols):
                ws.column_dimensions[get_column_letter(c)].width = 16
            ws.freeze_panes = "F5"
            ws.auto_filter.ref = f"A3:{get_column_letter(4 + ncols)}{r - 1}"

    # ── 3. Single-sheet long format ─────────────────────────
    if mode in ("single", "both"):
        ws = wb.create_sheet("Long Format (All)")
        ws.merge_cells("A1:N1")
        ws["A1"] = "All Sections × All States × All Months — Long Format"
        ws["A1"].font = TITLE; ws["A1"].fill = HDR_FILL
        ws["A1"].alignment = CENTER
        ws.row_dimensions[1].height = 28

        hdrs = ["S.No.", "State Code", "State Name", "GSTIN", "FY",
                "Section", "Tax Period", "Is Total?",
                "Column (Main Header)", "Sub Header (Tax Type)",
                "Column #", "Value"]
        for c, h in enumerate(hdrs, 1):
            cell = ws.cell(row=3, column=c, value=h)
            cell.font = WHITE_B; cell.fill = SUB_FILL
            cell.alignment = CENTER; cell.border = BORDER_ALL
        ws.row_dimensions[3].height = 32

        r = 4
        sn = 0
        for data in all_data:
            m = data["meta"]
            for sheet_key, ss in data["sheets"].items():
                if not ss or ss.get("missing"): continue
                for row_data in ss["rows"]:
                    for c, v in enumerate(row_data["values"], 1):
                        # Skip empty/blank values to reduce row count
                        if v is None or (isinstance(v, str) and not v.strip()):
                            continue
                        if isinstance(v, str) and v.strip() in ("-", "—"):
                            continue
                        sn += 1
                        main_h = ss["main_headers"][c] if c < len(ss["main_headers"]) else None
                        sub_h  = ss["sub_headers"][c]  if c < len(ss["sub_headers"]) else None
                        row_vals = [sn, m["state_code"], m["state_name"],
                                    m["gstin"], m["fy"], _SHORT_NAMES.get(sheet_key, sheet_key),
                                    row_data["month"],
                                    "Yes" if row_data["is_total"] else "",
                                    main_h, sub_h, c + 1, v]
                        for cc, vv in enumerate(row_vals, 1):
                            cell = ws.cell(row=r, column=cc, value=vv)
                            cell.font = REG; cell.border = BORDER_ALL
                            if cc == 1:
                                cell.alignment = CENTER
                                cell.number_format = "0"
                            elif cc in (2, 5, 7, 8, 11):
                                cell.alignment = CENTER
                            elif cc == 12 and isinstance(vv, (int, float)):
                                cell.alignment = RIGHT
                                cell.number_format = NUM_FMT
                            else:
                                cell.alignment = LEFT
                        r += 1

        for col, w in zip("ABCDEFGHIJKL", [7, 7, 20, 18, 8, 24, 10, 8, 50, 18, 8, 16]):
            ws.column_dimensions[col].width = w
        ws.freeze_panes = "A4"
        if r > 4:
            ws.auto_filter.ref = f"A3:L{r-1}"

    # ── 4. State-Summary Pivot ──────────────────────────────
    if mode == "both":
        ws = wb.create_sheet("State Summary")
        ws.merge_cells("A1:K1")
        ws["A1"] = ("State-wise Annual Totals  —  Tax Liability & ITC Summary")
        ws["A1"].font = TITLE; ws["A1"].fill = HDR_FILL
        ws["A1"].alignment = CENTER
        ws.row_dimensions[1].height = 28

        hdrs = ["S.No.", "State Code", "State Name", "GSTIN", "Legal Name",
                "Liability (GSTR-1) Total", "Liability (GSTR-3B) Total",
                "Shortfall/Excess",
                "ITC Claimed Total", "ITC (GSTR-2B) Total",
                "ITC Shortfall/Excess"]
        for c, h in enumerate(hdrs, 1):
            cell = ws.cell(row=3, column=c, value=h)
            cell.font = WHITE_B; cell.fill = SUB_FILL
            cell.alignment = CENTER; cell.border = BORDER_ALL
        ws.row_dimensions[3].height = 36

        r = 4
        for i, data in enumerate(all_data, 1):
            m = data["meta"]
            # Pull from "Comparison Summary" total row if available
            cs = data["sheets"].get("Comparison Summary")
            tot = None
            if cs and not cs.get("missing"):
                for row_data in cs["rows"]:
                    if row_data["is_total"]:
                        tot = row_data["values"]
                        break
            # values map (0-indexed) for Comparison Summary:
            # 0=GSTR-1, 1=GSTR-3B, 2=Shortfall/Excess, 3=cum, 4=cum%,
            # 5=ITC 3B, 6=ITC 2B, 7=Shortfall, 8=cum, 9=cum%, 10..=reversed cols
            if tot:
                def gv(i):
                    if i < len(tot):
                        v = tot[i]
                        if isinstance(v, (int, float)): return v
                    return None
                gstr1_tot = gv(0); gstr3b_tot = gv(1); short_liab = gv(2)
                itc_3b_tot = gv(5); itc_2b_tot = gv(6); short_itc = gv(7)
            else:
                gstr1_tot = gstr3b_tot = short_liab = itc_3b_tot = itc_2b_tot = short_itc = None

            row_vals = [i, m["state_code"], m["state_name"], m["gstin"], m["legal_name"],
                        gstr1_tot, gstr3b_tot, short_liab,
                        itc_3b_tot, itc_2b_tot, short_itc]
            for c, v in enumerate(row_vals, 1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.font = REG; cell.border = BORDER_ALL
                if c == 1:
                    cell.alignment = CENTER
                    cell.number_format = "0"
                elif c == 2:
                    cell.alignment = CENTER
                elif c in (3, 4, 5):
                    cell.alignment = LEFT
                else:
                    cell.alignment = RIGHT
                    if isinstance(v, (int, float)):
                        cell.number_format = NUM_FMT
            r += 1

        # Grand total row
        ws.cell(row=r, column=1).fill = TOTAL_FILL
        for c in range(1, 12):
            cell = ws.cell(row=r, column=c)
            cell.fill = TOTAL_FILL; cell.border = BORDER_ALL; cell.font = BOLD
        ws.cell(row=r, column=5, value="GRAND TOTAL").alignment = LEFT
        from openpyxl.utils import get_column_letter as gcl
        for c in range(6, 12):
            col_letter = gcl(c)
            ws.cell(row=r, column=c,
                    value=f"=SUM({col_letter}4:{col_letter}{r-1})")
            ws.cell(row=r, column=c).number_format = NUM_FMT
            ws.cell(row=r, column=c).alignment = RIGHT

        for col, w in zip("ABCDEFGHIJK",
                          [7, 7, 20, 18, 28, 22, 22, 18, 22, 22, 18]):
            ws.column_dimensions[col].width = w
        ws.freeze_panes = "A4"
        if r > 4:
            ws.auto_filter.ref = f"A3:K{r}"

    wb.save(out_path)


# ════════════════════════════════════════════════════════════════
#  GSTR-2B CONSOLIDATOR ENGINE
# ════════════════════════════════════════════════════════════════
# Consolidates GSTR-2B Excel files from multiple states/months into ONE
# file, preserving original GSTN layout (per-sheet stacked), flipping
# Credit Note rows to negative for proper netting in downstream analysis.

# Sheets to skip entirely
_G2B_SKIP_SHEETS = {"Read me"}

# Summary sheets — different structure (vertical "card" layout)
_G2B_SUMMARY_SHEETS = [
    "ITC Available", "ITC not available", "ITC Reversal", "ITC Rejected",
]

# Transactional sheets — list captures the typical full set
_G2B_TXN_SHEETS = [
    "B2B", "B2BA", "B2B-CDNR", "B2B-CDNRA",
    "ECO", "ECOA",
    "ISD", "ISDA",
    "IMPG", "IMPGA", "IMPGSEZ", "IMPGSEZA",
    "B2B (ITC Reversal)", "B2BA (ITC Reversal)",
    "B2B-DNR", "B2B-DNRA",
    "B2B(Rejected)", "B2BA(Rejected)",
    "B2B-CDNR(Rejected)", "B2B-CDNRA(Rejected)",
    "ECO(Rejected)", "ECOA(Rejected)",
    "ISD(Rejected)", "ISDA(Rejected)",
]

# Sheets where we should look for "Credit Note" in the Note type col
# and flip taxable + tax columns to negative
_G2B_CDN_SHEETS = {
    "B2B-CDNR", "B2B-CDNRA", "B2B-CDNR(Rejected)", "B2B-CDNRA(Rejected)",
}


def _g2b_extract_meta(wb, filepath):
    """Pull GSTIN, FY, period, legal name from 'Read me' sheet."""
    meta = {
        "source_file": os.path.basename(filepath),
        "gstin": "", "fy": "", "period": "",
        "state_code": "", "state_name": "",
        "legal_name": "", "trade_name": "", "gen_date": "",
    }
    if "Read me" in wb.sheetnames:
        ws = wb["Read me"]
        # Standard layout: row 4-9, label in col A, value in col C
        for r in range(1, min(ws.max_row + 1, 15)):
            label = ws.cell(r, 1).value
            value = ws.cell(r, 3).value
            if not label or value is None: continue
            label_lower = str(label).strip().lower()
            value_str = str(value).strip()
            if "financial year" in label_lower:
                meta["fy"] = value_str
            elif "tax period" in label_lower:
                meta["period"] = value_str
            elif label_lower == "gstin":
                meta["gstin"] = value_str
            elif "legal name" in label_lower:
                meta["legal_name"] = value_str
            elif "trade name" in label_lower:
                meta["trade_name"] = value_str
            elif "date of generation" in label_lower:
                meta["gen_date"] = value_str

    # Fallback to filename parsing if Read me missing
    if not meta["gstin"]:
        base = os.path.basename(filepath)
        # GSTIN structure: 2 digits + 5 letters + 4 digits + 1 letter + 1
        # alphanum + 'Z' + 1 alphanum.  Cannot use \b boundaries because
        # underscore is a word char (so '\b' fails before/after '_GSTIN_').
        # Use explicit non-alphanum lookarounds instead.
        m = re.search(
            r"(?<![A-Z0-9])(\d{2}[A-Z]{5}\d{4}[A-Z][A-Z0-9]Z[A-Z0-9])"
            r"(?![A-Z0-9])", base)
        if not m:
            # Looser fallback: any 15-char sequence starting with 2 digits.
            # Validate it has 'Z' at position 13 (standard GSTIN check).
            for cand in re.findall(r"(?<![A-Z0-9])(\d{2}[A-Z0-9]{13})(?![A-Z0-9])",
                                    base):
                if len(cand) == 15 and cand[13] == 'Z':
                    m = re.match(r"(.+)", cand)  # wrap as match
                    break
        if m: meta["gstin"] = m.group(1)
    if not meta["period"]:
        # Match 6-digit MMYYYY pattern: either at start, or between underscores.
        # Examples this must handle:
        #   1780133457797_022026_04AAACH4041D1ZC_GSTR2B_14032026.xlsx  (has _MMYYYY_)
        #   012026_06AAACH4041D1Z8_GSTR2B_14022026_1.xlsx              (starts with MMYYYY_)
        # Use the GSTIN's neighbouring digits as anchor to disambiguate from
        # other 6-digit numbers (like the 8-digit generation date).
        base = os.path.basename(filepath)
        # Best signal: 6 digits immediately followed by _GSTIN (2 digits + 13 alphanum)
        m = re.search(r"(?:^|_)(\d{6})_(?=\d{2}[A-Z0-9]{13}_)", base)
        if not m:
            # Fallback: 6 digits between underscores (old behavior)
            m = re.search(r"_(\d{6})_", base)
        if not m:
            # Last resort: any 6-digit MM-YYYY at start of basename
            m = re.match(r"^(\d{6})_", base)
        if m:
            mm, yy = m.group(1)[:2], m.group(1)[2:]
            # Validate: MM should be 01-12, YYYY should be 20XX
            if mm.isdigit() and yy.isdigit() and 1 <= int(mm) <= 12 \
               and 2000 <= int(yy) <= 2099:
                months = {"01":"January","02":"February","03":"March","04":"April",
                          "05":"May","06":"June","07":"July","08":"August",
                          "09":"September","10":"October","11":"November","12":"December"}
                meta["period"] = months.get(mm, mm)
                # Also derive FY if not set
                if not meta.get("fy"):
                    mm_int = int(mm); yy_int = int(yy)
                    if mm_int <= 3:
                        meta["fy"] = f"{yy_int - 1}-{yy_int % 100:02d}"
                    else:
                        meta["fy"] = f"{yy_int}-{(yy_int + 1) % 100:02d}"

    # Derive state from GSTIN
    if meta["gstin"]:
        meta["state_code"] = meta["gstin"][:2]
        meta["state_name"] = _STATE_CODES.get(meta["state_code"], "Unknown")

    # Build short Month abbreviation like 'Feb-26'
    if meta["period"] and meta["fy"]:
        months_abbr = {"January":"Jan","February":"Feb","March":"Mar","April":"Apr",
                       "May":"May","June":"Jun","July":"Jul","August":"Aug",
                       "September":"Sep","October":"Oct","November":"Nov","December":"Dec"}
        abbr = months_abbr.get(meta["period"], meta["period"][:3])
        # FY format "2025-26" → for Jan/Feb/Mar use "26", else "25"
        if "-" in meta["fy"]:
            yy1, yy2 = meta["fy"].split("-")[0][-2:], meta["fy"].split("-")[1]
            yy = yy2 if meta["period"] in ("January","February","March") else yy1
            meta["month_abbr"] = f"{abbr}-{yy}"
        else:
            meta["month_abbr"] = abbr
    else:
        meta["month_abbr"] = meta["period"]

    return meta


def _g2b_find_data_start_row(ws):
    """Detect first data row in a GSTR-2B transaction sheet.

    Layout varies:
      • 2-row header: row 5 = main headers, row 6 = sub-headers, row 7+ = data
      • 3-row header (amendment sheets): row 5 = group banner
        ('Original Details' / 'Revised Details'), row 6 = main headers,
        row 7 = sub-headers, row 8+ = data

    Returns the row number where actual data begins (or where it would begin
    if the sheet has no rows, so headers are still correctly placed).
    """
    HEADER_WORDS_A = {
        "gstin", "trade", "icegate", "port code", "place of supply",
        "supply attract", "isd document", "amendments to", "taxable inward",
        "debit/credit", "documents reported", "import of goods", "itc reversed",
        "itc rejected", "amendment", "credit notes", "isd credits",
        "amendments isd", "debit notes", "original details", "revised details",
        "original invoice number", "input tax",
    }
    HEADER_SUBSTR_ANY_CELL = (
        "integrated tax", "central tax", "state/ut tax", "cess(",
        "input tax distribution", "tax amount", "taxable value",
        "invoice details", "credit note/debit note", "bill of entry",
        "amount of tax", "credit note", "debit note details", "tax amount",
        "amount declared by taxpayer", "type of amendment",
    )

    def is_header_row(r):
        """Return True if this row looks like a header row anywhere across cells."""
        all_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        # Count non-empty cells
        non_empty = [v for v in all_vals if v is not None and
                     (not isinstance(v, str) or v.strip())]
        if not non_empty:
            return False  # truly blank — not a header
        # If every non-empty cell is a string and matches header keywords, it's a header
        all_str = all(isinstance(v, str) for v in non_empty)
        if not all_str:
            return False  # numeric data present → data row
        # Check if any cell contains a sub-header substring
        for v in non_empty:
            vl = v.strip().lower()
            if any(kw in vl for kw in HEADER_SUBSTR_ANY_CELL):
                return True
        # First-column starts-with check
        a = ws.cell(r, 1).value
        if a:
            a_str = str(a).strip().lower()
            if any(a_str.startswith(w) for w in HEADER_WORDS_A):
                return True
        return False

    # Walk rows 5 through 12 to find first row that is NOT a header
    for r in range(5, min(ws.max_row + 1, 13)):
        a = ws.cell(r, 1).value
        # GSTIN pattern or date pattern in col A → definitely data
        if a is not None:
            a_str = str(a).strip()
            if re.match(r"^\d{2}[A-Z0-9]{13}$", a_str):
                return r
            if re.match(r"^\d{2}[-/]\d{2}[-/]\d{4}$", a_str):
                return r
        if is_header_row(r):
            continue
        # Empty col A but other cells have content — could be data row (e.g. IMPG amendments)
        any_non_empty = any(
            ws.cell(r, c).value is not None and
            (not isinstance(ws.cell(r, c).value, str) or ws.cell(r, c).value.strip())
            for c in range(1, ws.max_column + 1)
        )
        if any_non_empty:
            return r
    # Empty sheet: figure out where data WOULD start by detecting last header row
    last_header = 4
    for r in range(5, min(ws.max_row + 1, 13)):
        if is_header_row(r):
            last_header = r
        else:
            break
    return last_header + 1


def _g2b_collect_headers(ws, data_start_row):
    """Return list of merged header strings, one per column.
    Handles both layouts:
      • 2-row header: rows [data_start-2, data_start-1] = (main, sub)
      • 3-row header: rows [data_start-3, data_start-2, data_start-1] =
                       (group banner, main, sub)
    We combine main + sub as 'Main — Sub' for display.
    """
    ncols = ws.max_column

    def _resolve(ws, row, col):
        if row < 1: return ""
        v = ws.cell(row, col).value
        if v is not None and (not isinstance(v, str) or v.strip()):
            return str(v).strip()
        for mr in ws.merged_cells.ranges:
            if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
                top_val = ws.cell(mr.min_row, mr.min_col).value
                if top_val is not None:
                    return str(top_val).strip()
        return ""

    # Detect 3-row header by checking if row data_start-3 contains 'Original Details'
    # or 'Revised Details' as a group banner
    has_group_banner = False
    if data_start_row >= 4:
        for c in range(1, min(ncols + 1, 10)):
            v = _resolve(ws, data_start_row - 3, c).lower()
            if v in ("original details", "revised details"):
                has_group_banner = True
                break

    if has_group_banner:
        # 3-row: group_row = data_start-3, main_row = data_start-2, sub_row = data_start-1
        main_row = data_start_row - 2
        sub_row  = data_start_row - 1
    else:
        # 2-row: main_row = data_start-2, sub_row = data_start-1
        main_row = data_start_row - 2
        sub_row  = data_start_row - 1

    headers = []
    for c in range(1, ncols + 1):
        m = _resolve(ws, main_row, c) if main_row >= 1 else ""
        s = _resolve(ws, sub_row, c)  if sub_row  >= 1 else ""
        # If sub-row value equals main-row value (merged through), use just main
        if m and s and m != s:
            combined = f"{m} — {s}"
        elif s:
            combined = s
        elif m:
            combined = m
        else:
            combined = f"Col {c}"
        headers.append(combined)
    return headers


def _g2b_parse_summary_sheet(ws):
    """Parse one of the four summary sheets. Returns list of dicts with
    section, sno, heading, gstr3b_table, igst, cgst, sgst, cess, advisory, level."""
    out = []
    if ws.max_row < 6: return out
    # Find header row by scanning for 'S.no.' in col A
    hdr_row = None
    for r in range(1, min(ws.max_row + 1, 15)):
        v = ws.cell(r, 1).value
        if v and str(v).strip().lower().startswith("s.no"):
            hdr_row = r
            break
    if hdr_row is None: return out
    data_start = hdr_row + 1

    current_part = ""        # 'Part A' / 'Part B'
    current_part_desc = ""
    current_section = ""     # Roman 'I', 'II', etc.
    current_section_desc = ""
    current_3b_ref = ""

    for r in range(data_start, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        c = ws.cell(r, 3).value
        igst = ws.cell(r, 4).value
        cgst = ws.cell(r, 5).value
        sgst = ws.cell(r, 6).value
        cess = ws.cell(r, 7).value
        advisory = ws.cell(r, 8).value

        if a is None and b is None: continue
        a_str = str(a).strip() if a else ""
        b_str = str(b).strip() if b else ""

        # Skip rows that announce a textual heading ("Credit which may be...")
        if a_str and not b_str and not isinstance(igst, (int, float)):
            # heading row with no values
            continue

        # Detect "Part A" / "Part B"
        if a_str.lower().startswith("part "):
            current_part = a_str
            current_part_desc = b_str
            continue

        # Detect roman section heading (I/II/III/IV)
        if a_str in ("I", "II", "III", "IV", "V"):
            current_section = a_str
            current_section_desc = b_str
            current_3b_ref = str(c).strip() if c else ""
            # This is a section TOTAL row — record it as level='section'
            out.append({
                "level": "Section",
                "part": current_part, "part_desc": current_part_desc,
                "section": current_section, "heading": current_section_desc,
                "gstr3b_table": current_3b_ref,
                "igst": igst if isinstance(igst, (int, float)) else None,
                "cgst": cgst if isinstance(cgst, (int, float)) else None,
                "sgst": sgst if isinstance(sgst, (int, float)) else None,
                "cess": cess if isinstance(cess, (int, float)) else None,
                "advisory": str(advisory).strip() if advisory else "",
            })
            continue

        # Detect "Details" row — sub-row under current section
        if a_str.lower() == "details":
            out.append({
                "level": "Detail",
                "part": current_part, "part_desc": current_part_desc,
                "section": current_section, "heading": b_str,
                "gstr3b_table": str(c).strip() if c else current_3b_ref,
                "igst": igst if isinstance(igst, (int, float)) else None,
                "cgst": cgst if isinstance(cgst, (int, float)) else None,
                "sgst": sgst if isinstance(sgst, (int, float)) else None,
                "cess": cess if isinstance(cess, (int, float)) else None,
                "advisory": str(advisory).strip() if advisory else "",
            })
            continue

        # Otherwise: continuation of detail rows (no S.no. label, only heading in B)
        if b_str and isinstance(igst, (int, float)):
            out.append({
                "level": "Detail",
                "part": current_part, "part_desc": current_part_desc,
                "section": current_section, "heading": b_str,
                "gstr3b_table": str(c).strip() if c else current_3b_ref,
                "igst": igst, "cgst": cgst, "sgst": sgst, "cess": cess,
                "advisory": str(advisory).strip() if advisory else "",
            })
    return out


def parse_gstr2b_file(filepath):
    """Parse one GSTR-2B Excel. Returns dict with meta + per-sheet data."""
    import openpyxl
    import warnings
    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=False)
    meta = _g2b_extract_meta(wb, filepath)

    result = {"meta": meta, "summary": {}, "transactions": {}}

    # Parse summary sheets
    for sn in _G2B_SUMMARY_SHEETS:
        if sn in wb.sheetnames:
            result["summary"][sn] = _g2b_parse_summary_sheet(wb[sn])

    # Parse transaction sheets
    for sn in wb.sheetnames:
        if sn in _G2B_SKIP_SHEETS or sn in _G2B_SUMMARY_SHEETS:
            continue
        ws = wb[sn]
        # Find data start
        ds = _g2b_find_data_start_row(ws)
        # Pull headers
        headers = _g2b_collect_headers(ws, ds)
        # Read data rows
        rows = []
        for r in range(ds, ws.max_row + 1):
            row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            # Skip entirely blank rows
            if all(v is None or (isinstance(v, str) and not v.strip()) for v in row_vals):
                continue
            rows.append(row_vals)
        result["transactions"][sn] = {
            "headers": headers,
            "rows": rows,
            "data_start_row": ds,
            "n_cols": ws.max_column,
        }

    wb.close()
    return result


def _g2b_group_split_files(all_data):
    """Merge "split" GSTR-2B files (big states downloaded as multiple parts)
    into a single logical entry per (GSTIN, Period, GenerationDate).

    GSTN portal splits 2B for large GSTINs into:
       <period>_<GSTIN>_GSTR2B_<gendate>.xlsx        ← part 1 (summary + start)
       <period>_<GSTIN>_GSTR2B_<gendate>_1.xlsx      ← part 2 (continuation)
       <period>_<GSTIN>_GSTR2B_<gendate>_2.xlsx      ← part 3 (more), etc.

    Without merging, these appear as DUPLICATE state-month entries in the
    Cover/Console and transaction rows are split across two output blocks.

    Returns dict:
        merged_list  — final list of entries (one per state-month-gendate)
        split_groups — list of dicts describing each multi-part merge
        redownloads  — list of dicts where same (GSTIN, Period) had MULTIPLE
                       generation dates (likely re-downloads — kept latest)
        no_meta_files — list of source files whose (GSTIN, Period) couldn't
                        be determined, kept as standalone (review needed)
    """
    import re

    def extract_gendate(source_file):
        """Pull the 8-digit DDMMYYYY generation date out of filename."""
        if not source_file: return ""
        m = re.search(r"_GSTR2B_(\d{8})", source_file, re.IGNORECASE)
        return m.group(1) if m else ""

    # Bucket by (gstin, month_abbr) — month_abbr like "Apr-25" / "Apr-26"
    # DISAMBIGUATES THE YEAR. Using just "period" ('April') would collide
    # April-2025 with April-2026 and incorrectly treat one as a re-download
    # of the other.
    by_period = {}
    no_meta_files = []
    for d in all_data:
        m = d["meta"]
        gstin = m.get("gstin", "")
        # Prefer month_abbr (year-aware) as key, fall back to period+fy
        month_key = m.get("month_abbr", "")
        if not month_key:
            month_key = f"{m.get('period', '')}_{m.get('fy', '')}"
        if gstin and month_key.strip("_"):
            by_period.setdefault((gstin, month_key), []).append(d)
        else:
            no_meta_files.append(d)

    merged_list = []
    split_groups = []
    redownloads = []

    for (gstin, month_key), entries in by_period.items():
        # Within each (GSTIN, Period), distinguish by generation date.
        # If multiple gen dates exist, that's a re-download — keep LATEST only.
        by_gendate = {}
        for d in entries:
            gd = extract_gendate(d["meta"].get("source_file", ""))
            by_gendate.setdefault(gd, []).append(d)

        # If only ONE gen date (or only "" gen date), proceed normally
        # If multiple gen dates, pick the latest (DDMMYYYY → re-arrange to YYYYMMDD)
        if len(by_gendate) > 1 and any(by_gendate.keys()):
            def gd_sortkey(gd):
                if len(gd) != 8: return ""
                return gd[4:8] + gd[2:4] + gd[0:2]
            sorted_gds = sorted(by_gendate.keys(), key=gd_sortkey, reverse=True)
            latest_gd = sorted_gds[0]
            discarded = []
            for gd in sorted_gds[1:]:
                for d in by_gendate[gd]:
                    discarded.append(d["meta"].get("source_file", "?"))
            kept = by_gendate[latest_gd]
            redownloads.append({
                "state": kept[0]["meta"].get("state_name", "?"),
                "period": month_key,
                "gstin": gstin,
                "all_gen_dates": sorted_gds,
                "kept_gen_date": latest_gd,
                "discarded_files": discarded,
            })
            files_to_merge = kept
        else:
            # Single gen date — all files are parts of one split
            files_to_merge = entries

        if len(files_to_merge) == 1:
            merged_list.append(files_to_merge[0])
        else:
            merged = _g2b_merge_parts(files_to_merge)
            merged_list.append(merged)
            total_rows = sum(len(t.get("rows", []))
                             for t in merged["transactions"].values())
            split_groups.append({
                "state": merged["meta"].get("state_name", "?"),
                "period": month_key,
                "gstin": gstin,
                "parts": len(files_to_merge),
                "total_rows": total_rows,
                "sources": [d["meta"].get("source_file", "?")
                            for d in files_to_merge],
            })

    # Append files with no parseable meta (cannot be grouped)
    merged_list.extend(no_meta_files)

    return {
        "merged_list": merged_list,
        "split_groups": split_groups,
        "redownloads": redownloads,
        "no_meta_files": [d["meta"].get("source_file", "?")
                          for d in no_meta_files],
    }


def _g2b_merge_parts(parts):
    """Merge multiple parsed files (split parts of one state-month) into one
    combined entry. Strategy:
      • Meta: take from first part, fill blanks from later parts
      • Summary: per sheet, take first non-empty version (parts typically share
        the same summary block)
      • Transactions: per sheet, concatenate rows from ALL parts
    """
    # Sort parts so 'part1' (no _N suffix) comes first, then _1, _2, ...
    # GSTN filenames look like:  ..._GSTR2B_<DDMMYYYY>.xlsx (part 1, no suffix)
    #                       OR:  ..._GSTR2B_<DDMMYYYY>_<N>.xlsx (part N+1)
    # We must NOT confuse the 8-digit generation date for a part number.
    def sort_key(d):
        sf = (d["meta"].get("source_file") or "").lower()
        import re
        # Look for the specific GSTN suffix pattern: _GSTR2B_<8 digits>_<part>.xlsx
        m = re.search(r"_gstr2b_\d{8}_(\d+)\.xlsx?$", sf)
        if m:
            return (int(m.group(1)), sf)
        # No part suffix → this is the original/part-1
        return (0, sf)
    parts = sorted(parts, key=sort_key)

    # Meta — start from first, fill blanks from later
    merged_meta = dict(parts[0]["meta"])
    for p in parts[1:]:
        for k, v in p["meta"].items():
            if v and not merged_meta.get(k):
                merged_meta[k] = v
    # Source file list (helpful for audit)
    src_files = [p["meta"].get("source_file", "?") for p in parts]
    merged_meta["source_file"] = "  +  ".join(src_files)

    # Summary — per sheet, prefer first non-empty version
    merged_summary = {}
    for p in parts:
        for sn, summ in p.get("summary", {}).items():
            if sn not in merged_summary or not merged_summary[sn]:
                merged_summary[sn] = summ

    # Transactions — concatenate rows across parts (per sheet)
    merged_tx = {}
    for p in parts:
        for sn, tx in p.get("transactions", {}).items():
            if sn not in merged_tx:
                merged_tx[sn] = {
                    "headers": tx.get("headers", []),
                    "rows": list(tx.get("rows", [])),
                    "data_start_row": tx.get("data_start_row", 1),
                    "n_cols": tx.get("n_cols", 0),
                }
            else:
                merged_tx[sn]["rows"].extend(tx.get("rows", []))
                # Adopt headers from a later part if the earlier one had none
                if not merged_tx[sn]["headers"] and tx.get("headers"):
                    merged_tx[sn]["headers"] = tx["headers"]

    return {
        "meta": merged_meta,
        "summary": merged_summary,
        "transactions": merged_tx,
    }


def _g2b_flip_cdn_row(sheet_name, headers, row):
    """If this row is a Credit Note row, return a new row with taxable + tax columns
    flipped to negative. Otherwise return the row unchanged.

    Detection:
      - For sheets in _G2B_CDN_SHEETS: column whose sub-header == 'Note type'
        and value == 'Credit Note'  →  flip
    """
    if sheet_name not in _G2B_CDN_SHEETS:
        return row, False
    # Find Note type column
    note_type_col = None
    for i, h in enumerate(headers):
        h_lower = (h or "").lower()
        if "note type" in h_lower:
            note_type_col = i
            break
    if note_type_col is None or note_type_col >= len(row):
        return row, False
    nt = row[note_type_col]
    if not (isinstance(nt, str) and nt.strip().lower() == "credit note"):
        return row, False

    # It's a credit note — flip these column types to negative
    new_row = list(row)
    FLIP_KEYWORDS = (
        "taxable value", "integrated tax", "central tax", "state/ut tax", "cess",
        "note value",
    )
    for i, h in enumerate(headers):
        if i >= len(new_row): break
        h_lower = (h or "").lower()
        if any(kw in h_lower for kw in FLIP_KEYWORDS):
            v = new_row[i]
            if isinstance(v, (int, float)) and v > 0:
                new_row[i] = -v
    return new_row, True


def _g2b_classify_sheet(sheet_name):
    """Classify a sheet into (sheet_type, category) for the Console.

    sheet_type:  B2B / CDNR / DNR / ECO / ISD / IMPG / IMPGSEZ
    category:    Original / Amendment / Rejected / Reversal / Rejected-Amendment / Reversal-Amendment
    """
    name = sheet_name
    # Determine category from suffix/qualifier
    if "(Rejected)" in name:
        cat = "Rejected-Amendment" if name.endswith("A(Rejected)") else "Rejected"
    elif "(ITC Reversal)" in name:
        cat = "Reversal-Amendment" if "BA" in name.split("(")[0] else "Reversal"
    elif name.endswith("A") or name.endswith("RA") or name.endswith("ZA"):
        cat = "Amendment"
    else:
        cat = "Original"

    # Determine sheet type
    base = name.replace("(Rejected)", "").replace("(ITC Reversal)", "").strip()
    if base.startswith("B2B-CDNR"):
        stype = "CDNR"
    elif base.startswith("B2B-DNR"):
        stype = "DNR"
    elif base.startswith("ECO"):
        stype = "ECO"
    elif base.startswith("ISD"):
        stype = "ISD"
    elif base.startswith("IMPGSEZ"):
        stype = "IMPGSEZ"
    elif base.startswith("IMPG"):
        stype = "IMPG"
    elif base.startswith("B2B"):
        stype = "B2B"
    else:
        stype = base
    return stype, cat


def _g2b_match_col(headers, *keyword_groups):
    """Find first column index whose header matches any keyword group.
    Each group is a tuple of (must_have_all_lowercase_keywords).
    Returns column index (0-based) or None."""
    for i, h in enumerate(headers):
        hl = (h or "").lower()
        if not hl: continue
        for grp in keyword_groups:
            if all(kw in hl for kw in grp):
                return i
        # Single-keyword shortcut
    return None


def _g2b_map_to_console(sheet_name, headers, row, is_credit_note):
    """Map a single transactional row from any GSTR-2B sheet to the unified
    Console schema. Returns dict with all Console columns."""
    stype, cat = _g2b_classify_sheet(sheet_name)

    def G(*groups, default=None):
        """Find value by header keyword groups."""
        idx = None
        # Try each group in order
        for grp in groups:
            for i, h in enumerate(headers):
                hl = (h or "").lower()
                if not hl: continue
                if all(kw in hl for kw in grp):
                    idx = i
                    break
            if idx is not None: break
        if idx is None or idx >= len(row): return default
        v = row[idx]
        if v is None: return default
        if isinstance(v, str) and not v.strip(): return default
        return v

    # Counterparty GSTIN — varies by sheet
    if stype == "ISD":
        gstin_cp = G(("gstin", "isd",))
    elif stype == "ECO":
        gstin_cp = G(("gstin", "eco",))
    elif stype in ("IMPG",):
        gstin_cp = None  # imports from overseas have no supplier GSTIN
    elif stype == "IMPGSEZ":
        gstin_cp = G(("gstin", "supplier"))
    else:
        gstin_cp = G(("gstin", "supplier"))

    trade_name = G(("trade",), ("legal",))

    # Document number, date, value, type — varies by sheet
    if stype == "CDNR":
        doc_num   = G(("note number",), ("note", "number"))
        doc_type  = G(("note type",))
        doc_supply = G(("note supply",))
        doc_date  = G(("note date",))
        doc_value = G(("note value",))
    elif stype == "DNR":
        doc_num   = G(("debit note", "note number"), ("note number",))
        doc_type  = G(("note type",))
        doc_supply = G(("note supply",))
        doc_date  = G(("note date",))
        doc_value = G(("note value",))
    elif stype == "ECO":
        doc_num   = G(("document number",))
        doc_type  = G(("document type",))
        doc_supply = None
        doc_date  = G(("document date",))
        doc_value = G(("document value",))
    elif stype == "ISD":
        doc_num   = G(("isd document number",), ("document number",))
        doc_type  = G(("isd document type",), ("document type",))
        doc_supply = G(("original invoice number",))
        doc_date  = G(("isd document date",), ("document date",))
        doc_value = None
    elif stype in ("IMPG", "IMPGSEZ"):
        doc_num   = G(("bill of entry", "number"), ("number",))
        doc_type  = G(("port code",))   # use port code as identifier
        doc_supply = G(("icegate", "date"))
        doc_date  = G(("bill of entry", "date"), ("date",))
        doc_value = None
    else:  # B2B and variants
        doc_num   = G(("invoice number",))
        doc_type  = G(("invoice type",))
        doc_supply = None
        doc_date  = G(("invoice date",))
        doc_value = G(("invoice value",))

    place = G(("place of supply",))
    rcm = G(("reverse charge",))

    # Tax + taxable
    taxable = G(("taxable value",), ("taxable",))
    igst    = G(("integrated",))
    cgst    = G(("central tax",))
    sgst    = G(("state/ut",), ("state",), ("ut tax",))
    cess    = G(("cess",))

    # Other fields
    itc_avail = G(("itc availability",), ("eligibility",))
    reason    = G(("reason",))
    source    = G(("source",))
    # Tax Rate column — only B2B-family sheets actually have this column.
    # The loose ('rate',) fallback was matching headers like 'Integrated Tax
    # (Rate)' on IMPG/IMPGSEZ sheets and putting IGST amounts in the Rate
    # column. Restrict matches to actual "Applicable % of Tax Rate" / "Rate(%)"
    # patterns, and force None for IMPG/IMPGSEZ where no rate column exists.
    if stype in ("IMPG", "IMPGSEZ"):
        rate = None
    else:
        rate = G(("applicable", "tax rate"),
                 ("rate(%)",),
                 ("applicable", "%"),
                 ("rate", "%"))
    irn       = G(("irn",))
    period    = G(("gstr-1",), ("gstr-6", "period"), ("period",))
    filing_dt = G(("filing date",))

    # If we marked the row as Credit Note (already flipped), the numbers are
    # already negative — no further work.
    return {
        "sheet_type":   stype,
        "category":     cat,
        "is_cn":        "Yes" if is_credit_note else "",
        "gstin_cp":     gstin_cp,
        "trade_name":   trade_name,
        "doc_num":      doc_num,
        "doc_type":     doc_type,
        "doc_supply":   doc_supply,
        "doc_date":     doc_date,
        "doc_value":    doc_value,
        "place":        place,
        "rcm":          rcm,
        "taxable":      taxable,
        "igst":         igst,
        "cgst":         cgst,
        "sgst":         sgst,
        "cess":         cess,
        "itc_avail":    itc_avail,
        "reason":       reason,
        "source":       source,
        "rate":         rate,
        "irn":          irn,
        "period":       period,
        "filing_dt":    filing_dt,
    }


def write_consolidated_gstr2b(all_data, out_path):
    """Build a consolidated workbook from all parsed GSTR-2B files."""
    import openpyxl
    from openpyxl.styles import Font as XLFont, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HDR_FILL = PatternFill("solid", start_color="1F4E78")
    SUB_FILL = PatternFill("solid", start_color="2E75B6")
    META_FILL = PatternFill("solid", start_color="DDEBF7")
    TOTAL_FILL = PatternFill("solid", start_color="FFE699")
    CDN_FILL = PatternFill("solid", start_color="FCE4D6")  # light orange — flipped rows
    SECTION_FILL = PatternFill("solid", start_color="FFF2CC")

    WHITE_B = XLFont(name="Calibri", bold=True, color="FFFFFF", size=10)
    BOLD = XLFont(name="Calibri", bold=True, size=10)
    REG = XLFont(name="Calibri", size=10)
    NEG_FONT = XLFont(name="Calibri", size=10, color="C00000")
    TITLE = XLFont(name="Calibri", bold=True, color="FFFFFF", size=14)

    thin = Side(border_style="thin", color="B4B4B4")
    BORDER_ALL = Border(left=thin, right=thin, top=thin, bottom=thin)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
    RIGHT = Alignment(horizontal="right", vertical="center")

    NUM_FMT = '#,##0.00;[Red](#,##0.00);"-"'

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ─── Cover sheet ──────────────────────────────────────────
    ws = wb.create_sheet("Cover")
    ws.merge_cells("A1:H1")
    ws["A1"] = "GSTR-2B — Multi-State Consolidated"
    ws["A1"].font = TITLE; ws["A1"].fill = HDR_FILL
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 30
    ws["A3"] = "Total state-month returns consolidated:"
    ws["B3"] = len(all_data)
    ws["A3"].font = BOLD; ws["B3"].font = BOLD
    ws["A4"] = "Note:"
    ws["A4"].font = BOLD
    ws["B4"] = ("Credit Note rows (in CDNR/CDNRA sheets) are flipped to NEGATIVE for "
                "correct netting. Original GSTN format is preserved; State/GSTIN/Month "
                "columns added on the left for filtering.")
    ws["B4"].alignment = LEFT
    ws.row_dimensions[4].height = 50
    ws.merge_cells("B4:H4")

    hdrs = ["S.No.", "Month", "FY", "State Code", "State Name", "GSTIN",
            "Legal Name", "Source File"]
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(row=6, column=c, value=h)
        cell.font = WHITE_B; cell.fill = SUB_FILL
        cell.alignment = CENTER; cell.border = BORDER_ALL
    ws.row_dimensions[6].height = 22

    for i, data in enumerate(all_data, 1):
        m = data["meta"]
        vals = [i, m["month_abbr"], m["fy"], m["state_code"], m["state_name"],
                m["gstin"], m["legal_name"], m["source_file"]]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=6 + i, column=c, value=v)
            cell.font = REG
            cell.alignment = LEFT if c >= 5 else CENTER
            cell.border = BORDER_ALL
            if c == 1: cell.number_format = "0"

    for col, w in zip("ABCDEFGH", [7, 9, 8, 7, 20, 20, 28, 42]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A7"

    # ─── Coverage Matrix sheet ────────────────────────────────
    # State × Month grid showing which (GSTIN, Month) combinations were
    # consolidated. Useful for verifying complete coverage (e.g. "did
    # all states get processed for Apr-26?").
    ws = wb.create_sheet("Coverage")

    def _month_sort_key(month_abbr):
        """Convert 'Apr-25' → '202504' for chronological sorting."""
        if not month_abbr or "-" not in month_abbr: return "9999-99"
        abbr, yy = month_abbr.split("-", 1)
        month_nums = {"Jan":"01","Feb":"02","Mar":"03","Apr":"04",
                      "May":"05","Jun":"06","Jul":"07","Aug":"08",
                      "Sep":"09","Oct":"10","Nov":"11","Dec":"12"}
        full_year = ("20" + yy.strip()) if len(yy.strip()) == 2 else yy.strip()
        return full_year + month_nums.get(abbr, "00")

    # Build GSTIN → {month → txn_count} map
    gstin_to_months = {}
    gstin_info = {}  # gstin → (state_code, state_name, legal_name)
    months_set = set()
    for data in all_data:
        m = data["meta"]
        gstin = m.get("gstin", "")
        month = m.get("month_abbr", "")
        if not gstin or not month: continue
        if gstin not in gstin_info:
            gstin_info[gstin] = (m.get("state_code", ""),
                                  m.get("state_name", ""),
                                  m.get("legal_name", ""))
        n_txn = sum(len(t.get("rows", []))
                    for t in data.get("transactions", {}).values())
        gstin_to_months.setdefault(gstin, {})[month] = n_txn
        months_set.add(month)

    # Sort months chronologically, GSTINs by state code
    sorted_months = sorted(months_set, key=_month_sort_key)
    sorted_gstins = sorted(gstin_to_months.keys(),
                            key=lambda g: (gstin_info[g][0], g))

    # Headers (row 1: title; row 3: column headers)
    ncols_cov = 4 + len(sorted_months) + 1  # 4 meta + months + Total
    ws.merge_cells(start_row=1, start_column=1, end_row=1,
                    end_column=max(ncols_cov, 4))
    ws["A1"] = ("Coverage Matrix — State × Month  "
                "(cell value = txn rows consolidated)")
    ws["A1"].font = TITLE; ws["A1"].fill = HDR_FILL
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 26

    hdrs = ["State Code", "State Name", "GSTIN", "Legal Name"] \
           + sorted_months + ["Total"]
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = WHITE_B; cell.fill = SUB_FILL
        cell.alignment = CENTER; cell.border = BORDER_ALL
    ws.row_dimensions[3].height = 30

    EMPTY_FILL = PatternFill("solid", start_color="FFD7D7")  # light red
    PRESENT_FILL = PatternFill("solid", start_color="E2EFDA") # light green

    r = 4
    month_totals = {m: 0 for m in sorted_months}
    for gstin in sorted_gstins:
        state_code, state_name, legal_name = gstin_info[gstin]
        ws.cell(row=r, column=1, value=state_code).alignment = CENTER
        ws.cell(row=r, column=2, value=state_name).alignment = LEFT
        ws.cell(row=r, column=3, value=gstin).alignment = CENTER
        ws.cell(row=r, column=4, value=legal_name).alignment = LEFT
        for c in (1, 2, 3, 4):
            ws.cell(row=r, column=c).font = REG
            ws.cell(row=r, column=c).border = BORDER_ALL

        row_total = 0
        for j, mo in enumerate(sorted_months, 5):
            n_txn = gstin_to_months[gstin].get(mo)
            cell = ws.cell(row=r, column=j, value=n_txn if n_txn else None)
            cell.border = BORDER_ALL
            cell.alignment = CENTER
            cell.font = REG
            if n_txn is None:
                cell.value = "—"
                cell.fill = EMPTY_FILL
                cell.font = XLFont(name="Calibri", color="C00000", size=10)
            else:
                cell.fill = PRESENT_FILL
                cell.number_format = "0"
                row_total += n_txn
                month_totals[mo] += n_txn
        # Row total
        cell = ws.cell(row=r, column=4 + len(sorted_months) + 1, value=row_total)
        cell.font = BOLD; cell.alignment = CENTER
        cell.border = BORDER_ALL; cell.number_format = "#,##0"
        cell.fill = TOTAL_FILL
        r += 1

    # Totals row
    ws.cell(row=r, column=1, value="TOTAL").font = BOLD
    ws.cell(row=r, column=1).fill = TOTAL_FILL
    ws.cell(row=r, column=1).alignment = CENTER
    ws.cell(row=r, column=1).border = BORDER_ALL
    for c in (2, 3, 4):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
        ws.cell(row=r, column=c).border = BORDER_ALL
    grand = 0
    for j, mo in enumerate(sorted_months, 5):
        cell = ws.cell(row=r, column=j, value=month_totals[mo])
        cell.font = BOLD; cell.alignment = CENTER
        cell.fill = TOTAL_FILL; cell.border = BORDER_ALL
        cell.number_format = "#,##0"
        grand += month_totals[mo]
    cell = ws.cell(row=r, column=4 + len(sorted_months) + 1, value=grand)
    cell.font = BOLD; cell.alignment = CENTER
    cell.fill = TOTAL_FILL; cell.border = BORDER_ALL
    cell.number_format = "#,##0"

    # Column widths
    ws.column_dimensions["A"].width = 7    # State Code
    ws.column_dimensions["B"].width = 18   # State Name
    ws.column_dimensions["C"].width = 18   # GSTIN
    ws.column_dimensions["D"].width = 26   # Legal Name
    for j in range(5, 5 + len(sorted_months)):
        ws.column_dimensions[get_column_letter(j)].width = 9
    ws.column_dimensions[get_column_letter(4 + len(sorted_months) + 1)].width = 11
    ws.freeze_panes = "E4"

    # ─── ITC Summary consolidated sheet ────────────────────────
    ws = wb.create_sheet("ITC Summary (All)")
    ws.merge_cells("A1:N1")
    ws["A1"] = "ITC Summary — All Four Cards (Available / Not Available / Reversal / Rejected) — All States"
    ws["A1"].font = TITLE; ws["A1"].fill = HDR_FILL
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 28

    hdrs = ["S.No.", "Month", "State Code", "State Name", "GSTIN",
            "Summary Type", "Level", "Part", "Section", "Heading",
            "GSTR-3B Table", "IGST (₹)", "CGST (₹)", "SGST (₹)", "Cess (₹)"]
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = WHITE_B; cell.fill = SUB_FILL
        cell.alignment = CENTER; cell.border = BORDER_ALL
    ws.row_dimensions[3].height = 36

    r = 4
    sn = 0
    for data in all_data:
        m = data["meta"]
        for summary_name in _G2B_SUMMARY_SHEETS:
            rows = data["summary"].get(summary_name, [])
            for row in rows:
                sn += 1
                level = row.get("level", "")
                vals = [
                    sn, m["month_abbr"], m["state_code"], m["state_name"], m["gstin"],
                    summary_name, level,
                    row.get("part", ""), row.get("section", ""),
                    row.get("heading", ""), row.get("gstr3b_table", ""),
                    row.get("igst"), row.get("cgst"), row.get("sgst"), row.get("cess"),
                ]
                is_section = (level == "Section")
                fill = SECTION_FILL if is_section else None
                font = BOLD if is_section else REG
                for c, v in enumerate(vals, 1):
                    cell = ws.cell(row=r, column=c, value=v)
                    cell.font = font; cell.border = BORDER_ALL
                    if fill: cell.fill = fill
                    if c == 1:
                        cell.alignment = CENTER; cell.number_format = "0"
                    elif c in (2, 3, 7, 8, 9, 11):
                        cell.alignment = CENTER
                    elif c in (4, 5, 6, 10):
                        cell.alignment = LEFT
                    elif c >= 12 and isinstance(v, (int, float)):
                        cell.alignment = RIGHT
                        cell.number_format = NUM_FMT
                r += 1

    for col, w in zip("ABCDEFGHIJKLMNO",
                      [6, 8, 6, 18, 18, 18, 8, 8, 8, 50, 12, 14, 14, 14, 12]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "F4"
    if r > 4:
        ws.auto_filter.ref = f"A3:O{r-1}"

    # ─── Console sheet — unified invoice-level across ALL txn sheets ──
    # Every row from every transaction sheet mapped to a common B2B-CDNR-like
    # schema (28 columns total: 5 meta + 23 unified data cols).
    ws = wb.create_sheet("Console")
    console_hdrs = [
        # Meta columns
        "S.No.", "Month", "State Code", "State Name", "GSTIN of Filer",
        # Sheet origin / classification
        "Sheet Type", "Category", "Is Credit Note?",
        # Counterparty
        "GSTIN of Counterparty", "Trade/Legal Name",
        # Document
        "Doc Number", "Doc Type", "Doc Supply Type", "Doc Date", "Doc Value (₹)",
        # Tax dimensions
        "Place of Supply", "Reverse Charge",
        # Values
        "Taxable Value (₹)", "IGST (₹)", "CGST (₹)", "SGST/UTGST (₹)", "Cess (₹)",
        # ITC / source / other
        "ITC Availability", "Reason", "Source", "Tax Rate", "IRN",
        "Filing Period", "Filing Date",
    ]
    ncols_console = len(console_hdrs)

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols_console)
    ws["A1"] = ("Console — Unified Invoice-level View  (every transaction row, "
                "every state, every sheet)")
    ws["A1"].font = TITLE; ws["A1"].fill = HDR_FILL
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 28

    # Sub-title with note
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols_console)
    ws["A2"] = ("Credit Note rows are negative & highlighted orange · "
                "Filter on Sheet Type / Category / State / Counterparty")
    ws["A2"].font = REG; ws["A2"].alignment = CENTER
    ws["A2"].fill = META_FILL

    # Header row at row 3
    for c, h in enumerate(console_hdrs, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = WHITE_B; cell.fill = SUB_FILL
        cell.alignment = CENTER; cell.border = BORDER_ALL
    ws.row_dimensions[3].height = 36

    # Discover sheets (same logic as below)
    all_txn_sheets = []
    seen = set()
    for data in all_data:
        for sn_name in data["transactions"]:
            if sn_name not in seen:
                seen.add(sn_name)
                all_txn_sheets.append(sn_name)
    ordered_for_console = [s for s in _G2B_TXN_SHEETS if s in seen]
    ordered_for_console += [s for s in all_txn_sheets if s not in ordered_for_console]

    # Walk every transaction row from every file and every sheet
    r = 4
    sn = 0
    console_flip_count = 0
    for data in all_data:
        m = data["meta"]
        for sheet_name in ordered_for_console:
            tx = data["transactions"].get(sheet_name)
            if not tx: continue
            for row in tx["rows"]:
                if all(v is None or (isinstance(v, str) and not v.strip())
                       for v in row):
                    continue
                flipped_row, was_flipped = _g2b_flip_cdn_row(
                    sheet_name, tx["headers"], row)
                if was_flipped:
                    console_flip_count += 1

                mapped = _g2b_map_to_console(
                    sheet_name, tx["headers"], flipped_row, was_flipped)

                sn += 1
                vals = [
                    sn,
                    m["month_abbr"], m["state_code"], m["state_name"], m["gstin"],
                    mapped["sheet_type"], mapped["category"], mapped["is_cn"],
                    mapped["gstin_cp"], mapped["trade_name"],
                    mapped["doc_num"], mapped["doc_type"], mapped["doc_supply"],
                    mapped["doc_date"], mapped["doc_value"],
                    mapped["place"], mapped["rcm"],
                    mapped["taxable"], mapped["igst"], mapped["cgst"],
                    mapped["sgst"], mapped["cess"],
                    mapped["itc_avail"], mapped["reason"],
                    mapped["source"], mapped["rate"], mapped["irn"],
                    mapped["period"], mapped["filing_dt"],
                ]
                for c, v in enumerate(vals, 1):
                    cell = ws.cell(row=r, column=c, value=v)
                    cell.border = BORDER_ALL
                    if was_flipped:
                        cell.fill = CDN_FILL
                    if c == 1:
                        cell.alignment = CENTER
                        cell.number_format = "0"
                        cell.font = BOLD if was_flipped else REG
                    elif c in (2, 3, 6, 7, 8):
                        cell.alignment = CENTER
                        cell.font = BOLD if (was_flipped and c == 8) else REG
                    elif isinstance(v, (int, float)) and c >= 15:
                        cell.alignment = RIGHT
                        cell.number_format = NUM_FMT
                        cell.font = NEG_FONT if (isinstance(v, (int, float)) and v < 0) else REG
                    else:
                        cell.alignment = LEFT
                        cell.font = REG
                r += 1

    # Column widths
    widths = [6, 8, 6, 18, 16,         # meta
              8, 12, 8,                # sheet/category/iscn
              18, 24,                  # gstin_cp, trade_name
              20, 12, 12, 11, 14,      # doc fields
              18, 8,                   # place, rcm
              14, 14, 14, 14, 12,      # taxable + tax
              12, 18, 11, 9, 28,       # itc/reason/source/rate/irn
              10, 12]                  # period/filing_dt
    for i, w in enumerate(widths[:ncols_console], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "I4"  # freeze meta + classification cols
    if r > 4:
        ws.auto_filter.ref = f"A3:{get_column_letter(ncols_console)}{r-1}"

    # ─── Per-transaction-sheet consolidated outputs ────────────
    # Discover all unique txn sheet names across all input files
    all_txn_sheets = []
    seen = set()
    for data in all_data:
        for sn_name in data["transactions"]:
            if sn_name not in seen:
                seen.add(sn_name)
                all_txn_sheets.append(sn_name)
    # Maintain canonical order
    ordered = [s for s in _G2B_TXN_SHEETS if s in seen]
    ordered += [s for s in all_txn_sheets if s not in ordered]

    cdn_flip_count = 0
    for sheet_name in ordered:
        # Excel sheet name length limit = 31
        out_name = sheet_name[:31]
        ws = wb.create_sheet(out_name)
        # Get canonical headers from the FIRST file that has this sheet
        canon_headers = None
        for data in all_data:
            tx = data["transactions"].get(sheet_name)
            if tx and tx["headers"]:
                canon_headers = tx["headers"]
                break
        if canon_headers is None: canon_headers = []
        n_orig_cols = len(canon_headers)

        # Title
        ncols_total = 6 + n_orig_cols  # 6 meta columns + data columns
        ws.merge_cells(start_row=1, start_column=1, end_row=1,
                       end_column=max(ncols_total, 6))
        ws["A1"] = f"{sheet_name} — All States Consolidated"
        ws["A1"].font = TITLE; ws["A1"].fill = HDR_FILL
        ws["A1"].alignment = CENTER
        ws.row_dimensions[1].height = 26

        # Header row at row 3
        meta_hdrs = ["S.No.", "Month", "State Code", "State Name", "GSTIN", "Is Credit Note?"]
        all_hdrs = meta_hdrs + canon_headers
        for c, h in enumerate(all_hdrs, 1):
            cell = ws.cell(row=3, column=c, value=h)
            cell.font = WHITE_B; cell.fill = SUB_FILL
            cell.alignment = CENTER; cell.border = BORDER_ALL
        ws.row_dimensions[3].height = 42

        # Write data rows
        r = 4
        sn = 0
        for data in all_data:
            m = data["meta"]
            tx = data["transactions"].get(sheet_name)
            if not tx: continue
            # If this file's headers differ in column count, align by index
            for row in tx["rows"]:
                # Skip empty rows (already filtered, but double-check)
                if all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
                    continue

                # Apply credit note flipping if applicable
                flipped_row, was_flipped = _g2b_flip_cdn_row(
                    sheet_name, tx["headers"], row)
                if was_flipped:
                    cdn_flip_count += 1

                sn += 1
                meta_vals = [sn, m["month_abbr"], m["state_code"],
                             m["state_name"], m["gstin"],
                             "Yes" if was_flipped else ""]
                for c, v in enumerate(meta_vals, 1):
                    cell = ws.cell(row=r, column=c, value=v)
                    cell.font = REG; cell.border = BORDER_ALL
                    cell.fill = META_FILL
                    if c == 1:
                        cell.alignment = CENTER; cell.number_format = "0"
                    elif c in (2, 3, 6):
                        cell.alignment = CENTER
                    else:
                        cell.alignment = LEFT
                    if was_flipped and c == 6:
                        cell.font = BOLD; cell.fill = CDN_FILL

                # Write data values
                for c, v in enumerate(flipped_row, 7):
                    if c - 7 >= n_orig_cols: break
                    cell = ws.cell(row=r, column=c, value=v)
                    cell.border = BORDER_ALL
                    if was_flipped:
                        cell.fill = CDN_FILL
                    if isinstance(v, (int, float)):
                        cell.alignment = RIGHT
                        cell.number_format = NUM_FMT
                        if v < 0:
                            cell.font = NEG_FONT
                        else:
                            cell.font = REG
                    else:
                        cell.alignment = LEFT
                        cell.font = REG
                r += 1

        # Column widths — meta cols + data cols
        widths_meta = [6, 8, 6, 18, 18, 12]
        for i, w in enumerate(widths_meta, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        # Data column widths — default 18, except common known columns
        for i in range(7, 7 + n_orig_cols):
            ws.column_dimensions[get_column_letter(i)].width = 18

        ws.freeze_panes = "G4"
        if r > 4 and n_orig_cols > 0:
            ws.auto_filter.ref = f"A3:{get_column_letter(6 + n_orig_cols)}{r-1}"

    wb.save(out_path)
    return cdn_flip_count


# ════════════════════════════════════════════════════════════════
#  GSTR-9 / 9C CONSOLIDATOR ENGINE
# ════════════════════════════════════════════════════════════════
# Extracts data from GSTR-9 (Annual Return) and GSTR-9C (Reconciliation)
# PDFs. Builds a unified "Console" sheet across all input files + per-form
# detail sheets. Designed for multi-state / multi-FY consolidation.

# Per-table value-column profile.
# For each (form_type, table_no): if the standard row has fewer values than
# 5, map them to the 5 standard tax columns [Taxable, CGST, SGST, IGST, Cess].
# Default for tables with 5 values: positions 0..4 = Taxable, CGST, SGST, IGST, Cess.

_G9_NUM_RE = re.compile(r'-?[\d,]+\.\d{1,2}|-?[\d,]+')


def _g9_parse_number(s):
    """Convert '25,21,705.81' or '0.00' or '-' to float, or None."""
    if not s or s in ('-', '—'): return None
    s = s.strip().replace(',', '').replace('₹', '').strip()
    if not s: return None
    try: return float(s)
    except: return None


def _g9_extract_pdf_text(filepath):
    """Extract full text from PDF as single string."""
    import pdfplumber
    pages = []
    with pdfplumber.open(filepath) as pdf:
        for page in pdf.pages:
            pages.append(page.extract_text() or "")
    return "\n".join(pages)


def _g9_clean_lines(text):
    """Strip 'FINAL' watermark single-letter lines and blanks."""
    out = []
    for line in text.split('\n'):
        line = line.strip()
        if not line: continue
        # Watermark "FINAL" letters
        if len(line) == 1 and line in 'FINAL': continue
        if line == 'FINAL': continue
        out.append(line)
    return out


def _g9_merge_wrapped(lines):
    """Merge continuation lines into their parent rows.
    A new row starts when the line begins with one of:
      - 'Pt.' / 'PART'
      - Sr.No / column header marker
      - <digit>+space+<capital letter> (table title)
      - <letter><opt-digit>+space (row label)
    Otherwise the line is appended to the previous one.
    """
    merged = []
    for line in lines:
        if re.match(r'^(Pt\.|PART|Sr\.)', line):
            merged.append(line); continue
        if re.match(r'^\d+\s+[A-Z][a-zA-Z]', line):
            merged.append(line); continue
        if line.startswith(('Description', 'Central Tax', 'Tax(₹)',
                            'Form GSTR', 'See rule', 'Annual Return',
                            'Reconciliation Statement', '1 2 3', '(Amount')):
            merged.append(line); continue
        if re.match(r'^[A-Z][0-9-]?\s+', line):
            merged.append(line); continue
        if re.match(r'^(Inputs|Capital Goods|Input Services)(\s|$)', line):
            if re.search(r'\d', line):
                merged.append(line); continue
        if merged:
            merged[-1] = merged[-1] + ' ' + line
        else:
            merged.append(line)
    return merged


def _g9_split_desc_values(text):
    """Find the trailing stretch of numeric tokens. Description = rest."""
    tokens = text.split()
    if not tokens: return text, []
    is_num = [bool(_G9_NUM_RE.fullmatch(t)) for t in tokens]
    n = len(tokens)
    # Find last number token
    last_num_idx = -1
    for i in range(n - 1, -1, -1):
        if is_num[i]:
            last_num_idx = i; break
    if last_num_idx == -1:
        return text.strip(), []
    # Find first number of trailing contiguous block
    first_num_idx = last_num_idx
    for i in range(last_num_idx - 1, -1, -1):
        if is_num[i]:
            first_num_idx = i
        else:
            break
    values = [_g9_parse_number(tokens[i])
              for i in range(first_num_idx, last_num_idx + 1)]
    desc_tokens = tokens[:first_num_idx] + tokens[last_num_idx + 1:]
    return ' '.join(desc_tokens).strip(), values


# Per-(form, table) profile: how N values map to (Taxable, CGST, SGST, IGST, Cess)
# None means leave that column blank. Position is 0-indexed into the value list.
def _g9_map_values(form_type, table_no, sr, values):
    """Return dict with taxable/cgst/sgst/igst/cess based on table & sr_no."""
    out = {"taxable": None, "cgst": None, "sgst": None, "igst": None,
           "cess": None, "extra": None}
    n = len(values)

    if form_type == "GSTR-9":
        if table_no == "4":
            # 5 vals: Taxable, CGST, SGST, IGST, Cess
            # Special rows C/D (Export/SEZ on payment) have 3 vals: Taxable, IGST, Cess
            if n == 5:
                out["taxable"], out["cgst"], out["sgst"], out["igst"], out["cess"] = values
            elif n == 3:
                out["taxable"], out["igst"], out["cess"] = values
            elif n == 4:
                # Sometimes Cess is dropped → 4 vals = Taxable, CGST, SGST, IGST
                out["taxable"], out["cgst"], out["sgst"], out["igst"] = values
        elif table_no == "5":
            # Most rows: 1 val (Taxable)
            # Row N has 5 vals (Total Turnover with tax)
            if n == 1:
                out["taxable"] = values[0]
            elif n == 5:
                out["taxable"], out["cgst"], out["sgst"], out["igst"], out["cess"] = values
        elif table_no == "6":
            # 4 vals: CGST, SGST, IGST, Cess
            # Row E (Import goods): 2 vals (IGST, Cess) — no CGST/SGST for imports
            # Row F (Import services): 2 vals (IGST, Cess)
            # Row G (ISD): 4 vals (CGST, SGST, IGST, Cess)
            if n == 4:
                out["cgst"], out["sgst"], out["igst"], out["cess"] = values
            elif n == 2:
                out["igst"], out["cess"] = values
        elif table_no in ("7", "8", "12", "13", "15", "16"):
            # 4 vals: CGST, SGST, IGST, Cess
            if n == 4:
                out["cgst"], out["sgst"], out["igst"], out["cess"] = values
            elif n == 5:
                out["taxable"], out["cgst"], out["sgst"], out["igst"], out["cess"] = values
        elif table_no == "9":
            # Tax Paid: variable column count per row.
            # Typical: [Tax Payable, Cash, CGST_ITC, SGST_ITC, IGST_ITC, Cess_ITC,
            #          Total Paid, Diff]
            # But rows often omit columns (e.g. own-tax-ITC is blank).
            # Strategy: put Tax Payable in 'taxable', stash full breakdown in 'extra'.
            if n >= 1:
                out["taxable"] = values[0]
            extra_parts = []
            if n >= 2:
                extra_parts.append(f"Cash: {_g9_fmt(values[1])}")
            if n >= 3:
                # Trailing values: depends on row, but last 2 are usually
                # Total Paid + Diff. The middle are ITC components.
                # Safest: dump all as positional
                pos_labels = ["Cash", "Col3", "Col4", "Col5", "Col6", "Col7", "Col8"]
                for idx, v in enumerate(values[1:], 1):
                    label = pos_labels[idx - 1] if idx - 1 < len(pos_labels) else f"V{idx+1}"
                # Better: identify the Total Paid as second-to-last,
                # Diff as last
                if n >= 4:
                    out["extra"] = (f"Cash: {_g9_fmt(values[1])}; "
                                    f"Total Paid: {_g9_fmt(values[-2])}; "
                                    f"Diff: {_g9_fmt(values[-1])}")
                    # Map middle values to CGST/SGST/IGST/Cess based on row
                    middle = values[2:-2]
                    if len(middle) >= 4:
                        out["cgst"], out["sgst"], out["igst"], out["cess"] = middle[:4]
                    elif len(middle) == 3:
                        out["cgst"], out["sgst"], out["igst"] = middle
                    elif len(middle) == 2:
                        out["cgst"], out["sgst"] = middle
        elif table_no in ("10", "11"):
            # 5 vals: Taxable, CGST, SGST, IGST, Cess
            if n == 5:
                out["taxable"], out["cgst"], out["sgst"], out["igst"], out["cess"] = values
        elif table_no in ("14",):
            # Differential tax
            if n >= 1: out["taxable"] = values[0]
            if n >= 2: out["extra"] = "Paid: " + _g9_fmt(values[1])
        elif table_no == "19":
            # Late fee — Payable, Paid
            if n >= 1: out["taxable"] = values[0]
            if n >= 2: out["extra"] = "Paid: " + _g9_fmt(values[1])
        else:
            # Default: 5-val mapping
            if n == 5:
                out["taxable"], out["cgst"], out["sgst"], out["igst"], out["cess"] = values
            elif n == 4:
                out["cgst"], out["sgst"], out["igst"], out["cess"] = values
            elif n == 1:
                out["taxable"] = values[0]

    elif form_type == "GSTR-9C":
        if table_no in ("5", "7", "12", "14", "16"):
            # Reconciliation tables: 1 value (Amount)
            if n == 1:
                out["taxable"] = values[0]
            elif n == 3:
                # Table 14: Value, Total ITC, Eligible ITC
                out["taxable"] = values[0]
                out["cgst"] = values[1]  # Total ITC
                out["sgst"] = values[2]  # Eligible ITC
                out["extra"] = "Col2=Total ITC; Col3=Eligible ITC"
        elif table_no in ("9", "11", "17"):
            # Rate-wise: Taxable, CGST, SGST, IGST, Cess
            if n == 5:
                out["taxable"], out["cgst"], out["sgst"], out["igst"], out["cess"] = values
            elif n == 4:
                # Some rows only have 4 (no Cess)
                out["taxable"], out["cgst"], out["sgst"], out["igst"] = values
            elif n == 3:
                # G 28%: Taxable, IGST, Cess only (no CGST/SGST)
                out["taxable"], out["igst"], out["cess"] = values
            elif n == 2:
                out["taxable"], out["igst"] = values

    return out


def _g9_fmt(v):
    """Format number for display."""
    if v is None: return ""
    if isinstance(v, float) and v == int(v):
        return f"{int(v):,}"
    return f"{v:,.2f}"


def _g9_extract_meta(text, filepath):
    """Extract GSTIN, FY, ARN, etc. from header text."""
    head = ' '.join(text.split('\n')[:30])
    meta = {"source_file": os.path.basename(filepath)}
    m = re.search(r'Financial Year\s+(\d{4}-\d{2})', head)
    meta["fy"] = m.group(1) if m else ""
    m = re.search(r'\b(\d{2}[A-Z0-9]{13})\b', head)
    meta["gstin"] = m.group(1) if m else ""
    meta["state_code"] = meta["gstin"][:2] if meta["gstin"] else ""
    meta["state_name"] = _STATE_CODES.get(meta["state_code"], "Unknown")
    m = re.search(r'Legal [Nn]ame[^:]*?(?:registered person)?\s+([A-Z][^\n0-9]+?)(?:\s+3\(b\)|\s+Trade|$)',
                  head)
    meta["legal_name"] = m.group(1).strip() if m else ""
    m = re.search(r'Trade [Nn]ame[^:]*?\s+([A-Z][^\n]+?)(?:\s+3\(c\)|\s+ARN|$)', head)
    meta["trade_name"] = m.group(1).strip() if m else ""
    m = re.search(r'ARN\s+([A-Z0-9]+)', head)
    meta["arn"] = m.group(1) if m else ""
    # Filing date: GSTR-9 uses "Date of Filing", GSTR-9C uses "ARN Date"
    m = re.search(r'(?:Date of Filing|ARN Date)\s+(\d{2}-\d{2}-\d{4})', head)
    meta["filing_date"] = m.group(1) if m else ""

    # Detect form type
    if 'GSTR-9C' in head or 'Reconciliation Statement' in head:
        meta["form_type"] = "GSTR-9C"
    else:
        meta["form_type"] = "GSTR-9"
    return meta


def _g9_parse_table6_with_subtypes(label, rest):
    """Handle GSTR-9 Table 6 rows that have Inputs/Capital Goods/Input Services
    sub-rows. Returns a list of row dicts (one per sub-type, or one if none).
    """
    parts = re.split(r'\s+(Inputs|Capital Goods|Input Services)\s+', rest)
    if len(parts) == 1:
        # No sub-types
        desc, values = _g9_split_desc_values(parts[0])
        return [{"sr": label, "desc": desc, "subtype": "", "values": values}]

    main_desc = parts[0].strip()
    out = []
    for i in range(1, len(parts), 2):
        subtype = parts[i]
        rest_str = parts[i + 1] if i + 1 < len(parts) else ""
        # Extract leading numbers; trailing tokens become description continuation
        toks = rest_str.split()
        values = []
        trail_desc = []
        in_values = True
        for tok in toks:
            if in_values and _G9_NUM_RE.fullmatch(tok):
                values.append(_g9_parse_number(tok))
            else:
                in_values = False
                trail_desc.append(tok)
        if trail_desc:
            main_desc = (main_desc + ' ' + ' '.join(trail_desc)).strip()
        out.append({"sr": label, "desc": main_desc,
                    "subtype": subtype, "values": values})
    # Final desc applies to all
    for r in out: r["desc"] = main_desc
    return out


def parse_gstr9_pdf(filepath):
    """Parse a GSTR-9 OR GSTR-9C PDF. Returns dict with meta + rows."""
    text = _g9_extract_pdf_text(filepath)
    lines = _g9_clean_lines(text)
    merged = _g9_merge_wrapped(lines)

    meta = _g9_extract_meta(text, filepath)
    form_type = meta["form_type"]

    rows = []
    current_pt = ""
    current_table = ""
    current_table_title = ""

    for line in merged:
        # Skip column headers / form headers
        if re.match(r'^(Sr\.|Description|Central Tax|Tax\(|1 2 3'
                    r'|Form GSTR|See rule|Annual Return|Reconciliation Statement'
                    r'|PART|\(Amount|Verification)', line):
            continue
        # Skip empty/short
        if len(line) < 3: continue

        # Pt header — capture Roman + descriptive text
        m = re.match(r'^Pt\.?\s*(I+|IV|V|VI|VII)\b\s*(.*)?', line)
        if m:
            current_pt = m.group(1)
            continue

        # Table header: digit + space + capital letter + alpha
        m = re.match(r'^(\d+)\s+([A-Z][a-zA-Z].+)', line)
        if m:
            current_table = m.group(1)
            current_table_title = m.group(2)[:120]
            continue

        # Row label: A, B, C, A1, A2, K-2 etc.
        m = re.match(r'^([A-Z][0-9]?(?:-[0-9])?)\s+(.+)', line)
        if m:
            label = m.group(1)
            rest = m.group(2)

            # Skip if rest looks like another section header word
            if rest in ('Nature of Supplies', 'Description', 'Details', 'Type'):
                continue

            # Table 6 sub-type splitting
            if form_type == "GSTR-9" and current_table == "6":
                sub_rows = _g9_parse_table6_with_subtypes(label, rest)
                for sr in sub_rows:
                    mapped = _g9_map_values(form_type, current_table,
                                            sr["sr"], sr["values"])
                    rows.append({
                        "pt": current_pt,
                        "table": current_table,
                        "table_title": current_table_title,
                        "sr": sr["sr"], "desc": sr["desc"],
                        "subtype": sr["subtype"],
                        **mapped,
                        "raw_values": sr["values"],
                    })
                continue

            # Regular row parse
            desc, values = _g9_split_desc_values(rest)
            mapped = _g9_map_values(form_type, current_table, label, values)
            rows.append({
                "pt": current_pt,
                "table": current_table,
                "table_title": current_table_title,
                "sr": label, "desc": desc,
                "subtype": "",
                **mapped,
                "raw_values": values,
            })

    return {"meta": meta, "rows": rows}


def parse_gstr9_or_9c_pdf(filepath):
    """Auto-detect and parse a GSTR-9 or GSTR-9C PDF."""
    return parse_gstr9_pdf(filepath)


def write_consolidated_gstr9_9c(all_data, out_path):
    """Build consolidated workbook for GSTR-9 + GSTR-9C data.
    all_data is a list of dicts from parse_gstr9_pdf(), each with
    meta.form_type indicating which form."""
    import openpyxl
    from openpyxl.styles import Font as XF, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Style constants
    HDR_FILL = PatternFill("solid", start_color="1F4E78")
    SUB_FILL = PatternFill("solid", start_color="2E75B6")
    META_FILL = PatternFill("solid", start_color="DDEBF7")
    GROUP_FILL_9 = PatternFill("solid", start_color="E2EFDA")   # light green
    GROUP_FILL_9C = PatternFill("solid", start_color="FFF2CC")  # light yellow

    WHITE_B = XF(name="Calibri", bold=True, color="FFFFFF", size=10)
    BOLD = XF(name="Calibri", bold=True, size=10)
    REG = XF(name="Calibri", size=10)
    NEG_FONT = XF(name="Calibri", size=10, color="C00000")
    TITLE = XF(name="Calibri", bold=True, color="FFFFFF", size=14)

    thin = Side(border_style="thin", color="B4B4B4")
    BORDER_ALL = Border(left=thin, right=thin, top=thin, bottom=thin)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
    RIGHT = Alignment(horizontal="right", vertical="center")
    NUM_FMT = '#,##0.00;[Red](#,##0.00);"-"'

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Split into GSTR-9 and GSTR-9C
    data_9 = [d for d in all_data if d["meta"].get("form_type") == "GSTR-9"]
    data_9c = [d for d in all_data if d["meta"].get("form_type") == "GSTR-9C"]

    # ─── Cover sheet ──────────────────────────────────────────
    ws = wb.create_sheet("Cover")
    ws.merge_cells("A1:I1")
    ws["A1"] = "GSTR-9 / GSTR-9C — Multi-State Consolidated"
    ws["A1"].font = TITLE; ws["A1"].fill = HDR_FILL
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 30
    ws["A3"] = "GSTR-9 returns:"
    ws["B3"] = len(data_9)
    ws["A4"] = "GSTR-9C statements:"
    ws["B4"] = len(data_9c)
    ws["A3"].font = BOLD; ws["A4"].font = BOLD
    ws["B3"].font = BOLD; ws["B4"].font = BOLD

    hdrs = ["S.No.", "Form Type", "FY", "State Code", "State Name", "GSTIN",
            "Legal Name", "Filing Date", "Source File"]
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(row=6, column=c, value=h)
        cell.font = WHITE_B; cell.fill = SUB_FILL
        cell.alignment = CENTER; cell.border = BORDER_ALL
    ws.row_dimensions[6].height = 22

    sn = 0
    for data in (data_9 + data_9c):
        sn += 1; m = data["meta"]
        vals = [sn, m["form_type"], m["fy"], m["state_code"], m["state_name"],
                m["gstin"], m["legal_name"], m["filing_date"], m["source_file"]]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=6 + sn, column=c, value=v)
            cell.font = REG; cell.border = BORDER_ALL
            cell.alignment = LEFT if c >= 5 else CENTER
            if c == 1: cell.number_format = "0"
    for col, w in zip("ABCDEFGHI", [6, 10, 8, 7, 20, 20, 28, 12, 42]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A7"

    # ─── Helper to write a Console sheet ──────────────────────
    def _write_console(sheet_name, data_list, form_type, group_fill):
        if not data_list: return
        ws = wb.create_sheet(sheet_name)
        cols = ["S.No.", "State Code", "State Name", "GSTIN", "Legal Name",
                "FY", "Filing Date", "Form Type",
                "Pt", "Table", "Sr.No",
                "Description", "Sub-Type",
                "Taxable Value (₹)", "CGST (₹)", "SGST/UTGST (₹)",
                "IGST (₹)", "Cess (₹)", "Extra Info"]
        ncols = len(cols)

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        ws["A1"] = (f"{form_type} Console — every row from every filing  "
                    f"({len(data_list)} files)")
        ws["A1"].font = TITLE; ws["A1"].fill = HDR_FILL
        ws["A1"].alignment = CENTER
        ws.row_dimensions[1].height = 28

        for c, h in enumerate(cols, 1):
            cell = ws.cell(row=3, column=c, value=h)
            cell.font = WHITE_B; cell.fill = SUB_FILL
            cell.alignment = CENTER; cell.border = BORDER_ALL
        ws.row_dimensions[3].height = 38

        r = 4; sn = 0
        for data in data_list:
            m = data["meta"]
            for row in data["rows"]:
                sn += 1
                vals = [
                    sn, m["state_code"], m["state_name"], m["gstin"],
                    m["legal_name"], m["fy"], m["filing_date"], m["form_type"],
                    row["pt"], row["table"], row["sr"],
                    row["desc"], row["subtype"],
                    row.get("taxable"), row.get("cgst"), row.get("sgst"),
                    row.get("igst"), row.get("cess"), row.get("extra"),
                ]
                for c, v in enumerate(vals, 1):
                    cell = ws.cell(row=r, column=c, value=v)
                    cell.border = BORDER_ALL
                    if c == 1:
                        cell.alignment = CENTER; cell.number_format = "0"
                        cell.font = REG
                    elif c in (2, 3, 8, 9, 10, 11):
                        cell.alignment = CENTER; cell.font = REG
                    elif isinstance(v, (int, float)) and c >= 14:
                        cell.alignment = RIGHT
                        cell.number_format = NUM_FMT
                        cell.font = NEG_FONT if v < 0 else REG
                    else:
                        cell.alignment = LEFT; cell.font = REG
                r += 1

        widths = [6, 6, 18, 18, 26,
                  10, 12, 10,
                  6, 6, 6,
                  50, 14,
                  18, 16, 16, 16, 12, 36]
        for i, w in enumerate(widths[:ncols], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "L4"
        if r > 4:
            ws.auto_filter.ref = f"A3:{get_column_letter(ncols)}{r-1}"

    # ─── GSTR-9 Console ───────────────────────────────────────
    if data_9:
        _write_console("GSTR-9 Console", data_9, "GSTR-9", GROUP_FILL_9)

    # ─── GSTR-9C Console ──────────────────────────────────────
    if data_9c:
        _write_console("GSTR-9C Console", data_9c, "GSTR-9C", GROUP_FILL_9C)

    wb.save(out_path)
    return len(data_9), len(data_9c)


# ════════════════════════════════════════════════════════════════
#  ECRRS — Electronic Credit Reversal and Reclaimed Statement
# ════════════════════════════════════════════════════════════════
# Parses GSTN portal CSV "Electronic Credit Reversal and Re-claimed
# Statement" exports. Multi-GSTIN/state consolidation with:
#   • Cover sheet
#   • Console sheet — every monthly row across all GSTINs, long format
#   • Summary sheet — period totals per GSTIN with health indicators
#   • Per-GSTIN detail sheets — original portal layout preserved


def _ecrrs_parse_num(s):
    """Convert '54,144,976' / '0' / '-' / '' → float or None."""
    if s is None: return None
    s = str(s).strip().replace(',', '').replace('₹', '').strip()
    if not s or s in ('-', '—', 'NA', 'N/A'): return None
    try: return float(s)
    except: return None


def parse_ecrrs_csv(filepath):
    """Parse one ECRRS CSV file. Returns dict with meta + rows."""
    import csv
    raw_rows = []
    with open(filepath, encoding='utf-8-sig', errors='replace') as f:
        reader = csv.reader(f)
        for row in reader:
            raw_rows.append(row)

    meta = {
        "source_file": os.path.basename(filepath),
        "gstin": "", "legal_name": "", "from_date": "", "to_date": "",
        "state_code": "", "state_name": "",
    }
    # First few rows have meta in column-pair format
    for r in raw_rows[:10]:
        if len(r) < 5: continue
        label = (r[3] or "").strip() if len(r) > 3 else ""
        value = (r[4] or "").strip() if len(r) > 4 else ""
        if not label or not value: continue
        if label == "GSTIN": meta["gstin"] = value
        elif label == "Legal Name": meta["legal_name"] = value
        elif label == "From": meta["from_date"] = value
        elif label == "To": meta["to_date"] = value
    if meta["gstin"]:
        meta["state_code"] = meta["gstin"][:2]
        meta["state_name"] = _STATE_CODES.get(meta["state_code"], "Unknown")

    # Find header rows
    # Row with 'S.No.' is the main header; the next row has tax-type sub-headers
    hdr_idx = None
    for i, r in enumerate(raw_rows):
        if r and r[0] and r[0].strip().startswith("S.No"):
            hdr_idx = i; break
    if hdr_idx is None:
        return {"meta": meta, "rows": []}

    data_start = hdr_idx + 2
    rows = []
    for r in raw_rows[data_start:]:
        if not r or not r[0] or not str(r[0]).strip():
            continue
        # Skip rows with fewer than 21 columns (data has 21 columns including S.No)
        # but allow Opening/Closing Balance rows with dashes
        def col(i):
            return r[i].strip() if i < len(r) and r[i] else ""

        sr_no = col(0)
        date = col(1)
        ref_no = col(2)
        period = col(3)
        desc = col(4)

        # ITC Claimed (Table 4A(5)) — IGST, CGST, SGST, Cess
        claimed_igst = _ecrrs_parse_num(col(5))
        claimed_cgst = _ecrrs_parse_num(col(6))
        claimed_sgst = _ecrrs_parse_num(col(7))
        claimed_cess = _ecrrs_parse_num(col(8))

        # ITC Reversal (Table 4B(2))
        rev_igst = _ecrrs_parse_num(col(9))
        rev_cgst = _ecrrs_parse_num(col(10))
        rev_sgst = _ecrrs_parse_num(col(11))
        rev_cess = _ecrrs_parse_num(col(12))

        # ITC Reclaimed (Table 4D(1))
        rec_igst = _ecrrs_parse_num(col(13))
        rec_cgst = _ecrrs_parse_num(col(14))
        rec_sgst = _ecrrs_parse_num(col(15))
        rec_cess = _ecrrs_parse_num(col(16))

        # Closing Balance
        clo_igst = _ecrrs_parse_num(col(17))
        clo_cgst = _ecrrs_parse_num(col(18))
        clo_sgst = _ecrrs_parse_num(col(19))
        clo_cess = _ecrrs_parse_num(col(20))

        # Classify row type
        if "Opening Balance" in desc:
            row_type = "Opening"
        elif "Closing Balance" in desc:
            row_type = "Closing"
        else:
            row_type = "Monthly"

        rows.append({
            "sr_no": sr_no, "date": date, "ref_no": ref_no,
            "period": period, "desc": desc, "row_type": row_type,
            # Claimed
            "claimed_igst": claimed_igst, "claimed_cgst": claimed_cgst,
            "claimed_sgst": claimed_sgst, "claimed_cess": claimed_cess,
            # Reversal
            "rev_igst": rev_igst, "rev_cgst": rev_cgst,
            "rev_sgst": rev_sgst, "rev_cess": rev_cess,
            # Reclaimed
            "rec_igst": rec_igst, "rec_cgst": rec_cgst,
            "rec_sgst": rec_sgst, "rec_cess": rec_cess,
            # Closing
            "clo_igst": clo_igst, "clo_cgst": clo_cgst,
            "clo_sgst": clo_sgst, "clo_cess": clo_cess,
        })

    return {"meta": meta, "rows": rows}


def _ecrrs_dedup_by_gstin(all_data):
    """Combine multiple ECRRS files for the same GSTIN and remove duplicate
    rows by Reference No.

    Real-world need: users often download overlapping periods (e.g. Apr-25 to
    Mar-26 AND 20-Mar-26 to 20-Apr-26). The Feb-26 GSTR-3B filing (20-Mar-26)
    appears in BOTH. Reference No is GSTN's unique transaction ID, so we use
    that as the dedup key.

    Returns (combined_list, dedup_stats).
    `combined_list` has one entry per GSTIN with all unique monthly rows
    sorted by date. `dedup_stats` is a list of dicts logging what was removed.
    """
    from datetime import datetime

    # Group by GSTIN
    by_gstin = {}
    for data in all_data:
        gstin = data["meta"].get("gstin", "")
        if not gstin: continue
        by_gstin.setdefault(gstin, []).append(data)

    combined = []
    stats = []

    def parse_date(s):
        """Parse '20/03/2026' or '19/04/2025' to a sortable datetime, or None."""
        if not s or s in ('-', '—'): return None
        for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
            try: return datetime.strptime(s.strip(), fmt)
            except: continue
        return None

    for gstin, files in by_gstin.items():
        # Sort files by from_date so earliest comes first
        files_sorted = sorted(files,
                              key=lambda d: parse_date(d["meta"].get("from_date", ""))
                                            or datetime.min)

        # Collect all monthly rows, keep unique by reference number
        # If duplicates found, prefer the row from the earliest file
        # (which is usually the more complete year statement)
        seen_refs = {}              # ref_no → row
        all_monthly = []
        duplicates_skipped = 0
        dup_examples = []
        opening_row = None          # from earliest file
        closing_row = None          # from latest file

        for data in files_sorted:
            for row in data["rows"]:
                rt = row["row_type"]
                if rt == "Opening":
                    if opening_row is None:
                        opening_row = dict(row)  # take FIRST opening
                elif rt == "Closing":
                    closing_row = dict(row)      # take LAST closing
                else:
                    ref = (row.get("ref_no") or "").strip()
                    # Fallback dedup key for rows missing ref_no
                    if not ref or ref == "-":
                        ref = (f"{row.get('date','')}_{row.get('period','')}"
                               f"_{row.get('desc','')[:20]}")
                    if ref in seen_refs:
                        duplicates_skipped += 1
                        if len(dup_examples) < 3:
                            dup_examples.append(
                                f"{row.get('period','?')} "
                                f"(Ref: {row.get('ref_no','?')[:20]})")
                        continue
                    seen_refs[ref] = row
                    all_monthly.append(dict(row))

        # Sort monthly rows by date so they appear chronologically
        all_monthly.sort(key=lambda r: parse_date(r.get("date", ""))
                                       or datetime.min)

        # Re-number Sr.No
        rows_out = []
        sn = 0
        if opening_row:
            sn += 1
            opening_row["sr_no"] = str(sn)
            rows_out.append(opening_row)
        for row in all_monthly:
            sn += 1
            row["sr_no"] = str(sn)
            rows_out.append(row)
        if closing_row:
            sn += 1
            closing_row["sr_no"] = str(sn)
            rows_out.append(closing_row)

        # Build the merged meta — combine periods of all source files
        first_meta = files_sorted[0]["meta"]
        last_meta = files_sorted[-1]["meta"]
        merged_meta = dict(first_meta)
        merged_meta["from_date"] = first_meta.get("from_date", "")
        merged_meta["to_date"] = last_meta.get("to_date", "")
        if len(files_sorted) > 1:
            merged_meta["source_file"] = (
                f"{len(files_sorted)} files merged: "
                + ", ".join(d["meta"].get("source_file", "?")
                            for d in files_sorted))

        combined.append({"meta": merged_meta, "rows": rows_out})

        if len(files_sorted) > 1 or duplicates_skipped > 0:
            stats.append({
                "gstin": gstin,
                "files_merged": len(files_sorted),
                "duplicates_skipped": duplicates_skipped,
                "duplicate_examples": dup_examples,
                "unique_monthly": len(all_monthly),
            })

    return combined, stats


def write_consolidated_ecrrs(all_data, out_path):
    """Build consolidated workbook for ECRRS data from multiple GSTINs."""
    import openpyxl
    from openpyxl.styles import (Font as XF, PatternFill, Alignment,
                                  Border, Side)
    from openpyxl.utils import get_column_letter

    # Step 1: Dedup by GSTIN + Reference No (handles overlapping period downloads)
    all_data, dedup_stats = _ecrrs_dedup_by_gstin(all_data)

    HDR_FILL = PatternFill("solid", start_color="1F4E78")
    SUB_FILL = PatternFill("solid", start_color="2E75B6")
    GRP_FILL_CLAIM = PatternFill("solid", start_color="E2EFDA")   # green
    GRP_FILL_REV = PatternFill("solid", start_color="FCE4D6")     # orange
    GRP_FILL_REC = PatternFill("solid", start_color="DDEBF7")     # blue
    GRP_FILL_BAL = PatternFill("solid", start_color="FFF2CC")     # yellow
    META_FILL = PatternFill("solid", start_color="F2F2F2")
    OPEN_FILL = PatternFill("solid", start_color="E7E6E6")
    CLOSE_FILL = PatternFill("solid", start_color="FFE699")
    GOOD_FILL = PatternFill("solid", start_color="C6EFCE")
    WARN_FILL = PatternFill("solid", start_color="FFEB9C")
    BAD_FILL = PatternFill("solid", start_color="FFC7CE")

    WHITE_B = XF(name="Calibri", bold=True, color="FFFFFF", size=10)
    BOLD = XF(name="Calibri", bold=True, size=10)
    REG = XF(name="Calibri", size=10)
    NEG_FONT = XF(name="Calibri", size=10, color="C00000")
    TITLE = XF(name="Calibri", bold=True, color="FFFFFF", size=14)
    SUBTITLE = XF(name="Calibri", italic=True, size=10, color="404040")

    thin = Side(border_style="thin", color="B4B4B4")
    BORDER_ALL = Border(left=thin, right=thin, top=thin, bottom=thin)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
    RIGHT = Alignment(horizontal="right", vertical="center")
    NUM_FMT = '#,##0.00;[Red](#,##0.00);"-"'

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ─── 1. Cover ─────────────────────────────────────────────
    ws = wb.create_sheet("Cover")
    ws.merge_cells("A1:H1")
    ws["A1"] = "Electronic Credit Reversal & Re-claimed Statement — Consolidated"
    ws["A1"].font = TITLE; ws["A1"].fill = HDR_FILL
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws["A2"] = ("Multi-GSTIN consolidation · ITC Claimed / Reversed / Reclaimed "
                "/ Closing Balance · long-format Console + per-GSTIN sheets")
    ws["A2"].font = SUBTITLE; ws["A2"].alignment = CENTER

    hdrs = ["S.No.", "GSTIN", "State Code", "State Name", "Legal Name",
            "Period From", "Period To", "Source File"]
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = WHITE_B; cell.fill = SUB_FILL
        cell.alignment = CENTER; cell.border = BORDER_ALL
    ws.row_dimensions[4].height = 22

    for i, data in enumerate(all_data, 1):
        m = data["meta"]
        vals = [i, m["gstin"], m["state_code"], m["state_name"],
                m["legal_name"], m["from_date"], m["to_date"],
                m["source_file"]]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=4 + i, column=c, value=v)
            cell.font = REG; cell.border = BORDER_ALL
            cell.alignment = LEFT if c >= 4 else CENTER
            if c == 1: cell.number_format = "0"

    for col, w in zip("ABCDEFGH", [6, 18, 7, 18, 28, 12, 12, 42]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A5"

    # ─── 2. Console (long format) ─────────────────────────────
    ws = wb.create_sheet("Console")
    cols = ["S.No.", "GSTIN", "State Code", "State Name", "Legal Name",
            "Row Type", "Sr.No.", "Date", "Reference No.", "Return Period",
            "Description",
            # 4 groups × 4 tax types = 16 numeric cols
            "Claimed IGST", "Claimed CGST", "Claimed SGST/UTGST", "Claimed Cess",
            "Reversal IGST", "Reversal CGST", "Reversal SGST/UTGST", "Reversal Cess",
            "Reclaimed IGST", "Reclaimed CGST", "Reclaimed SGST/UTGST", "Reclaimed Cess",
            "Closing IGST", "Closing CGST", "Closing SGST/UTGST", "Closing Cess"]
    ncols = len(cols)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws["A1"] = ("Console — every row, every GSTIN, every month  "
                "(filter by State / GSTIN / Period / Row Type)")
    ws["A1"].font = TITLE; ws["A1"].fill = HDR_FILL
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 28

    # Group header row 2 — coloured banners
    def grp(c_start, c_end, label, fill):
        ws.merge_cells(start_row=2, start_column=c_start, end_row=2,
                       end_column=c_end)
        cell = ws.cell(row=2, column=c_start, value=label)
        cell.font = BOLD; cell.fill = fill
        cell.alignment = CENTER; cell.border = BORDER_ALL

    grp(1, 11, "Metadata", META_FILL)
    grp(12, 15, "ITC Claimed  [Table 4A(5)]", GRP_FILL_CLAIM)
    grp(16, 19, "ITC Reversal  [Table 4B(2)]", GRP_FILL_REV)
    grp(20, 23, "ITC Reclaimed  [Table 4D(1)]", GRP_FILL_REC)
    grp(24, 27, "Closing Balance", GRP_FILL_BAL)
    ws.row_dimensions[2].height = 22

    # Column header row 3
    for c, h in enumerate(cols, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = WHITE_B; cell.fill = SUB_FILL
        cell.alignment = CENTER; cell.border = BORDER_ALL
    ws.row_dimensions[3].height = 36

    r = 4; sn = 0
    for data in all_data:
        m = data["meta"]
        for row in data["rows"]:
            sn += 1
            vals = [
                sn, m["gstin"], m["state_code"], m["state_name"],
                m["legal_name"],
                row["row_type"], row["sr_no"], row["date"],
                row["ref_no"], row["period"], row["desc"],
                row["claimed_igst"], row["claimed_cgst"],
                row["claimed_sgst"], row["claimed_cess"],
                row["rev_igst"], row["rev_cgst"],
                row["rev_sgst"], row["rev_cess"],
                row["rec_igst"], row["rec_cgst"],
                row["rec_sgst"], row["rec_cess"],
                row["clo_igst"], row["clo_cgst"],
                row["clo_sgst"], row["clo_cess"],
            ]
            row_bold = row["row_type"] in ("Opening", "Closing")
            row_fill = (OPEN_FILL if row["row_type"] == "Opening"
                        else CLOSE_FILL if row["row_type"] == "Closing"
                        else None)
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.border = BORDER_ALL
                if row_fill: cell.fill = row_fill
                if c == 1:
                    cell.alignment = CENTER
                    cell.number_format = "0"
                    cell.font = BOLD if row_bold else REG
                elif c in (2, 3, 6, 7, 8, 10):
                    cell.alignment = CENTER
                    cell.font = BOLD if row_bold else REG
                elif c in (4, 5, 9, 11):
                    cell.alignment = LEFT
                    cell.font = BOLD if row_bold else REG
                elif isinstance(v, (int, float)) and c >= 12:
                    cell.alignment = RIGHT
                    cell.number_format = NUM_FMT
                    if v < 0:
                        cell.font = NEG_FONT
                    else:
                        cell.font = BOLD if row_bold else REG
                else:
                    cell.alignment = LEFT
                    cell.font = BOLD if row_bold else REG
            r += 1

    widths = [6, 18, 6, 16, 22,
              9, 6, 11, 18, 10, 26,
              14, 14, 16, 11,
              14, 14, 16, 11,
              14, 14, 16, 11,
              14, 14, 16, 11]
    for i, w in enumerate(widths[:ncols], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "L4"
    if r > 4:
        ws.auto_filter.ref = f"A3:{get_column_letter(ncols)}{r-1}"

    # ─── 3. Summary (per GSTIN totals + health) ───────────────
    ws = wb.create_sheet("Summary")
    ws.merge_cells("A1:R1")
    ws["A1"] = ("Summary — Total Claimed / Reversed / Reclaimed / Closing  "
                "per GSTIN  (with reversal-reclaim balance check)")
    ws["A1"].font = TITLE; ws["A1"].fill = HDR_FILL
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 28

    # Group banner row 2
    def grp2(c_start, c_end, label, fill):
        ws.merge_cells(start_row=2, start_column=c_start, end_row=2,
                       end_column=c_end)
        cell = ws.cell(row=2, column=c_start, value=label)
        cell.font = BOLD; cell.fill = fill
        cell.alignment = CENTER; cell.border = BORDER_ALL

    grp2(1, 5, "GSTIN Details", META_FILL)
    grp2(6, 8, "Period & Counts", META_FILL)
    grp2(9, 11, "Total Reversal (sum of Monthly Reversal)", GRP_FILL_REV)
    grp2(12, 14, "Total Reclaimed (sum of Monthly Reclaimed)", GRP_FILL_REC)
    grp2(15, 17, "Closing Balance (latest)", GRP_FILL_BAL)
    ws.merge_cells(start_row=2, start_column=18, end_row=2, end_column=18)
    cell = ws.cell(row=2, column=18, value="Status")
    cell.font = BOLD; cell.fill = META_FILL
    cell.alignment = CENTER; cell.border = BORDER_ALL
    ws.row_dimensions[2].height = 22

    sum_hdrs = ["S.No.", "GSTIN", "State", "Legal Name", "FY/Period",
                "From", "To", "Months",
                "Reversal IGST", "Reversal CGST+SGST", "Reversal Cess",
                "Reclaimed IGST", "Reclaimed CGST+SGST", "Reclaimed Cess",
                "Closing IGST", "Closing CGST+SGST", "Closing Cess",
                "Health"]
    for c, h in enumerate(sum_hdrs, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = WHITE_B; cell.fill = SUB_FILL
        cell.alignment = CENTER; cell.border = BORDER_ALL
    ws.row_dimensions[3].height = 36

    r = 4
    for i, data in enumerate(all_data, 1):
        m = data["meta"]
        monthly = [row for row in data["rows"] if row["row_type"] == "Monthly"]
        closing = next((row for row in data["rows"]
                        if row["row_type"] == "Closing"), None)

        def sum_field(field):
            return sum((row[field] or 0) for row in monthly)

        rev_i = sum_field("rev_igst")
        rev_c = sum_field("rev_cgst") + sum_field("rev_sgst")
        rev_x = sum_field("rev_cess")
        rec_i = sum_field("rec_igst")
        rec_c = sum_field("rec_cgst") + sum_field("rec_sgst")
        rec_x = sum_field("rec_cess")
        clo_i = (closing["clo_igst"] if closing else 0) or 0
        clo_c = ((closing["clo_cgst"] or 0) + (closing["clo_sgst"] or 0)
                 if closing else 0)
        clo_x = (closing["clo_cess"] if closing else 0) or 0

        # Health: total closing should equal opening + reversal - reclaimed
        # Simple status:
        #   Green if closing balance > 0 and reversal ≈ reclaimed
        #   Yellow if reversal > reclaimed (still parked)
        #   Red if closing < 0 (impossible — likely data issue)
        total_rev = rev_i + rev_c + rev_x
        total_rec = rec_i + rec_c + rec_x
        total_clo = clo_i + clo_c + clo_x

        if total_clo < 0:
            health = "⚠  Negative balance"
            health_fill = BAD_FILL
        elif total_rev == 0:
            health = "—  No activity"
            health_fill = META_FILL
        elif total_rec >= total_rev * 0.95:
            health = "✅  Mostly reclaimed"
            health_fill = GOOD_FILL
        elif total_rec >= total_rev * 0.5:
            health = "○  Partial reclaim"
            health_fill = WARN_FILL
        else:
            health = "⚠  Heavy parked balance"
            health_fill = WARN_FILL

        vals = [i, m["gstin"], m["state_name"], m["legal_name"], "",
                m["from_date"], m["to_date"], len(monthly),
                rev_i, rev_c, rev_x,
                rec_i, rec_c, rec_x,
                clo_i, clo_c, clo_x,
                health]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = REG; cell.border = BORDER_ALL
            if c == 1:
                cell.alignment = CENTER; cell.number_format = "0"
            elif c in (2, 3, 6, 7, 8):
                cell.alignment = CENTER
            elif c == 4:
                cell.alignment = LEFT
            elif c >= 9 and c <= 17 and isinstance(v, (int, float)):
                cell.alignment = RIGHT
                cell.number_format = NUM_FMT
                if v < 0: cell.font = NEG_FONT
            elif c == 18:
                cell.alignment = CENTER
                cell.fill = health_fill
                cell.font = BOLD
            else:
                cell.alignment = LEFT
        r += 1

    widths_sum = [6, 18, 16, 26, 10, 12, 12, 7,
                  14, 16, 11, 14, 16, 11, 14, 16, 11, 22]
    for i, w in enumerate(widths_sum, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"
    if r > 4:
        ws.auto_filter.ref = f"A3:R{r-1}"

    # ─── 4. Per-GSTIN detail sheets (portal layout) ───────────
    for data in all_data:
        m = data["meta"]
        # Sheet name: "<StateCode>_<last4 of GSTIN>"  (≤31 chars)
        sname = f"{m['state_code']}_{m['gstin'][-4:]}" if m['gstin'] else f"GSTIN_{len(wb.sheetnames)}"
        sname = sname[:31]
        # Handle duplicates
        base_sname = sname
        idx = 1
        while sname in wb.sheetnames:
            idx += 1
            sname = f"{base_sname[:28]}_{idx}"
        ws = wb.create_sheet(sname)

        # Title and meta
        ws.merge_cells("A1:U1")
        ws["A1"] = f"ECRRS — {m['gstin']}  ({m['state_name']})"
        ws["A1"].font = TITLE; ws["A1"].fill = HDR_FILL
        ws["A1"].alignment = CENTER
        ws.row_dimensions[1].height = 26

        ws.merge_cells("A2:U2")
        ws["A2"] = (f"{m['legal_name']}  |  Period: {m['from_date']} to {m['to_date']}")
        ws["A2"].font = SUBTITLE
        ws["A2"].alignment = CENTER
        ws.row_dimensions[2].height = 18

        # Group banner row 4
        gh = ["S.No.", "Date", "Reference No.", "Return Period", "Description"]
        for c, h in enumerate(gh, 1):
            ws.merge_cells(start_row=4, start_column=c, end_row=5, end_column=c)
            cell = ws.cell(row=4, column=c, value=h)
            cell.font = WHITE_B; cell.fill = SUB_FILL
            cell.alignment = CENTER; cell.border = BORDER_ALL

        # 4 grouped columns × 4 tax types
        group_specs = [
            (6, 9,  "ITC Claimed  [Table 4A(5)]",  GRP_FILL_CLAIM),
            (10, 13, "ITC Reversal  [Table 4B(2)]", GRP_FILL_REV),
            (14, 17, "ITC Reclaimed  [Table 4D(1)]", GRP_FILL_REC),
            (18, 21, "Closing Balance", GRP_FILL_BAL),
        ]
        for cs, ce, lab, fl in group_specs:
            ws.merge_cells(start_row=4, start_column=cs, end_row=4, end_column=ce)
            cell = ws.cell(row=4, column=cs, value=lab)
            cell.font = BOLD; cell.fill = fl
            cell.alignment = CENTER; cell.border = BORDER_ALL

        # Sub-headers row 5: IGST / CGST / SGST / Cess × 4
        tax_subs = ["IGST", "CGST", "SGST/UTGST", "Cess"]
        for grp_idx in range(4):
            for sub_idx, sub_label in enumerate(tax_subs):
                col = 6 + grp_idx * 4 + sub_idx
                cell = ws.cell(row=5, column=col, value=sub_label)
                cell.font = WHITE_B; cell.fill = SUB_FILL
                cell.alignment = CENTER; cell.border = BORDER_ALL
        ws.row_dimensions[4].height = 22
        ws.row_dimensions[5].height = 22

        # Data rows
        r = 6
        for row in data["rows"]:
            is_balance = row["row_type"] in ("Opening", "Closing")
            row_fill = (OPEN_FILL if row["row_type"] == "Opening"
                        else CLOSE_FILL if row["row_type"] == "Closing"
                        else None)
            vals = [
                row["sr_no"], row["date"], row["ref_no"],
                row["period"], row["desc"],
                row["claimed_igst"], row["claimed_cgst"],
                row["claimed_sgst"], row["claimed_cess"],
                row["rev_igst"], row["rev_cgst"],
                row["rev_sgst"], row["rev_cess"],
                row["rec_igst"], row["rec_cgst"],
                row["rec_sgst"], row["rec_cess"],
                row["clo_igst"], row["clo_cgst"],
                row["clo_sgst"], row["clo_cess"],
            ]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.border = BORDER_ALL
                if row_fill: cell.fill = row_fill
                if c <= 5:
                    cell.alignment = CENTER if c != 5 else LEFT
                    cell.font = BOLD if is_balance else REG
                elif isinstance(v, (int, float)):
                    cell.alignment = RIGHT
                    cell.number_format = NUM_FMT
                    if v < 0:
                        cell.font = NEG_FONT
                    else:
                        cell.font = BOLD if is_balance else REG
                else:
                    cell.alignment = CENTER
                    cell.font = BOLD if is_balance else REG
            r += 1

        # Column widths
        det_widths = [6, 11, 18, 10, 26,
                      14, 14, 16, 11,
                      14, 14, 16, 11,
                      14, 14, 16, 11,
                      14, 14, 16, 11]
        for i, w in enumerate(det_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "F6"
        if r > 6:
            ws.auto_filter.ref = f"A5:U{r-1}"

    wb.save(out_path)
    return len(all_data), dedup_stats


# ════════════════════════════════════════════════════════════════
#  PDF MERGE / SPLIT ENGINES
# ════════════════════════════════════════════════════════════════
# Used by Tab 2 (PDF Tools) — merge multiple PDFs to one, split one
# into many with auto-fit-under-N-MB intelligence for GST portal
# (typical limit: 5 MB per file).


def get_pdf_info(filepath):
    """Return (page_count, size_bytes) for a PDF.
    page_count is 0 if PyMuPDF/pikepdf is unavailable."""
    try:
        size_bytes = os.path.getsize(filepath)
    except Exception:
        return (0, 0)
    pages = 0
    if HAS_PYMUPDF:
        try:
            import fitz
            with fitz.open(filepath) as doc:
                pages = doc.page_count
        except Exception:
            pages = 0
    elif HAS_PIKEPDF:
        try:
            import pikepdf
            with pikepdf.open(filepath) as pdf:
                pages = len(pdf.pages)
        except Exception:
            pages = 0
    return (pages, size_bytes)


def merge_pdfs(filepaths, output_path, progress_cb=None):
    """Merge multiple PDFs in given order. Returns output size in bytes.
    progress_cb(idx, total, name) called per file."""
    if not filepaths:
        raise ValueError("No PDFs to merge.")
    # Try pikepdf first (slightly better metadata handling); fall back to PyMuPDF
    if HAS_PIKEPDF:
        import pikepdf
        merged = pikepdf.Pdf.new()
        try:
            for i, fp in enumerate(filepaths, 1):
                if progress_cb:
                    progress_cb(i, len(filepaths), os.path.basename(fp))
                src = pikepdf.open(fp)
                try:
                    merged.pages.extend(src.pages)
                finally:
                    src.close()
            merged.save(output_path,
                        compress_streams=True,
                        object_stream_mode=pikepdf.ObjectStreamMode.generate)
        finally:
            merged.close()
    elif HAS_PYMUPDF:
        import fitz
        merged = fitz.open()
        try:
            for i, fp in enumerate(filepaths, 1):
                if progress_cb:
                    progress_cb(i, len(filepaths), os.path.basename(fp))
                src = fitz.open(fp)
                merged.insert_pdf(src)
                src.close()
            merged.save(output_path, garbage=4, deflate=True, clean=True)
        finally:
            merged.close()
    else:
        raise RuntimeError("Need PyMuPDF or pikepdf installed for merging.")
    return os.path.getsize(output_path)


def compute_merge_buckets(filepaths, target_mb, preserve_order=False):
    """Bin-pack PDFs into buckets such that each bucket's total size ≤ target_mb.

    Uses First-Fit-Decreasing (sort by size desc, then greedy fit) for best
    packing. With preserve_order=True, packs files in given order, starting a
    new bucket whenever the current one would overflow.

    Returns (buckets, oversize_files) where:
      buckets        — list of buckets, each a list of (filepath, size_bytes)
      oversize_files — list of (filepath, size_bytes) for files that alone
                       exceed target. These each become their own bucket
                       (and need pre-compression to actually fit target).
    """
    target_bytes = target_mb * 1024 * 1024
    items = [(fp, os.path.getsize(fp)) for fp in filepaths]
    # Separate oversize (each alone > target) from normal
    oversize = [(fp, sz) for fp, sz in items if sz > target_bytes]
    normal = [(fp, sz) for fp, sz in items if sz <= target_bytes]

    if not preserve_order:
        # First-Fit Decreasing: pack big first for best fill ratio
        normal.sort(key=lambda x: -x[1])

    buckets = []
    if preserve_order:
        # Sequential pack: add to current bucket if fits, else start new bucket
        current = []
        current_total = 0
        for fp, sz in normal:
            if current and current_total + sz > target_bytes:
                buckets.append(current)
                current = [(fp, sz)]
                current_total = sz
            else:
                current.append((fp, sz))
                current_total += sz
        if current:
            buckets.append(current)
    else:
        # First-Fit: try each existing bucket; create new if none fits
        for fp, sz in normal:
            placed = False
            for bucket in buckets:
                bucket_total = sum(b[1] for b in bucket)
                if bucket_total + sz <= target_bytes:
                    bucket.append((fp, sz))
                    placed = True
                    break
            if not placed:
                buckets.append([(fp, sz)])

    # Each oversize file gets its own bucket
    for fp, sz in oversize:
        buckets.append([(fp, sz)])
    return buckets, oversize


def split_pdf_by_chunk(input_path, output_dir, pages_per_chunk, prefix=None):
    """Split a PDF into chunks of N pages each. Returns list of output paths."""
    if prefix is None:
        prefix = os.path.splitext(os.path.basename(input_path))[0]
    os.makedirs(output_dir, exist_ok=True)
    outputs = []
    if HAS_PIKEPDF:
        import pikepdf
        with pikepdf.open(input_path) as src:
            total = len(src.pages)
            for chunk_idx, start in enumerate(range(0, total, pages_per_chunk), 1):
                end = min(start + pages_per_chunk, total)
                out_path = os.path.join(
                    output_dir,
                    f"{prefix}_part{chunk_idx}_p{start+1}-{end}.pdf")
                out_pdf = pikepdf.Pdf.new()
                try:
                    for i in range(start, end):
                        out_pdf.pages.append(src.pages[i])
                    out_pdf.save(out_path, compress_streams=True,
                                 object_stream_mode=pikepdf.ObjectStreamMode.generate)
                finally:
                    out_pdf.close()
                outputs.append(out_path)
    elif HAS_PYMUPDF:
        import fitz
        with fitz.open(input_path) as src:
            total = src.page_count
            for chunk_idx, start in enumerate(range(0, total, pages_per_chunk), 1):
                end = min(start + pages_per_chunk, total)
                out_path = os.path.join(
                    output_dir,
                    f"{prefix}_part{chunk_idx}_p{start+1}-{end}.pdf")
                out = fitz.open()
                try:
                    out.insert_pdf(src, from_page=start, to_page=end - 1)
                    out.save(out_path, garbage=4, deflate=True, clean=True)
                finally:
                    out.close()
                outputs.append(out_path)
    else:
        raise RuntimeError("Need PyMuPDF or pikepdf installed for splitting.")
    return outputs


def split_pdf_by_ranges(input_path, output_dir, ranges, prefix=None):
    """Split by explicit (start, end) 1-indexed inclusive page ranges.
    e.g. [(1, 5), (6, 10), (11, 20)]. Returns list of output paths."""
    if prefix is None:
        prefix = os.path.splitext(os.path.basename(input_path))[0]
    os.makedirs(output_dir, exist_ok=True)
    outputs = []
    if HAS_PIKEPDF:
        import pikepdf
        with pikepdf.open(input_path) as src:
            total = len(src.pages)
            for chunk_idx, (s1, e1) in enumerate(ranges, 1):
                if s1 < 1 or e1 > total or s1 > e1:
                    raise ValueError(
                        f"Range {s1}-{e1} invalid (PDF has {total} pages).")
                out_path = os.path.join(
                    output_dir, f"{prefix}_part{chunk_idx}_p{s1}-{e1}.pdf")
                out_pdf = pikepdf.Pdf.new()
                try:
                    for i in range(s1 - 1, e1):
                        out_pdf.pages.append(src.pages[i])
                    out_pdf.save(out_path, compress_streams=True,
                                 object_stream_mode=pikepdf.ObjectStreamMode.generate)
                finally:
                    out_pdf.close()
                outputs.append(out_path)
    elif HAS_PYMUPDF:
        import fitz
        with fitz.open(input_path) as src:
            total = src.page_count
            for chunk_idx, (s1, e1) in enumerate(ranges, 1):
                if s1 < 1 or e1 > total or s1 > e1:
                    raise ValueError(
                        f"Range {s1}-{e1} invalid (PDF has {total} pages).")
                out_path = os.path.join(
                    output_dir, f"{prefix}_part{chunk_idx}_p{s1}-{e1}.pdf")
                out = fitz.open()
                try:
                    out.insert_pdf(src, from_page=s1 - 1, to_page=e1 - 1)
                    out.save(out_path, garbage=4, deflate=True, clean=True)
                finally:
                    out.close()
                outputs.append(out_path)
    else:
        raise RuntimeError("Need PyMuPDF or pikepdf installed for splitting.")
    return outputs


def parse_page_ranges(ranges_str, total_pages):
    """Parse '1-5, 6-10, 11-20' or '1,3,5-7' → [(1,5), (6,10), (11,20)]."""
    ranges = []
    parts = [p.strip() for p in ranges_str.replace(';', ',').split(',') if p.strip()]
    for p in parts:
        if '-' in p:
            try:
                s, e = p.split('-', 1)
                s, e = int(s.strip()), int(e.strip())
            except Exception:
                raise ValueError(f"Bad range: {p!r}")
        else:
            try:
                s = e = int(p.strip())
            except Exception:
                raise ValueError(f"Bad page number: {p!r}")
        ranges.append((s, e))
    return ranges


def split_pdf_to_fit(input_path, output_dir, target_mb, prefix=None,
                     progress_cb=None):
    """Smart split — divide PDF into chunks so each output is ≤ target_mb.
    Returns list of output paths. Uses iterative refinement: estimate
    per-page size, chunk accordingly, then re-split any chunk that
    overshoots the target."""
    if prefix is None:
        prefix = os.path.splitext(os.path.basename(input_path))[0]
    os.makedirs(output_dir, exist_ok=True)
    pages, size_bytes = get_pdf_info(input_path)
    target_bytes = int(target_mb * 1024 * 1024)
    if pages <= 0:
        raise RuntimeError("Could not read page count from PDF.")
    # Already fits
    if size_bytes <= target_bytes:
        import shutil
        out_path = os.path.join(
            output_dir, f"{prefix}_fits_{size_bytes/1024/1024:.2f}MB.pdf")
        shutil.copy2(input_path, out_path)
        return [out_path]

    # Estimate per-page size + 15% safety buffer
    per_page = size_bytes / pages
    pages_per_chunk = max(1, int((target_bytes * 0.85) / per_page))
    if progress_cb:
        progress_cb(
            f"Estimate: per-page ≈ {per_page/1024:.1f} KB, "
            f"chunk = {pages_per_chunk} pages")

    initial = split_pdf_by_chunk(input_path, output_dir, pages_per_chunk, prefix)

    # Verify each chunk, sub-split if needed
    refined = []
    for out_path in initial:
        out_size = os.path.getsize(out_path)
        if out_size > target_bytes:
            # Re-split this chunk
            sub_prefix = os.path.splitext(os.path.basename(out_path))[0] + "_sub"
            sub_outputs = split_pdf_to_fit(out_path, output_dir, target_mb,
                                            prefix=sub_prefix,
                                            progress_cb=progress_cb)
            try: os.remove(out_path)
            except Exception: pass
            refined.extend(sub_outputs)
        else:
            refined.append(out_path)

    # Renumber outputs so they look clean (part1, part2, ...)
    final = []
    for i, fp in enumerate(refined, 1):
        # Get page range from original filename if present (best-effort rename)
        base = os.path.basename(fp)
        # Insert a clean ordinal prefix
        new_name = f"{prefix}_part{i:02d}_{base}"
        new_path = os.path.join(output_dir, new_name)
        try:
            os.rename(fp, new_path)
            final.append(new_path)
        except Exception:
            final.append(fp)
    return final


# ════════════════════════════════════════════════════════════════
#  ECL — Electronic Credit Ledger consolidator
# ════════════════════════════════════════════════════════════════
# Parses GSTN portal CSV export of "Electronic Credit Ledger" and
# consolidates multi-GSTIN/multi-period downloads into one workbook
# with Cover, Console (long-format), Summary, and per-GSTIN detail.


def _ledger_parse_num(s):
    """Convert '54,144,976' / '9,81,534.00' / '0' / '-' / '' → float | None."""
    if s is None: return None
    s = str(s).strip().replace(',', '').replace('₹', '').strip()
    if not s or s in ('-', '—', 'NA', 'N/A'): return None
    try: return float(s)
    except: return None


def parse_ecl_csv(filepath):
    """Parse one Electronic Credit Ledger CSV. Returns dict with meta + rows.

    Structure of the CSV (typical GSTN portal export):
       R0:  Title 'Electronic Credit Ledger'
       R2-5: meta (label col 3, value col 4) — GSTIN, Legal Name, From, To
       R6:  main header row (Sr.No, Date, Reference No., Tax Period, ...)
       R7:  sub-header row — Integrated/Central/State/CESS/Total × 2 groups
       R8+: data rows  (Opening Balance, txns, optional Closing Balance)

    Data row columns:
       0: Sr.No
       1: Date
       2: Reference No.
       3: Tax Period
       4: Description
       5: Transaction Type ('Credit'/'Debit'/'-')
       6-10:  Credit/Debit amount  (IGST / CGST / SGST / CESS / Total)
       11-15: Balance after txn   (IGST / CGST / SGST / CESS / Total)
    """
    import csv
    raw_rows = []
    with open(filepath, encoding='utf-8-sig', errors='replace') as f:
        for row in csv.reader(f):
            raw_rows.append(row)

    meta = {
        "source_file": os.path.basename(filepath),
        "gstin": "", "legal_name": "",
        "from_date": "", "to_date": "",
        "state_code": "", "state_name": "",
    }
    # Meta rows: try label cols (3 then 6) and value in next non-empty cell
    for r in raw_rows[:10]:
        for li in (3, 6):
            if li >= len(r): continue
            label = (r[li] or "").strip()
            if not label: continue
            # value in li+1 (or li+2 if li+1 blank)
            value = ""
            for vi in (li + 1, li + 2):
                if vi < len(r) and r[vi] and r[vi].strip():
                    value = r[vi].strip(); break
            if not value: continue
            ll = label.lower()
            if label == "GSTIN" and not meta["gstin"]:
                meta["gstin"] = value
            elif ("legal" in ll or "name" in ll) and not meta["legal_name"]:
                meta["legal_name"] = value
            elif label == "From" and not meta["from_date"]:
                meta["from_date"] = value
            elif label == "To" and not meta["to_date"]:
                meta["to_date"] = value
    if meta["gstin"]:
        meta["state_code"] = meta["gstin"][:2]
        meta["state_name"] = _STATE_CODES.get(meta["state_code"], "Unknown")

    # Find Sr.No header row to locate data start
    hdr_idx = None
    for i, r in enumerate(raw_rows):
        if r and r[0] and str(r[0]).strip().startswith("Sr.No"):
            hdr_idx = i; break
    if hdr_idx is None:
        return {"meta": meta, "rows": []}
    data_start = hdr_idx + 2   # skip sub-header row

    rows = []
    for r in raw_rows[data_start:]:
        if not r or not r[0] or not str(r[0]).strip():
            continue

        def col(i):
            return r[i].strip() if i < len(r) and r[i] else ""

        sr_no = col(0)
        date = col(1)
        ref_no = col(2)
        tax_period = col(3)
        desc = col(4)
        ttype = col(5)
        amt_igst = _ledger_parse_num(col(6))
        amt_cgst = _ledger_parse_num(col(7))
        amt_sgst = _ledger_parse_num(col(8))
        amt_cess = _ledger_parse_num(col(9))
        amt_total = _ledger_parse_num(col(10))
        bal_igst = _ledger_parse_num(col(11))
        bal_cgst = _ledger_parse_num(col(12))
        bal_sgst = _ledger_parse_num(col(13))
        bal_cess = _ledger_parse_num(col(14))
        bal_total = _ledger_parse_num(col(15))

        # Row classification
        dlow = desc.lower()
        if "opening balance" in dlow: row_type = "Opening"
        elif "closing balance" in dlow: row_type = "Closing"
        elif ttype.lower() == "credit": row_type = "Credit"
        elif ttype.lower() == "debit":  row_type = "Debit"
        else: row_type = "Other"

        rows.append({
            "sr_no": sr_no, "date": date, "ref_no": ref_no,
            "tax_period": tax_period, "desc": desc, "ttype": ttype,
            "row_type": row_type,
            "amt_igst": amt_igst, "amt_cgst": amt_cgst,
            "amt_sgst": amt_sgst, "amt_cess": amt_cess,
            "amt_total": amt_total,
            "bal_igst": bal_igst, "bal_cgst": bal_cgst,
            "bal_sgst": bal_sgst, "bal_cess": bal_cess,
            "bal_total": bal_total,
        })

    return {"meta": meta, "rows": rows}


def _ledger_parse_date(s):
    """Parse DD/MM/YYYY → date object, or None on failure."""
    from datetime import datetime
    if not s: return None
    s = str(s).strip()
    if not s or s == '-': return None
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y"):
        try: return datetime.strptime(s, fmt).date()
        except ValueError: pass
    return None


def _ledger_txn_category(row):
    """Classify a ledger transaction. Returns 'Regular' or 'Extra'.

    Regular: standard GSTR-3B activity — ITC accrual, output discharge, RCM.
    Extra:   non-3B events — voluntary payments (DRC-03), refund claims,
             demand orders (DRC-07/01), recovery, anything not part of the
             regular monthly filing cycle.

    Used to keep Monthly Balances clean (only regular 3B impact per period)
    while surfacing extras in a dedicated 'Extra Payments' sheet.
    """
    desc = (row.get('desc') or '').lower().strip()
    if not desc or desc == '-':
        return 'Regular'
    extra_patterns = [
        'voluntary payment',
        'refund claimed',
        'refund from',
        'drc-03', 'drc 03', 'drc03',
        'drc-07', 'drc 07', 'drc07',
        'drc-01', 'drc 01', 'drc01',
        'on account of order',
        'demand', 'recovery',
    ]
    for pattern in extra_patterns:
        if pattern in desc:
            return 'Extra'
    return 'Regular'


def _ledger_txn_subcategory(row):
    """For 'Extra' category, identify subtype for Extra Payments sheet."""
    desc = (row.get('desc') or '').lower().strip()
    if 'voluntary' in desc: return 'Voluntary Payment'
    if 'refund' in desc:    return 'Refund Claim'
    if 'drc' in desc:       return 'DRC Payment'
    if 'demand' in desc:    return 'Demand Payment'
    if 'recovery' in desc:  return 'Recovery'
    if 'order' in desc:     return 'Order-based Payment'
    return 'Other Extra'


def dedup_ledger_files(all_data):
    """Merge multiple CSVs of the same GSTIN (overlapping date ranges) into
    one logical entry. Dedup transactions by Reference No.

    Use-case: user downloads Credit Ledger as "Apr-Feb" + "Feb-Mar" (because
    GSTN portal date range limit). Feb transactions appear in both files.
    Without dedup, they get double-counted.

    Returns (deduped_list, stats_list).
    stats: per-GSTIN dict with files_merged, duplicates_skipped, unique_txns.
    """
    by_gstin = {}
    for d in all_data:
        gstin = d['meta'].get('gstin', '')
        if not gstin:
            # Files without GSTIN — keep as standalone, can't dedup
            by_gstin.setdefault(f"_NO_GSTIN_{len(by_gstin)}", []).append(d)
        else:
            by_gstin.setdefault(gstin, []).append(d)

    deduped = []
    stats = []

    for gstin, files in by_gstin.items():
        if len(files) == 1:
            deduped.append(files[0])
            continue

        # Multiple files for same GSTIN — merge them
        def _sort_by_from(f):
            d = _ledger_parse_date(f['meta'].get('from_date', ''))
            return d if d else _ledger_parse_date('01/01/2000')
        files_sorted = sorted(files, key=_sort_by_from)

        merged_meta = dict(files_sorted[0]['meta'])
        # Widen date range to cover all files
        from_dates = [_ledger_parse_date(f['meta'].get('from_date', ''))
                      for f in files_sorted]
        to_dates = [_ledger_parse_date(f['meta'].get('to_date', ''))
                    for f in files_sorted]
        from_dates = [d for d in from_dates if d]
        to_dates = [d for d in to_dates if d]
        if from_dates:
            merged_meta['from_date'] = min(from_dates).strftime('%d/%m/%Y')
        if to_dates:
            merged_meta['to_date'] = max(to_dates).strftime('%d/%m/%Y')
        merged_meta['source_file'] = ' + '.join(
            f['meta'].get('source_file', '?') for f in files_sorted)

        # Dedup rows by Reference No within Credit/Debit txns.
        # Opening Balance: keep from EARLIEST file (true period start).
        # Closing Balance: keep from LATEST file (true period end).
        opening_row = None
        latest_closing = None
        seen_refs = set()
        merged_txns = []
        dup_count = 0

        for f in files_sorted:
            for row in f['rows']:
                if row['row_type'] == 'Opening':
                    if opening_row is None:
                        opening_row = row
                elif row['row_type'] == 'Closing':
                    latest_closing = row
                else:
                    ref = (row.get('ref_no') or '').strip()
                    if ref and ref != '-':
                        if ref in seen_refs:
                            dup_count += 1
                            continue
                        seen_refs.add(ref)
                    merged_txns.append(row)

        merged_rows = []
        if opening_row: merged_rows.append(opening_row)
        merged_rows.extend(merged_txns)
        if latest_closing: merged_rows.append(latest_closing)

        deduped.append({'meta': merged_meta, 'rows': merged_rows})
        stats.append({
            'gstin': gstin,
            'state_name': merged_meta.get('state_name', '?'),
            'files_merged': len(files_sorted),
            'duplicates_skipped': dup_count,
            'unique_txns': len(merged_txns),
        })

    return deduped, stats


def _tp_sort_key(tp):
    """Sort 'Apr-25' chronologically. Returns (year, month) tuple.
    Unknown/blank values sort last."""
    if not tp: return (9999, 99)
    s = str(tp).strip()
    if not s or s in ('-', '—', '(no tax period)'): return (9999, 99)
    months_map = {"Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
                  "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12}
    parts = s.replace('/', '-').split('-')
    if len(parts) != 2: return (9999, 99)
    m = parts[0].strip()[:3].capitalize()
    yy = parts[1].strip()
    if m not in months_map: return (9999, 99)
    try:
        full_year = 2000 + int(yy) if len(yy) <= 2 else int(yy)
    except ValueError:
        return (9999, 99)
    return (full_year, months_map[m])


def compute_tax_period_balances(data, ledger_type='ecl', regular_only=True):
    """For each Tax Period, compute opening + closing balance.

    regular_only=True (DEFAULT): excludes Voluntary/Refund/DRC/etc Extra
                                 payments — gives clean monthly view of pure
                                 GSTR-3B activity.
    regular_only=False: includes all txns (legacy behaviour).

    Opening of Tax Period X = balance just BEFORE the FIRST regular txn
                              tagged with Tax Period = X (chronological).
    Closing of Tax Period X = balance AFTER the LAST regular such txn.
    """
    from datetime import datetime
    rows = data["rows"]
    meta = data["meta"]

    def _extract_bal(row):
        if not row:
            return {"igst": 0, "cgst": 0, "sgst": 0, "cess": 0, "total": 0}
        if ledger_type == 'ecl':
            return {
                "igst":  row["bal_igst"]  or 0,
                "cgst":  row["bal_cgst"]  or 0,
                "sgst":  row["bal_sgst"]  or 0,
                "cess":  row["bal_cess"]  or 0,
                "total": row["bal_total"] or 0,
            }
        else:  # cash
            bal = {
                "igst": (row["bal_igst"]["total"] or 0)
                        if isinstance(row["bal_igst"], dict) else 0,
                "cgst": (row["bal_cgst"]["total"] or 0)
                        if isinstance(row["bal_cgst"], dict) else 0,
                "sgst": (row["bal_sgst"]["total"] or 0)
                        if isinstance(row["bal_sgst"], dict) else 0,
                "cess": (row["bal_cess"]["total"] or 0)
                        if isinstance(row["bal_cess"], dict) else 0,
            }
            bal["total"] = bal["igst"] + bal["cgst"] + bal["sgst"] + bal["cess"]
            return bal

    opening_row = next((r for r in rows if r["row_type"] == "Opening"), None)
    initial_bal = _extract_bal(opening_row)

    # Sort ALL non-meta txns chronologically (need full sequence for
    # "balance just before" calculations even when filtering to Regular)
    full_seq = []
    for i, r in enumerate(rows):
        if r["row_type"] not in ("Credit", "Debit"): continue
        d = _ledger_parse_date(r["date"])
        sort_d = d if d else _ledger_parse_date(meta.get("from_date"))
        if not sort_d:
            sort_d = datetime.min.date()
        full_seq.append((sort_d, i, r))
    full_seq.sort(key=lambda x: (x[0], x[1]))

    # bal_after_full_seq[k] = balance after the k-th chronological txn
    bal_after_full = [_extract_bal(r) for (_, _, r) in full_seq]

    # Filter to category if needed; track original full-sequence index
    filtered_indices = []
    for k, (_, _, r) in enumerate(full_seq):
        if regular_only and _ledger_txn_category(r) != 'Regular':
            continue
        filtered_indices.append(k)

    # Group filtered indices by Tax Period
    by_tp = {}
    for k in filtered_indices:
        row = full_seq[k][2]
        tp = (row.get("tax_period", "") or "").strip()
        if not tp or tp == '-':
            tp = "(no tax period)"
        by_tp.setdefault(tp, []).append(k)

    results = []
    for tp, k_list in by_tp.items():
        first_k = min(k_list)
        last_k = max(k_list)
        # Opening = balance just BEFORE first_k. That's bal_after of (first_k-1)
        # in the FULL chronological sequence (even if it was an Extra event).
        # This matches the actual ledger state at that moment.
        opening = (initial_bal if first_k == 0 else bal_after_full[first_k - 1])
        closing = bal_after_full[last_k]
        results.append({
            "tax_period": tp,
            "sort_key": _tp_sort_key(tp),
            "opening": dict(opening),
            "closing": dict(closing),
            "txn_count": len(k_list),
        })
    results.sort(key=lambda x: x["sort_key"])
    return results


def compute_extra_payments(data, ledger_type='ecl'):
    """Return list of extra-payment rows (voluntary, refund, DRC, etc.)
    with full context for the 'Extra Payments' summary sheet.

    Each extra row also includes 'effect_tax_period': the Tax Period of the
    NEXT regular 3B transaction (chronologically) after this extra event.
    This is the period where a gap will appear when comparing the user's
    internal monthly 3B working against the actual ECL — because the next
    regular 3B filing's opening balance will reflect this extra payment.
    """
    from datetime import datetime
    rows = data["rows"]

    # Build chronological index of ALL real transactions (Regular + Extra)
    # so we can look forward to find the next Regular after each Extra.
    indexed = []
    for i, r in enumerate(rows):
        if r["row_type"] not in ("Credit", "Debit"): continue
        d = _ledger_parse_date(r.get("date", ""))
        sort_d = d if d else datetime.min.date()
        indexed.append((sort_d, i, r))
    indexed.sort(key=lambda x: (x[0], x[1]))

    extras = []
    for k, (_, _, r) in enumerate(indexed):
        if _ledger_txn_category(r) != 'Extra': continue

        # Find next REGULAR txn AFTER this extra event → its tax period is
        # the one where the gap will manifest in user-vs-ECL comparison.
        effect_tp = ""
        for j in range(k + 1, len(indexed)):
            next_r = indexed[j][2]
            if _ledger_txn_category(next_r) == 'Regular':
                effect_tp = (next_r.get("tax_period", "") or "").strip()
                break
        # If no future regular txn exists (extra is the last activity), mark
        # as 'next period' — affects the period after the latest one filed
        if not effect_tp:
            effect_tp = "(no future 3B filed)"

        if ledger_type == 'ecl':
            amt_breakdown = {
                "igst": r["amt_igst"] or 0,
                "cgst": r["amt_cgst"] or 0,
                "sgst": r["amt_sgst"] or 0,
                "cess": r["amt_cess"] or 0,
                "total": r["amt_total"] or 0,
            }
            bal_total = r["bal_total"] or 0
        else:
            amt_breakdown = {
                "igst": (r["amt_igst"]["total"] or 0)
                        if isinstance(r["amt_igst"], dict) else 0,
                "cgst": (r["amt_cgst"]["total"] or 0)
                        if isinstance(r["amt_cgst"], dict) else 0,
                "sgst": (r["amt_sgst"]["total"] or 0)
                        if isinstance(r["amt_sgst"], dict) else 0,
                "cess": (r["amt_cess"]["total"] or 0)
                        if isinstance(r["amt_cess"], dict) else 0,
            }
            amt_breakdown["total"] = (amt_breakdown["igst"] + amt_breakdown["cgst"]
                                       + amt_breakdown["sgst"] + amt_breakdown["cess"])
            bal_total = ((r["bal_igst"]["total"] or 0) if isinstance(r["bal_igst"], dict) else 0) \
                        + ((r["bal_cgst"]["total"] or 0) if isinstance(r["bal_cgst"], dict) else 0) \
                        + ((r["bal_sgst"]["total"] or 0) if isinstance(r["bal_sgst"], dict) else 0) \
                        + ((r["bal_cess"]["total"] or 0) if isinstance(r["bal_cess"], dict) else 0)
        extras.append({
            "date": r.get("date", ""),
            "txn_month": _ledger_txn_month_label(r.get("date", "")),
            "ref_no": r.get("ref_no", ""),
            "tax_period": r.get("tax_period", ""),
            "effect_tax_period": effect_tp,
            "desc": r.get("desc", ""),
            "ttype": r.get("ttype", ""),
            "subcategory": _ledger_txn_subcategory(r),
            "amt": amt_breakdown,
            "bal_total_after": bal_total,
        })
    return extras


def _ledger_txn_month_label(date_str):
    """Convert a DD/MM/YYYY date string into 'Apr-25' style month label.
    Used to add a 'Txn Month' column in Extra Payments sheet, mirroring the
    Tax Period format but derived from the actual transaction Date column."""
    d = _ledger_parse_date(date_str)
    if not d: return ""
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    return f"{months[d.month - 1]}-{d.year % 100:02d}"


def _ledger_month_label(d):
    """[deprecated for monthly balances — kept for back-compat] Format 'Apr-25'."""
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    return f"{months[d.month - 1]}-{d.year % 100:02d}"


def _ledger_add_months(d, n=1):
    """[deprecated for monthly balances — kept for back-compat]."""
    from datetime import date
    total = d.year * 12 + (d.month - 1) + n
    new_year, new_month = divmod(total, 12)
    return date(new_year, new_month + 1, 1)


def compute_monthly_balances(data, ledger_type='ecl'):
    """[Back-compat wrapper] Now delegates to compute_tax_period_balances
    with regular_only=True (clean monthly view, excludes voluntary/refund)."""
    results = compute_tax_period_balances(data, ledger_type, regular_only=True)
    return [{"month": r["tax_period"],
             "year": r["sort_key"][0],
             "month_num": r["sort_key"][1],
             "sort_key": r["sort_key"][0] * 12 + r["sort_key"][1],
             "opening": r["opening"],
             "closing": r["closing"],
             "txn_count": r["txn_count"]}
            for r in results]


def write_consolidated_ecl(all_data, out_path):
    """Build consolidated workbook for Electronic Credit Ledger data.
    Returns count of GSTINs."""
    import openpyxl
    from openpyxl.styles import (Font as XF, PatternFill, Alignment,
                                  Border, Side)
    from openpyxl.utils import get_column_letter

    HDR_FILL = PatternFill("solid", start_color="1F4E78")
    SUB_FILL = PatternFill("solid", start_color="2E75B6")
    META_FILL = PatternFill("solid", start_color="F2F2F2")
    AMT_FILL = PatternFill("solid", start_color="E2EFDA")
    BAL_FILL = PatternFill("solid", start_color="FFF2CC")
    OPEN_FILL = PatternFill("solid", start_color="E7E6E6")
    CLOSE_FILL = PatternFill("solid", start_color="FFE699")
    TOT_FILL = PatternFill("solid", start_color="D9E1F2")

    WHITE_B = XF(name="Calibri", bold=True, color="FFFFFF", size=10)
    BOLD = XF(name="Calibri", bold=True, size=10)
    REG = XF(name="Calibri", size=10)
    NEG_FONT = XF(name="Calibri", size=10, color="C00000")
    TITLE = XF(name="Calibri", bold=True, color="FFFFFF", size=14)
    SUBTITLE = XF(name="Calibri", italic=True, size=10, color="404040")

    thin = Side(border_style="thin", color="B4B4B4")
    BORDER_ALL = Border(left=thin, right=thin, top=thin, bottom=thin)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
    RIGHT = Alignment(horizontal="right", vertical="center")
    NUM_FMT = '#,##0.00;[Red](#,##0.00);"-"'

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ─── Cover ────────────────────────────────────────────
    ws = wb.create_sheet("Cover")
    ws.merge_cells("A1:H1")
    ws["A1"] = "Electronic Credit Ledger — Multi-GSTIN Consolidated"
    ws["A1"].font = TITLE; ws["A1"].fill = HDR_FILL
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A2:H2")
    ws["A2"] = ("ITC Credit / Debit / Closing Balance per GSTIN  ·  "
                "Console + Summary + per-GSTIN detail")
    ws["A2"].font = SUBTITLE; ws["A2"].alignment = CENTER

    hdrs = ["S.No.", "GSTIN", "State Code", "State Name", "Legal Name",
            "From", "To", "Source File"]
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = WHITE_B; cell.fill = SUB_FILL
        cell.alignment = CENTER; cell.border = BORDER_ALL
    ws.row_dimensions[4].height = 22

    for i, data in enumerate(all_data, 1):
        m = data["meta"]
        for c, v in enumerate([i, m["gstin"], m["state_code"], m["state_name"],
                                m["legal_name"], m["from_date"], m["to_date"],
                                m["source_file"]], 1):
            cell = ws.cell(row=4 + i, column=c, value=v)
            cell.font = REG; cell.border = BORDER_ALL
            cell.alignment = LEFT if c >= 4 else CENTER
            if c == 1: cell.number_format = "0"

    for col, w in zip("ABCDEFGH", [6, 18, 7, 18, 28, 12, 12, 42]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A5"

    # ─── Monthly Balances — per GSTIN × per Tax Period ────
    # Walks every GSTIN's transactions, groups by Tax Period column (NOT
    # calendar month — important for GST workflow). Opening of Tax Period X
    # = balance just before the first transaction tagged with Tax Period X
    # in chronological (Date) order. Example: Apr-25 opening = balance just
    # before the first Apr-25 entry (typically dated 20-May when GSTR-3B for
    # Apr-25 was filed) — i.e. balance after any preceding May entries.
    ws = wb.create_sheet("Monthly Balances")
    cols = ["S.No.", "State Code", "State Name", "GSTIN", "Legal Name",
            "Tax Period", "FY Year", "Txn Count",
            "Opening IGST", "Opening CGST", "Opening SGST/UTGST",
            "Opening CESS", "Opening Total",
            "Closing IGST", "Closing CGST", "Closing SGST/UTGST",
            "Closing CESS", "Closing Total",
            "Net Change"]
    ncols_m = len(cols)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols_m)
    ws["A1"] = ("Monthly Balances — Opening + Closing per GSTIN, per Tax Period   "
                "(opening = balance just before first txn of that tax period)")
    ws["A1"].font = TITLE; ws["A1"].fill = HDR_FILL
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 28

    def banner(c1, c2, label, fill):
        ws.merge_cells(start_row=2, start_column=c1, end_row=2, end_column=c2)
        cell = ws.cell(row=2, column=c1, value=label)
        cell.font = BOLD; cell.fill = fill
        cell.alignment = CENTER; cell.border = BORDER_ALL
    banner(1, 8, "Metadata", META_FILL)
    banner(9, 13, "Opening Balance (before first txn of tax period)", AMT_FILL)
    banner(14, 18, "Closing Balance (after last txn of tax period)", BAL_FILL)
    banner(19, 19, "Δ", META_FILL)
    ws.row_dimensions[2].height = 22

    for c, h in enumerate(cols, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = WHITE_B; cell.fill = SUB_FILL
        cell.alignment = CENTER; cell.border = BORDER_ALL
    ws.row_dimensions[3].height = 36

    GAIN_FILL = PatternFill("solid", start_color="C6EFCE")  # green
    LOSS_FILL = PatternFill("solid", start_color="FFC7CE")  # red

    r = 4; sn = 0
    for data in all_data:
        m = data["meta"]
        monthly_list = compute_monthly_balances(data, ledger_type='ecl')
        for mb in monthly_list:
            sn += 1
            opg, clo = mb["opening"], mb["closing"]
            net_change = (clo["total"] or 0) - (opg["total"] or 0)
            vals = [
                sn, m["state_code"], m["state_name"], m["gstin"], m["legal_name"],
                mb["month"], mb["year"], mb["txn_count"],
                opg["igst"], opg["cgst"], opg["sgst"], opg["cess"], opg["total"],
                clo["igst"], clo["cgst"], clo["sgst"], clo["cess"], clo["total"],
                net_change,
            ]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.font = REG; cell.border = BORDER_ALL
                if c == 1:
                    cell.alignment = CENTER; cell.number_format = "0"
                elif c in (2, 4, 6, 7, 8):
                    cell.alignment = CENTER
                elif c in (3, 5):
                    cell.alignment = LEFT
                elif isinstance(v, (int, float)) and c >= 9:
                    cell.alignment = RIGHT; cell.number_format = NUM_FMT
                    if v < 0: cell.font = NEG_FONT
                # Colour the Net Change cell green/red
                if c == 19 and isinstance(v, (int, float)) and v != 0:
                    cell.fill = GAIN_FILL if v > 0 else LOSS_FILL
                    cell.font = BOLD if v > 0 else NEG_FONT
            r += 1

    widths_m = [6, 7, 16, 18, 22,
                10, 7, 8,
                13, 13, 15, 11, 13,
                13, 13, 15, 11, 13,
                14]
    for i, w in enumerate(widths_m, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "I4"
    if r > 4:
        ws.auto_filter.ref = f"A3:{get_column_letter(ncols_m)}{r-1}"

    # ─── Extra Payments sheet (Voluntary / Refund / DRC / etc.) ──
    # Tracks non-regular-3B events that affected the credit ledger.
    # These are EXCLUDED from Monthly Balances (which shows only regular
    # GSTR-3B activity per tax period). Listing them here gives full audit
    # trail of voluntary payments, refund claims, DRC payments, etc.
    #
    # Three "period" dimensions per extra payment:
    #   • Txn Month  = calendar month of the payment (Date column)
    #   • Tax Period = period the payment is FOR (relates to)
    #   • Effect TP  = period whose 3B comparison will show a gap (the next
    #                  regular 3B filing's tax period — useful when comparing
    #                  user's internal 3B working vs the actual ECL balance)
    ws = wb.create_sheet("Extra Payments")
    cols_e = ["S.No.", "State Code", "State Name", "GSTIN", "Legal Name",
              "Date", "Txn Month", "Reference No.", "Tax Period", "Effect TP",
              "Subcategory", "Description", "Txn Type",
              "Amt IGST", "Amt CGST", "Amt SGST/UTGST", "Amt CESS", "Amt Total",
              "Bal Total After"]
    ncols_e = len(cols_e)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols_e)
    ws["A1"] = ("Extra Payments — Voluntary / Refund / DRC / Demand / Recovery   "
                "(Effect TP = the tax period whose 3B comparison will show a "
                "gap from this extra payment)")
    ws["A1"].font = TITLE; ws["A1"].fill = HDR_FILL
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 28

    SUBCAT_FILL = PatternFill("solid", start_color="FFE6CC")     # orange
    TXN_MONTH_FILL = PatternFill("solid", start_color="D9E1F2")  # blue
    EFFECT_TP_FILL = PatternFill("solid", start_color="C6EFCE")  # green
    for c, h in enumerate(cols_e, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = WHITE_B; cell.fill = SUB_FILL
        cell.alignment = CENTER; cell.border = BORDER_ALL
    ws.row_dimensions[3].height = 30

    r = 4; sn = 0
    subcategory_totals = {}
    txn_month_totals = {}
    effect_tp_totals = {}
    for data in all_data:
        m = data["meta"]
        extras = compute_extra_payments(data, ledger_type='ecl')
        for ex in extras:
            sn += 1
            sub = ex["subcategory"]
            txm = ex["txn_month"] or "(unknown)"
            etp = ex["effect_tax_period"] or "(unknown)"
            subcategory_totals.setdefault(sub, {"count": 0, "total": 0})
            subcategory_totals[sub]["count"] += 1
            subcategory_totals[sub]["total"] += ex["amt"]["total"] or 0
            txn_month_totals.setdefault(txm, {"count": 0, "total": 0})
            txn_month_totals[txm]["count"] += 1
            txn_month_totals[txm]["total"] += ex["amt"]["total"] or 0
            effect_tp_totals.setdefault(etp, {"count": 0, "total": 0})
            effect_tp_totals[etp]["count"] += 1
            effect_tp_totals[etp]["total"] += ex["amt"]["total"] or 0
            vals = [
                sn, m["state_code"], m["state_name"], m["gstin"], m["legal_name"],
                ex["date"], ex["txn_month"], ex["ref_no"], ex["tax_period"],
                ex["effect_tax_period"], sub,
                ex["desc"], ex["ttype"],
                ex["amt"]["igst"], ex["amt"]["cgst"],
                ex["amt"]["sgst"], ex["amt"]["cess"], ex["amt"]["total"],
                ex["bal_total_after"],
            ]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.font = REG; cell.border = BORDER_ALL
                if c == 1:
                    cell.alignment = CENTER; cell.number_format = "0"
                elif c in (2, 4, 6, 8, 9, 13):
                    cell.alignment = CENTER
                elif c == 7:   # Txn Month
                    cell.alignment = CENTER
                    cell.fill = TXN_MONTH_FILL
                    cell.font = BOLD
                elif c == 10:  # Effect TP — gap indicator
                    cell.alignment = CENTER
                    cell.fill = EFFECT_TP_FILL
                    cell.font = BOLD
                elif c == 11:  # Subcategory
                    cell.alignment = CENTER
                    cell.fill = SUBCAT_FILL
                    cell.font = BOLD
                elif c in (3, 5, 12):
                    cell.alignment = LEFT
                elif isinstance(v, (int, float)) and c >= 14:
                    cell.alignment = RIGHT; cell.number_format = NUM_FMT
                    if v < 0: cell.font = NEG_FONT
                else:
                    cell.alignment = LEFT
            r += 1

    # Footer — THREE summary tables side by side
    if subcategory_totals or txn_month_totals or effect_tp_totals:
        r += 1
        # Row of section titles
        ws.cell(row=r, column=1, value="SUMMARY BY SUBCATEGORY").font = BOLD
        ws.cell(row=r, column=1).alignment = LEFT
        ws.cell(row=r, column=6, value="SUMMARY BY EFFECT TP (3B gap)").font = BOLD
        ws.cell(row=r, column=6).alignment = LEFT
        ws.cell(row=r, column=11, value="SUMMARY BY TXN MONTH").font = BOLD
        ws.cell(row=r, column=11).alignment = LEFT
        r += 1
        # Column header row
        for c, h in enumerate(["Subcategory", "# Events", "Total Amount"], 1):
            cell = ws.cell(row=r, column=c, value=h)
            cell.font = WHITE_B; cell.fill = SUB_FILL
            cell.alignment = CENTER; cell.border = BORDER_ALL
        for c, h in enumerate(["Effect TP", "# Events", "Total Amount"], 6):
            cell = ws.cell(row=r, column=c, value=h)
            cell.font = WHITE_B; cell.fill = SUB_FILL
            cell.alignment = CENTER; cell.border = BORDER_ALL
        for c, h in enumerate(["Txn Month", "# Events", "Total Amount"], 11):
            cell = ws.cell(row=r, column=c, value=h)
            cell.font = WHITE_B; cell.fill = SUB_FILL
            cell.alignment = CENTER; cell.border = BORDER_ALL
        r += 1
        subcat_sorted = sorted(subcategory_totals.items())
        etp_sorted = sorted(effect_tp_totals.items(),
                            key=lambda x: _tp_sort_key(x[0]))
        txm_sorted = sorted(txn_month_totals.items(),
                            key=lambda x: _tp_sort_key(x[0]))
        max_rows = max(len(subcat_sorted), len(etp_sorted), len(txm_sorted))
        for i in range(max_rows):
            if i < len(subcat_sorted):
                k, vs = subcat_sorted[i]
                ws.cell(row=r, column=1, value=k).font = BOLD
                ws.cell(row=r, column=1).fill = SUBCAT_FILL
                ws.cell(row=r, column=1).border = BORDER_ALL
                ws.cell(row=r, column=1).alignment = CENTER
                ws.cell(row=r, column=2, value=vs["count"]).alignment = CENTER
                ws.cell(row=r, column=2).border = BORDER_ALL
                ws.cell(row=r, column=3, value=vs["total"]).number_format = NUM_FMT
                ws.cell(row=r, column=3).alignment = RIGHT
                ws.cell(row=r, column=3).border = BORDER_ALL
            if i < len(etp_sorted):
                k, vs = etp_sorted[i]
                ws.cell(row=r, column=6, value=k).font = BOLD
                ws.cell(row=r, column=6).fill = EFFECT_TP_FILL
                ws.cell(row=r, column=6).border = BORDER_ALL
                ws.cell(row=r, column=6).alignment = CENTER
                ws.cell(row=r, column=7, value=vs["count"]).alignment = CENTER
                ws.cell(row=r, column=7).border = BORDER_ALL
                ws.cell(row=r, column=8, value=vs["total"]).number_format = NUM_FMT
                ws.cell(row=r, column=8).alignment = RIGHT
                ws.cell(row=r, column=8).border = BORDER_ALL
            if i < len(txm_sorted):
                k, vs = txm_sorted[i]
                ws.cell(row=r, column=11, value=k).font = BOLD
                ws.cell(row=r, column=11).fill = TXN_MONTH_FILL
                ws.cell(row=r, column=11).border = BORDER_ALL
                ws.cell(row=r, column=11).alignment = CENTER
                ws.cell(row=r, column=12, value=vs["count"]).alignment = CENTER
                ws.cell(row=r, column=12).border = BORDER_ALL
                ws.cell(row=r, column=13, value=vs["total"]).number_format = NUM_FMT
                ws.cell(row=r, column=13).alignment = RIGHT
                ws.cell(row=r, column=13).border = BORDER_ALL
            r += 1

    widths_e = [6, 7, 16, 18, 22,
                11, 10, 18, 10, 10, 18, 30, 8,
                13, 13, 15, 11, 13, 16]
    for i, w in enumerate(widths_e, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "I4"
    if sn > 0:
        ws.auto_filter.ref = f"A3:{get_column_letter(ncols_e)}{3 + sn}"

    # ─── Console (long format, every row from every GSTIN) ─
    ws = wb.create_sheet("Console")
    cols = ["S.No.", "GSTIN", "State Code", "State Name", "Legal Name",
            "Row Type", "Sr.No.", "Date", "Reference No.", "Tax Period",
            "Description", "Transaction Type",
            "Amt IGST", "Amt CGST", "Amt SGST/UTGST", "Amt CESS", "Amt Total",
            "Bal IGST", "Bal CGST", "Bal SGST/UTGST", "Bal CESS", "Bal Total"]
    ncols = len(cols)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws["A1"] = ("Console — every row, every GSTIN  "
                "(filter by State / GSTIN / Row Type / Tax Period)")
    ws["A1"].font = TITLE; ws["A1"].fill = HDR_FILL
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 28

    def banner(c1, c2, label, fill):
        ws.merge_cells(start_row=2, start_column=c1, end_row=2, end_column=c2)
        cell = ws.cell(row=2, column=c1, value=label)
        cell.font = BOLD; cell.fill = fill
        cell.alignment = CENTER; cell.border = BORDER_ALL

    banner(1, 12, "Metadata", META_FILL)
    banner(13, 17, "Amount (Credit/Debit)", AMT_FILL)
    banner(18, 22, "Balance after transaction", BAL_FILL)
    ws.row_dimensions[2].height = 22

    for c, h in enumerate(cols, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = WHITE_B; cell.fill = SUB_FILL
        cell.alignment = CENTER; cell.border = BORDER_ALL
    ws.row_dimensions[3].height = 36

    r = 4; sn = 0
    for data in all_data:
        m = data["meta"]
        for row in data["rows"]:
            sn += 1
            vals = [
                sn, m["gstin"], m["state_code"], m["state_name"],
                m["legal_name"],
                row["row_type"], row["sr_no"], row["date"], row["ref_no"],
                row["tax_period"], row["desc"], row["ttype"],
                row["amt_igst"], row["amt_cgst"], row["amt_sgst"],
                row["amt_cess"], row["amt_total"],
                row["bal_igst"], row["bal_cgst"], row["bal_sgst"],
                row["bal_cess"], row["bal_total"],
            ]
            row_bold = row["row_type"] in ("Opening", "Closing")
            row_fill = (OPEN_FILL if row["row_type"] == "Opening"
                        else CLOSE_FILL if row["row_type"] == "Closing"
                        else None)
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.border = BORDER_ALL
                if row_fill: cell.fill = row_fill
                if c == 1:
                    cell.alignment = CENTER; cell.number_format = "0"
                    cell.font = BOLD if row_bold else REG
                elif c in (2, 3, 6, 7, 8, 10, 12):
                    cell.alignment = CENTER
                    cell.font = BOLD if row_bold else REG
                elif c in (4, 5, 9, 11):
                    cell.alignment = LEFT
                    cell.font = BOLD if row_bold else REG
                elif isinstance(v, (int, float)) and c >= 13:
                    cell.alignment = RIGHT; cell.number_format = NUM_FMT
                    if v < 0: cell.font = NEG_FONT
                    else: cell.font = BOLD if row_bold else REG
                else:
                    cell.alignment = LEFT
                    cell.font = BOLD if row_bold else REG
            r += 1

    widths = [6, 18, 6, 16, 22,
              9, 6, 11, 18, 10, 28, 11,
              13, 13, 15, 11, 13,
              13, 13, 15, 11, 13]
    for i, w in enumerate(widths[:ncols], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "M4"
    if r > 4:
        ws.auto_filter.ref = f"A3:{get_column_letter(ncols)}{r-1}"

    # ─── Summary — per-GSTIN totals ───────────────────────
    ws = wb.create_sheet("Summary")
    ws.merge_cells("A1:N1")
    ws["A1"] = "Summary — Total Credit / Total Debit / Closing Balance per GSTIN"
    ws["A1"].font = TITLE; ws["A1"].fill = HDR_FILL
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 28

    sum_hdrs = ["S.No.", "GSTIN", "State", "Legal Name", "From", "To",
                "Txn Count",
                "Total Credit IGST", "Total Credit CGST+SGST", "Total Credit CESS",
                "Total Debit IGST", "Total Debit CGST+SGST", "Total Debit CESS",
                "Closing Total"]
    for c, h in enumerate(sum_hdrs, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = WHITE_B; cell.fill = SUB_FILL
        cell.alignment = CENTER; cell.border = BORDER_ALL
    ws.row_dimensions[3].height = 36

    r = 4
    grand_totals = [0] * 7  # cr_i, cr_c, cr_x, dr_i, dr_c, dr_x, closing
    for i, data in enumerate(all_data, 1):
        m = data["meta"]
        txns = [row for row in data["rows"]
                if row["row_type"] in ("Credit", "Debit")]
        closing = next((row for row in data["rows"]
                        if row["row_type"] == "Closing"), None)

        def sum_field(rows, key):
            return sum((row[key] or 0) for row in rows)

        credits = [r2 for r2 in txns if r2["row_type"] == "Credit"]
        debits  = [r2 for r2 in txns if r2["row_type"] == "Debit"]
        cr_i = sum_field(credits, "amt_igst")
        cr_c = sum_field(credits, "amt_cgst") + sum_field(credits, "amt_sgst")
        cr_x = sum_field(credits, "amt_cess")
        dr_i = sum_field(debits, "amt_igst")
        dr_c = sum_field(debits, "amt_cgst") + sum_field(debits, "amt_sgst")
        dr_x = sum_field(debits, "amt_cess")
        clo_total = (closing["bal_total"] if closing else 0) or 0

        for k, val in enumerate([cr_i, cr_c, cr_x, dr_i, dr_c, dr_x, clo_total]):
            grand_totals[k] += val or 0

        vals = [i, m["gstin"], m["state_name"], m["legal_name"],
                m["from_date"], m["to_date"], len(txns),
                cr_i, cr_c, cr_x, dr_i, dr_c, dr_x, clo_total]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = REG; cell.border = BORDER_ALL
            if c == 1:
                cell.alignment = CENTER; cell.number_format = "0"
            elif c in (2, 3, 5, 6, 7):
                cell.alignment = CENTER
            elif c == 4:
                cell.alignment = LEFT
            elif c >= 8 and isinstance(v, (int, float)):
                cell.alignment = RIGHT; cell.number_format = NUM_FMT
                if v < 0: cell.font = NEG_FONT
            else:
                cell.alignment = LEFT
        r += 1

    # Grand total row
    if all_data:
        cell = ws.cell(row=r, column=1, value="TOTAL")
        cell.font = BOLD; cell.fill = TOT_FILL
        cell.alignment = CENTER; cell.border = BORDER_ALL
        for c in (2, 3, 4, 5, 6, 7):
            ws.cell(row=r, column=c).fill = TOT_FILL
            ws.cell(row=r, column=c).border = BORDER_ALL
        for k, val in enumerate(grand_totals):
            cell = ws.cell(row=r, column=8 + k, value=val)
            cell.font = BOLD; cell.fill = TOT_FILL
            cell.alignment = RIGHT
            cell.border = BORDER_ALL; cell.number_format = NUM_FMT
            if val < 0: cell.font = XF(name="Calibri", bold=True, color="C00000", size=10)
        r += 1

    widths_sum = [6, 18, 16, 26, 12, 12, 8, 14, 16, 11, 14, 16, 11, 14]
    for i, w in enumerate(widths_sum, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"

    # ─── Per-GSTIN detail sheets ──────────────────────────
    for data in all_data:
        m = data["meta"]
        sname = (f"{m['state_code']}_{m['gstin'][-4:]}"
                 if m['gstin'] else f"GSTIN_{len(wb.sheetnames)}")
        sname = sname[:31]
        base = sname; idx = 1
        while sname in wb.sheetnames:
            idx += 1; sname = f"{base[:28]}_{idx}"
        ws = wb.create_sheet(sname)

        ws.merge_cells("A1:P1")
        ws["A1"] = f"Credit Ledger — {m['gstin']}  ({m['state_name']})"
        ws["A1"].font = TITLE; ws["A1"].fill = HDR_FILL
        ws["A1"].alignment = CENTER
        ws.row_dimensions[1].height = 26
        ws.merge_cells("A2:P2")
        ws["A2"] = (f"{m['legal_name']}  |  Period: {m['from_date']} to "
                    f"{m['to_date']}")
        ws["A2"].font = SUBTITLE; ws["A2"].alignment = CENTER
        ws.row_dimensions[2].height = 18

        gh = ["S.No.", "Date", "Reference No.", "Tax Period", "Description",
              "Transaction Type"]
        for c, h in enumerate(gh, 1):
            ws.merge_cells(start_row=4, start_column=c, end_row=5, end_column=c)
            cell = ws.cell(row=4, column=c, value=h)
            cell.font = WHITE_B; cell.fill = SUB_FILL
            cell.alignment = CENTER; cell.border = BORDER_ALL

        ws.merge_cells(start_row=4, start_column=7, end_row=4, end_column=11)
        cell = ws.cell(row=4, column=7, value="Amount (Credit/Debit)")
        cell.font = BOLD; cell.fill = AMT_FILL
        cell.alignment = CENTER; cell.border = BORDER_ALL
        ws.merge_cells(start_row=4, start_column=12, end_row=4, end_column=16)
        cell = ws.cell(row=4, column=12, value="Balance after transaction")
        cell.font = BOLD; cell.fill = BAL_FILL
        cell.alignment = CENTER; cell.border = BORDER_ALL

        subs = ["IGST", "CGST", "SGST/UTGST", "CESS", "Total"]
        for gi in range(2):
            for j, sub in enumerate(subs):
                col = 7 + gi * 5 + j
                cell = ws.cell(row=5, column=col, value=sub)
                cell.font = WHITE_B; cell.fill = SUB_FILL
                cell.alignment = CENTER; cell.border = BORDER_ALL
        ws.row_dimensions[4].height = 22
        ws.row_dimensions[5].height = 22

        r = 6
        for row in data["rows"]:
            is_balance = row["row_type"] in ("Opening", "Closing")
            row_fill = (OPEN_FILL if row["row_type"] == "Opening"
                        else CLOSE_FILL if row["row_type"] == "Closing"
                        else None)
            vals = [
                row["sr_no"], row["date"], row["ref_no"], row["tax_period"],
                row["desc"], row["ttype"],
                row["amt_igst"], row["amt_cgst"], row["amt_sgst"],
                row["amt_cess"], row["amt_total"],
                row["bal_igst"], row["bal_cgst"], row["bal_sgst"],
                row["bal_cess"], row["bal_total"],
            ]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.border = BORDER_ALL
                if row_fill: cell.fill = row_fill
                if c <= 6:
                    cell.alignment = CENTER if c != 5 else LEFT
                    cell.font = BOLD if is_balance else REG
                elif isinstance(v, (int, float)):
                    cell.alignment = RIGHT; cell.number_format = NUM_FMT
                    if v < 0: cell.font = NEG_FONT
                    else: cell.font = BOLD if is_balance else REG
                else:
                    cell.alignment = CENTER
                    cell.font = BOLD if is_balance else REG
            r += 1

        det_widths = [6, 11, 18, 10, 30, 10,
                      13, 13, 15, 11, 13,
                      13, 13, 15, 11, 13]
        for i, w in enumerate(det_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "G6"
        if r > 6:
            ws.auto_filter.ref = f"A5:P{r-1}"

    wb.save(out_path)
    return len(all_data)


# ════════════════════════════════════════════════════════════════
#  ECashL — Electronic Cash Ledger consolidator
# ════════════════════════════════════════════════════════════════
# Parses GSTN portal CSV export of "Electronic Cash Ledger". Cash Ledger
# has FINER granularity than Credit — every tax type (IGST / CGST / SGST /
# CESS) has 6 sub-cols (Tax / Interest / Penalty / Fee / Others / Total),
# giving 4×6 = 24 amount cols + 4×6 = 24 balance cols (48 numeric cols).
# Console shows Total per tax type for compact view; per-GSTIN sheet has
# the full breakdown.


def parse_ecashl_csv(filepath):
    """Parse one Electronic Cash Ledger CSV. Returns dict with meta + rows.

    Structure:
       R0:  Title 'Electronic Cash Ledger'
       R2-5: meta (label col 6, value col 7) — GSTIN, Name(Legal), From, To
       R6:  main header row
       R7:  sub-header row — Tax/Interest/Penalty/Fee/Others/Total × 8 groups
       R8+: data rows

    Data row columns:
       0: Sr.No
       1: Date of deposit/Debit
       2: Time of deposit
       3: Reporting date (by bank)
       4: Reference No.
       5: Tax Period
       6: Description
       7: Transaction Type ('Debit'/'Credit')
       8-13:  IGST amount  (Tax/Interest/Penalty/Fee/Others/Total)
       14-19: CGST amount
       20-25: SGST amount
       26-31: CESS amount
       32-37: IGST balance
       38-43: CGST balance
       44-49: SGST balance
       50-55: CESS balance
    """
    import csv
    raw_rows = []
    with open(filepath, encoding='utf-8-sig', errors='replace') as f:
        for row in csv.reader(f):
            raw_rows.append(row)

    meta = {
        "source_file": os.path.basename(filepath),
        "gstin": "", "legal_name": "",
        "from_date": "", "to_date": "",
        "state_code": "", "state_name": "",
    }
    # Cash Ledger meta is in column 6 (label) + col 7 (value); fall back to 3
    for r in raw_rows[:10]:
        for li in (6, 3):
            if li >= len(r): continue
            label = (r[li] or "").strip()
            if not label: continue
            value = ""
            for vi in (li + 1, li + 2):
                if vi < len(r) and r[vi] and r[vi].strip():
                    value = r[vi].strip(); break
            if not value: continue
            ll = label.lower()
            if label == "GSTIN" and not meta["gstin"]:
                meta["gstin"] = value
            elif ("legal" in ll or "name" in ll) and not meta["legal_name"]:
                meta["legal_name"] = value
            elif label == "From" and not meta["from_date"]:
                meta["from_date"] = value
            elif label == "To" and not meta["to_date"]:
                meta["to_date"] = value
    if meta["gstin"]:
        meta["state_code"] = meta["gstin"][:2]
        meta["state_name"] = _STATE_CODES.get(meta["state_code"], "Unknown")

    # Locate Sr.No header row
    hdr_idx = None
    for i, r in enumerate(raw_rows):
        if r and r[0] and str(r[0]).strip().startswith("Sr.No"):
            hdr_idx = i; break
    if hdr_idx is None:
        return {"meta": meta, "rows": []}
    data_start = hdr_idx + 2

    rows = []
    for r in raw_rows[data_start:]:
        if not r or not r[0] or not str(r[0]).strip():
            continue

        def col(i):
            return r[i].strip() if i < len(r) and r[i] else ""

        def block(start):
            return {
                "tax":      _ledger_parse_num(col(start + 0)),
                "interest": _ledger_parse_num(col(start + 1)),
                "penalty":  _ledger_parse_num(col(start + 2)),
                "fee":      _ledger_parse_num(col(start + 3)),
                "others":   _ledger_parse_num(col(start + 4)),
                "total":    _ledger_parse_num(col(start + 5)),
            }

        sr_no = col(0); date = col(1); time = col(2)
        rep_date = col(3); ref_no = col(4); tax_period = col(5)
        desc = col(6); ttype = col(7)

        amt_igst = block(8);  amt_cgst = block(14)
        amt_sgst = block(20); amt_cess = block(26)
        bal_igst = block(32); bal_cgst = block(38)
        bal_sgst = block(44); bal_cess = block(50)

        dlow = desc.lower()
        if "opening balance" in dlow: row_type = "Opening"
        elif "closing balance" in dlow: row_type = "Closing"
        elif ttype.lower() == "credit": row_type = "Credit"
        elif ttype.lower() == "debit":  row_type = "Debit"
        else: row_type = "Other"

        rows.append({
            "sr_no": sr_no, "date": date, "time": time,
            "rep_date": rep_date, "ref_no": ref_no,
            "tax_period": tax_period, "desc": desc, "ttype": ttype,
            "row_type": row_type,
            "amt_igst": amt_igst, "amt_cgst": amt_cgst,
            "amt_sgst": amt_sgst, "amt_cess": amt_cess,
            "bal_igst": bal_igst, "bal_cgst": bal_cgst,
            "bal_sgst": bal_sgst, "bal_cess": bal_cess,
        })

    return {"meta": meta, "rows": rows}


def write_consolidated_ecashl(all_data, out_path):
    """Build consolidated workbook for Electronic Cash Ledger.
    Returns count of GSTINs."""
    import openpyxl
    from openpyxl.styles import (Font as XF, PatternFill, Alignment,
                                  Border, Side)
    from openpyxl.utils import get_column_letter

    HDR_FILL = PatternFill("solid", start_color="1F4E78")
    SUB_FILL = PatternFill("solid", start_color="2E75B6")
    META_FILL = PatternFill("solid", start_color="F2F2F2")
    IGST_FILL = PatternFill("solid", start_color="DDEBF7")
    CGST_FILL = PatternFill("solid", start_color="E2EFDA")
    SGST_FILL = PatternFill("solid", start_color="FCE4D6")
    CESS_FILL = PatternFill("solid", start_color="FFF2CC")
    BAL_FILL = PatternFill("solid", start_color="F4B084")
    OPEN_FILL = PatternFill("solid", start_color="E7E6E6")
    CLOSE_FILL = PatternFill("solid", start_color="FFE699")
    TOT_FILL = PatternFill("solid", start_color="D9E1F2")

    WHITE_B = XF(name="Calibri", bold=True, color="FFFFFF", size=10)
    BOLD = XF(name="Calibri", bold=True, size=10)
    REG = XF(name="Calibri", size=10)
    NEG_FONT = XF(name="Calibri", size=10, color="C00000")
    TITLE = XF(name="Calibri", bold=True, color="FFFFFF", size=14)
    SUBTITLE = XF(name="Calibri", italic=True, size=10, color="404040")

    thin = Side(border_style="thin", color="B4B4B4")
    BORDER_ALL = Border(left=thin, right=thin, top=thin, bottom=thin)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
    RIGHT = Alignment(horizontal="right", vertical="center")
    NUM_FMT = '#,##0.00;[Red](#,##0.00);"-"'

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ─── Cover ────────────────────────────────────────────
    ws = wb.create_sheet("Cover")
    ws.merge_cells("A1:H1")
    ws["A1"] = "Electronic Cash Ledger — Multi-GSTIN Consolidated"
    ws["A1"].font = TITLE; ws["A1"].fill = HDR_FILL
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A2:H2")
    ws["A2"] = ("Cash deposit / debit / closing balance per GSTIN  ·  "
                "Tax / Interest / Penalty / Fee / Others breakdown")
    ws["A2"].font = SUBTITLE; ws["A2"].alignment = CENTER

    hdrs = ["S.No.", "GSTIN", "State Code", "State Name", "Legal Name",
            "From", "To", "Source File"]
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = WHITE_B; cell.fill = SUB_FILL
        cell.alignment = CENTER; cell.border = BORDER_ALL
    ws.row_dimensions[4].height = 22

    for i, data in enumerate(all_data, 1):
        m = data["meta"]
        for c, v in enumerate([i, m["gstin"], m["state_code"], m["state_name"],
                                m["legal_name"], m["from_date"], m["to_date"],
                                m["source_file"]], 1):
            cell = ws.cell(row=4 + i, column=c, value=v)
            cell.font = REG; cell.border = BORDER_ALL
            cell.alignment = LEFT if c >= 4 else CENTER
            if c == 1: cell.number_format = "0"
    for col, w in zip("ABCDEFGH", [6, 18, 7, 18, 28, 12, 12, 42]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A5"

    # ─── Monthly Balances — per GSTIN × per Tax Period (Cash) ──
    # Cash Ledger: shows Totals per tax type. Full Tax/Interest/Penalty/Fee/
    # Others breakdown remains in per-GSTIN detail sheets. Same tax-period
    # logic as Credit Ledger — opening = balance just before first txn of
    # that tax period.
    ws = wb.create_sheet("Monthly Balances")
    cols = ["S.No.", "State Code", "State Name", "GSTIN", "Legal Name",
            "Tax Period", "FY Year", "Txn Count",
            "Opening IGST", "Opening CGST", "Opening SGST/UTGST",
            "Opening CESS", "Opening Total",
            "Closing IGST", "Closing CGST", "Closing SGST/UTGST",
            "Closing CESS", "Closing Total",
            "Net Change"]
    ncols_m = len(cols)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols_m)
    ws["A1"] = ("Monthly Balances — Cash Ledger Opening + Closing per GSTIN, "
                "per Tax Period   (opening = balance just before first txn of "
                "that tax period)")
    ws["A1"].font = TITLE; ws["A1"].fill = HDR_FILL
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 28

    def banner(c1, c2, label, fill):
        ws.merge_cells(start_row=2, start_column=c1, end_row=2, end_column=c2)
        cell = ws.cell(row=2, column=c1, value=label)
        cell.font = BOLD; cell.fill = fill
        cell.alignment = CENTER; cell.border = BORDER_ALL
    banner(1, 8, "Metadata", META_FILL)
    banner(9, 13, "Opening Balance (before first txn of tax period)", CGST_FILL)
    banner(14, 18, "Closing Balance (after last txn of tax period)", BAL_FILL)
    banner(19, 19, "Δ", META_FILL)
    ws.row_dimensions[2].height = 22

    for c, h in enumerate(cols, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = WHITE_B; cell.fill = SUB_FILL
        cell.alignment = CENTER; cell.border = BORDER_ALL
    ws.row_dimensions[3].height = 36

    GAIN_FILL = PatternFill("solid", start_color="C6EFCE")
    LOSS_FILL = PatternFill("solid", start_color="FFC7CE")

    r = 4; sn = 0
    for data in all_data:
        m = data["meta"]
        monthly_list = compute_monthly_balances(data, ledger_type='cash')
        for mb in monthly_list:
            sn += 1
            opg, clo = mb["opening"], mb["closing"]
            net_change = (clo["total"] or 0) - (opg["total"] or 0)
            vals = [
                sn, m["state_code"], m["state_name"], m["gstin"], m["legal_name"],
                mb["month"], mb["year"], mb["txn_count"],
                opg["igst"], opg["cgst"], opg["sgst"], opg["cess"], opg["total"],
                clo["igst"], clo["cgst"], clo["sgst"], clo["cess"], clo["total"],
                net_change,
            ]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.font = REG; cell.border = BORDER_ALL
                if c == 1:
                    cell.alignment = CENTER; cell.number_format = "0"
                elif c in (2, 4, 6, 7, 8):
                    cell.alignment = CENTER
                elif c in (3, 5):
                    cell.alignment = LEFT
                elif isinstance(v, (int, float)) and c >= 9:
                    cell.alignment = RIGHT; cell.number_format = NUM_FMT
                    if v < 0: cell.font = NEG_FONT
                if c == 19 and isinstance(v, (int, float)) and v != 0:
                    cell.fill = GAIN_FILL if v > 0 else LOSS_FILL
                    cell.font = BOLD if v > 0 else NEG_FONT
            r += 1

    widths_m = [6, 7, 16, 18, 22,
                10, 7, 8,
                13, 13, 15, 11, 13,
                13, 13, 15, 11, 13,
                14]
    for i, w in enumerate(widths_m, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "I4"
    if r > 4:
        ws.auto_filter.ref = f"A3:{get_column_letter(ncols_m)}{r-1}"

    # ─── Extra Payments sheet (Cash — Voluntary / DRC / Refund / etc.) ──
    # Same logic as Credit Ledger Extra Payments — Txn Month is calendar
    # month of payment, Tax Period is period it relates to, Effect TP is
    # the tax period whose 3B comparison will show a gap.
    ws = wb.create_sheet("Extra Payments")
    cols_e = ["S.No.", "State Code", "State Name", "GSTIN", "Legal Name",
              "Date", "Txn Month", "Reference No.", "Tax Period", "Effect TP",
              "Subcategory", "Description", "Txn Type",
              "Amt IGST", "Amt CGST", "Amt SGST/UTGST", "Amt CESS", "Amt Total",
              "Bal Total After"]
    ncols_e = len(cols_e)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols_e)
    ws["A1"] = ("Extra Payments — Voluntary / Refund / DRC / Demand / Recovery   "
                "(Cash Ledger; Effect TP = tax period whose 3B comparison "
                "will show a gap from this payment)")
    ws["A1"].font = TITLE; ws["A1"].fill = HDR_FILL
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 28

    SUBCAT_FILL = PatternFill("solid", start_color="FFE6CC")
    TXN_MONTH_FILL = PatternFill("solid", start_color="D9E1F2")
    EFFECT_TP_FILL = PatternFill("solid", start_color="C6EFCE")
    for c, h in enumerate(cols_e, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = WHITE_B; cell.fill = SUB_FILL
        cell.alignment = CENTER; cell.border = BORDER_ALL
    ws.row_dimensions[3].height = 30

    r = 4; sn = 0
    subcategory_totals = {}
    txn_month_totals = {}
    effect_tp_totals = {}
    for data in all_data:
        m = data["meta"]
        extras = compute_extra_payments(data, ledger_type='cash')
        for ex in extras:
            sn += 1
            sub = ex["subcategory"]
            txm = ex["txn_month"] or "(unknown)"
            etp = ex["effect_tax_period"] or "(unknown)"
            subcategory_totals.setdefault(sub, {"count": 0, "total": 0})
            subcategory_totals[sub]["count"] += 1
            subcategory_totals[sub]["total"] += ex["amt"]["total"] or 0
            txn_month_totals.setdefault(txm, {"count": 0, "total": 0})
            txn_month_totals[txm]["count"] += 1
            txn_month_totals[txm]["total"] += ex["amt"]["total"] or 0
            effect_tp_totals.setdefault(etp, {"count": 0, "total": 0})
            effect_tp_totals[etp]["count"] += 1
            effect_tp_totals[etp]["total"] += ex["amt"]["total"] or 0
            vals = [
                sn, m["state_code"], m["state_name"], m["gstin"], m["legal_name"],
                ex["date"], ex["txn_month"], ex["ref_no"], ex["tax_period"],
                ex["effect_tax_period"], sub,
                ex["desc"], ex["ttype"],
                ex["amt"]["igst"], ex["amt"]["cgst"],
                ex["amt"]["sgst"], ex["amt"]["cess"], ex["amt"]["total"],
                ex["bal_total_after"],
            ]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.font = REG; cell.border = BORDER_ALL
                if c == 1:
                    cell.alignment = CENTER; cell.number_format = "0"
                elif c in (2, 4, 6, 8, 9, 13):
                    cell.alignment = CENTER
                elif c == 7:
                    cell.alignment = CENTER
                    cell.fill = TXN_MONTH_FILL
                    cell.font = BOLD
                elif c == 10:
                    cell.alignment = CENTER
                    cell.fill = EFFECT_TP_FILL
                    cell.font = BOLD
                elif c == 11:
                    cell.alignment = CENTER
                    cell.fill = SUBCAT_FILL
                    cell.font = BOLD
                elif c in (3, 5, 12):
                    cell.alignment = LEFT
                elif isinstance(v, (int, float)) and c >= 14:
                    cell.alignment = RIGHT; cell.number_format = NUM_FMT
                    if v < 0: cell.font = NEG_FONT
                else:
                    cell.alignment = LEFT
            r += 1

    if subcategory_totals or txn_month_totals or effect_tp_totals:
        r += 1
        ws.cell(row=r, column=1, value="SUMMARY BY SUBCATEGORY").font = BOLD
        ws.cell(row=r, column=1).alignment = LEFT
        ws.cell(row=r, column=6, value="SUMMARY BY EFFECT TP (3B gap)").font = BOLD
        ws.cell(row=r, column=6).alignment = LEFT
        ws.cell(row=r, column=11, value="SUMMARY BY TXN MONTH").font = BOLD
        ws.cell(row=r, column=11).alignment = LEFT
        r += 1
        for c, h in enumerate(["Subcategory", "# Events", "Total Amount"], 1):
            cell = ws.cell(row=r, column=c, value=h)
            cell.font = WHITE_B; cell.fill = SUB_FILL
            cell.alignment = CENTER; cell.border = BORDER_ALL
        for c, h in enumerate(["Effect TP", "# Events", "Total Amount"], 6):
            cell = ws.cell(row=r, column=c, value=h)
            cell.font = WHITE_B; cell.fill = SUB_FILL
            cell.alignment = CENTER; cell.border = BORDER_ALL
        for c, h in enumerate(["Txn Month", "# Events", "Total Amount"], 11):
            cell = ws.cell(row=r, column=c, value=h)
            cell.font = WHITE_B; cell.fill = SUB_FILL
            cell.alignment = CENTER; cell.border = BORDER_ALL
        r += 1
        subcat_sorted = sorted(subcategory_totals.items())
        etp_sorted = sorted(effect_tp_totals.items(),
                            key=lambda x: _tp_sort_key(x[0]))
        txm_sorted = sorted(txn_month_totals.items(),
                            key=lambda x: _tp_sort_key(x[0]))
        max_rows = max(len(subcat_sorted), len(etp_sorted), len(txm_sorted))
        for i in range(max_rows):
            if i < len(subcat_sorted):
                k, vs = subcat_sorted[i]
                ws.cell(row=r, column=1, value=k).font = BOLD
                ws.cell(row=r, column=1).fill = SUBCAT_FILL
                ws.cell(row=r, column=1).border = BORDER_ALL
                ws.cell(row=r, column=1).alignment = CENTER
                ws.cell(row=r, column=2, value=vs["count"]).alignment = CENTER
                ws.cell(row=r, column=2).border = BORDER_ALL
                ws.cell(row=r, column=3, value=vs["total"]).number_format = NUM_FMT
                ws.cell(row=r, column=3).alignment = RIGHT
                ws.cell(row=r, column=3).border = BORDER_ALL
            if i < len(etp_sorted):
                k, vs = etp_sorted[i]
                ws.cell(row=r, column=6, value=k).font = BOLD
                ws.cell(row=r, column=6).fill = EFFECT_TP_FILL
                ws.cell(row=r, column=6).border = BORDER_ALL
                ws.cell(row=r, column=6).alignment = CENTER
                ws.cell(row=r, column=7, value=vs["count"]).alignment = CENTER
                ws.cell(row=r, column=7).border = BORDER_ALL
                ws.cell(row=r, column=8, value=vs["total"]).number_format = NUM_FMT
                ws.cell(row=r, column=8).alignment = RIGHT
                ws.cell(row=r, column=8).border = BORDER_ALL
            if i < len(txm_sorted):
                k, vs = txm_sorted[i]
                ws.cell(row=r, column=11, value=k).font = BOLD
                ws.cell(row=r, column=11).fill = TXN_MONTH_FILL
                ws.cell(row=r, column=11).border = BORDER_ALL
                ws.cell(row=r, column=11).alignment = CENTER
                ws.cell(row=r, column=12, value=vs["count"]).alignment = CENTER
                ws.cell(row=r, column=12).border = BORDER_ALL
                ws.cell(row=r, column=13, value=vs["total"]).number_format = NUM_FMT
                ws.cell(row=r, column=13).alignment = RIGHT
                ws.cell(row=r, column=13).border = BORDER_ALL
            r += 1

    widths_e = [6, 7, 16, 18, 22,
                11, 10, 18, 10, 10, 18, 30, 8,
                13, 13, 15, 11, 13, 16]
    for i, w in enumerate(widths_e, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "I4"
    if sn > 0:
        ws.auto_filter.ref = f"A3:{get_column_letter(ncols_e)}{3 + sn}"

    # ─── Console (Totals view) ────────────────────────────
    ws = wb.create_sheet("Console")
    cols = ["S.No.", "GSTIN", "State Code", "State Name", "Legal Name",
            "Row Type", "Sr.No.", "Date", "Reference No.", "Tax Period",
            "Description", "Transaction Type",
            "IGST Amt Total", "CGST Amt Total",
            "SGST Amt Total", "CESS Amt Total",
            "IGST Bal Total", "CGST Bal Total",
            "SGST Bal Total", "CESS Bal Total"]
    ncols = len(cols)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws["A1"] = ("Console — every row, every GSTIN  "
                "(Totals shown; full Tax/Interest/Penalty/Fee/Others "
                "breakdown in per-GSTIN detail sheets)")
    ws["A1"].font = TITLE; ws["A1"].fill = HDR_FILL
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 28

    def banner(c1, c2, label, fill):
        ws.merge_cells(start_row=2, start_column=c1, end_row=2, end_column=c2)
        cell = ws.cell(row=2, column=c1, value=label)
        cell.font = BOLD; cell.fill = fill
        cell.alignment = CENTER; cell.border = BORDER_ALL
    banner(1, 12, "Metadata", META_FILL)
    banner(13, 16, "Cash Amount (Credit/Debit) — Totals", CGST_FILL)
    banner(17, 20, "Cash Balance after transaction — Totals", BAL_FILL)
    ws.row_dimensions[2].height = 22

    for c, h in enumerate(cols, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = WHITE_B; cell.fill = SUB_FILL
        cell.alignment = CENTER; cell.border = BORDER_ALL
    ws.row_dimensions[3].height = 36

    r = 4; sn = 0
    for data in all_data:
        m = data["meta"]
        for row in data["rows"]:
            sn += 1
            vals = [
                sn, m["gstin"], m["state_code"], m["state_name"],
                m["legal_name"],
                row["row_type"], row["sr_no"], row["date"], row["ref_no"],
                row["tax_period"], row["desc"], row["ttype"],
                row["amt_igst"]["total"], row["amt_cgst"]["total"],
                row["amt_sgst"]["total"], row["amt_cess"]["total"],
                row["bal_igst"]["total"], row["bal_cgst"]["total"],
                row["bal_sgst"]["total"], row["bal_cess"]["total"],
            ]
            row_bold = row["row_type"] in ("Opening", "Closing")
            row_fill = (OPEN_FILL if row["row_type"] == "Opening"
                        else CLOSE_FILL if row["row_type"] == "Closing"
                        else None)
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.border = BORDER_ALL
                if row_fill: cell.fill = row_fill
                if c == 1:
                    cell.alignment = CENTER; cell.number_format = "0"
                    cell.font = BOLD if row_bold else REG
                elif c in (2, 3, 6, 7, 8, 10, 12):
                    cell.alignment = CENTER
                    cell.font = BOLD if row_bold else REG
                elif c in (4, 5, 9, 11):
                    cell.alignment = LEFT
                    cell.font = BOLD if row_bold else REG
                elif isinstance(v, (int, float)) and c >= 13:
                    cell.alignment = RIGHT; cell.number_format = NUM_FMT
                    if v < 0: cell.font = NEG_FONT
                    else: cell.font = BOLD if row_bold else REG
                else:
                    cell.alignment = LEFT
                    cell.font = BOLD if row_bold else REG
            r += 1

    widths = [6, 18, 6, 16, 22,
              9, 6, 11, 18, 10, 28, 11,
              14, 14, 14, 13,
              14, 14, 14, 13]
    for i, w in enumerate(widths[:ncols], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "M4"
    if r > 4:
        ws.auto_filter.ref = f"A3:{get_column_letter(ncols)}{r-1}"

    # ─── Summary (per-GSTIN totals) ───────────────────────
    ws = wb.create_sheet("Summary")
    ws.merge_cells("A1:N1")
    ws["A1"] = "Summary — Cash Credit / Debit / Closing Balance per GSTIN"
    ws["A1"].font = TITLE; ws["A1"].fill = HDR_FILL
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 28

    sum_hdrs = ["S.No.", "GSTIN", "State", "Legal Name", "From", "To",
                "Txn Count",
                "Cr IGST", "Cr CGST+SGST", "Cr CESS",
                "Dr IGST", "Dr CGST+SGST", "Dr CESS",
                "Closing (4-head total)"]
    for c, h in enumerate(sum_hdrs, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = WHITE_B; cell.fill = SUB_FILL
        cell.alignment = CENTER; cell.border = BORDER_ALL
    ws.row_dimensions[3].height = 36

    r = 4
    grand_totals = [0] * 7
    for i, data in enumerate(all_data, 1):
        m = data["meta"]
        txns = [row for row in data["rows"]
                if row["row_type"] in ("Credit", "Debit")]
        closing = next((row for row in data["rows"]
                        if row["row_type"] == "Closing"), None)

        def sum_total(rows, key):
            return sum((row[key]["total"] or 0) for row in rows)

        credits = [r2 for r2 in txns if r2["row_type"] == "Credit"]
        debits  = [r2 for r2 in txns if r2["row_type"] == "Debit"]
        cr_i = sum_total(credits, "amt_igst")
        cr_c = sum_total(credits, "amt_cgst") + sum_total(credits, "amt_sgst")
        cr_x = sum_total(credits, "amt_cess")
        dr_i = sum_total(debits, "amt_igst")
        dr_c = sum_total(debits, "amt_cgst") + sum_total(debits, "amt_sgst")
        dr_x = sum_total(debits, "amt_cess")
        clo_total = 0
        if closing:
            clo_total = ((closing["bal_igst"]["total"] or 0)
                         + (closing["bal_cgst"]["total"] or 0)
                         + (closing["bal_sgst"]["total"] or 0)
                         + (closing["bal_cess"]["total"] or 0))

        for k, val in enumerate([cr_i, cr_c, cr_x, dr_i, dr_c, dr_x, clo_total]):
            grand_totals[k] += val or 0

        vals = [i, m["gstin"], m["state_name"], m["legal_name"],
                m["from_date"], m["to_date"], len(txns),
                cr_i, cr_c, cr_x, dr_i, dr_c, dr_x, clo_total]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = REG; cell.border = BORDER_ALL
            if c == 1:
                cell.alignment = CENTER; cell.number_format = "0"
            elif c in (2, 3, 5, 6, 7):
                cell.alignment = CENTER
            elif c == 4:
                cell.alignment = LEFT
            elif c >= 8 and isinstance(v, (int, float)):
                cell.alignment = RIGHT; cell.number_format = NUM_FMT
                if v < 0: cell.font = NEG_FONT
            else:
                cell.alignment = LEFT
        r += 1

    if all_data:
        cell = ws.cell(row=r, column=1, value="TOTAL")
        cell.font = BOLD; cell.fill = TOT_FILL
        cell.alignment = CENTER; cell.border = BORDER_ALL
        for c in (2, 3, 4, 5, 6, 7):
            ws.cell(row=r, column=c).fill = TOT_FILL
            ws.cell(row=r, column=c).border = BORDER_ALL
        for k, val in enumerate(grand_totals):
            cell = ws.cell(row=r, column=8 + k, value=val)
            cell.font = BOLD; cell.fill = TOT_FILL
            cell.alignment = RIGHT
            cell.border = BORDER_ALL; cell.number_format = NUM_FMT
            if val < 0:
                cell.font = XF(name="Calibri", bold=True, color="C00000", size=10)
        r += 1

    widths_sum = [6, 18, 16, 26, 12, 12, 8,
                  14, 16, 11, 14, 16, 11, 18]
    for i, w in enumerate(widths_sum, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"

    # ─── Per-GSTIN detail sheets (full 48-col layout) ─────
    SUBS = ["Tax", "Interest", "Penalty", "Fee", "Others", "Total"]
    GROUPS = [
        ("IGST",     "amt_igst", IGST_FILL),
        ("CGST",     "amt_cgst", CGST_FILL),
        ("SGST",     "amt_sgst", SGST_FILL),
        ("CESS",     "amt_cess", CESS_FILL),
        ("IGST Bal", "bal_igst", BAL_FILL),
        ("CGST Bal", "bal_cgst", BAL_FILL),
        ("SGST Bal", "bal_sgst", BAL_FILL),
        ("CESS Bal", "bal_cess", BAL_FILL),
    ]
    for data in all_data:
        m = data["meta"]
        sname = (f"{m['state_code']}_{m['gstin'][-4:]}"
                 if m['gstin'] else f"G_{len(wb.sheetnames)}")
        sname = sname[:31]
        base = sname; idx = 1
        while sname in wb.sheetnames:
            idx += 1; sname = f"{base[:28]}_{idx}"
        ws = wb.create_sheet(sname)

        total_cols = 8 + 8 * 6  # 8 meta + 48 numeric = 56
        last_col = get_column_letter(total_cols)
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=total_cols)
        ws["A1"] = f"Cash Ledger — {m['gstin']}  ({m['state_name']})"
        ws["A1"].font = TITLE; ws["A1"].fill = HDR_FILL
        ws["A1"].alignment = CENTER
        ws.row_dimensions[1].height = 26
        ws.merge_cells(start_row=2, start_column=1,
                       end_row=2, end_column=total_cols)
        ws["A2"] = (f"{m['legal_name']}  |  Period: {m['from_date']} to "
                    f"{m['to_date']}")
        ws["A2"].font = SUBTITLE; ws["A2"].alignment = CENTER
        ws.row_dimensions[2].height = 18

        meta_h = ["S.No.", "Date", "Time", "Ref. Date", "Reference No.",
                  "Tax Period", "Description", "Transaction Type"]
        for c, h in enumerate(meta_h, 1):
            ws.merge_cells(start_row=4, start_column=c, end_row=5, end_column=c)
            cell = ws.cell(row=4, column=c, value=h)
            cell.font = WHITE_B; cell.fill = SUB_FILL
            cell.alignment = CENTER; cell.border = BORDER_ALL

        for gi, (label, _, fill) in enumerate(GROUPS):
            cs = 9 + gi * 6
            ce = cs + 5
            ws.merge_cells(start_row=4, start_column=cs,
                           end_row=4, end_column=ce)
            cell = ws.cell(row=4, column=cs, value=label)
            cell.font = BOLD; cell.fill = fill
            cell.alignment = CENTER; cell.border = BORDER_ALL
            for si, sub in enumerate(SUBS):
                col = cs + si
                cell = ws.cell(row=5, column=col, value=sub)
                cell.font = WHITE_B; cell.fill = SUB_FILL
                cell.alignment = CENTER; cell.border = BORDER_ALL
        ws.row_dimensions[4].height = 22
        ws.row_dimensions[5].height = 22

        r = 6
        for row in data["rows"]:
            is_balance = row["row_type"] in ("Opening", "Closing")
            row_fill = (OPEN_FILL if row["row_type"] == "Opening"
                        else CLOSE_FILL if row["row_type"] == "Closing"
                        else None)
            base_vals = [row["sr_no"], row["date"], row["time"],
                         row["rep_date"], row["ref_no"], row["tax_period"],
                         row["desc"], row["ttype"]]
            for c, v in enumerate(base_vals, 1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.border = BORDER_ALL
                if row_fill: cell.fill = row_fill
                cell.alignment = CENTER if c != 7 else LEFT
                cell.font = BOLD if is_balance else REG

            for gi, (_, key, _) in enumerate(GROUPS):
                blk = row[key]
                for si, sub in enumerate(SUBS):
                    v = blk[sub.lower()]
                    col = 9 + gi * 6 + si
                    cell = ws.cell(row=r, column=col, value=v)
                    cell.border = BORDER_ALL
                    if row_fill: cell.fill = row_fill
                    if isinstance(v, (int, float)):
                        cell.alignment = RIGHT; cell.number_format = NUM_FMT
                        if v < 0: cell.font = NEG_FONT
                        else: cell.font = BOLD if is_balance else REG
                    else:
                        cell.alignment = CENTER
                        cell.font = BOLD if is_balance else REG
            r += 1

        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 11
        ws.column_dimensions["C"].width = 9
        ws.column_dimensions["D"].width = 11
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 10
        ws.column_dimensions["G"].width = 26
        ws.column_dimensions["H"].width = 11
        for c in range(9, 9 + 8 * 6):
            ws.column_dimensions[get_column_letter(c)].width = 11
        ws.freeze_panes = "I6"
        if r > 6:
            ws.auto_filter.ref = f"A5:{last_col}{r-1}"

    wb.save(out_path)
    return len(all_data)


# ════════════════════════════════════════════════════════════════
